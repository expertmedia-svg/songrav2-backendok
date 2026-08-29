"""
SONGRA - Backend API avec Computer Vision LOCALE
Version FINALE - Avec analyse IA complète
"""

from fastapi import FastAPI, HTTPException, Depends, Header, Query, UploadFile, File, Form, BackgroundTasks, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, HTMLResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime, timedelta
import os
import jwt
from sqlalchemy import create_engine, Column, Integer, String, Float, Boolean, DateTime, Text, func, or_
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
import hashlib
import unicodedata
import json
import re
import base64
import mimetypes
import math
from io import BytesIO
from PIL import Image
import numpy as np
import time
import shutil
import zipfile
import csv
import secrets
import bcrypt
import urllib.parse
import urllib.request
import urllib.error
from dotenv import load_dotenv

# Charger d'abord backend/.env, quel que soit le dossier depuis lequel Uvicorn
# est lancé. Sans cela, un démarrage depuis la racine pouvait ignorer Groq.
_MODULE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(_MODULE_DIR, ".env"), override=False)
load_dotenv(override=False)

from openai import OpenAI
try:
    import google.generativeai as genai
except ImportError:
    genai = None
try:
    from gemini_vision import GeminiVisionEngine
except ImportError:
    GeminiVisionEngine = None
import v2_services
import agri_services
import yingr_ai_api
from burkina_translator import (
    dictionary_stats,
    import_dictionary_file,
    translate_fields,
    translate_fields_and_voice_summary,
    SONGRA_TEXT_FIELDS as _TRANSLATOR_TEXT_FIELDS,
    VALID_LANGS as _TRANSLATOR_VALID_LANGS,
    LANG_NAMES as _TRANSLATOR_LANG_NAMES,
)


os.makedirs("uploads", exist_ok=True)
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "resolvehub.db")
    DATABASE_URL = f"sqlite:///{db_path}"

AI_PROVIDER = os.getenv("AI_PROVIDER", "openai").lower()

engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
groq_client = OpenAI(
    api_key=GROQ_API_KEY,
    base_url="https://api.groq.com/openai/v1",
) if GROQ_API_KEY else None

# Gemini API Key pour analyse photo
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY and genai is not None:
    genai.configure(api_key=GEMINI_API_KEY)
    print(f"[OK] Gemini API configuree")
else:
    print("[WARN] Gemini indisponible - Songra utilisera Groq/OpenAI et les replis locaux")

APP_ENV = os.getenv("APP_ENV", "development").strip().lower()
JWT_SECRET = os.getenv("JWT_SECRET", "").strip()
if not JWT_SECRET:
    if APP_ENV == "production":
        raise RuntimeError("JWT_SECRET doit etre configure en production")
    JWT_SECRET = "songra-mobile-dev-secret-change-me"
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_HOURS = int(os.getenv("JWT_EXPIRE_HOURS", "72"))

BACKEND_DIR = os.path.dirname(__file__)
EXPERT_LOCAL_KNOWLEDGE_SOURCE = "expert_local_knowledge"
EXPERT_AUDIO_MAP_PATH = os.getenv(
    "EXPERT_AUDIO_MAP_PATH",
    os.path.join(BACKEND_DIR, "audio_map.json"),
)
EXPERT_AUDIO_MAP_LEGACY_SEED = os.path.abspath(
    os.path.join(BACKEND_DIR, "..", "backend-node", "data", "audioMap.json")
)
EXPERT_LOCAL_KNOWLEDGE_LEGACY_SEED = os.path.abspath(
    os.path.join(BACKEND_DIR, "..", "backend-node", "data", "localKnowledgeBase.json")
)
EXPERT_AUDIO_UPLOAD_DIR = os.path.join("uploads", "audio")
COMMUNITY_AUDIO_UPLOAD_DIR = os.path.join("uploads", "community_audio")

os.makedirs(EXPERT_AUDIO_UPLOAD_DIR, exist_ok=True)
os.makedirs(COMMUNITY_AUDIO_UPLOAD_DIR, exist_ok=True)

# ==========================================
# DÉTECTION D'URGENCE - SOS/ACCIDENTS CRITIQUES
# ==========================================

def detect_emergency(text: str, category: Optional[str] = None) -> Dict[str, Any]:
    """Détecte automatiquement les situations d'urgence/critiques.
    
    Retourne:
    {
        "is_emergency": True/False,
        "severity": "critical" | "high" | "medium" | "normal",
        "emergency_type": "accident" | "saignement" | "inconscience" | "empoisonnement" | etc.,
        "protocol": "protocole de premiers secours spécifique",
        "immediate_actions": ["action 1", "action 2", ...],
        "call_emergency_number": "15 ou 17 ou 112 (FR: SAMU/Police/Ambulance)",
        "warning_alert": "Alerte à afficher à l'utilisateur"
    }
    """
    text_lower = text.lower().strip()
    
    # Mots-clés CRITIQUES - VIE OU MORT
    critical_keywords = {
        # Accidents
        "accident": "accident",
        "chute": "accident", 
        "collision": "accident",
        "fracture": "accident",
        "coup": "accident",
        "blessure grave": "accident",
        
        # Saignements graves
        "saigne": "saignement",
        "saignement": "saignement",
        "hémorragie": "saignement",
        "saigne de la tete": "saignement_tete",
        "saigne de la face": "saignement_tete",
        
        # État de conscience
        "inconscient": "inconscience",
        "perd connaissance": "inconscience",
        "évanoui": "inconscience",
        "coma": "inconscience",
        
        # Respiratoire
        "asphyxie": "respiratoire",
        "étouffement": "respiratoire",
        "ne respire pas": "respiratoire",
        "difficulté respirer": "respiratoire",
        "étranglement": "respiratoire",
        
        # Empoisonnement/Ingestion
        "poison": "empoisonnement",
        "empoisonnement": "empoisonnement",
        "intoxication": "empoisonnement",
        "avalé": "empoisonnement",
        "intoxiqué": "empoisonnement",
        
        # Brûlures graves
        "brûlure grave": "brulure",
        "brûlé": "brulure",
        "embrasé": "brulure",
        
        # Électrocution
        "électrocuté": "electrocution",
        "électrochoc": "electrocution",
        "courant électrique": "electrocution",
        
        # Choc/Arrêt cardiaque
        "arrêt cardiaque": "arrêt_cardiaque",
        "crise cardiaque": "arrêt_cardiaque",
        "infarctus": "arrêt_cardiaque",
        "choc": "choc",
        
        # Convulsions
        "convulsion": "convulsion",
        "crise": "convulsion",
        "tremblements": "convulsion",
    }
    
    # Vérifier les mots-clés
    detected_type = None
    for keyword, emergency_type in critical_keywords.items():
        if keyword in text_lower:
            detected_type = emergency_type
            break
    
    if not detected_type:
        # Pas d'urgence détectée
        return {
            "is_emergency": False,
            "severity": "normal",
            "emergency_type": None,
            "protocol": None,
            "immediate_actions": [],
            "call_emergency_number": None,
            "warning_alert": None
        }
    
    # PROTOCOLES DE PREMIERS SECOURS
    protocols = {
        "accident": {
            "severity": "critical",
            "call": "112 (SAMU/Ambulance)",
            "immediate_actions": [
                "🚨 APPELEZ IMMÉDIATEMENT le 112 ou l'ambulance locale",
                "1️⃣ Placez la personne en position stable sur le côté (si inconsciente mais respire)",
                "2️⃣ Vérifiez la respiration et le pouls",
                "3️⃣ Maintenez la personne au chaud avec vêtements/couvertures",
                "4️⃣ NE BOUGEZ PAS la colonne vertébrale (risque de paralysie)",
                "5️⃣ Contrôlez les saignements visibles avec compresses propres",
                "6️⃣ Restez avec la personne jusqu'à l'arrivée des secours"
            ],
            "warning": "🚨 SITUATIONS D'URGENCE CRITIQUE - APPELEZ IMMÉDIATEMENT LES SECOURS (112)"
        },
        
        "saignement": {
            "severity": "high",
            "call": "112 (SAMU)",
            "immediate_actions": [
                "🚨 SAIGNEMENT GRAVE - Appelez le 112 immédiatement",
                "1️⃣ Écartez la personne du danger (si possible)",
                "2️⃣ Nettoyez avec EAU PROPRE si possible",
                "3️⃣ Appuyez FERMEMENT avec gaze/tissu propre PENDANT 10-15 minutes",
                "4️⃣ NE RETIREZ PAS la gaze, ajoutez-la par-dessus si saignement continue",
                "5️⃣ Surélevez la zone blessée au-dessus du cœur si possible",
                "6️⃣ Maintenez la pression jusqu'à l'arrivée des secours"
            ],
            "warning": "🚨 SAIGNEMENT GRAVE - APPELEZ LES SECOURS (112) ET APPUYEZ FERMEMENT"
        },
        
        "saignement_tete": {
            "severity": "critical",
            "call": "112 (SAMU)",
            "immediate_actions": [
                "🚨 BLESSURE À LA TÊTE AVEC SAIGNEMENT - Appelez le 112 IMMÉDIATEMENT",
                "1️⃣ Installez la personne en position semi-assise (45°) pour éviter étouffement",
                "2️⃣ Appuyez DOUCEMENT sur la plaie avec gaze propre (attention: ne pas appuyer trop si fracture du crâne)",
                "3️⃣ Si perte de conscience: mettez en position latérale stable",
                "4️⃣ Nettoyez doucement avec eau propre SANS APPUYER",
                "5️⃣ Surveillez la conscience et la respiration - le saignement peut continuer longtemps",
                "6️⃣ ATTENTION: Si sommeil anormal, vomissements ou troubles vision → très grave, ambulance URGENTE"
            ],
            "warning": "🚨 BLESSURE À LA TÊTE CRITIQUE - APPELEZ LE 112 URGENTEMENT"
        },
        
        "inconscience": {
            "severity": "critical",
            "call": "112 (SAMU/Ambulance)",
            "immediate_actions": [
                "🚨 PERTE DE CONSCIENCE - Appelez le 112 IMMÉDIATEMENT",
                "1️⃣ Vérifiez si la personne RESPIRE (regardez le thorax, écoutez)",
                "2️⃣ SI RESPIRE: Position latérale stable (couché sur le côté - tête vers l'arrière)",
                "3️⃣ SI NE RESPIRE PAS: Commencez le massage cardiaque",
                "   - Appuyer au centre du thorax, 100-120 compressions/minute",
                "   - Alterner: 30 compressions, 2 respirations bouche-à-bouche",
                "4️⃣ Continuez jusqu'à l'arrivée des secours",
                "5️⃣ Réchauffez avec couvertures"
            ],
            "warning": "🚨 PERSONNE INCONSCIENTE - APPELEZ LE 112 ET METTEZ EN POSITION LATÉRALE STABLE"
        },
        
        "respiratoire": {
            "severity": "critical",
            "call": "112 (Ambulance)",
            "immediate_actions": [
                "🚨 PROBLÈME RESPIRATOIRE GRAVE - Appelez le 112 MAINTENANT",
                "1️⃣ Desserrez les vêtements autour du cou/poitrine",
                "2️⃣ Mettez la personne EN POSITION ASSISE (penché vers l'avant)",
                "3️⃣ Ouvrez grand la bouche pour vérifier obstruction (débris, aliment)",
                "4️⃣ Si objet visible: enlevez DOUCEMENT avec doigt",
                "5️⃣ Apaisez la personne - respiration calme aide",
                "6️⃣ Si perte de conscience + pas de respiration: faire massage cardiaque"
            ],
            "warning": "🚨 ASPHYXIE/ÉTOUFFEMENT - APPELEZ LE 112 ET LIBÉREZ LES VOIES RESPIRATOIRES"
        },
        
        "empoisonnement": {
            "severity": "critical",
            "call": "112 ou Centre antipoison",
            "immediate_actions": [
                "🚨 EMPOISONNEMENT/INTOXICATION - Appelez le 112 ou 15",
                "1️⃣ Notez EXACTEMENT ce qui a été ingéré (nom, quantité)",
                "2️⃣ NE FAITES PAS VOMIR sauf si indicateur sur l'étiquette",
                "3️⃣ Placez en position assise si conscient, latérale si inconscient",
                "4️⃣ Si conscience perdue: vérifiez respiration",
                "5️⃣ Apportez le flacon/emballage du poison à l'ambulance",
                "6️⃣ Surveillez signes: vomissements, convulsions, difficulté respirer"
            ],
            "warning": "🚨 EMPOISONNEMENT - APPELEZ LE 112/15 ET NOTEZ LE POISON INGÉRÉ"
        },
        
        "brulure": {
            "severity": "high",
            "call": "112 (SAMU) si brûlure grave",
            "immediate_actions": [
                "🚨 BRÛLURE GRAVE - Pour brûlures importantes, appelez le 112",
                "1️⃣ Refroidissez avec EAU FROIDE (ou laissez sous douche) PENDANT 10-20 minutes",
                "2️⃣ Ne pas utiliser eau glacée - juste froide",
                "3️⃣ Enlevez les vêtements non collants (ATTENTION: ne pas arracher si collé)",
                "4️⃣ Couvrez avec tissu PROPRE (gaze, coton stérile)",
                "5️⃣ NE METTEZ PAS de beurre, huile, pommade",
                "6️⃣ Donnez eau à boire (petites gorgées) si conscient"
            ],
            "warning": "🚨 BRÛLURE GRAVE - REFROIDISSEZ À L'EAU FROIDE ET APPELEZ LE 112"
        },
        
        "electrocution": {
            "severity": "critical",
            "call": "112 (SAMU)",
            "immediate_actions": [
                "🚨 ÉLECTROCUTION - Appelez le 112 IMMÉDIATEMENT",
                "1️⃣ ÉLOIGNEZ-VOUS du danger électrique (coupez l'électricité si sûr)",
                "2️⃣ NE TOUCHEZ PAS la personne si elle touche le courant",
                "3️⃣ Utilisez un objet NON MÉTALLIQUE pour dégager le contact",
                "4️⃣ Vérifiez respiration et perte de conscience",
                "5️⃣ Si inconscient mais respire: position latérale stable",
                "6️⃣ Si ne respire pas: massage cardiaque immédiat"
            ],
            "warning": "🚨 ÉLECTROCUTION - COUPEZ L'ÉLECTRICITÉ ET APPELEZ LE 112"
        },
        
        "arrêt_cardiaque": {
            "severity": "critical",
            "call": "112 (SAMU)",
            "immediate_actions": [
                "🚨 ARRÊT CARDIAQUE - Appelez le 112 IMMÉDIATEMENT",
                "1️⃣ Vérifiez respiration et pouls (10 secondes max)",
                "2️⃣ COMMENCEZ LE MASSAGE CARDIAQUE:",
                "   - Position: coeur (centre poitrine)",
                "   - Force: appuyer fermement, 100-120 compressions/minute",
                "3️⃣ Alterner: 30 compressions + 2 respirations bouche-à-bouche",
                "4️⃣ Continuez SANS ARRÊT jusqu'à:",
                "   - Arrivée ambulance",
                "   - Reprise conscience/respiration",
                "5️⃣ Si DEA (défibrillateur) disponible: utilisez immédiatement"
            ],
            "warning": "🚨 ARRÊT CARDIAQUE - APPELEZ LE 112 ET COMMENCEZ MASSAGE CARDIAQUE"
        },
        
        "choc": {
            "severity": "high",
            "call": "112 si état s'aggrave",
            "immediate_actions": [
                "🚨 CHOC (état grave) - Appelez le 112",
                "1️⃣ Allongez la personne sur le dos",
                "2️⃣ Soulevez les jambes (30cm) pour irriguer le cerveau",
                "3️⃣ Maintenez au chaud avec couvertures",
                "4️⃣ Nettoyez les plaies visibles, arrêtez saignements",
                "5️⃣ Rassurez la personne - parlez calmement",
                "6️⃣ Si vomissements: tournez la tête sur le côté"
            ],
            "warning": "🚨 CHOC - ALLONGEZ À PLAT, ÉLEVEZ LES JAMBES, APPELEZ LE 112"
        },
        
        "convulsion": {
            "severity": "high",
            "call": "112 après convulsions",
            "immediate_actions": [
                "🚨 CONVULSIONS - Ne paniquez pas",
                "1️⃣ ÉCARTEZ les objets tranchants/dangereux autour",
                "2️⃣ NE RETENEZ PAS les mouvements, laissez convulser",
                "3️⃣ Placez quelque chose de mou SOUS la tête",
                "4️⃣ Tournez la tête sur le côté (écume peut s'écouler)",
                "5️⃣ Une fois convulsions terms: position latérale stable",
                "6️⃣ Appelez le 112 et restez avec la personne"
            ],
            "warning": "🚨 CONVULSIONS - ÉCARTEZ LES OBJETS DANGEREUX, LAISSER CONVULSER, POSITION LATÉRALE APRÈS"
        }
    }
    
    protocol = protocols.get(detected_type, protocols["accident"])
    
    return {
        "is_emergency": True,
        "severity": protocol["severity"],
        "emergency_type": detected_type,
        "protocol": protocol.get("warning"),
        "immediate_actions": protocol.get("immediate_actions", []),
        "call_emergency_number": protocol.get("call"),
        "warning_alert": protocol.get("warning")
    }

# ==========================================
# MODÈLES (Compatibles avec base existante)
# ==========================================

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    phone_number = Column(String, unique=True, nullable=False)
    is_active = Column(Boolean, default=True)
    role = Column(String, default="utilisateur")
    password_hash = Column(String, nullable=True)
    name = Column(String, nullable=True)
    location = Column(String, nullable=True)
    organization = Column(String, nullable=True, index=True)
    organization_id = Column(Integer, nullable=True, index=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    is_anonymized = Column(Boolean, default=False)
    is_premium = Column(Boolean, default=False)
    premium_expires_at = Column(DateTime, nullable=True)
    messages_used = Column(Integer, default=0)
    messages_limit = Column(Integer, default=1)  # 1 gratuit, 10 pour premium
    failed_pin_attempts = Column(Integer, default=0)
    pin_locked_until = Column(DateTime, nullable=True)
    wallet_balance = Column(Integer, default=0)
    subscription_plan = Column(String, nullable=True)
    subscription_started_at = Column(DateTime, nullable=True)
    subscription_expires_at = Column(DateTime, nullable=True)
    analysis_credits = Column(Integer, default=0)
    ticket_credits = Column(Integer, default=0)

class Expert(Base):
    __tablename__ = "experts"
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String, unique=True, nullable=False)
    password_hash = Column(String, nullable=False)
    full_name = Column(String, nullable=False)
    specialization = Column(String, nullable=True)
    role = Column(String, default="expert")
    is_active = Column(Boolean, default=True)
    zone = Column(String, nullable=True)
    project = Column(String, nullable=True)
    language = Column(String, nullable=True)
    institution = Column(String, nullable=True)
    organization_id = Column(Integer, nullable=True, index=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class Organization(Base):
    __tablename__ = "organizations"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, nullable=False, index=True)
    code = Column(String, unique=True, nullable=True, index=True)
    description = Column(Text, nullable=True)
    region = Column(String, nullable=True)
    phone_number = Column(String, nullable=True)
    email = Column(String, nullable=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class Ticket(Base):
    __tablename__ = "tickets"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, nullable=False)
    expert_id = Column(Integer, nullable=True)
    category = Column(String, nullable=True)
    urgency = Column(String, nullable=True)
    status = Column(String, default="open")
    ai_confidence_score = Column(Float, nullable=True)
    ai_extracted_keywords = Column(String, nullable=True)
    ai_photo_analysis = Column(Text, nullable=True)
    photo_path = Column(String, nullable=True)  # NOM ORIGINAL - NE PAS CHANGER
    photo_paths_json = Column(Text, nullable=True)
    internal_notes = Column(Text, nullable=True)
    resolution_notes = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    resolved_at = Column(DateTime, nullable=True)
    passed_expert_ids_json = Column(Text, nullable=True)
    preferred_language = Column(String, nullable=True)

class PhotoAnalysisHistoryDB(Base):
    __tablename__ = "photo_analysis_history"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, nullable=False)
    client_record_id = Column(String, nullable=True)
    category = Column(String, nullable=True)
    prompt = Column(Text, nullable=True)
    analysis_json = Column(Text, nullable=True)
    photo_paths_json = Column(Text, nullable=True)
    photo_labels_json = Column(Text, nullable=True)
    source_ticket_id = Column(Integer, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class Message(Base):
    __tablename__ = "messages"
    id = Column(Integer, primary_key=True, index=True)
    ticket_id = Column(Integer, nullable=False)
    sender_type = Column(String, nullable=False)
    sender_id = Column(Integer, nullable=True)
    content = Column(Text, nullable=False)
    channel = Column(String, nullable=False)
    sent_at = Column(DateTime, default=datetime.utcnow)
    is_read = Column(Boolean, default=False)
    audio_url = Column(String, nullable=True)  # URL du fichier audio de réponse expert
    language = Column(String, nullable=True)

class KnowledgeItem(Base):
    __tablename__ = "knowledge_items"
    id = Column(Integer, primary_key=True, index=True)
    domain = Column(String, nullable=False)  # agriculture / elevage / cybersecurity / health
    title = Column(String, nullable=False)
    question = Column(Text, nullable=True)
    answer = Column(Text, nullable=False)
    tags = Column(String, nullable=True)  # JSON list of tags
    language = Column(String, default="fr")
    source = Column(String, nullable=True)
    media = Column(Text, nullable=True)  # JSON list of media items (images / videos)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class ExpertLocalKnowledgeDB(Base):
    __tablename__ = "expert_local_knowledge"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String, nullable=False)
    category = Column(String, nullable=False, default="agriculture")
    question_fr = Column(Text, nullable=False)
    resolution_fr = Column(Text, nullable=False)
    tags_json = Column(Text, nullable=True)
    status = Column(String, nullable=False, default="validated")
    origin = Column(String, nullable=False, default="expert_manual")
    translations_json = Column(Text, nullable=True)
    audio_json = Column(Text, nullable=True)
    expert_id = Column(Integer, nullable=True, index=True)
    reviewer_id = Column(Integer, nullable=True, index=True)
    review_notes = Column(Text, nullable=True)
    reviewed_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class AcademyCourseDB(Base):
    __tablename__ = "academy_courses"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String, nullable=False)
    course_type = Column(String, default="culture")  # culture | technique
    crop = Column(String, nullable=True)
    summary = Column(Text, nullable=False)
    cover_url = Column(String, nullable=True)
    steps_json = Column(Text, nullable=True)
    audio_json = Column(Text, nullable=True)
    status = Column(String, default="published")
    created_by = Column(Integer, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class BillingTransaction(Base):
    __tablename__ = "billing_transactions"
    id = Column(String, primary_key=True)
    user_id = Column(Integer, nullable=False, index=True)
    reference = Column(String, unique=True, nullable=False, index=True)
    provider = Column(String, default="yengapay")
    provider_intent_id = Column(String, nullable=True, index=True)
    kind = Column(String, nullable=False)
    target_id = Column(String, nullable=True)
    amount = Column(Integer, nullable=False)
    status = Column(String, default="pending", index=True)
    checkout_url = Column(Text, nullable=True)
    provider_payload_json = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    paid_at = Column(DateTime, nullable=True)


class UsageEventDB(Base):
    __tablename__ = "usage_events"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, nullable=False, index=True)
    resource = Column(String, nullable=False, index=True)
    source = Column(String, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, index=True)


class CourseAccessDB(Base):
    __tablename__ = "academy_course_access"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, nullable=False, index=True)
    course_id = Column(Integer, nullable=False, index=True)
    source = Column(String, nullable=False)
    period_key = Column(String, nullable=True, index=True)
    permanent = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.utcnow)


class OfflineKnowledgeEntryDB(Base):
    __tablename__ = "offline_knowledge_entries"
    id = Column(Integer, primary_key=True, index=True)
    fingerprint = Column(String, unique=True, index=True, nullable=False)
    user_id = Column(Integer, nullable=True)
    domain = Column(String, nullable=False)
    source_kind = Column(String, nullable=False)
    title = Column(String, nullable=False)
    question = Column(Text, nullable=True)
    answer = Column(Text, nullable=False)
    tags_json = Column(Text, nullable=True)
    response_json = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class EmergencyNumber(Base):
    __tablename__ = "emergency_numbers"
    id = Column(Integer, primary_key=True, index=True)
    label = Column(String, nullable=False)
    number = Column(String, nullable=False)
    description = Column(Text, nullable=True)
    display_order = Column(Integer, default=0)
    is_active = Column(Boolean, default=True)

class RuralContactDB(Base):
    __tablename__ = "rural_contacts"
    id = Column(String, primary_key=True, index=True)
    user_id = Column(Integer, nullable=False, index=True)
    name = Column(String, nullable=False)
    actor_type = Column(String, nullable=False)
    phone_number = Column(String, nullable=False)
    location_label = Column(String, nullable=False)
    organization = Column(String, nullable=True)
    market_name = Column(String, nullable=True)
    notes = Column(Text, nullable=True)
    tags_json = Column(Text, nullable=True)
    crop_labels_json = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class PhoneOtpDB(Base):
    __tablename__ = "phone_otps"
    id = Column(Integer, primary_key=True, index=True)
    phone_number = Column(String, nullable=False, index=True)
    code_hash = Column(String, nullable=False)
    purpose = Column(String, nullable=False, default="authentication")
    attempts = Column(Integer, nullable=False, default=0)
    expires_at = Column(DateTime, nullable=False)
    consumed_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, index=True)

# class EntreprendreHistoryDB(Base):
#     __tablename__ = "entreprendre_history"
#     id = Column(String, primary_key=True, index=True)
#     user_id = Column(Integer, nullable=False, index=True)
#     category = Column(String, nullable=False)
#     user_query = Column(Text, nullable=False)
#     response_json = Column(Text, nullable=False)
#     photo_path = Column(String, nullable=True)
#     plot_id = Column(String, nullable=True)
#     translations_json = Column(Text, nullable=True)
#     created_at = Column(DateTime, default=datetime.utcnow)
#     updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
# 
# 
Base.metadata.create_all(bind=engine)
#
#
def _is_sqlite_engine() -> bool:
    return str(engine.url).startswith("sqlite")


def _table_columns(conn, table: str) -> List[str]:
    """Liste les colonnes d'une table, compatible SQLite et PostgreSQL."""
    if _is_sqlite_engine():
        return [row[1] for row in conn.exec_driver_sql(f"PRAGMA table_info({table})")]
    result = conn.exec_driver_sql(
        f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table}'"
    )
    return [row[0] for row in result]


def _add_column_if_missing(
    conn, table: str, column: str, sqlite_ddl: str, postgres_ddl: Optional[str] = None
) -> None:
    """Ajoute une colonne a une table si elle est absente.

    SQLite n'a pas de "ADD COLUMN IF NOT EXISTS" : on verifie via PRAGMA avant.
    PostgreSQL le supporte nativement. Les deux moteurs different aussi sur la
    syntaxe des types/DEFAULT (ex: DATETIME vs TIMESTAMP, DEFAULT 0 vs DEFAULT
    FALSE) d'ou les deux variantes de DDL.
    """
    if _is_sqlite_engine():
        if column not in _table_columns(conn, table):
            conn.exec_driver_sql(f"ALTER TABLE {table} ADD COLUMN {column} {sqlite_ddl}")
    else:
        conn.exec_driver_sql(
            f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {postgres_ddl or sqlite_ddl}"
        )


def _ensure_user_auth_columns() -> None:
    """Ajouter les colonnes d'auth mobile si elles sont absentes (SQLite et PostgreSQL).

    IMPORTANT : tant que ces colonnes manquent sur la base reellement utilisee
    par le serveur (ex: PostgreSQL en production), TOUTE requete authentifiee
    (login, register, scanner, assistant vocal, ...) echoue avec une erreur 500
    non geree des que l'ORM lit la table users - avant meme d'atteindre l'IA.
    """
    try:
        with engine.connect() as conn:
            migrations = [
                ("password_hash", "TEXT", "TEXT"),
                ("is_anonymized", "BOOLEAN DEFAULT 0", "BOOLEAN DEFAULT FALSE"),
                ("is_premium", "BOOLEAN DEFAULT 0", "BOOLEAN DEFAULT FALSE"),
                ("premium_expires_at", "DATETIME", "TIMESTAMP"),
                ("messages_used", "INTEGER DEFAULT 0", "INTEGER DEFAULT 0"),
                ("messages_limit", "INTEGER DEFAULT 1", "INTEGER DEFAULT 1"),
                ("is_active", "BOOLEAN DEFAULT 1", "BOOLEAN DEFAULT TRUE"),
                ("role", "VARCHAR DEFAULT 'utilisateur'", "VARCHAR DEFAULT 'utilisateur'"),
                ("organization", "TEXT", "TEXT"),
                ("organization_id", "INTEGER", "INTEGER"),
                ("failed_pin_attempts", "INTEGER DEFAULT 0", "INTEGER DEFAULT 0"),
                ("wallet_balance", "INTEGER DEFAULT 0", "INTEGER DEFAULT 0"),
                ("subscription_plan", "TEXT", "TEXT"),
                ("subscription_started_at", "DATETIME", "TIMESTAMP"),
                ("subscription_expires_at", "DATETIME", "TIMESTAMP"),
                ("analysis_credits", "INTEGER DEFAULT 0", "INTEGER DEFAULT 0"),
                ("ticket_credits", "INTEGER DEFAULT 0", "INTEGER DEFAULT 0"),
                ("pin_locked_until", "DATETIME", "TIMESTAMP"),
            ]
            for col, sqlite_ddl, postgres_ddl in migrations:
                _add_column_if_missing(conn, "users", col, sqlite_ddl, postgres_ddl)
            conn.commit()
            print("  [OK] Colonnes d'auth utilisateur verifiees sur users")
    except Exception as e:
        print(f"[WARN] Impossible d'ajouter les colonnes d'auth utilisateur: {e}")


_ensure_user_auth_columns()


def _ensure_expert_profile_columns() -> None:
    """Ajouter les colonnes de profil expert si elles sont absentes (SQLite et PostgreSQL).

    Meme piege que pour users : le modele Expert a recu zone/project/language/
    institution sans migration associee, donc toute requete sur experts (dont
    /api/auth/login) plantait avec "no such column: experts.zone".
    """
    try:
        with engine.connect() as conn:
            migrations = [
                ("role", "VARCHAR DEFAULT 'expert'", "VARCHAR DEFAULT 'expert'"),
                ("zone", "TEXT", "TEXT"),
                ("project", "TEXT", "TEXT"),
                ("language", "TEXT", "TEXT"),
                ("institution", "TEXT", "TEXT"),
                ("organization_id", "INTEGER", "INTEGER"),
            ]
            for col, sqlite_ddl, postgres_ddl in migrations:
                _add_column_if_missing(conn, "experts", col, sqlite_ddl, postgres_ddl)
            conn.commit()
            print("  [OK] Colonnes de profil expert verifiees sur experts")
    except Exception as e:
        print(f"[WARN] Impossible d'ajouter les colonnes de profil expert: {e}")


_ensure_expert_profile_columns()


def _ensure_media_column_for_knowledge_items() -> None:
    """S'assurer que la colonne 'media' existe dans la table knowledge_items.

    Utile quand la base existait avant l'ajout de ce champ : on ajoute
    simplement la colonne manquante sans casser les données existantes.
    """
    try:
        with engine.connect() as conn:
            _add_column_if_missing(conn, "knowledge_items", "media", "TEXT", "TEXT")
            conn.commit()
    except Exception as e:
        # On loggue mais on ne bloque pas le démarrage de l'API
        print(f"[WARN] Impossible d'ajouter la colonne 'media' à knowledge_items: {e}")


_ensure_media_column_for_knowledge_items()


def _ensure_ticket_photo_columns() -> None:
    try:
        with engine.connect() as conn:
            _add_column_if_missing(conn, "tickets", "photo_paths_json", "TEXT", "TEXT")
            _add_column_if_missing(conn, "tickets", "internal_notes", "TEXT", "TEXT")
            _add_column_if_missing(conn, "tickets", "passed_expert_ids_json", "TEXT", "TEXT")
            _add_column_if_missing(conn, "tickets", "preferred_language", "TEXT", "TEXT")
            _add_column_if_missing(conn, "messages", "language", "TEXT", "TEXT")
            conn.commit()
    except Exception as e:
        print(f"[WARN] Impossible d'ajouter des colonnes à tickets: {e}")


_ensure_ticket_photo_columns()


def _ensure_expert_local_knowledge_workflow_columns() -> None:
    """Ajoute le suivi auteur/validation aux anciennes bases sans perte de données."""
    try:
        with engine.connect() as conn:
            migrations = [
                ("expert_id", "INTEGER", "INTEGER"),
                ("reviewer_id", "INTEGER", "INTEGER"),
                ("review_notes", "TEXT", "TEXT"),
                ("reviewed_at", "DATETIME", "TIMESTAMP"),
            ]
            for col, sqlite_ddl, postgres_ddl in migrations:
                _add_column_if_missing(conn, "expert_local_knowledge", col, sqlite_ddl, postgres_ddl)
            conn.commit()
    except Exception as exc:
        print(f"[WARN] Migration workflow connaissances impossible: {exc}")


_ensure_expert_local_knowledge_workflow_columns()

# ==========================================
# MODÈLES PYDANTIC
# ==========================================

class ConversationTurn(BaseModel):
    role: str
    content: str

class MessageCreate(BaseModel):
    content: str
    phone_number: str
    channel: str = "app"
    category: Optional[str] = None  # catégorie choisie côté app (agriculture, elevage, sos_accident, cybersecurity)
    photo_base64: Optional[str] = None
    photo_base64_list: Optional[List[str]] = None
    conversation_context: Optional[List[ConversationTurn]] = None
    generate_media: Optional[bool] = False
    target_lang: Optional[str] = None  # Langue locale cible


class ExpertLogin(BaseModel):
    email: str
    password: str


class UserRegister(BaseModel):
    phone_number: str
    password: str
    name: str
    location: Optional[str] = None


class UserLogin(BaseModel):
    phone_number: str
    password: str


class MobileQuestionCreate(BaseModel):
    content: str
    category: Optional[str] = None
    photo_base64: Optional[str] = None
    photo_base64_list: Optional[List[str]] = None
    conversation_context: Optional[List[ConversationTurn]] = None
    target_lang: Optional[str] = None  # Langue locale cible : "moore", "dioula", "fulfulde"

class ReplyMessage(BaseModel):
    message: str
    language: Optional[str] = None


class AcademyCourseIn(BaseModel):
    title: str
    course_type: str = "culture"
    crop: Optional[str] = None
    summary: str
    steps: List[Dict[str, Any]] = []
    status: str = "published"


class PhotoAnalysisHistoryIn(BaseModel):
    phone_number: str
    client_record_id: Optional[str] = None
    category: Optional[str] = None
    prompt: Optional[str] = None
    analysis: Dict[str, Any]
    photo_base64_list: List[str] = []
    photo_labels: List[str] = []


class KnowledgeMedia(BaseModel):
    type: str
    url: str
    title: Optional[str] = None


class KnowledgeItemIn(BaseModel):
    domain: str
    title: str
    question: Optional[str] = None
    answer: str
    tags: List[str] = []
    language: str = "fr"
    source: Optional[str] = None
    media: Optional[List[KnowledgeMedia]] = None


class KnowledgeBulkImport(BaseModel):
    items: List[KnowledgeItemIn]


class LocalizationTranslateIn(BaseModel):
    question_fr: str
    resolution_fr: str
    category: str = "agriculture"
    actions_fr: Optional[List[str]] = None

class RuralContactSyncIn(BaseModel):
    id: str
    name: str
    actor_type: str
    phone_number: str
    location_label: str
    organization: Optional[str] = None
    market_name: Optional[str] = None
    notes: Optional[str] = None
    tags: List[str] = []
    crop_labels: List[str] = []
    updated_at: str

class RuralSyncPayload(BaseModel):
    contacts: List[RuralContactSyncIn]

class EntreprendreSyncIn(BaseModel):
    id: str
    category: str
    user_query: str
    response_json: str
    photo_path: Optional[str] = None
    plot_id: Optional[str] = None
    translations_json: Optional[str] = None
    created_at: str
    updated_at: str

class EntreprendreSyncPayload(BaseModel):
    records: List[EntreprendreSyncIn]


class ExpertLocalKnowledgeIn(BaseModel):
    title: str
    category: str = "agriculture"
    question_fr: str
    resolution_fr: str
    tags: List[str] = []
    translations: Dict[str, Any] = {}
    audio: Dict[str, Any] = {}
    status: str = "validated"
    origin: Optional[str] = None


class ExpertLocalKnowledgeReviewIn(BaseModel):
    status: str
    review_notes: Optional[str] = None


class EmergencyNumberIn(BaseModel):
    label: str
    number: str
    description: Optional[str] = None
    display_order: int = 0


class PhoneAuthStartIn(BaseModel):
    phone_number: str
    force_otp: bool = False


class PhoneOtpVerifyIn(BaseModel):
    phone_number: str
    code: str
    pin: Optional[str] = None
    name: Optional[str] = None
    location: Optional[str] = None


class PhonePinLoginIn(BaseModel):
    phone_number: str
    pin: str


# ==========================================
# MODULE IA PHOTO LOCALE (Computer Vision) - RESTAURÉ
# ==========================================

class LocalComputerVision:
    """
    Système de Computer Vision 100% LOCAL
    Détection de maladies des plantes sans API externe
    """
    
    def __init__(self):
        # Base de connaissances des maladies courantes au Burkina Faso
        self.diseases_database = {
            "mais_taches_jaunes": {
                "name": "Carence en Azote",
                "confidence_keywords": ["jaune", "feuille", "maïs", "sécher"],
                "symptoms": ["Jaunissement des feuilles du bas vers le haut", "Croissance ralentie"],
                "treatment": "Appliquer engrais NPK (10-10-10) à 50kg/ha. Améliorer drainage. Arrosage régulier matin/soir.",
                "urgency": "medium",
                "prevention": "Rotation des cultures, compost organique, analyse sol annuelle"
            },
            "mais_rouille": {
                "name": "Rouille du Maïs",
                "confidence_keywords": ["tache", "orange", "rouille", "poudre"],
                "symptoms": ["Pustules orange/brunes sur feuilles", "Aspect poudreuse"],
                "treatment": "Fongicide naturel (purin d'ortie dilué 1:10). Retirer feuilles infectées. Espacer plants.",
                "urgency": "high",
                "prevention": "Variétés résistantes, rotation, bon espacement"
            },
            "tomate_mildiou": {
                "name": "Mildiou de la Tomate",
                "confidence_keywords": ["tache", "brun", "noir", "tomate", "pourrir"],
                "symptoms": ["Taches brunes/noires sur feuilles", "Fruits pourrissent"],
                "treatment": "URGENT: Retirer plants infectés. Bouillie bordelaise. Éviter arrosage feuilles.",
                "urgency": "high",
                "prevention": "Paillage, arrosage au pied, aération"
            },
            "sorgho_charbon": {
                "name": "Charbon du Sorgho",
                "confidence_keywords": ["noir", "poudre", "épi", "sorgho"],
                "symptoms": ["Masse noire poudreuse remplace grains"],
                "treatment": "Détruire plants infectés (brûler). Traiter semences. Rotation 3 ans.",
                "urgency": "high",
                "prevention": "Semences certifiées traitées, rotation cultures"
            },
            "manioc_mosaique": {
                "name": "Mosaïque du Manioc",
                "confidence_keywords": ["mosaïque", "déformation", "feuille", "manioc"],
                "symptoms": ["Motif mosaïque jaune/vert sur feuilles", "Déformation"],
                "treatment": "Pas de traitement. Arracher et détruire. Utiliser boutures saines certifiées.",
                "urgency": "high",
                "prevention": "Boutures certifiées, contrôle pucerons, éliminer plants malades"
            },
            "animal_fievre": {
                "name": "Fièvre Animale (suspicion)",
                "confidence_keywords": ["bétail", "fièvre", "faible", "animal"],
                "symptoms": ["Température élevée", "Perte appétit", "Faiblesse"],
                "treatment": "CONSULTER vétérinaire RAPIDEMENT. Isoler animal. Eau fraîche disponible.",
                "urgency": "high",
                "prevention": "Vaccination, vermifugation, abri ombragé"
            },
            "animal_plaie": {
                "name": "Plaie ou traumatisme animal",
                "confidence_keywords": ["plaie", "blessure", "saigne", "boite", "coupure", "peau"],
                "symptoms": ["Rougeur, saignement ou lésion visible", "Douleur ou gêne au déplacement"],
                "treatment": "Nettoyer la zone avec une solution antiseptique adaptée, limiter les mouches et isoler l'animal si nécessaire.",
                "urgency": "high",
                "prevention": "Inspecter les clôtures, retirer les objets coupants et surveiller les parasites.",
            },
            "animal_infection_cutanee": {
                "name": "Infection cutanée animale (suspicion)",
                "confidence_keywords": ["croûte", "peau", "purulent", "plaque", "démangeaison"],
                "symptoms": ["Croûtes, inflammation ou dépilation locale", "Zone cutanée anormale persistante"],
                "treatment": "Nettoyer la zone, éviter le léchage ou frottement et demander un avis vétérinaire pour confirmer le traitement.",
                "urgency": "medium",
                "prevention": "Hygiène de l'abri, contrôle des parasites et isolement des animaux atteints.",
            },
            "animal_oculaire": {
                "name": "Atteinte oculaire ou nasale animale",
                "confidence_keywords": ["oeil", "œil", "ecoulement", "nez", "narine", "crête"],
                "symptoms": ["Écoulement, irritation ou gonflement visible", "Atteinte possible des voies respiratoires ou des yeux"],
                "treatment": "Isoler l'animal, nettoyer délicatement les sécrétions externes et consulter rapidement un agent d'élevage ou vétérinaire.",
                "urgency": "high",
                "prevention": "Ventilation correcte, réduction de la promiscuité et surveillance du lot.",
            },
            "human_plaie_hemorragique": {
                "name": "Plaie ouverte ou saignement",
                "confidence_keywords": ["sang", "saigne", "plaie", "coupure", "blessure", "accident"],
                "symptoms": ["Plaie ouverte ou saignement visible", "Atteinte cutanée nécessitant compression ou pansement"],
                "treatment": "Comprimer la plaie avec un tissu propre, surélever si possible la zone touchée et consulter d'urgence si le saignement persiste.",
                "urgency": "high",
                "prevention": "Port de protections et désinfection rapide des petites coupures.",
            },
            "human_brule": {
                "name": "Brûlure ou irritation thermique",
                "confidence_keywords": ["brulure", "brûlure", "chaud", "huile", "feu", "peau rouge"],
                "symptoms": ["Rougeur diffuse, cloque ou surface brûlée", "Douleur et inflammation locale"],
                "treatment": "Refroidir immédiatement à l'eau propre tempérée pendant 10 à 20 minutes. Ne pas appliquer de produit agressif ni percer les cloques.",
                "urgency": "high",
                "prevention": "Manipuler chaleur et liquides bouillants avec protection adaptée.",
            },
            "human_infection_plaie": {
                "name": "Plaie infectée ou inflammatoire",
                "confidence_keywords": ["pus", "infecte", "infectée", "gonfle", "rouge", "chaud"],
                "symptoms": ["Rougeur persistante, gonflement ou écoulement", "Suspicion d'infection locale"],
                "treatment": "Nettoyer la plaie à l'eau propre, couvrir avec un pansement propre et consulter un soignant si douleur, fièvre ou pus apparaissent.",
                "urgency": "high",
                "prevention": "Désinfecter tôt les plaies et renouveler les pansements propres.",
            },
            "human_contusion": {
                "name": "Contusion ou hématome",
                "confidence_keywords": ["choc", "tombe", "bleu", "gonfle", "douleur", "coup"],
                "symptoms": ["Coloration sombre ou tuméfaction visible", "Douleur localisée après choc"],
                "treatment": "Appliquer du froid enveloppé, surélever la zone si possible et surveiller douleur intense ou incapacité à bouger.",
                "urgency": "medium",
                "prevention": "Protéger les zones exposées et sécuriser les zones de travail ou de déplacement.",
            },
            "mais_helminthosporiose": {
                "name": "Helminthosporiose / brûlure foliaire du maïs",
                "confidence_keywords": ["mais", "maïs", "taches allongees", "brun", "feuille", "brule"],
                "symptoms": ["Taches allongées brunes sur feuilles", "Dessèchement progressif du feuillage"],
                "treatment": "Retirer les feuilles très atteintes, améliorer l'aération et utiliser un traitement fongique adapté si disponible localement.",
                "urgency": "high",
                "prevention": "Rotation culturale, destruction des résidus malades et semences saines.",
            },
            "manioc_bacteriose": {
                "name": "Brûlure bactérienne du manioc (suspicion)",
                "confidence_keywords": ["manioc", "brulure", "brûlure", "feuille", "dessèchement", "bacteriose"],
                "symptoms": ["Brunissement et dessèchement foliaire", "Atteinte progressive des feuilles ou tiges"],
                "treatment": "Éliminer les plants très atteints, désinfecter les outils et éviter les boutures issues des plants suspects.",
                "urgency": "high",
                "prevention": "Utiliser des boutures saines, éviter la propagation mécanique et pratiquer la rotation.",
            },
            "animal_pied_lesion": {
                "name": "Lésion du pied ou du sabot",
                "confidence_keywords": ["pied", "sabot", "boiterie", "boite", "patte", "plaie"],
                "symptoms": ["Boiterie ou douleur à l'appui", "Lésion visible au pied ou au sabot"],
                "treatment": "Nettoyer le pied, limiter les déplacements et faire vérifier rapidement si l'animal ne pose plus correctement le membre.",
                "urgency": "high",
                "prevention": "Assainir les sols humides, inspecter régulièrement les sabots et retirer les objets blessants.",
            },
            "volaille_variole": {
                "name": "Variole aviaire (suspicion)",
                "confidence_keywords": ["volaille", "croute", "crete", "crête", "face", "bouton"],
                "symptoms": ["Croûtes ou nodules sur tête, crête ou autour des yeux", "Atteinte cutanée évocatrice chez la volaille"],
                "treatment": "Isoler la volaille, désinfecter l'abri et consulter rapidement un technicien d'élevage pour confirmer la conduite à tenir.",
                "urgency": "high",
                "prevention": "Lutter contre les moustiques, isoler les sujets atteints et renforcer l'hygiène du poulailler.",
            }
        }
        
        # Maladies par culture pour reconnaissance rapide
        self.crop_diseases = {
            "maïs": ["mais_taches_jaunes", "mais_rouille"],
            "tomate": ["tomate_mildiou"],
            "sorgho": ["sorgho_charbon"],
            "manioc": ["manioc_mosaïque"],
            "bétail": ["animal_fievre"]
        }
        self.subject_profiles = {
            "mais": {
                "label": "Maïs",
                "keywords": ["mais", "maïs", "epi", "épi"],
                "capture_guidance": [
                    "Vue générale de la parcelle ou du plant",
                    "Gros plan des feuilles jaunies ou tachées",
                    "Photo du revers de la feuille ou de l'épi",
                ],
            },
            "tomate": {
                "label": "Tomate",
                "keywords": ["tomate", "fruit", "tige"],
                "capture_guidance": [
                    "Vue générale du plant de tomate",
                    "Gros plan des feuilles touchées",
                    "Photo des fruits ou de la tige atteinte",
                ],
            },
            "manioc": {
                "label": "Manioc",
                "keywords": ["manioc", "bouture"],
                "capture_guidance": [
                    "Vue générale du plant de manioc",
                    "Gros plan d'une feuille entière",
                    "Photo du revers des feuilles ou des jeunes pousses",
                ],
            },
            "sorgho": {
                "label": "Sorgho",
                "keywords": ["sorgho", "panicule", "epi", "épi"],
                "capture_guidance": [
                    "Vue générale du plant de sorgho",
                    "Photo de l'épi ou panicule",
                    "Gros plan de la zone noircie ou poudreuse",
                ],
            },
            "oignon": {
                "label": "Oignon",
                "keywords": ["oignon", "bulbe"],
                "capture_guidance": [
                    "Vue générale du rang d'oignons",
                    "Gros plan des feuilles ou du collet",
                    "Photo du bulbe si possible",
                ],
            },
            "arachide": {
                "label": "Arachide",
                "keywords": ["arachide", "cacahuete", "cacahuète"],
                "capture_guidance": [
                    "Vue générale du plant d'arachide",
                    "Gros plan des folioles tachées",
                    "Photo du pied et du sol autour",
                ],
            },
            "betail": {
                "label": "Bétail",
                "keywords": ["betail", "bétail", "vache", "boeuf", "bovin", "veau"],
                "capture_guidance": [
                    "Vue générale de l'animal",
                    "Gros plan de la zone touchée",
                    "Photo des yeux, de la bouche ou du museau si anormal",
                ],
            },
            "petit_ruminant": {
                "label": "Petit ruminant",
                "keywords": ["chevre", "chèvre", "mouton", "brebis"],
                "capture_guidance": [
                    "Vue générale de l'animal",
                    "Gros plan de la peau ou de la lésion",
                    "Photo des yeux ou de la bouche si écoulement",
                ],
            },
            "volaille": {
                "label": "Volaille",
                "keywords": ["volaille", "poule", "coq", "poulet", "canard"],
                "capture_guidance": [
                    "Vue générale de la volaille",
                    "Gros plan de la tête, des yeux ou de la crête",
                    "Photo de la zone plumage ou peau touchée",
                ],
            },
            "lapin": {
                "label": "Lapin",
                "keywords": ["lapin", "lapins", "lapereau", "lapereaux", "clapier"],
                "capture_guidance": [
                    "Vue générale du lapin dans le clapier",
                    "Gros plan de la zone touchée ou de la tête",
                    "Photo des yeux, du nez, des oreilles ou des pattes si anormaux",
                ],
            },
            "humain": {
                "label": "Blessure humaine",
                "keywords": ["main", "bras", "jambe", "pied", "doigt", "peau", "plaie", "brulure", "brûlure", "blessure", "sang"],
                "capture_guidance": [
                    "Vue générale de la zone touchée",
                    "Gros plan net de la blessure ou brûlure",
                    "Photo latérale montrant le gonflement, la profondeur ou l'étendue",
                ],
            },
        }

    def _normalize_text(self, text: str) -> str:
        normalized = unicodedata.normalize("NFD", (text or "").lower())
        return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")

    def _prepare_image(self, image_data: bytes) -> Image.Image:
        image = Image.open(BytesIO(image_data)).convert("RGB")
        image.thumbnail((256, 256))
        return image

    def _extract_visual_features(self, image: Image.Image) -> Dict[str, float]:
        rgb = np.asarray(image, dtype=np.float32)
        hsv = np.asarray(image.convert("HSV"), dtype=np.float32)

        hue = hsv[..., 0] * (360.0 / 255.0)
        saturation = hsv[..., 1] / 255.0
        value = hsv[..., 2] / 255.0
        gray = rgb.mean(axis=2) / 255.0

        green_mask = (hue >= 45) & (hue <= 150) & (saturation > 0.20) & (value > 0.16)
        yellow_mask = (hue >= 32) & (hue <= 72) & (saturation > 0.22) & (value > 0.28)
        orange_mask = (hue >= 8) & (hue <= 30) & (saturation > 0.32) & (value > 0.20)
        brown_mask = (hue >= 10) & (hue <= 40) & (saturation > 0.25) & (value >= 0.10) & (value <= 0.65)
        dark_mask = value < 0.22
        white_mask = (saturation < 0.15) & (value > 0.72)
        red_mask = ((hue <= 12) | (hue >= 340)) & (saturation > 0.32) & (value > 0.20)
        blue_purple_mask = (hue >= 210) & (hue <= 300) & (saturation > 0.18) & (value > 0.12)
        skin_mask = (
            (((hue >= 0) & (hue <= 35)) | ((hue >= 340) & (hue <= 360)))
            & (saturation >= 0.12)
            & (saturation <= 0.65)
            & (value >= 0.28)
            & (value <= 0.96)
        )

        horizontal_diff = np.abs(np.diff(gray, axis=1))
        vertical_diff = np.abs(np.diff(gray, axis=0))
        edge_density = float(
            (
                (horizontal_diff > 0.16).mean()
                + (vertical_diff > 0.16).mean()
            ) / 2.0
        )

        lesion_mask = brown_mask | (dark_mask & (saturation > 0.18))

        return {
            "green_ratio": float(green_mask.mean()),
            "yellow_ratio": float(yellow_mask.mean()),
            "orange_ratio": float(orange_mask.mean()),
            "brown_ratio": float(brown_mask.mean()),
            "dark_ratio": float(dark_mask.mean()),
            "white_ratio": float(white_mask.mean()),
            "red_ratio": float(red_mask.mean()),
            "blue_purple_ratio": float(blue_purple_mask.mean()),
            "skin_ratio": float(skin_mask.mean()),
            "lesion_ratio": float(lesion_mask.mean()),
            "brightness": float(value.mean()),
            "saturation": float(saturation.mean()),
            "texture": float(gray.std()),
            "edge_density": edge_density,
        }

    def _keyword_score(self, text: str, keywords: List[str]) -> float:
        if not keywords:
            return 0.0
        hits = sum(1 for keyword in keywords if self._normalize_text(keyword) in text)
        return min(0.18, hits * 0.06)

    def _build_visual_observations(self, features: Dict[str, float]) -> List[str]:
        observations: List[str] = []

        if features["green_ratio"] > 0.18:
            observations.append("La photo contient une forte présence de feuillage vert exploitable pour le diagnostic.")
        if features["yellow_ratio"] > 0.14:
            observations.append("Présence notable de jaunissement sur la zone analysée.")
        if features["orange_ratio"] > 0.04:
            observations.append("Des zones orange/brun clair ressemblant à des pustules ou taches sont visibles.")
        if features["brown_ratio"] > 0.10 or features["lesion_ratio"] > 0.12:
            observations.append("Des nécroses ou taches sombres/brunes sont détectées.")
        if features["white_ratio"] > 0.08:
            observations.append("Des zones pâles ou blanchâtres sont visibles sur l'image.")
        if features["red_ratio"] > 0.06:
            observations.append("Des zones rouges ou inflammatoires ressortent sur la photo.")
        if features.get("blue_purple_ratio", 0.0) > 0.05:
            observations.append("Des zones bleu-violet évoquant un hématome ou une contusion sont visibles.")
        if features.get("skin_ratio", 0.0) > 0.16:
            observations.append("La photo contient une zone cutanée bien visible, utile pour une analyse de blessure.")
        if features["texture"] < 0.08:
            observations.append("L'image semble peu contrastée; un diagnostic plus précis demanderait une photo plus nette.")

        return observations[:4]

    def _detect_subject_profile(self, normalized_text: str, normalized_category: str) -> Optional[Dict[str, Any]]:
        for key, profile in self.subject_profiles.items():
            if any(self._normalize_text(keyword) in normalized_text for keyword in profile["keywords"]):
                return {"key": key, **profile}

        if normalized_category == "elevage":
            return {"key": "betail", **self.subject_profiles["betail"]}
        if normalized_category == "sos_accident":
            return {"key": "humain", **self.subject_profiles["humain"]}
        if normalized_category == "agriculture":
            return {"key": "mais", **self.subject_profiles["mais"]}
        return None

    def _infer_diagnosis_type(self, disease: str, subject_profile: Optional[Dict[str, Any]]) -> Dict[str, str]:
        disease_key = self._normalize_text(disease)
        profile_key = subject_profile.get("key") if subject_profile else None

        if profile_key == "humain" or any(token in disease_key for token in ["brulure", "brûlure", "plaie", "contusion", "hematome", "hématome"]):
            return {
                "diagnosis_type": "human_first_aid",
                "diagnosis_type_label": "Blessure humaine / premiers secours",
            }
        if profile_key in {"betail", "petit_ruminant", "volaille", "lapin"} or any(token in disease_key for token in ["animale", "volaille", "sabot", "veterinaire", "vétérinaire", "lapin"]):
            return {
                "diagnosis_type": "animal_health_injury",
                "diagnosis_type_label": "Maladie ou blessure animale",
            }
        return {
            "diagnosis_type": "plant_disease_stress",
            "diagnosis_type_label": "Maladie ou stress de plante",
        }

    def _build_critical_alert(self, disease: str, urgency: str, diagnosis_type: str) -> Dict[str, Any]:
        disease_key = self._normalize_text(disease)
        critical_alert = None
        emergency_actions: List[str] = []
        severity_label = "Surveillance"

        if diagnosis_type == "human_first_aid" and any(token in disease_key for token in ["saignement", "plaie ouverte", "brulure", "brûlure"]):
            critical_alert = "Cas potentiellement urgent: appliquez immédiatement les premiers gestes et cherchez une aide médicale si l'état est grave ou s'aggrave."
            emergency_actions = [
                "Comprimer ou refroidir la zone selon le type de blessure.",
                "Utiliser uniquement de l'eau propre et un tissu propre si disponible.",
                "Contacter les secours ou un centre de santé si la douleur, le saignement ou l'étendue est importante.",
            ]
            severity_label = "Alerte immédiate"
        elif diagnosis_type == "animal_health_injury" and any(token in disease_key for token in ["plaie", "sabot", "oculaire", "variole", "fievre", "fièvre"]):
            critical_alert = "Suspicion de cas animal sérieux: isolez l'animal et faites confirmer rapidement par un agent d'élevage ou vétérinaire."
            emergency_actions = [
                "Isoler l'animal ou la volaille atteinte si possible.",
                "Limiter les déplacements et surveiller écoulement, boiterie, abattement ou difficulté respiratoire.",
                "Désinfecter le matériel et éviter le contact rapproché avec le reste du troupeau.",
            ]
            severity_label = "A surveiller d'urgence"
        elif diagnosis_type == "plant_disease_stress" and urgency == "high":
            critical_alert = "Risque d'aggravation rapide de la culture: isolez ou retirez les parties très atteintes et confirmez vite sur le terrain."
            emergency_actions = [
                "Éviter la propagation par contact ou arrosage sur le feuillage.",
                "Retirer les parties ou plants très atteints si la maladie se diffuse rapidement.",
                "Prendre une seconde photo de confirmation des feuilles, tiges ou fruits.",
            ]
            severity_label = "Intervention rapide"
        elif urgency == "high":
            severity_label = "Intervention rapide"
        elif urgency == "medium":
            severity_label = "À surveiller"

        return {
            "severity_label": severity_label,
            "critical_alert": critical_alert,
            "emergency_actions": emergency_actions,
        }

    def _build_local_context(self, disease: str, diagnosis_type: str, subject_profile: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        disease_key = self._normalize_text(disease)
        subject_label = subject_profile["label"] if subject_profile else None

        if diagnosis_type == "human_first_aid":
            return {
                "local_context_note": "Conseils adaptés à un contexte terrain Burkina: priorité à l'eau propre, au tissu propre, à la protection contre la poussière et à l'accès rapide au centre de santé le plus proche.",
                "local_examples": [
                    "Au champ ou au marché, une coupure sale doit être rincée vite avec eau propre avant pansement propre.",
                    "Pour une brûlure à l'huile ou au feu de cuisson, refroidir à l'eau propre sans appliquer de poudre ou de pâte agressive.",
                    "Si la blessure continue à saigner ou empêche de bouger, cherchez une prise en charge médicale sans attendre.",
                ],
            }

        if diagnosis_type == "animal_health_injury":
            examples = [
                "En élevage villageois, isolez vite l'animal atteint pour limiter la contagion ou l'aggravation.",
                "Nettoyez la zone touchée avec un antiseptique adapté si disponible et gardez l'abri plus sec et propre.",
                "Si l'animal ne mange plus, boite fort ou présente écoulement/fièvre, faites intervenir rapidement un agent d'élevage ou vétérinaire.",
            ]
            if subject_label == "Volaille":
                examples[1] = "Pour la volaille, séparez immédiatement les sujets atteints et désinfectez mangeoires, abreuvoirs et poulailler."
            return {
                "local_context_note": "Conseils orientés élevage Burkina: gestion du troupeau ou du lot, isolement rapide, hygiène de l'abri et recours à l'agent d'élevage local.",
                "local_examples": examples,
            }

        plant_examples = [
            "En saison humide, évitez l'arrosage direct du feuillage déjà taché et retirez vite les parties très atteintes.",
            "Au champ, observez aussi les plants voisins pour voir si le problème se diffuse sur la ligne ou la parcelle.",
            "Si possible, combinez traitement local, aération et rotation culturale au prochain cycle.",
        ]
        if "mais" in disease_key:
            plant_examples[0] = "Sur maïs, comparez plusieurs feuilles du bas et du haut pour distinguer carence et maladie foliaire." 
        elif "manioc" in disease_key:
            plant_examples[0] = "Sur manioc, évitez d'utiliser comme boutures des tiges venant de plants suspects ou déjà desséchés."
        elif "tomate" in disease_key:
            plant_examples[0] = "Sur tomate, aérez davantage les plants et évitez de mouiller les feuilles en fin de journée."

        return {
            "local_context_note": "Conseils formulés pour des pratiques agricoles courantes au Burkina: observation de parcelle, retrait ciblé, hygiène culturale et confirmation par agent agricole si propagation rapide.",
            "local_examples": plant_examples,
        }

    def _analyze_human_condition(self, normalized_text: str, features: Dict[str, float], observations: List[str],
                                 subject_profile: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        bleeding_score = 0.28 + min(0.34, features["red_ratio"] * 2.3) + min(0.16, features["edge_density"] * 0.8)
        bleeding_score += self._keyword_score(normalized_text, self.diseases_database["human_plaie_hemorragique"]["confidence_keywords"])

        burn_score = 0.24 + min(0.22, features["red_ratio"] * 1.5) + min(0.16, features["white_ratio"] * 1.2) + min(0.08, features["orange_ratio"] * 0.9)
        burn_score += self._keyword_score(normalized_text, self.diseases_database["human_brule"]["confidence_keywords"])

        infected_score = 0.22 + min(0.22, features["red_ratio"] * 1.2) + min(0.22, features["yellow_ratio"] * 1.0) + min(0.10, features["brown_ratio"] * 0.8)
        infected_score += self._keyword_score(normalized_text, self.diseases_database["human_infection_plaie"]["confidence_keywords"])

        bruise_score = 0.20 + min(0.28, features.get("blue_purple_ratio", 0.0) * 2.1) + min(0.10, features["dark_ratio"] * 0.6)
        bruise_score += self._keyword_score(normalized_text, self.diseases_database["human_contusion"]["confidence_keywords"])

        if bleeding_score >= max(burn_score, infected_score, bruise_score) and bleeding_score >= 0.42:
            disease = self.diseases_database["human_plaie_hemorragique"]
            return self._base_result(
                disease=disease["name"],
                confidence=bleeding_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Les indices visuels et le contexte évoquent une plaie ouverte ou un saignement nécessitant des gestes de premiers secours rapides.",
                recommendations="Si le saignement est abondant, non contrôlé, ou si la plaie est profonde, contactez immédiatement les secours ou un soignant.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if burn_score >= max(infected_score, bruise_score) and burn_score >= 0.40:
            disease = self.diseases_database["human_brule"]
            return self._base_result(
                disease=disease["name"],
                confidence=burn_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="L'image évoque davantage une brûlure superficielle ou intermédiaire qu'une simple coupure mécanique.",
                recommendations="Ajoutez une photo montrant l'étendue complète et précisez la cause (eau chaude, feu, produit chimique, huile, métal chaud).",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if infected_score >= bruise_score and infected_score >= 0.38:
            disease = self.diseases_database["human_infection_plaie"]
            return self._base_result(
                disease=disease["name"],
                confidence=infected_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="La rougeur, les tons jaunâtres ou brunâtres et le contexte texte suggèrent une plaie inflammatoire ou déjà infectée.",
                recommendations="Consultez rapidement s'il y a fièvre, pus, douleur croissante ou extension de la rougeur.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if bruise_score >= 0.34:
            disease = self.diseases_database["human_contusion"]
            return self._base_result(
                disease=disease["name"],
                confidence=bruise_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="L'aspect bleu-violet ou sombre détecté ressemble à une contusion ou un hématome après choc.",
                recommendations="Si la zone ne peut plus bouger normalement, si la douleur est très forte ou si un os semble touché, cherchez une prise en charge médicale.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        return self._base_result(
            disease="Blessure humaine non caractérisée",
            confidence=0.36,
            symptoms=["Lésion visible mais difficile à classer précisément sur photo seule"],
            treatment="Nettoyez à l'eau propre si possible, protégez la zone et consultez si douleur, saignement ou brûlure importante.",
            prevention="Utiliser une protection adaptée et éviter de manipuler la plaie avec des mains sales.",
            urgency="medium",
            analysis="La photo montre bien une atteinte cutanée humaine, mais le type exact de blessure ne peut pas être confirmé avec assez de certitude.",
            recommendations="Ajoutez une autre vue plus nette et précisez la cause, l'heure de l'accident et l'intensité de la douleur.",
            requires_expert=True,
            features=features,
            observations=observations,
            subject_profile=subject_profile,
        )

    def _analyze_animal_condition(self, normalized_text: str, features: Dict[str, float], observations: List[str],
                                  subject_profile: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        subject_key = subject_profile["key"] if subject_profile else None
        wound_score = 0.26 + min(0.24, features["red_ratio"] * 1.8) + min(0.18, features["edge_density"] * 0.7)
        wound_score += self._keyword_score(normalized_text, self.diseases_database["animal_plaie"]["confidence_keywords"])

        skin_infection_score = 0.22 + min(0.18, features["brown_ratio"] * 1.0) + min(0.14, features["yellow_ratio"] * 0.9) + min(0.12, features["texture"] * 0.8)
        skin_infection_score += self._keyword_score(normalized_text, self.diseases_database["animal_infection_cutanee"]["confidence_keywords"])

        ocular_score = 0.21 + min(0.16, features["white_ratio"] * 1.1) + min(0.16, features["red_ratio"] * 1.0)
        ocular_score += self._keyword_score(normalized_text, self.diseases_database["animal_oculaire"]["confidence_keywords"])
        if subject_profile and subject_profile.get("key") == "volaille":
            ocular_score += 0.06

        fever_score = 0.34 + self._keyword_score(normalized_text, self.diseases_database["animal_fievre"]["confidence_keywords"])
        foot_score = 0.24 + min(0.18, features["red_ratio"] * 1.1) + min(0.20, features["brown_ratio"] * 1.1) + min(0.12, features["edge_density"] * 0.6)
        foot_score += self._keyword_score(normalized_text, self.diseases_database["animal_pied_lesion"]["confidence_keywords"])
        if subject_key in {"betail", "petit_ruminant"}:
            foot_score += 0.06

        pox_score = 0.22 + min(0.20, features["brown_ratio"] * 1.0) + min(0.16, features["red_ratio"] * 0.9) + min(0.10, features["white_ratio"] * 0.8)
        pox_score += self._keyword_score(normalized_text, self.diseases_database["volaille_variole"]["confidence_keywords"])
        if subject_key == "volaille":
            pox_score += 0.10

        if foot_score >= max(wound_score, skin_infection_score, ocular_score, fever_score, pox_score) and foot_score >= 0.40:
            disease = self.diseases_database["animal_pied_lesion"]
            return self._base_result(
                disease=disease["name"],
                confidence=foot_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Le contexte de boiterie ou de pied/sabot associé aux indices visuels évoque une lésion localisée du membre.",
                recommendations="Ajoutez une photo de dessous et de profil du pied ou sabot si l'animal accepte de se laisser observer.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if pox_score >= max(wound_score, skin_infection_score, ocular_score, fever_score) and pox_score >= 0.39:
            disease = self.diseases_database["volaille_variole"]
            return self._base_result(
                disease=disease["name"],
                confidence=pox_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Chez la volaille, les croûtes visibles au niveau de la tête ou de la crête évoquent une suspicion de variole aviaire.",
                recommendations="Ajoutez une photo de face et précisez si plusieurs volailles présentent des croûtes similaires.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if wound_score >= max(skin_infection_score, ocular_score, fever_score, foot_score, pox_score) and wound_score >= 0.40:
            disease = self.diseases_database["animal_plaie"]
            return self._base_result(
                disease=disease["name"],
                confidence=wound_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="La photo montre des signes compatibles avec une blessure, une plaie ouverte ou un traumatisme local chez l'animal.",
                recommendations="Vérifiez s'il y a boiterie, écoulement, odeur inhabituelle ou infestation par les mouches.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if skin_infection_score >= max(ocular_score, fever_score, foot_score, pox_score) and skin_infection_score >= 0.38:
            disease = self.diseases_database["animal_infection_cutanee"]
            return self._base_result(
                disease=disease["name"],
                confidence=skin_infection_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Les croûtes, taches brunâtres ou altérations de texture orientent vers une atteinte cutanée infectieuse ou parasitaire.",
                recommendations="Ajoutez une photo du contour de la lésion et précisez s'il y a démangeaison, chute de poils ou mauvaise odeur.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if ocular_score >= max(fever_score, foot_score, pox_score) and ocular_score >= 0.35:
            disease = self.diseases_database["animal_oculaire"]
            return self._base_result(
                disease=disease["name"],
                confidence=ocular_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Le contexte et les indices visuels évoquent une atteinte des yeux, des narines ou de la tête chez l'animal.",
                recommendations="Prenez une vue de face en lumière naturelle et signalez toux, abattement ou difficulté respiratoire si présents.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        disease = self.diseases_database["animal_fievre"]
        return self._base_result(
            disease=disease["name"],
            confidence=fever_score,
            symptoms=disease["symptoms"],
            treatment=disease["treatment"],
            prevention=disease["prevention"],
            urgency=disease["urgency"],
            analysis="Le contexte évoque un problème sanitaire animal plus général, sans signe photo assez spécifique pour une confirmation visuelle forte.",
            recommendations="Ajoutez la température, l'appétit, la durée d'évolution et une vue plus nette de la zone anormale.",
            requires_expert=True,
            features=features,
            observations=observations,
            subject_profile=subject_profile,
        )

    def _analyze_plant_condition(self, normalized_text: str, features: Dict[str, float], observations: List[str],
                                 subject_profile: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        subject_key = subject_profile["key"] if subject_profile else None
        nitrogen_score = 0.28 + min(0.28, features["yellow_ratio"] * 1.4) + min(0.10, features["green_ratio"] * 0.4)
        nitrogen_score += self._keyword_score(normalized_text, self.diseases_database["mais_taches_jaunes"]["confidence_keywords"])

        rust_score = 0.24 + min(0.30, features["orange_ratio"] * 2.2) + min(0.12, features["lesion_ratio"] * 0.9)
        rust_score += self._keyword_score(normalized_text, self.diseases_database["mais_rouille"]["confidence_keywords"])

        blight_score = 0.24 + min(0.26, features["brown_ratio"] * 1.5) + min(0.14, features["dark_ratio"] * 0.8)
        blight_score += self._keyword_score(normalized_text, self.diseases_database["tomate_mildiou"]["confidence_keywords"])

        mosaic_score = 0.22 + min(0.24, features["yellow_ratio"] * 0.9) + min(0.12, features["texture"] * 0.8)
        mosaic_score += 0.08 if "manioc" in normalized_text else 0.0

        drought_score = 0.20 + min(0.26, features["brown_ratio"] * 1.1) + min(0.16, (1 - features["brightness"]) * 0.5)
        drought_score += 0.06 if any(token in normalized_text for token in ["seche", "fletri", "manque d'eau", "sol"]) else 0.0

        charbon_score = 0.18 + min(0.34, features["dark_ratio"] * 1.9) + min(0.12, features["texture"] * 0.7)
        charbon_score += self._keyword_score(normalized_text, self.diseases_database["sorgho_charbon"]["confidence_keywords"])

        maize_blight_score = 0.22 + min(0.24, features["brown_ratio"] * 1.3) + min(0.18, features["lesion_ratio"] * 1.1) + min(0.10, features["edge_density"] * 0.5)
        maize_blight_score += self._keyword_score(normalized_text, self.diseases_database["mais_helminthosporiose"]["confidence_keywords"])

        cassava_blight_score = 0.21 + min(0.22, features["brown_ratio"] * 1.0) + min(0.18, features["yellow_ratio"] * 0.8) + min(0.12, features["dark_ratio"] * 0.6)
        cassava_blight_score += self._keyword_score(normalized_text, self.diseases_database["manioc_bacteriose"]["confidence_keywords"])

        oignon_score = 0.20 + min(0.22, features["yellow_ratio"] * 0.8) + min(0.18, features["brown_ratio"] * 1.1)
        arachide_score = 0.19 + min(0.22, features["brown_ratio"] * 1.0) + min(0.10, features["orange_ratio"] * 0.9)

        if subject_key == "tomate":
            blight_score += 0.08
        if subject_key == "manioc":
            mosaic_score += 0.10
            cassava_blight_score += 0.08
        if subject_key == "sorgho":
            charbon_score += 0.10
        if subject_key == "mais":
            rust_score += 0.06
            nitrogen_score += 0.05
            maize_blight_score += 0.09
        if subject_key == "oignon":
            oignon_score += 0.10
        if subject_key == "arachide":
            arachide_score += 0.10

        if maize_blight_score >= max(nitrogen_score, rust_score, blight_score, mosaic_score, drought_score, charbon_score, oignon_score, arachide_score, cassava_blight_score) and maize_blight_score >= 0.40:
            disease = self.diseases_database["mais_helminthosporiose"]
            return self._base_result(
                disease=disease["name"],
                confidence=maize_blight_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Les taches brunes et allongées sur le maïs orientent plutôt vers une brûlure foliaire fongique qu'une simple rouille ou carence.",
                recommendations="Ajoutez une photo d'une feuille entière, idéalement avec plusieurs lésions alignées, pour confirmer l'aspect allongé des taches.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if cassava_blight_score >= max(nitrogen_score, rust_score, blight_score, mosaic_score, drought_score, charbon_score, oignon_score, arachide_score) and cassava_blight_score >= 0.39:
            disease = self.diseases_database["manioc_bacteriose"]
            return self._base_result(
                disease=disease["name"],
                confidence=cassava_blight_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Le brunissement et le dessèchement visibles sur le manioc évoquent davantage une brûlure bactérienne qu'une simple mosaïque.",
                recommendations="Prenez aussi une photo des tiges et précisez si le dessèchement progresse rapidement sur plusieurs plants.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if charbon_score >= max(nitrogen_score, rust_score, blight_score, mosaic_score, drought_score, oignon_score, arachide_score, maize_blight_score, cassava_blight_score) and charbon_score >= 0.42:
            disease = self.diseases_database["sorgho_charbon"]
            return self._base_result(
                disease=disease["name"],
                confidence=charbon_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Les zones noires et poudreuses détectées sont compatibles avec une suspicion de charbon du sorgho.",
                recommendations="Prenez une photo de l'épi entier pour confirmer si les grains sont remplacés par une masse noire poudreuse.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if oignon_score >= max(nitrogen_score, rust_score, blight_score, mosaic_score, drought_score, arachide_score, maize_blight_score, cassava_blight_score) and oignon_score >= 0.38:
            return self._base_result(
                disease="Suspicion de mildiou ou brûlure foliaire de l'oignon",
                confidence=oignon_score,
                symptoms=[
                    "Jaunissement ou brunissement des feuilles",
                    "Affaiblissement progressif du feuillage",
                ],
                treatment="Réduire l'humidité sur le feuillage, améliorer l'aération et retirer les feuilles très atteintes.",
                prevention="Espacer les plants, éviter l'arrosage tardif sur les feuilles et pratiquer la rotation.",
                urgency="medium",
                analysis="Le profil visuel des feuilles d'oignon suggère une atteinte foliaire plutôt qu'une simple carence.",
                recommendations="Ajoutez une photo du collet et précisez si l'humidité est élevée ou si l'attaque se diffuse rapidement.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if arachide_score >= max(nitrogen_score, rust_score, blight_score, mosaic_score, drought_score, maize_blight_score, cassava_blight_score) and arachide_score >= 0.36:
            return self._base_result(
                disease="Suspicion de cercosporiose ou taches foliaires de l'arachide",
                confidence=arachide_score,
                symptoms=[
                    "Petites taches brunes sur folioles",
                    "Dégradation progressive du feuillage",
                ],
                treatment="Retirer les feuilles très atteintes et appliquer un traitement fongique adapté si disponible localement.",
                prevention="Rotation culturale, semences saines et limitation de l'humidité stagnante.",
                urgency="medium",
                analysis="Les taches détectées sur le feuillage d'arachide évoquent une maladie foliaire fongique courante.",
                recommendations="Prenez un gros plan des deux faces de la foliole pour mieux distinguer tache foliaire et carence.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if rust_score >= max(nitrogen_score, blight_score, mosaic_score, drought_score, maize_blight_score, cassava_blight_score) and rust_score >= 0.45:
            disease = self.diseases_database["mais_rouille"]
            return self._base_result(
                disease=disease["name"],
                confidence=rust_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Les zones orange/brunes détectées sur le feuillage sont compatibles avec une suspicion de rouille.",
                recommendations="Photographiez aussi le revers des feuilles pour confirmer la présence de pustules poudreuses.",
                requires_expert=rust_score < 0.62,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if blight_score >= max(nitrogen_score, mosaic_score, drought_score, maize_blight_score, cassava_blight_score) and blight_score >= 0.43:
            disease = self.diseases_database["tomate_mildiou"]
            return self._base_result(
                disease=disease["name"],
                confidence=blight_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Les taches sombres et brunes détectées évoquent une atteinte foliaire de type mildiou ou nécrose avancée.",
                recommendations="Isolez les plants atteints et prenez une seconde photo des tiges et fruits pour confirmer l'extension.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if mosaic_score >= max(nitrogen_score, drought_score, cassava_blight_score, maize_blight_score) and mosaic_score >= 0.40:
            return self._base_result(
                disease="Suspicion de mosaïque foliaire",
                confidence=mosaic_score,
                symptoms=[
                    "Alternance de zones vertes et jaunâtres",
                    "Aspect irrégulier ou marbré du feuillage",
                ],
                treatment="Éliminer les plants très atteints et éviter toute bouture issue de plants suspects.",
                prevention="Utiliser du matériel végétal sain, contrôler les insectes vecteurs et désinfecter les outils.",
                urgency="high",
                analysis="Le motif visuel mêlant vert et jaune sur le feuillage fait penser à une mosaïque virale ou un stress foliaire sévère.",
                recommendations="Ajoutez une photo d'une feuille entière sur fond neutre pour distinguer mosaïque virale et carence.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if nitrogen_score >= max(drought_score, maize_blight_score, cassava_blight_score) and nitrogen_score >= 0.42:
            disease = self.diseases_database["mais_taches_jaunes"]
            return self._base_result(
                disease=disease["name"],
                confidence=nitrogen_score,
                symptoms=disease["symptoms"],
                treatment=disease["treatment"],
                prevention=disease["prevention"],
                urgency=disease["urgency"],
                analysis="Le jaunissement dominant détecté sur l'image évoque davantage une carence nutritive qu'une brûlure localisée.",
                recommendations="Vérifiez si le jaunissement commence sur les feuilles basses et ajoutez si possible l'âge de la culture.",
                requires_expert=nitrogen_score < 0.58,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        if drought_score >= 0.36:
            return self._base_result(
                disease="Stress hydrique ou problème de sol",
                confidence=drought_score,
                symptoms=[
                    "Brunissement ou dessèchement des bords",
                    "Perte de vigueur visuelle",
                ],
                treatment="Contrôler l'humidité du sol, améliorer le paillage et ajuster l'irrigation selon le stade de la culture.",
                prevention="Maintenir une humidité plus stable, apporter de la matière organique et éviter le compactage du sol.",
                urgency="medium",
                analysis="L'image suggère surtout un stress abiotiques: manque d'eau, chaleur ou déséquilibre du sol.",
                recommendations="Prenez aussi une photo du sol au pied de la plante et précisez la fréquence d'arrosage.",
                requires_expert=False,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )

        return self._base_result(
            disease="Indéterminé",
            confidence=0.34,
            symptoms=["Informations visuelles insuffisantes pour identifier précisément la maladie"],
            treatment="Un expert peut confirmer le diagnostic si vous ajoutez une photo plus proche et une description plus détaillée.",
            prevention="Prendre plusieurs photos: vue générale, gros plan des zones atteintes et face inférieure des feuilles.",
            urgency="medium",
            analysis="Une photo a bien été analysée, mais les indices visuels ne permettent pas encore de conclure avec un niveau de confiance suffisant.",
            recommendations="Ajoutez le nom de la culture ou de l'animal, l'ancienneté du problème et une photo plus nette pour améliorer la détection.",
            requires_expert=True,
            features=features,
            observations=observations,
            subject_profile=subject_profile,
        )

    def _base_result(self, *, disease: str, confidence: float, symptoms: List[str], treatment: str,
                     prevention: str, urgency: str, analysis: str, recommendations: str,
                     requires_expert: bool, features: Dict[str, float], observations: List[str],
                     subject_profile: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        diagnosis_meta = self._infer_diagnosis_type(disease, subject_profile)
        alert_meta = self._build_critical_alert(disease, urgency, diagnosis_meta["diagnosis_type"])
        local_context = self._build_local_context(disease, diagnosis_meta["diagnosis_type"], subject_profile)
        return {
            "disease_detected": disease,
            "confidence": round(max(0.0, min(confidence, 0.92)), 2),
            "symptoms": symptoms,
            "treatment": treatment,
            "prevention": prevention,
            "urgency": urgency,
            "analysis": analysis,
            "recommendations": recommendations,
            "requires_expert": requires_expert,
            "analysis_mode": "hybrid_local_image_text",
            "visual_observations": observations,
            "visual_features": {key: round(value, 3) for key, value in features.items()},
            "detected_subject": subject_profile["label"] if subject_profile else None,
            "capture_guidance": subject_profile["capture_guidance"] if subject_profile else None,
            "diagnosis_type": diagnosis_meta["diagnosis_type"],
            "diagnosis_type_label": diagnosis_meta["diagnosis_type_label"],
            "severity_label": alert_meta["severity_label"],
            "critical_alert": alert_meta["critical_alert"],
            "emergency_actions": alert_meta["emergency_actions"],
            "local_context_note": local_context["local_context_note"],
            "local_examples": local_context["local_examples"],
        }

    def analyze_image_simple(self, image_data: bytes, text_description: str = "", category: Optional[str] = None) -> dict:
        """
        Analyse locale hybride: indices visuels extraits de l'image + contexte texte.
        """
        try:
            normalized_text = self._normalize_text(text_description)
            normalized_category = self._normalize_text(category or "")
            image = self._prepare_image(image_data)
            features = self._extract_visual_features(image)
            observations = self._build_visual_observations(features)
            subject_profile = self._detect_subject_profile(normalized_text, normalized_category)

            is_human_context = normalized_category == "sos_accident" or any(
                token in normalized_text for token in ["blessure", "plaie", "coupure", "brulure", "brûlure", "sang", "main", "bras", "jambe", "pied", "doigt"]
            )
            is_animal_context = not is_human_context and (normalized_category == "elevage" or any(
                token in normalized_text for token in ["animal", "betail", "vache", "mouton", "chevre", "volaille", "peau"]
            ))
            is_plant_context = not is_human_context and not is_animal_context and (
                normalized_category in {"agriculture", ""}
                or any(token in normalized_text for token in ["feuille", "plante", "culture", "mais", "tomate", "manioc", "sorgho", "tache"])
                or features["green_ratio"] > 0.12
            )

            if is_human_context:
                return self._analyze_human_condition(
                    normalized_text,
                    features,
                    observations,
                    subject_profile,
                )

            if is_animal_context:
                return self._analyze_animal_condition(
                    normalized_text,
                    features,
                    observations,
                    subject_profile,
                )

            if is_plant_context:
                return self._analyze_plant_condition(
                    normalized_text,
                    features,
                    observations,
                    subject_profile,
                )

            return self._base_result(
                disease="Indéterminé",
                confidence=0.34,
                symptoms=["Informations visuelles insuffisantes pour identifier précisément la maladie"],
                treatment="Un expert peut confirmer le diagnostic si vous ajoutez une photo plus proche et une description plus détaillée.",
                prevention="Prendre plusieurs photos: vue générale, gros plan des zones atteintes et face inférieure des feuilles.",
                urgency="medium",
                analysis="Une photo a bien été analysée, mais les indices visuels ne permettent pas encore de conclure avec un niveau de confiance suffisant.",
                recommendations="Ajoutez le nom de la culture ou de l'animal, l'ancienneté du problème et une photo plus nette pour améliorer la détection.",
                requires_expert=True,
                features=features,
                observations=observations,
                subject_profile=subject_profile,
            )
        except Exception as e:
            return {
                "disease_detected": "Erreur d'analyse",
                "confidence": 0.0,
                "symptoms": ["Erreur technique"],
                "treatment": "Veuillez décrire le problème par texte.",
                "prevention": "Réessayez avec une photo plus claire.",
                "urgency": "low",
                "analysis": f"Erreur technique lors de l'analyse: {str(e)}",
                "recommendations": "Veuillez réessayer ou décrire le problème par texte.",
                "requires_expert": True,
                "analysis_mode": "hybrid_local_image_text",
            }

    def analyze_images(self, images_data: List[bytes], text_description: str = "", category: Optional[str] = None) -> Dict[str, Any]:
        valid_images = [image for image in images_data if image][:3]
        if not valid_images:
            raise ValueError("Aucune photo exploitable fournie")

        results = [
            self.analyze_image_simple(image, text_description, category)
            for image in valid_images
        ]

        ranked = sorted(
            enumerate(results, start=1),
            key=lambda item: (
                item[1].get("confidence", 0.0),
                0 if item[1].get("requires_expert") else 1,
                len(item[1].get("visual_observations") or []),
            ),
            reverse=True,
        )
        best_view_index, best_result = ranked[0]
        aggregated = dict(best_result)

        same_disease_count = sum(
            1
            for result in results
            if result.get("disease_detected") == best_result.get("disease_detected")
        )
        if same_disease_count >= 2:
            aggregated["confidence"] = round(
                min(0.95, aggregated.get("confidence", 0.0) + 0.05),
                2,
            )

        aggregated["photo_count"] = len(valid_images)
        aggregated["best_view_index"] = best_view_index
        aggregated["analyzed_views"] = [
            {
                "view_index": index,
                "disease_detected": result.get("disease_detected"),
                "confidence": result.get("confidence"),
                "analysis": result.get("analysis"),
            }
            for index, result in enumerate(results, start=1)
        ]
        if aggregated.get("photo_count", 1) > 1:
            aggregated["analysis"] = (
                f"{aggregated.get('analysis', '')} Analyse consolidée sur {aggregated['photo_count']} vues; "
                f"la vue {best_view_index} apporte les indices les plus nets."
            ).strip()

        return aggregated


def _collect_photo_payloads(primary_photo: Optional[str], photo_list: Optional[List[str]]) -> List[str]:
    payloads: List[str] = []
    for payload in ([primary_photo] if primary_photo else []) + (photo_list or []):
        if not payload:
            continue
        if payload not in payloads:
            payloads.append(payload)
    return payloads[:3]


def _collect_audio_payloads(primary_audio: Optional[str], audio_list: Optional[List[str]]) -> List[str]:
    payloads: List[str] = []
    for payload in ([primary_audio] if primary_audio else []) + (audio_list or []):
        if not payload:
            continue
        if payload not in payloads:
            payloads.append(payload)
    return payloads[:3]


def _decode_photo_payload(photo_string: str) -> bytes:
    if "," in photo_string:
        photo_string = photo_string.split(",", 1)[1]
    return base64.b64decode(photo_string)


def _extract_data_url_mime_type(payload: str) -> Optional[str]:
    if not payload or "," not in payload:
        return None
    header = payload.split(",", 1)[0]
    match = re.match(r"data:(.*?);base64$", header, re.IGNORECASE)
    if not match:
        return None
    return match.group(1).strip().lower() or None


def _decode_base64_media_payload(payload: str) -> bytes:
    if "," in payload:
        payload = payload.split(",", 1)[1]
    return base64.b64decode(payload)


def _load_json_list(raw: Optional[str]) -> List[Any]:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def _load_json_dict(raw: Optional[str]) -> Dict[str, Any]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _build_upload_url(path: Optional[str]) -> Optional[str]:
    if not path:
        return None
    normalized_path = str(path).replace("\\", "/").lstrip("/")
    public_base = os.getenv(
        "PUBLIC_API_BASE_URL", "https://songraback.yingr-ai.com"
    ).strip().rstrip("/")
    # Une ancienne configuration de VM contenait "http//localhost" sans ':'.
    # Ne jamais propager cette URL invalide aux applications clientes.
    if not re.match(r"^https?://", public_base, re.IGNORECASE):
        public_base = "https://songraback.yingr-ai.com"
    return f"{public_base}/{normalized_path}"


def _normalize_expert_local_language(language: Optional[str]) -> str:
    normalized = str(language or "fr").strip().lower()
    normalized = "".join(
        character
        for character in unicodedata.normalize("NFD", normalized)
        if unicodedata.category(character) != "Mn"
    )
    normalized = re.sub(r"[^a-z]", "", normalized)
    if normalized in {"moore", "mooree", "more", "mossi"}:
        return "moore"
    if normalized in {"jula", "dyula"}:
        return "dioula"
    if normalized in {"fula", "fulani", "peul", "peulh"}:
        return "fulfulde"
    if normalized not in {"fr", "moore", "dioula", "fulfulde"}:
        return "fr"
    return normalized


def _ensure_parent_dir(file_path: str) -> None:
    os.makedirs(os.path.dirname(file_path), exist_ok=True)


def _ensure_audio_map_store() -> None:
    _ensure_parent_dir(EXPERT_AUDIO_MAP_PATH)
    if os.path.exists(EXPERT_AUDIO_MAP_PATH):
        return
    seed_data: Dict[str, Any] = {}
    if os.path.exists(EXPERT_AUDIO_MAP_LEGACY_SEED):
        try:
            with open(EXPERT_AUDIO_MAP_LEGACY_SEED, "r", encoding="utf-8") as handle:
                parsed = json.load(handle)
                if isinstance(parsed, dict):
                    seed_data = parsed
        except Exception:
            seed_data = {}
    with open(EXPERT_AUDIO_MAP_PATH, "w", encoding="utf-8") as handle:
        json.dump(seed_data, handle, ensure_ascii=False, indent=2)


def _load_audio_map_store() -> Dict[str, Any]:
    _ensure_audio_map_store()
    try:
        with open(EXPERT_AUDIO_MAP_PATH, "r", encoding="utf-8") as handle:
            parsed = json.load(handle)
            return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _save_audio_map_store(items: Dict[str, Any]) -> None:
    _ensure_audio_map_store()
    with open(EXPERT_AUDIO_MAP_PATH, "w", encoding="utf-8") as handle:
        json.dump(items, handle, ensure_ascii=False, indent=2)


def _normalize_audio_map_audios(raw: Any) -> Dict[str, Dict[str, Any]]:
    if not isinstance(raw, dict):
        return {}
    audios: Dict[str, Dict[str, Any]] = {}
    for language, value in raw.items():
        if not isinstance(value, dict):
            continue
        normalized_language = _normalize_expert_local_language(language)
        audios[normalized_language] = {
            "url": str(value.get("url") or "").strip(),
            "mime_type": str(value.get("mime_type") or "").strip() or None,
            "updated_at": str(value.get("updated_at") or datetime.utcnow().isoformat()),
        }
    return audios


def _sanitize_audio_map_entry(raw: Dict[str, Any], existing: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    current = existing or {}
    next_audios = {
        **_normalize_audio_map_audios(current.get("audios")),
        **_normalize_audio_map_audios(raw.get("audios")),
    }
    return {
        "label": str(raw.get("label") or current.get("label") or "").strip(),
        "action": str(raw.get("action") or current.get("action") or "").strip(),
        "screen": str(raw.get("screen") or current.get("screen") or "").strip(),
        "fallback_text": str(raw.get("fallback_text") or current.get("fallback_text") or "").strip(),
        "audio": str(raw.get("audio") or current.get("audio") or "").strip() or None,
        "mime_type": str(raw.get("mime_type") or current.get("mime_type") or "").strip() or None,
        "audios": next_audios,
        "updated_at": datetime.utcnow().isoformat(),
    }


def _normalize_expert_local_translations(raw: Any) -> Dict[str, Dict[str, Any]]:
    if not isinstance(raw, dict):
        return {}
    translations: Dict[str, Dict[str, Any]] = {}
    for language, value in raw.items():
        if not isinstance(value, dict):
            continue
        normalized_language = _normalize_expert_local_language(language)
        if normalized_language == "fr":
            continue
        translations[normalized_language] = {
            "question": str(value.get("question") or "").strip(),
            "text": str(value.get("text") or value.get("resolution") or "").strip(),
            "speech_text": str(value.get("speech_text") or "").strip() or None,
            "summary": str(value.get("summary") or "").strip(),
            "actions": value.get("actions") if isinstance(value.get("actions"), list) else [],
            "audio_url": str(value.get("audio_url") or "").strip() or None,
            "audio_mime_type": str(value.get("audio_mime_type") or "").strip() or None,
            "updated_at": str(value.get("updated_at") or datetime.utcnow().isoformat()),
        }
    return translations


def _google_translate_tts_chunks(text: str, max_len: int = 200) -> List[str]:
    """Découpe le texte en morceaux compatibles avec la limite non documentée
    (~200 caractères) de l'API Google Translate TTS, sans couper au milieu d'un mot."""
    text = (text or "").strip()
    if not text:
        return []
    if len(text) <= max_len:
        return [text]

    chunks: List[str] = []
    remaining = text
    while remaining:
        if len(remaining) <= max_len:
            chunks.append(remaining.strip())
            break
        window = remaining[:max_len]
        cut = max(window.rfind(". "), window.rfind("! "), window.rfind("? "), window.rfind(", "))
        if cut < max_len // 2:
            cut = window.rfind(" ")
        cut = cut + 1 if cut > 0 else max_len
        chunks.append(remaining[:cut].strip())
        remaining = remaining[cut:].strip()
    return [c for c in chunks if c]


def _synthesize_google_translate_tts(text: str, absolute_path: str) -> bool:
    """Synthèse vocale via Google Translate TTS (gratuit, sans clé API, HTTPS
    simple - contrairement à l'API Azure "tts.speech.microsoft.com" qui exige
    une clé d'abonnement payante, et au trick WebSocket edge-tts qui échoue
    de façon non fiable sur cet hébergement).

    Ne supporte pas les langues locales (moore/dioula/...) directement : on
    lit toujours la transcription phonétique francisée (speech_text) avec une
    voix française - c'est exactement l'usage prévu de ce champ.
    """
    import urllib.parse
    import urllib.request

    chunks = _google_translate_tts_chunks(text)
    if not chunks:
        return False

    audio_bytes = bytearray()
    try:
        for chunk in chunks:
            params = urllib.parse.urlencode({
                "ie": "UTF-8",
                "q": chunk,
                "tl": "fr",
                "client": "tw-ob",
            })
            req = urllib.request.Request(
                f"https://translate.google.com/translate_tts?{params}",
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                audio_bytes.extend(resp.read())
        if not audio_bytes:
            return False
        _ensure_parent_dir(absolute_path)
        with open(absolute_path, "wb") as f:
            f.write(bytes(audio_bytes))
        return True
    except Exception as exc:
        print(f"[WARN] Google Translate TTS echec: {exc}")
        return False


def _synthesize_local_translation_audio_bg(text: str, language: str, absolute_path: str) -> None:
    normalized_language = _normalize_expert_local_language(language)
    cleaned_text = str(text or "").strip()
    if not cleaned_text:
        return

    if _synthesize_google_translate_tts(cleaned_text[:1200], absolute_path):
        print(f"[INFO] Synthese Google Translate TTS en arriere-plan terminee pour {normalized_language} : {absolute_path}")
        return

    if not openai_client or not OPENAI_API_KEY:
        return
    try:
        response = openai_client.audio.speech.create(
            model=os.getenv("OPENAI_TTS_MODEL", "tts-1-hd"),
            voice=os.getenv("OPENAI_TTS_VOICE", "onyx"),
            input=cleaned_text[:1200],
        )
        _ensure_parent_dir(absolute_path)
        response.stream_to_file(absolute_path)
        print(f"[INFO] Synthese TTS (repli OpenAI) en arriere-plan terminee pour {normalized_language} : {absolute_path}")
    except Exception as exc:
        print(f"[WARN] Echec de synthese TTS en arriere-plan pour {normalized_language}: {exc}")


def _synthesize_local_translation_audio(
    text: str,
    language: str,
    background_tasks: Optional[BackgroundTasks] = None,
) -> Optional[Dict[str, Any]]:
    normalized_language = _normalize_expert_local_language(language)
    cleaned_text = str(text or "").strip()
    if not cleaned_text:
        return None

    digest = hashlib.sha1(f"{normalized_language}:{cleaned_text}".encode("utf-8")).hexdigest()[:20]
    file_name = f"guided-{normalized_language}-{digest}.mp3"
    relative_path = os.path.join(EXPERT_AUDIO_UPLOAD_DIR, file_name).replace("\\", "/")
    absolute_path = os.path.abspath(relative_path)
    audio_url = _build_upload_url(relative_path)

    if not os.path.exists(absolute_path):
        if background_tasks is not None:
            background_tasks.add_task(
                _synthesize_local_translation_audio_bg,
                cleaned_text,
                language,
                absolute_path,
            )
            print(f"[INFO] Synthese TTS pour {normalized_language} planifiee en arriere-plan")
        else:
            if not _synthesize_google_translate_tts(cleaned_text[:1200], absolute_path):
                if not openai_client or not OPENAI_API_KEY:
                    return None
                _ensure_parent_dir(absolute_path)
                try:
                    response = openai_client.audio.speech.create(
                        model=os.getenv("OPENAI_TTS_MODEL", "tts-1-hd"),
                        voice=os.getenv("OPENAI_TTS_VOICE", "onyx"),
                        input=cleaned_text[:1200],
                    )
                    response.stream_to_file(absolute_path)
                except Exception as exc:
                    print(f"[WARN] Audio local indisponible synchrone pour {normalized_language}: {exc}")
                    return None

    return {
        "audio_url": audio_url,
        "audio_mime_type": "audio/mpeg",
    }


def _attach_local_translation_audio(
    translations: Dict[str, Dict[str, Any]],
    background_tasks: Optional[BackgroundTasks] = None,
) -> Dict[str, Dict[str, Any]]:
    for language, payload in translations.items():
        if payload.get("audio_url"):
            continue
        text_to_speak = payload.get("speech_text") or payload.get("text") or payload.get("summary") or ""
        audio_meta = _synthesize_local_translation_audio(text_to_speak, language, background_tasks)
        if not audio_meta:
            continue
        payload.update(audio_meta)
    return translations


def _synthesize_voice_audio(
    voice_summary: Optional[Dict[str, Any]],
    target_lang: str,
    background_tasks: Optional[BackgroundTasks] = None,
) -> Optional[Dict[str, Any]]:
    """Synthetise l'audio a partir d'un resume vocal DEJA traduit.

    La traduction + le resume oral court doivent avoir ete produits en amont par
    burkina_translator.translate_fields_and_voice_summary() (UNE SEULE requete
    LLM couvrant a la fois le texte affiche et le resume vocal) : ne jamais
    refaire un appel de traduction ici, ca doublerait le nombre d'appels
    Gemini/OpenAI par reponse et ferait tomber plus vite sur le quota gratuit
    Gemini (429 Too Many Requests).
    """
    if not voice_summary or target_lang not in _TRANSLATOR_VALID_LANGS:
        return None

    speech_text = (voice_summary.get("speech_text") or voice_summary.get("summary") or "").strip()
    if not speech_text:
        return None

    audio_meta = _synthesize_local_translation_audio(speech_text, target_lang, background_tasks)
    if not audio_meta:
        return None

    return {
        "voice_summary": voice_summary.get("summary"),
        "speech_text": speech_text,
        "audio_url": audio_meta.get("audio_url"),
        "audio_mime_type": audio_meta.get("audio_mime_type"),
        "confidence": voice_summary.get("confidence"),
    }


def _translate_v2_response_with_voice(
    final_response: Dict[str, Any],
    target_lang: str,
    category: Optional[str],
) -> Tuple[Dict[str, Any], Optional[Dict[str, Any]]]:
    """Traduit un final_response v2 (module Songra) ET produit l'audio vocal,
    en UNE SEULE requete LLM (cf. translate_fields_and_voice_summary)."""
    to_translate: Dict[str, str] = {}
    for field in _TRANSLATOR_TEXT_FIELDS:
        val = final_response.get(field)
        if isinstance(val, str) and val.strip():
            to_translate[field] = val.strip()
        elif isinstance(val, list):
            joined = " | ".join(str(v) for v in val if v)
            if joined:
                to_translate[field] = joined

    if not to_translate:
        return final_response, None

    combined = translate_fields_and_voice_summary(
        to_translate, target_lang, GEMINI_API_KEY, category, voice_source_field="message",
    )
    final_response = dict(final_response)
    final_response["local_translation"] = {
        "target_lang": target_lang,
        "lang_name": _TRANSLATOR_LANG_NAMES.get(target_lang),
        "fields": combined.get("translations", {}),
    }
    voice_payload = _synthesize_voice_audio(combined.get("voice_summary"), target_lang)
    return final_response, voice_payload


def _normalize_expert_local_audio(raw: Any) -> Dict[str, Dict[str, Any]]:
    if not isinstance(raw, dict):
        return {}
    audio_map: Dict[str, Dict[str, Any]] = {}
    for language, value in raw.items():
        if not isinstance(value, dict):
            continue
        normalized_language = _normalize_expert_local_language(language)
        audio_map[normalized_language] = {
            "url": str(value.get("url") or "").strip(),
            "mime_type": str(value.get("mime_type") or "").strip() or None,
            "uploaded_at": str(value.get("uploaded_at") or datetime.utcnow().isoformat()),
        }
    return audio_map


def _normalize_expert_local_status(value: Optional[str]) -> str:
    normalized = str(value or "validated").strip().lower()
    allowed = {"validated", "resolved", "expert_verified", "pending_review"}
    return normalized if normalized in allowed else "validated"


def _normalize_expert_local_category(value: Optional[str]) -> str:
    normalized = str(value or "agriculture").strip().lower()
    normalized_ascii = "".join(
        char
        for char in unicodedata.normalize("NFD", normalized)
        if unicodedata.category(char) != "Mn"
    ).replace("-", " ").replace("_", " ")
    if "cyber" in normalized_ascii or (
        "securite" in normalized_ascii
        and any(term in normalized_ascii for term in ("numerique", "informatique", "internet"))
    ):
        return "cybersecurity"
    mapping = {
        "health": "urgence",
        "urgence": "urgence",
        "sos_accident": "urgence",
        "premiers_secours": "urgence",
        "premier_secours": "urgence",
        "premiers soins": "urgence",
        "premiers_soins": "urgence",
        "cybersecurity": "cybersecurity",
        "cybersecurite": "cybersecurity",
        "cybersécurité": "cybersecurity",
        "cyber securite": "cybersecurity",
        "cyber sécurité": "cybersecurity",
        "securite numerique": "cybersecurity",
        "sécurité numérique": "cybersecurity",
        "securite informatique": "cybersecurity",
        "sécurité informatique": "cybersecurity",
        "elevage": "elevage",
        "agriculture": "agriculture",
    }
    return mapping.get(normalized, "agriculture")


def _infer_expert_local_category(value: Optional[str], *content_values: Any) -> str:
    """Déduit la catégorie depuis le contenu quand un import est mal étiqueté.

    Les documents JSON structurés ne passent pas par le LLM. Cette détection
    empêche donc une valeur absente, générique ou erronée ``agriculture`` de
    ranger automatiquement les fiches de cybersécurité au mauvais endroit.
    """
    category = _normalize_expert_local_category(value)
    content_parts: List[str] = []
    for content in content_values:
        if isinstance(content, (list, tuple, set)):
            content_parts.extend(str(part) for part in content if part is not None)
        elif content is not None:
            content_parts.append(str(content))
    normalized_content = "".join(
        char
        for char in unicodedata.normalize("NFD", " ".join(content_parts).lower())
        if unicodedata.category(char) != "Mn"
    )
    cyber_markers = (
        "cyber", "hameconnage", "phishing", "arnaque", "fraude", "pirat",
        "mot de passe", "code otp", "code secret", "mobile money",
        "orange money", "moov money", "sim swap", "carte sim", "whatsapp",
        "facebook", "compte vole", "compte bloque", "lien suspect",
        "sms suspect", "application suspecte", "malware", "rancongiciel",
        "securite numerique", "securite informatique",
    )
    if any(marker in normalized_content for marker in cyber_markers):
        return "cybersecurity"
    return category


def _serialize_expert_local_knowledge_item(item: ExpertLocalKnowledgeDB) -> Dict[str, Any]:
    return {
        "id": item.id,
        "title": item.title,
        "category": item.category,
        "question_fr": item.question_fr,
        "resolution_fr": item.resolution_fr,
        "tags": _load_json_list(item.tags_json),
        "status": item.status,
        "origin": item.origin,
        "translations": _load_json_dict(item.translations_json),
        # Les anciennes fiches peuvent contenir des clés comme "Mooré".
        # Le scanner consomme toujours les codes canoniques moore/dioula/fulfulde.
        "audio": _normalize_expert_local_audio(_load_json_dict(item.audio_json)),
        "expert_id": getattr(item, "expert_id", None),
        "reviewer_id": getattr(item, "reviewer_id", None),
        "review_notes": getattr(item, "review_notes", None),
        "reviewed_at": item.reviewed_at.isoformat() if getattr(item, "reviewed_at", None) else None,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


def _upsert_expert_local_knowledge_from_legacy_seed(db: Session) -> int:
    if not os.path.exists(EXPERT_LOCAL_KNOWLEDGE_LEGACY_SEED):
        return 0
    if db.query(ExpertLocalKnowledgeDB).count() > 0:
        return 0
    try:
        with open(EXPERT_LOCAL_KNOWLEDGE_LEGACY_SEED, "r", encoding="utf-8") as handle:
            parsed = json.load(handle)
    except Exception:
        return 0
    if not isinstance(parsed, list):
        return 0

    inserted = 0
    for raw in parsed:
        if not isinstance(raw, dict):
            continue
        question_fr = str(raw.get("question_fr") or raw.get("question") or "").strip()
        resolution_fr = str(raw.get("resolution_fr") or raw.get("answer") or "").strip()
        if not question_fr or not resolution_fr:
            continue
        db.add(
            ExpertLocalKnowledgeDB(
                title=str(raw.get("title") or question_fr[:120]).strip(),
                category=_normalize_expert_local_category(raw.get("category")),
                question_fr=question_fr,
                resolution_fr=resolution_fr,
                tags_json=json.dumps(raw.get("tags") or [], ensure_ascii=False),
                status=_normalize_expert_local_status(raw.get("status")),
                origin=str(raw.get("origin") or "expert_manual"),
                translations_json=json.dumps(
                    _normalize_expert_local_translations(raw.get("translations")),
                    ensure_ascii=False,
                ),
                audio_json=json.dumps(
                    _normalize_expert_local_audio(raw.get("audio")),
                    ensure_ascii=False,
                ),
            )
        )
        inserted += 1

    if inserted > 0:
        db.commit()
    return inserted


def _extract_json_object_from_text(raw_text: str) -> Optional[str]:
    cleaned = str(raw_text or "").replace("```json", "").replace("```", "").strip()
    start = cleaned.find("{")
    if start < 0:
        return None

    depth = 0
    in_string = False
    escape = False

    for index in range(start, len(cleaned)):
        char = cleaned[index]
        if in_string:
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            continue

        if char == '"':
            in_string = True
        elif char == '{':
            depth += 1
        elif char == '}':
            depth -= 1
            if depth == 0:
                return cleaned[start : index + 1]

    return None


def _parse_json_object_from_text(raw_text: str) -> Dict[str, Any]:
    cleaned = str(raw_text or "").replace("```json", "").replace("```", "").strip()
    payload = _extract_json_object_from_text(cleaned) or cleaned
    parsed = json.loads(payload)
    if not isinstance(parsed, dict):
        raise ValueError("Réponse JSON invalide")
    return parsed


def _build_local_translation_prompt(question_fr: str, resolution_fr: str, category: str, actions_fr: Optional[List[str]] = None) -> str:
    actions_str = ""
    if actions_fr:
        actions_str = "\nACTIONS À TRADUIRE:\n" + "\n".join([f"- {a}" for a in actions_fr])

    return (
        "Tu es un traducteur expert et natif des trois langues nationales prises en charge par Songra (Mooré, Dioula et Fulfuldé).\n"
        "Ta mission est de traduire une fiche de conseil agricole/santé/sécurité pour des producteurs ruraux.\n"
        "Pour ce faire, mobilise activement les dictionnaires bilingues Songra et les règles grammaticales du Mooré, du Dioula et du Fulfuldé afin d'obtenir la traduction la plus précise possible.\n"
        "CRITÈRES DE HAUTE QUALITÉ (OBJECTIF 90%+ DE FIDÉLITÉ NATURELLE ET PRÉCISION GRAMMATICALE) :\n"
        "1. **Règles Grammaticales, Conjugaison & Base de Données Externe** : Appuie-toi sur les règles académiques officielles et ton savoir encyclopédique de la morphologie et de la syntaxe du Mooré, Dioula et Fulfuldé. Utilise des expressions idiomatiques authentiques du Burkina Faso plutôt que des calques mot-à-mot du français. Assure-toi que les termes techniques (médicaux, agricoles, de sécurité) soient traduits par leur équivalent culturel exact.\n"
        "2. **Pas de traduction littérale** : Ne traduis SURTOUT PAS mot-à-mot (pas de traduction littérale). Adapte le sens en utilisant les expressions et termes les plus naturels possibles en langue locale sans altérer le sens original.\n"
        "3. **Gestion des mots difficiles / Synonymes** : Si un mot spécifique (terme technique, moderne ou peu usité) n'a pas de traduction littérale directe reconnue, utilise des synonymes, des paraphrases ou des équivalents imagés naturels en langue locale plutôt que de le laisser en français ou d'inventer un calque artificiel.\n"
        "4. **Ton de Prononciation & Clarté** : Le ton doit être CHALEUREUX, RASSURANT, CONSEILLER et ORAL (adapté à l'écoute par des personnes analphabètes).\n"
        "5. **Double format de texte** :\n"
        "   - `text` : La traduction textuelle officielle, écrite selon l'orthographe correcte standardisée de la langue (avec les caractères spéciaux appropriés si nécessaires).\n"
        "   - `speech_text` : Une transcription phonétique francophone simplifiée spécifiquement optimisée pour la synthèse vocale. Utilisez impérativement un découpage syllabique et phonétique en séparant les syllabes complexes par des tirets '-' ou des espaces (ex: pour le Mooré, écris 'ou-ain-dé' au lieu de 'wẽndé', 'yee-kee' pour 'yiki', 'koa-a-da' pour 'koaada'). Insère des virgules ou points pour forcer des pauses naturelles et respecter la tonalité burkinabè, favorisant une lecture à voix très lente et claire.\n"
        "\n"
        "Structure JSON stricte attendue (retourne uniquement ce JSON, pas de markdown, pas de ```json) :\n"
        "{\n"
        '  "moore": {\n'
        '    "question": "Traduction de la question (orthographe standard)",\n'
        '    "text": "Traduction de la résolution (orthographe standard)",\n'
        '    "speech_text": "Traduction de la résolution adaptée phonétiquement avec découpage syllabique (ex: ou-ain-dé, ko-no-ko) pour la lecture TTS en français",\n'
        '    "summary": "Résumé court (orthographe standard)",\n'
        '    "actions": ["Action 1 traduite (standard)", "Action 2 traduite (standard)", ...]\n'
        '  },\n'
        '  "dioula": {\n'
        '    "question": "...", "text": "...", "speech_text": "...", "summary": "...", "actions": [...] \n'
        '  },\n'
        '  "fulfulde": {\n'
        '    "question": "...", "text": "...", "speech_text": "...", "summary": "...", "actions": [...]\n'
        '  }\n'
        "}\n"
        "Contraintes : pas de markdown, pas d'explication.\n"
        f"Catégorie : {category}\n"
        f"Question FR : {question_fr}\n"
        f"Résolution FR : {resolution_fr}"
        f"{actions_str}"
    )

def _build_entreprendre_translation_prompt(data: dict) -> str:
    # On extrait les parties textuelles clés pour la traduction
    propositions = "\n".join([f"- {p.get('titre')}: {p.get('description')}" for p in data.get('propositions', [])])
    calendrier = "\n".join([f"- {c.get('mois')}: {c.get('activite')} ({c.get('details')})" for c in data.get('calendrier_cultural', [])])
    gestion_eau = f"Techniques: {', '.join(data.get('gestion_eau', {}).get('techniques', []))}. Conseils: {data.get('gestion_eau', {}).get('conseils', '')}"
    
    return f"""Tu es un traducteur expert émérite et natif en Mooré, Dioula et Fulfuldé au Burkina Faso.
TA MISSION : Traduire les points clés d'un plan d'exploitation pour un producteur local.

EXIGENCE DE QUALITÉ (OBJECTIF 90%+ DE PROXIMITÉ NATURELLE ET PHONÉTIQUE DE LECTURE) :
1. Respecte rigoureusement les règles grammaticales, de conjugaison et d'accord de chaque langue locale du Burkina Faso. Évite les calques du français.
2. **Pas de traduction littérale** : Ne traduis SURTOUT PAS mot-à-mot (pas de traduction littérale). Adapte le sens en utilisant les expressions et termes les plus naturels et usuels possibles en langue locale sans altérer le sens original.
3. **Gestion des mots difficiles / Synonymes** : Si un terme n'a pas d'équivalent direct, utilise des synonymes ou des périphrases imagées naturelles en langue locale.
4. Pour le champ `speech_text` : Rédige un résumé vocal continu de tout le plan d'exploitation (synthèse du terrain, des propositions, du découpage, du calendrier et de la gestion de l'eau) dans la langue locale, mais écrit sous forme de transcription phonétique francophone simplifiée, avec un découpage syllabique et phonétique très clair en séparant les syllabes complexes par des tirets '-' ou des espaces légers (ex: 'ou-ain-dé' pour Mooré wẽndé, 'yee-kee' pour yiki, 'koa-a-da' pour koaada). Cela permet à un moteur de synthèse vocale (TTS) français de le lire à voix très lente, claire et décomposée pour être parfaitement intelligible en zone rurale.
5. Utilise le ton de prononciation approprié (conseiller, motivateur, respectueux du savoir paysan).
6. Adapte le vocabulaire technique agricole pour qu'il soit immédiatement compris par un locuteur natif rural.

DONNÉES À TRADUIRE :
1. Terrain : {data.get('description_terrain')}
2. Propositions : {propositions}
3. Découpage : {data.get('decoupage_terrain')}
4. Calendrier : {calendrier}
5. Gestion Eau : {gestion_eau}

RETOURNE UNIQUEMENT un objet JSON avec cette structure :
{{
  "moore": {{
    "description_terrain": "...",
    "propositions": ["titre: description", ...],
    "decoupage_terrain": "...",
    "calendrier": ["mois: activite", ...],
    "gestion_eau": "...",
    "speech_text": "Plan d'exploitation complet phonétisé avec découpage syllabique séparé par des tirets"
  }},
  "dioula": {{
    "description_terrain": "...",
    "propositions": ["..."],
    "decoupage_terrain": "...",
    "calendrier": ["..."],
    "gestion_eau": "...",
    "speech_text": "..."
  }},
  "fulfulde": {{
    "description_terrain": "...",
    "propositions": ["..."],
    "decoupage_terrain": "...",
    "calendrier": ["..."],
    "gestion_eau": "...",
    "speech_text": "..."
  }}
}}

RETOURNE SEULEMENT LE JSON PUR."""


def _generate_local_translations_with_gemini(prompt: str) -> Dict[str, Any]:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY non configuree")
    model = genai.GenerativeModel("gemini-2.5-flash")
    result = model.generate_content(prompt)
    return _parse_json_object_from_text(result.text)


def _generate_local_translations_with_openai(prompt: str) -> Dict[str, Any]:
    if not openai_client or not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY non configuree")
    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": prompt,
            }
        ],
        temperature=0.2,
        max_tokens=2500,
    )
    content = response.choices[0].message.content if response.choices else ""
    return _parse_json_object_from_text(content or "")


def _generate_local_translations(
    question_fr: str,
    resolution_fr: str,
    category: str,
    background_tasks: Optional[BackgroundTasks] = None,
    actions_fr: Optional[List[str]] = None,
) -> Dict[str, Dict[str, Any]]:
    # Une passe indépendante par langue permet d'injecter le dictionnaire
    # agricole correspondant et évite qu'un modèle mélange trois langues dans
    # une même réponse JSON.
    source_fields = {
        "question": question_fr,
        "text": resolution_fr,
    }
    if actions_fr:
        source_fields["actions"] = " | ".join(str(action) for action in actions_fr if action)

    normalized: Dict[str, Dict[str, Any]] = {}
    for language in ("moore", "dioula", "fulfulde"):
        translated = translate_fields(
            source_fields,
            language,
            GEMINI_API_KEY,
            category=category,
        )
        question_result = translated.get("question", {})
        text_result = translated.get("text", {})
        actions_result = translated.get("actions", {})
        action_text = str(actions_result.get("translation") or "").strip()
        normalized[language] = {
            "question": question_result.get("translation", question_fr),
            "text": text_result.get("translation", resolution_fr),
            "speech_text": text_result.get("speech_text", text_result.get("translation", resolution_fr)),
            "summary": text_result.get("translation", resolution_fr),
            "actions": [part.strip() for part in action_text.split("|") if part.strip()],
            "updated_at": datetime.utcnow().isoformat(),
        }

    if not normalized:
        raise HTTPException(status_code=502, detail="La traduction locale renvoyée est vide")
    return _attach_local_translation_audio(normalized, background_tasks)


def _find_studio_knowledge_match(
    db: Session,
    *,
    category: str,
    query_text: str,
    photo_analysis: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    """Retourne la fiche validée du Studio la plus proche du texte/diagnostic."""
    normalized_category = _normalize_expert_local_category(category)
    candidates = (
        db.query(ExpertLocalKnowledgeDB)
        .filter(
            ExpertLocalKnowledgeDB.category == normalized_category,
            ExpertLocalKnowledgeDB.status.in_(["validated", "resolved", "expert_verified"]),
        )
        .order_by(ExpertLocalKnowledgeDB.updated_at.desc())
        .limit(1000)
        .all()
    )
    analysis = photo_analysis or {}
    search_parts = [
        query_text,
        str(analysis.get("problem_label") or ""),
        str(analysis.get("disease_detected") or ""),
        str(analysis.get("diagnosis") or ""),
        # Contrat du pipeline V2 (v2_services._validate_analysis).
        str(analysis.get("diagnostic") or ""),
        str(analysis.get("description_visuelle") or ""),
        " ".join(str(item) for item in (analysis.get("causes_probables") or [])),
        " ".join(str(item) for item in (analysis.get("actions_immediates") or [])),
        " ".join(str(item) for item in (analysis.get("actions_detaillees") or [])),
        str(analysis.get("situation_type") or ""),
        str(analysis.get("threat_type") or ""),
        str(analysis.get("analysis") or ""),
    ]
    search_blob = " ".join(part for part in search_parts if part)
    search_tokens = set(_tokenize(search_blob))
    # Les modèles Vision peuvent employer les noms internationaux même quand
    # la fiche Studio est rédigée en français. Ces équivalences ne traduisent
    # pas la réponse : elles servent uniquement au rapprochement des fiches.
    normalized_blob = _normalize_search_text(search_blob)
    studio_match_aliases = {
        "armyworm": "chenille legionnaire automne",
        "fall armyworm": "chenille legionnaire automne",
        "corn": "mais",
        "maize": "mais",
        "pest": "ravageur",
        "pests": "ravageurs",
        "caterpillar": "chenille",
    }
    for source_term, studio_terms in studio_match_aliases.items():
        if source_term in normalized_blob:
            search_tokens.update(_tokenize(studio_terms))
    if not search_tokens:
        return None

    best: Optional[Tuple[float, ExpertLocalKnowledgeDB]] = None
    normalized_problem = _normalize_search_text(
        str(analysis.get("problem_label") or analysis.get("disease_detected") or query_text)
    )
    for item in candidates:
        tags = [str(tag) for tag in _load_json_list(item.tags_json)]
        score = (
            4.0 * len(search_tokens & set(_tokenize(item.title or "")))
            + 3.5 * len(search_tokens & set(_tokenize(" ".join(tags))))
            + 2.0 * len(search_tokens & set(_tokenize(item.question_fr or "")))
            + 0.25 * len(search_tokens & set(_tokenize(item.resolution_fr or "")))
        )
        normalized_title = _normalize_search_text(item.title or "")
        if normalized_title and normalized_problem and (
            normalized_title in normalized_problem or normalized_problem in normalized_title
        ):
            score += 15.0
        if best is None or score > best[0]:
            best = (score, item)

    if best is not None:
        print(
            f"[STUDIO-MATCH] diagnostic='{str(analysis.get('diagnostic') or analysis.get('disease_detected') or '')[:120]}' "
            f"meilleure_fiche=#{best[1].id} '{best[1].title}' score={round(best[0], 2)}"
        )
    if best is None or best[0] < 4.0:
        return None
    result = _serialize_expert_local_knowledge_item(best[1])
    result["match_score"] = round(best[0], 2)
    result["source"] = "studio_connaissances"
    return result


def _apply_studio_match_to_v2_response(
    final_response: Dict[str, Any],
    studio_match: Dict[str, Any],
    target_lang: Optional[str],
) -> Dict[str, Any]:
    """Remplace la réponse générique par la fiche Studio et ses vraies voix."""
    result = dict(final_response)
    result["message"] = studio_match["resolution_fr"]
    diagnostic = dict(result.get("diagnostic") or {})
    diagnostic["description"] = studio_match["title"]
    diagnostic["type"] = studio_match["category"]
    # Une fois la fiche validée choisie, ne plus mélanger ses informations
    # avec les causes/actions générées auparavant par le modèle Vision.
    diagnostic["causes"] = []
    diagnostic["description_visuelle"] = None
    result["diagnostic"] = diagnostic
    result["actions"] = []
    result["knowledge_mode"] = "studio_knowledge"
    result["knowledge_card"] = studio_match

    audio_map = _normalize_expert_local_audio(studio_match.get("audio") or {})
    translations = studio_match.get("translations") or {}
    localizations: Dict[str, Any] = dict(result.get("localizations") or {})
    french_audio = audio_map.get("fr") if isinstance(audio_map.get("fr"), dict) else {}
    localizations["fr"] = {
        "question": studio_match["title"],
        "text": studio_match["resolution_fr"],
        "speech_text": studio_match["resolution_fr"],
        "audio_url": french_audio.get("url"),
        "audio_mime_type": french_audio.get("mime_type"),
    }

    for language in _TRANSLATOR_VALID_LANGS:
        translated = translations.get(language) if isinstance(translations.get(language), dict) else {}
        local_audio = audio_map.get(language) if isinstance(audio_map.get(language), dict) else {}
        translated_text = str(translated.get("text") or "").strip()
        if translated_text or local_audio.get("url"):
            localizations[language] = {
                "question": str(translated.get("question") or studio_match["title"]),
                "text": translated_text or studio_match["resolution_fr"],
                "speech_text": str(translated.get("speech_text") or translated_text),
                "summary": str(translated.get("summary") or ""),
                "audio_url": local_audio.get("url"),
                "audio_mime_type": local_audio.get("mime_type"),
            }

    if target_lang in _TRANSLATOR_VALID_LANGS:
        local_audio = audio_map.get(target_lang) if isinstance(audio_map.get(target_lang), dict) else {}
        has_local_audio = bool(str(local_audio.get("url") or "").strip())
        result["local_audio_available"] = has_local_audio
        result["local_audio_message"] = None if has_local_audio else (
            "La voix dédiée à cette fiche est indisponible dans cette langue."
        )
        result["french_fallback_available"] = True
        result["target_lang"] = target_lang
        result["lang_name"] = _TRANSLATOR_LANG_NAMES.get(target_lang)
        result["audio_url"] = local_audio.get("url") if has_local_audio else None
        result["audio_mime_type"] = local_audio.get("mime_type") if has_local_audio else None
        available_audio_languages = sorted(
            language
            for language, audio_data in audio_map.items()
            if isinstance(audio_data, dict) and str(audio_data.get("url") or "").strip()
        )
        print(
            f"[STUDIO-AUDIO] fiche=#{studio_match.get('id')} langue_demandee={target_lang} "
            f"langues_disponibles={available_audio_languages} voix_trouvee={has_local_audio} "
            f"url={str(local_audio.get('url') or '')[:180]}"
        )
    result["localizations"] = localizations
    return result


def _history_sort_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _store_photo_payloads(owner_id: int, photo_data_list: List[bytes], prefix: str = "photo") -> List[str]:
    stored_paths: List[str] = []
    timestamp = int(datetime.utcnow().timestamp())
    for index, photo_data in enumerate(photo_data_list, start=1):
        filename = f"{owner_id}_{timestamp}_{prefix}_{index}.jpg"
        path = f"uploads/{filename}"
        with open(path, "wb") as f:
            f.write(photo_data)
        stored_paths.append(path)
    return stored_paths


def _guess_audio_extension_from_payload(payload: str) -> str:
    mime_type = _extract_data_url_mime_type(payload)
    mapping = {
        "audio/webm": ".webm",
        "audio/mp4": ".m4a",
        "audio/x-m4a": ".m4a",
        "audio/mpeg": ".mp3",
        "audio/mp3": ".mp3",
        "audio/wav": ".wav",
        "audio/x-wav": ".wav",
        "audio/ogg": ".ogg",
        "audio/aac": ".aac",
    }
    if mime_type in mapping:
        return mapping[mime_type]
    guessed = mimetypes.guess_extension(mime_type or "")
    return guessed or ".webm"


def _store_audio_payloads(owner_id: int, audio_payloads: List[str], prefix: str = "audio") -> List[str]:
    stored_paths: List[str] = []
    timestamp = int(datetime.utcnow().timestamp())
    for index, payload in enumerate(audio_payloads, start=1):
        audio_bytes = _decode_base64_media_payload(payload)
        extension = _guess_audio_extension_from_payload(payload)
        filename = f"{owner_id}_{timestamp}_{prefix}_{index}{extension}"
        path = os.path.join(COMMUNITY_AUDIO_UPLOAD_DIR, filename).replace("\\", "/")
        with open(path, "wb") as handle:
            handle.write(audio_bytes)
        stored_paths.append(path)
    return stored_paths


def _serialize_community_audio_urls(raw_paths_json: Optional[str]) -> List[Dict[str, Any]]:
    items = _load_json_list(raw_paths_json)
    serialized: List[Dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        relative_path = item.get("path") if isinstance(item, dict) else item
        if not relative_path:
            continue
        serialized.append(
            {
                "url": _build_upload_url(str(relative_path)),
                "mime_type": mimetypes.guess_type(str(relative_path))[0] or "audio/webm",
                "label": f"Vocal {index}",
            }
        )
    return serialized


def _build_community_case_title(title: str, description: str, room: str) -> str:
    cleaned_title = str(title or "").strip()
    if cleaned_title:
        return cleaned_title[:180]
    normalized_description = re.sub(r"\s+", " ", str(description or "")).strip()
    if normalized_description:
        return normalized_description[:180]
    room_label = COMMUNITY_ROOM_LABELS.get(_normalize_community_room(room), "terrain")
    return f"Signalement {room_label.lower()}"


def _serialize_photo_history_record(record: PhotoAnalysisHistoryDB) -> Dict[str, Any]:
    photo_paths = _load_json_list(record.photo_paths_json)
    photo_labels = _load_json_list(record.photo_labels_json)
    photos = [
        {
            "url": _build_upload_url(path),
            "label": photo_labels[index] if index < len(photo_labels) else f"Vue {index + 1}",
        }
        for index, path in enumerate(photo_paths)
        if path
    ]

    analysis: Dict[str, Any] = {}
    if record.analysis_json:
        try:
            parsed = json.loads(record.analysis_json)
            if isinstance(parsed, dict):
                analysis = parsed
        except Exception:
            analysis = {}

    return {
        "id": record.client_record_id or str(record.id),
        "server_id": record.id,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "category": record.category or "agriculture",
        "prompt": record.prompt or "",
        "analysis": analysis,
        "photos": photos,
        "source_ticket_id": record.source_ticket_id,
    }


def _normalize_offline_domain(category: Optional[str]) -> str:
    normalized = (category or "").lower().strip()
    if normalized in ("elevage", "élevage"):
        return "elevage"
    if normalized in ("urgence", "sos_accident", "sos", "health"):
        return "health"
    if normalized == "cybersecurity":
        return "cybersecurity"
    return "agriculture"


def _normalize_search_text(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text or "")
    normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return normalized.lower()


def _clean_assistant_text(value: Any) -> str:
    """Transforme le Markdown LLM en texte propre pour écran et synthèse vocale."""
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r"```(?:[a-zA-Z0-9_+-]+)?\s*", "", text)
    text = text.replace("```", "").replace("`", "")
    text = re.sub(r"!\[([^\]]*)\]\([^)]*\)", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", text)
    text = re.sub(r"(?m)^\s{0,3}#{1,6}\s*", "", text)
    text = re.sub(r"(?m)^\s*>\s?", "", text)
    text = re.sub(r"(?m)^\s*[-+*]\s+", "", text)
    text = re.sub(r"(?m)^\s*\d+[.)]\s+", "", text)
    text = re.sub(r"(\*\*|__)(.*?)\1", r"\2", text)
    text = re.sub(r"(?<!\w)[*_~]+|[*_~]+(?!\w)", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _extract_offline_keywords(*values: Any, limit: int = 18) -> List[str]:
    stopwords = {
        "avec", "dans", "pour", "sans", "mais", "dont", "cette", "cette", "votre", "leurs",
        "vous", "nous", "elles", "comme", "plus", "tres", "trop", "etre", "fait", "faire",
        "faut", "cela", "alors", "apres", "avant", "entre", "sous", "chez", "vers", "aussi",
        "cest", "cette", "quand", "quoi", "parce", "dans", "pourquoi", "comment", "the", "and",
    }
    ordered: List[str] = []
    seen = set()
    for value in values:
        if value is None:
            continue
        if isinstance(value, list):
            iterable = value
        else:
            iterable = [value]
        for item in iterable:
            text = _normalize_search_text(str(item))
            for token in re.findall(r"[a-z0-9_]{3,}", text):
                if token in stopwords or token in seen:
                    continue
                seen.add(token)
                ordered.append(token)
                if len(ordered) >= limit:
                    return ordered
    return ordered


def _compact_response_for_offline(payload: Dict[str, Any], source_kind: Optional[str] = None) -> Dict[str, Any]:
    try:
        compact = json.loads(json.dumps(payload, ensure_ascii=False))
    except Exception:
        return {}

    fields_to_clear = [
        "video_base64",
        "image_url",
    ]

    if source_kind == "entreprendre":
        fields_to_clear = [
            "video_base64",
            "image_url",
        ]

    for field in fields_to_clear:
        if field in compact:
            compact[field] = None
    return compact


def _build_entrepreneurship_offline_answer(payload: Dict[str, Any]) -> str:
    parts: List[str] = []
    if payload.get("description_terrain"):
        parts.append(f"Terrain: {payload['description_terrain']}")

    propositions = payload.get("propositions") or []
    if propositions:
        snippets = []
        for item in propositions[:3]:
            if not isinstance(item, dict):
                continue
            titre = item.get("titre") or "Projet"
            description = item.get("description") or ""
            investissement = item.get("investissement") or ""
            revenu = item.get("revenu_estime") or ""
            snippets.append(
                f"{titre}: {description} Investissement {investissement}. Revenu estime {revenu}."
            )
        if snippets:
            parts.append("Propositions: " + " ".join(snippets))

    if payload.get("decoupage_terrain"):
        parts.append(f"Decoupage: {payload['decoupage_terrain']}")

    gestion_eau = payload.get("gestion_eau") or {}
    if isinstance(gestion_eau, dict) and gestion_eau.get("conseils"):
        parts.append(f"Eau: {gestion_eau['conseils']}")

    risques = payload.get("risques") or []
    if risques:
        parts.append("Risques: " + "; ".join(str(risk) for risk in risques[:3]))

    return "\n\n".join(part for part in parts if part).strip()


def _build_offline_entry_payload(
    *,
    source_kind: str,
    category: str,
    question_text: str,
    response_payload: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    domain = _normalize_offline_domain(category)
    clean_question = (question_text or "").strip()
    compact_payload = _compact_response_for_offline(response_payload, source_kind=source_kind)

    if source_kind == "entreprendre":
        answer = _build_entrepreneurship_offline_answer(compact_payload)
        propositions = compact_payload.get("propositions") or []
        first_title = ""
        if propositions and isinstance(propositions[0], dict):
            first_title = str(propositions[0].get("titre") or "").strip()
        title = first_title or (compact_payload.get("description_terrain") or "Plan terrain Songra")
        keyword_inputs = [
            category,
            clean_question,
            compact_payload.get("description_terrain"),
            [item.get("titre") for item in propositions if isinstance(item, dict)],
            compact_payload.get("decoupage_terrain"),
            compact_payload.get("risques"),
        ]
    else:
        diagnostic = compact_payload.get("diagnostic") or {}
        if not isinstance(diagnostic, dict):
            diagnostic = {}
        answer = str(compact_payload.get("message") or diagnostic.get("description") or "").strip()
        title = str(
            diagnostic.get("description")
            or clean_question
            or f"Conseil Songra {domain}"
        ).strip()
        keyword_inputs = [
            category,
            clean_question,
            title,
            answer,
            diagnostic.get("type"),
            diagnostic.get("causes"),
            [action.get("texte") for action in (compact_payload.get("actions") or []) if isinstance(action, dict)],
        ]

    if not answer:
        return None

    title = title[:180] if title else f"Conseil Songra {domain}"
    tags = _extract_offline_keywords(*keyword_inputs)
    return {
        "domain": domain,
        "title": title,
        "question": clean_question or None,
        "answer": answer,
        "tags": tags,
        "response_json": compact_payload,
    }


def _persist_offline_knowledge_entry(
    *,
    db: Session,
    user_id: Optional[int],
    source_kind: str,
    category: str,
    question_text: str,
    response_payload: Dict[str, Any],
) -> None:
    payload = _build_offline_entry_payload(
        source_kind=source_kind,
        category=category,
        question_text=question_text,
        response_payload=response_payload,
    )
    if not payload:
        return

    fingerprint_source = "|".join(
        [
            payload["domain"],
            source_kind,
            _normalize_search_text(payload.get("question") or payload["title"]),
            _normalize_search_text(payload["answer"][:400]),
        ]
    )
    fingerprint = hashlib.sha1(fingerprint_source.encode("utf-8")).hexdigest()

    existing = (
        db.query(OfflineKnowledgeEntryDB)
        .filter(OfflineKnowledgeEntryDB.fingerprint == fingerprint)
        .first()
    )

    if existing is None:
        existing = OfflineKnowledgeEntryDB(
            fingerprint=fingerprint,
            user_id=user_id,
            source_kind=source_kind,
            domain=payload["domain"],
            title=payload["title"],
            question=payload.get("question"),
            answer=payload["answer"],
            tags_json=json.dumps(payload["tags"], ensure_ascii=False),
            response_json=json.dumps(payload["response_json"], ensure_ascii=False),
        )
        db.add(existing)
    else:
        existing.user_id = user_id or existing.user_id
        existing.source_kind = source_kind
        existing.domain = payload["domain"]
        existing.title = payload["title"]
        existing.question = payload.get("question")
        existing.answer = payload["answer"]
        existing.tags_json = json.dumps(payload["tags"], ensure_ascii=False)
        existing.response_json = json.dumps(payload["response_json"], ensure_ascii=False)
        existing.updated_at = datetime.utcnow()

    db.commit()


def _serialize_offline_entry_for_rag(entry: OfflineKnowledgeEntryDB) -> Dict[str, Any]:
    response_payload = _parse_offline_response_json(entry)
    return {
        "id": f"generated-{entry.id}",
        "domain": entry.domain,
        "title": entry.title,
        "question": entry.question,
        "answer": entry.answer,
        "tags": _load_json_list(entry.tags_json),
        "source": f"generated:{entry.source_kind}",
        "response_json": response_payload,
        "source_kind": entry.source_kind,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
    }


def _trusted_shared_rag_source_kinds() -> List[str]:
    # Garde-fou anti-boucle: seules les réponses validées par résolution humaine
    # sont réinjectées dans le RAG partagé général.
    return ["resolved_ticket"]


def _trusted_shared_reuse_source_kinds() -> List[str]:
    # Réutilisation directe en ligne: limiter aux cas validés humainement.
    return ["resolved_ticket"]


def _find_reusable_offline_entry(
    db: Session,
    *,
    domain: str,
    source_kinds: List[str],
    question_text: str,
    limit: int = 120,
) -> Optional[OfflineKnowledgeEntryDB]:
    normalized_question = _normalize_search_text(question_text)
    if not normalized_question:
        return None

    question_tokens = set(_tokenize(question_text))
    candidates = (
        db.query(OfflineKnowledgeEntryDB)
        .filter(
            OfflineKnowledgeEntryDB.domain == domain,
            OfflineKnowledgeEntryDB.source_kind.in_(source_kinds),
        )
        .order_by(OfflineKnowledgeEntryDB.updated_at.desc(), OfflineKnowledgeEntryDB.id.desc())
        .limit(max(1, min(limit, 300)))
        .all()
    )

    best_entry: Optional[OfflineKnowledgeEntryDB] = None
    best_score = 0.0

    for entry in candidates:
        candidate_text = entry.question or entry.title or ""
        normalized_candidate = _normalize_search_text(candidate_text)
        if not normalized_candidate:
            continue

        score = 0.0
        if normalized_candidate == normalized_question:
            score += 100.0
        elif normalized_question in normalized_candidate or normalized_candidate in normalized_question:
            score += 60.0

        candidate_tokens = set(_tokenize(
            f"{entry.title or ''} {entry.question or ''} {entry.answer or ''} {' '.join(_load_json_list(entry.tags_json))}"
        ))
        overlap = len(question_tokens & candidate_tokens)
        if overlap:
            score += overlap * 6.0
            score += overlap / max(len(question_tokens), 1)

        if score > best_score:
            best_score = score
            best_entry = entry

    if best_entry is None:
        return None

    if best_score >= 100.0:
        return best_entry

    min_overlap_score = max(12.0, min(24.0, len(question_tokens) * 4.0))
    return best_entry if best_score >= min_overlap_score else None


def _find_previously_answered_question(
    db: Session, *, domain: str, question_text: str
) -> Optional[OfflineKnowledgeEntryDB]:
    """Réutilise d'abord une question identique, puis un cas expert proche.

    Les réponses IA ordinaires ne sont reprises que si la question normalisée
    est identique. Les tickets résolus par un humain gardent la recherche
    sémantique plus souple existante.
    """
    normalized_question = _normalize_search_text(question_text).strip()
    if not normalized_question:
        return None

    recent = (
        db.query(OfflineKnowledgeEntryDB)
        .filter(
            OfflineKnowledgeEntryDB.domain == domain,
            OfflineKnowledgeEntryDB.source_kind.in_([
                "assistant_query", "resolved_ticket"
            ]),
        )
        .order_by(
            OfflineKnowledgeEntryDB.updated_at.desc(),
            OfflineKnowledgeEntryDB.id.desc(),
        )
        .limit(300)
        .all()
    )
    for entry in recent:
        if _normalize_search_text(entry.question or "").strip() == normalized_question:
            return entry

    return _find_reusable_offline_entry(
        db,
        domain=domain,
        source_kinds=_trusted_shared_reuse_source_kinds(),
        question_text=question_text,
    )


def _build_media_cache_question(diagnostic: str, steps: List[str]) -> str:
    clean_steps = [step.strip() for step in steps if step and step.strip()]
    if clean_steps:
        return f"{diagnostic.strip()}\n" + "\n".join(clean_steps[:5])
    return diagnostic.strip()


def _build_media_offline_payload(
    *,
    diagnostic: str,
    steps: List[str],
    image_result: Optional[Dict[str, Any]] = None,
    video_result: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "message": diagnostic,
        "diagnostic": {
            "description": diagnostic,
            "type": "illustration",
            "causes": steps[:5],
        },
        "steps": steps[:5],
    }

    if image_result:
        payload.update(
            {
                "image_base64": image_result.get("image_base64"),
                "image_mime_type": image_result.get("mime_type") or image_result.get("image_mime_type"),
                "image_description": image_result.get("fallback_description")
                or image_result.get("image_description")
                or diagnostic,
            }
        )

    if video_result:
        payload.update(
            {
                "video_base64": video_result.get("video_base64"),
                "video_url": video_result.get("video_url"),
                "video_mime_type": video_result.get("mime_type") or video_result.get("video_mime_type"),
                "video_duration": video_result.get("duration_sec") or video_result.get("video_duration"),
                "video_description": video_result.get("video_description") or diagnostic,
                "video_steps": video_result.get("steps_visuelles") or video_result.get("video_steps") or steps[:5],
            }
        )
        if video_result.get("type") == "image_steps":
            payload["steps"] = video_result.get("steps") or payload["steps"]

    return payload


def _serialize_offline_generated_entry(entry: OfflineKnowledgeEntryDB) -> Dict[str, Any]:
    tags = _load_json_list(entry.tags_json)
    response_payload = _parse_offline_response_json(entry)
    keywords = _extract_offline_keywords(tags, entry.title, entry.question, entry.answer)
    
    # Check media presence status before stripping
    has_image = bool(response_payload.get("image_base64") or response_payload.get("imageBase64"))
    has_video = bool(response_payload.get("video_url") or response_payload.get("video_base64") or response_payload.get("videoBase64") or response_payload.get("video_steps"))
    
    # Strip massive base64 strings to prevent client OutOfMemoryError
    keys_to_strip = ["image_base64", "imageBase64", "video_base64", "videoBase64", "input_photo_base64"]
    for key in keys_to_strip:
        if key in response_payload:
            del response_payload[key]
        if "diagnostic" in response_payload and isinstance(response_payload["diagnostic"], dict):
            if key in response_payload["diagnostic"]:
                del response_payload["diagnostic"][key]

    return {
        "id": f"generated-{entry.id}",
        "domain": entry.domain,
        "title": entry.title,
        "question": entry.question,
        "answer": entry.answer,
        "tags": tags,
        "keywords": keywords,
        "language": "fr",
        "source": f"generated:{entry.source_kind}",
        "source_kind": entry.source_kind,
        "has_image": has_image,
        "has_video": has_video,
        "image_description": response_payload.get("image_description"),
        "video_description": response_payload.get("video_description"),
        "video_steps": response_payload.get("video_steps"),
        "response_json": response_payload,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
    }


def _parse_offline_response_json(entry: OfflineKnowledgeEntryDB) -> Dict[str, Any]:
    if not entry.response_json:
        return {}
    try:
        parsed = json.loads(entry.response_json)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _serialize_entreprendre_history_entry(entry: OfflineKnowledgeEntryDB) -> Dict[str, Any]:
    response_payload = _parse_offline_response_json(entry)
    return {
        "id": f"entreprendre-{entry.id}",
        "history_type": "entreprendre",
        "category": entry.domain,
        "urgency": "low",
        "status": "shared",
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
        "last_message": entry.title or entry.answer,
        "has_photo": bool(response_payload.get("input_photo_base64")),
        "photo_url": response_payload.get("input_photo_base64"),
        "photo_urls": [response_payload.get("input_photo_base64")] if response_payload.get("input_photo_base64") else [],
        "problem": entry.question,
        "result": response_payload,
    }


def _serialize_v2_history_entry(entry: OfflineKnowledgeEntryDB) -> Dict[str, Any]:
    response_payload = _parse_offline_response_json(entry)
    diagnostic = response_payload.get("diagnostic") or {}
    if not isinstance(diagnostic, dict):
        diagnostic = {}

    domain_to_category = {
        "agriculture": "agriculture",
        "elevage": "elevage",
        "health": "sos_accident",
        "urgence": "sos_accident",
        "cybersecurity": "cybersecurity",
    }
    category = domain_to_category.get(entry.domain, entry.domain or "agriculture")

    return {
        "id": f"consultation-{entry.id}",
        "history_type": "consultation",
        "category": category,
        "urgency": "high" if response_payload.get("urgence") else ("medium" if response_payload.get("priorite") == 2 else "low"),
        "status": "shared",
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
        "last_message": diagnostic.get("description") or response_payload.get("message") or entry.title,
        "has_photo": bool(response_payload.get("input_photo_base64")),
        "photo_url": response_payload.get("input_photo_base64"),
        "photo_urls": [response_payload.get("input_photo_base64")] if response_payload.get("input_photo_base64") else [],
        "problem": entry.question,
        "v2_response": response_payload,
        "result": response_payload,
    }


def _serialize_generated_media_history_entry(entry: OfflineKnowledgeEntryDB) -> Dict[str, Any]:
    response_payload = _parse_offline_response_json(entry)
    diagnostic = response_payload.get("diagnostic") or {}
    if not isinstance(diagnostic, dict):
        diagnostic = {}

    domain_to_category = {
        "agriculture": "agriculture",
        "elevage": "elevage",
        "health": "sos_accident",
        "urgence": "sos_accident",
        "cybersecurity": "cybersecurity",
    }
    category = domain_to_category.get(entry.domain, entry.domain or "agriculture")

    return {
        "id": f"media-{entry.id}",
        "history_type": "media",
        "category": category,
        "urgency": "low",
        "status": "shared",
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
        "last_message": diagnostic.get("description") or response_payload.get("message") or entry.title,
        "problem": entry.question,
        "source_kind": entry.source_kind,
        "result": response_payload,
    }


def _build_offline_cache_payload(
    db: Session,
    *,
    domain: Optional[str] = None,
    language: str = "fr",
    limit: int = 100,
) -> Dict[str, Any]:
    safe_limit = max(1, min(limit, 500))
    normalized_domain = _normalize_offline_domain(domain) if domain else None

    knowledge_query = db.query(KnowledgeItem).filter(KnowledgeItem.language == language)
    generated_query = db.query(OfflineKnowledgeEntryDB)

    if normalized_domain:
        knowledge_query = knowledge_query.filter(KnowledgeItem.domain == normalized_domain)
        generated_query = generated_query.filter(OfflineKnowledgeEntryDB.domain == normalized_domain)

    knowledge_items = knowledge_query.order_by(KnowledgeItem.updated_at.desc()).limit(safe_limit).all()
    generated_items = generated_query.order_by(OfflineKnowledgeEntryDB.updated_at.desc()).limit(min(safe_limit, 250)).all()
    
    # Intégrer le corpus offline validé (ExpertLocalKnowledgeDB)
    expert_query = db.query(ExpertLocalKnowledgeDB)
    if normalized_domain:
        expert_query = expert_query.filter(ExpertLocalKnowledgeDB.category == normalized_domain)
    expert_items = expert_query.order_by(ExpertLocalKnowledgeDB.updated_at.desc()).limit(100).all()

    serialized_items: List[Dict[str, Any]] = []
    dedupe_keys = set()

    for item in generated_items:
        serialized = _serialize_offline_generated_entry(item)
        dedupe_key = (
            serialized.get("domain"),
            _normalize_search_text(serialized.get("question") or ""),
            _normalize_search_text(serialized.get("answer") or ""),
        )
        if dedupe_key in dedupe_keys:
            continue
        dedupe_keys.add(dedupe_key)
        serialized_items.append(serialized)

    for item in knowledge_items:
        serialized = {
            "id": item.id,
            "domain": item.domain,
            "title": item.title,
            "question": item.question,
            "answer": item.answer,
            "tags": _load_json_list(item.tags),
            "keywords": _extract_offline_keywords(item.title, item.question, item.answer, _load_json_list(item.tags)),
            "language": item.language or "fr",
            "source": item.source or "knowledge_base",
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        }
        dedupe_key = (
            serialized.get("domain"),
            _normalize_search_text(serialized.get("title") or ""),
            _normalize_search_text(serialized.get("answer") or ""),
        )
        if dedupe_key not in dedupe_keys:
            dedupe_keys.add(dedupe_key)
            serialized_items.append(serialized)

    for item in expert_items:
        # On utilise le sérialiseur existant pour récupérer translations et audio
        item_data = _serialize_expert_local_knowledge_item(item)
        serialized = {
            "id": f"expert_{item.id}",
            "domain": item.category,
            "title": item.title,
            "question": item.question_fr,
            "answer": item.resolution_fr,
            "tags": _load_json_list(item.tags_json),
            "keywords": _extract_offline_keywords(item.title, item.question_fr, item.resolution_fr, _load_json_list(item.tags_json)),
            "language": "fr", # Base est fr, mais contient des translations
            "translations": item_data.get("translations"),
            "audio": item_data.get("audio"),
            "source": "expert_validated",
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        }
        dedupe_key = (
            serialized.get("domain"),
            _normalize_search_text(serialized.get("title") or ""),
            _normalize_search_text(serialized.get("answer") or ""),
        )
        if dedupe_key in dedupe_keys:
            continue
        dedupe_keys.add(dedupe_key)
        serialized_items.append(serialized)

    return {
        "items": serialized_items,
        "total": len(serialized_items),
        "cached_at": datetime.utcnow().isoformat(),
    }


class GPTVisionEngine:
    """Analyse d'images via GPT-4 Vision API (OpenAI)
    
    Remplace le Computer Vision local pour bénéficier des capacités
    d'analyse avancée de ChatGPT avec sa propre connaissance.
    """
    
    def __init__(self, openai_client, model: str = "gpt-4o", provider_name: str = "OpenAI"):
        self.client = openai_client
        self.model = model
        self.provider_name = provider_name
    
    def analyze_images(self, images_data: List[bytes], text_description: str = "", category: Optional[str] = None) -> Dict[str, Any]:
        """Analyser les images via GPT-4 Vision API"""
        valid_images = [image for image in images_data if image][:3]
        if not valid_images:
            raise ValueError("Aucune photo exploitable fournie")
        
        try:
            # Convertir les images en base64
            images_base64 = [base64.b64encode(img).decode('utf-8') for img in valid_images]
            
            # Créer le prompt contextuel
            context_prompt = ""
            if category == "agriculture":
                context_prompt = """TÂCHE: Analyser une photo agricole pour identifier les maladies des cultures.

INSTRUCTIONS IMPORTANTES:
- Analysez la photo avec ATTENTION aux détails
- Identifiez chaque culture visible
- Détectez TOUTE maladie, ravageur ou problème visible
- Si AUCUN problème: écrivez "Aucune maladie détectée"
- Donnez une confiance entre 0.0 (aucune certitude) et 1.0 (certitude totale)
- IMPORTANT: Répondez UNIQUEMENT avec du JSON valide, rien d'autre

FORMAT JSON REQUIS (copiez-collez et remplissez):
{
    "disease_detected": "Maladie identifiée ou 'Aucune'",
    "confidence": 0.85,
    "symptoms": ["Symptôme 1", "Symptôme 2"],
    "treatment": "Action recommandée",
    "urgency": "low|medium|high",
    "prevents": "Prévention",
    "visual_observations": ["Détail observé"],
    "analysis": "Explication détaillée en français"
}"""
            elif category == "elevage":
                context_prompt = """TÂCHE: Analyser une photo d'animal pour identifier les maladies et problèmes de santé.

INSTRUCTIONS IMPORTANTES:
- Analysez la photo avec ATTENTION
- Identifiez l'espèce et l'état de l'animal
- Détectez TOUTE maladie, blessure ou anomalie
- Si AUCUN problème visible: écrivez "Aucun"
- Donnez une confiance entre 0.0 et 1.0
- IMPORTANT: Répondez UNIQUEMENT avec du JSON valide, rien d'autre

FORMAT JSON REQUIS (copiez-collez et remplissez):
{
    "disease_detected": "Maladie ou 'Aucun'",
    "confidence": 0.85,
    "symptoms": ["Symptôme visible"],
    "treatment": "Action de traitement/aide",
    "urgency": "low|medium|high",
    "prevents": "Prévention future",
    "visual_observations": ["Observation"],
    "analysis": "Analyse détaillée en français"
}"""
            else:
                context_prompt = """TÂCHE: Analyser cette image pour identifier tout problème de santé/maladie.

INSTRUCTIONS:
- Analysez soigneusement
- Identifiez problèmes visibles
- Donnez confiance 0.0 à 1.0
- RÉPONDEZ UNIQUEMENT EN JSON, PAS DE TEXTE SUPPLÉMENTAIRE

FORMAT JSON:
{
    "disease_detected": "Problème ou 'Aucun'",
    "confidence": 0.85,
    "symptoms": ["Symptôme"],
    "treatment": "Recommandation",
    "urgency": "low|medium|high",
    "prevents": "Prévention",
    "visual_observations": ["Observation"],
    "analysis": "Analyse en français"
}"""
            
            context_prompt += f"""

CONTROLE OBLIGATOIRE DE LA PHOTO (catégorie demandée: {category or 'non précisée'}):
- Identifiez d'abord le type réel de contenu; ne supposez pas qu'il correspond à la catégorie.
- Si la photo est hors catégorie, floue ou inexploitable, ne fabriquez aucun diagnostic.
- Ajoutez obligatoirement au JSON: image_category (agriculture|elevage|sos_accident|cybersecurity|other|unclear), category_match (booléen), image_usable (booléen), problem_identified (booléen), problem_status (identified|not_identified|uncertain|wrong_category|unusable), problem_label et validation_message.
- problem_identified doit être false lorsqu'aucun signe anormal n'est réellement visible.
"""

            # Construire le message pour GPT-4 Vision
            content = [
                {
                    "type": "text",
                    "text": context_prompt + (f"\n\nContext supplémentaire: {text_description}" if text_description else "")
                }
            ]
            
            # Ajouter les images
            for idx, img_b64 in enumerate(images_base64):
                print(f"📸 Image {idx + 1}: {len(img_b64)} bytes ({len(img_b64)/1024:.1f}KB) - {category} mode")
                content.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/jpeg;base64,{img_b64}"
                    }
                })
            
            print(f"📤 Envoi à {self.provider_name}/{self.model}: {len(content)} éléments (1 texte + {len(images_base64)} images)")

            # Appeler le modèle avec vision.
            # NOTE: certains modèles Groq (ex: qwen/qwen3.6-27b) sont des
            # modèles "thinking" qui émettent un bloc <think>...</think> de
            # raisonnement AVANT la réponse finale. Avec un max_tokens trop
            # bas, ce raisonnement consommait tout le budget et la réponse
            # JSON n'était jamais générée (réponse tronquée en plein <think>).
            # reasoning_format="hidden" demande à Groq de ne renvoyer que la
            # réponse finale (sans le raisonnement) dans le contenu du
            # message, et on augmente max_tokens pour laisser la place au
            # raisonnement interne + au JSON complet.
            extra_kwargs: Dict[str, Any] = {}
            if self.provider_name == "Groq":
                extra_kwargs["extra_body"] = {
                    "reasoning_effort": "none",
                    "reasoning_format": "hidden",
                }
                extra_kwargs["response_format"] = {"type": "json_object"}

            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "user",
                        "content": content
                    }
                ],
                max_tokens=6000,
                temperature=0.1,
                **extra_kwargs,
            )

            print(f"✅ Réponse {self.provider_name} Vision reçue")

            # Parser la réponse
            raw_response_text = response.choices[0].message.content or ""
            response_text = raw_response_text
            # Filet de sécurité : si le modèle a quand même renvoyé un bloc
            # <think>...</think> malgré reasoning_format="hidden" (ou pour
            # d'autres providers), on l'ignore pour ne parser que la partie
            # utile de la réponse.
            response_text = re.sub(r"<think>.*?</think>", "", response_text, flags=re.DOTALL).strip()
            print(f"📝 Réponse brute {self.provider_name}: {response_text[:300]}...")
            
            # Extraire le JSON de la réponse
            analysis_json = None
            try:
                # Chercher le JSON dans la réponse
                json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response_text, re.DOTALL)
                if json_match:
                    json_str = json_match.group()
                    analysis_json = json.loads(json_str)
                    print(f"✅ JSON parsé avec succès: {analysis_json.get('disease_detected', 'N/A')}")
                else:
                    print(f"[WARN] Pas de JSON trouvé dans la réponse")
                    
            except json.JSONDecodeError as je:
                print(f"[WARN] JSON parsing error: {je}")
                # Essayer de nettoyer et re-parser
                try:
                    # Remplacer characteres problématiques
                    cleaned = response_text.replace('\n', ' ').replace('  ', ' ')
                    json_match = re.search(r'\{.*\}', cleaned, re.DOTALL)
                    if json_match:
                        analysis_json = json.loads(json_match.group())
                        print(f"✅ JSON parsé après nettoyage: {analysis_json.get('disease_detected', 'N/A')}")
                except:
                    print(f"❌ Impossible de parser JSON après nettoyage")
            
            # Si toujours pas de JSON valide, créer response par défaut
            if not analysis_json:
                # Essayer d'extraire quelques mots-clés utiles
                disease_keywords = ['maladie', 'malade', 'blessure', 'infection', 'aucun', 'aucune', 'normal', 'sain']
                detected = 'Non identifiée'
                for keyword in disease_keywords:
                    if keyword in response_text.lower():
                        detected = 'Détecté' if keyword not in ['aucun', 'aucune', 'normal', 'sain'] else 'Aucune maladie'
                        break
                
                print(f"[WARN] Utilisant response par défaut avec keyword matching")
                analysis_json = {
                    "disease_detected": detected,
                    "confidence": 0.4,
                    # Conserver le raisonnement visuel utile pour la recherche
                    # Studio (culture, ravageur, symptômes). L'ancienne limite
                    # de 300 caractères supprimait souvent "armyworm" ou
                    # "chenille légionnaire" avant le matching des fiches.
                    "analysis": (response_text or raw_response_text)[:6000]
                    if (response_text or raw_response_text)
                    else "Analyse incomplète",
                    "urgency": "medium",
                    "requires_expert": False
                }
            
            # Enrichir avec métadonnées
            analysis_json["photo_count"] = len(valid_images)
            analysis_json["best_view_index"] = 1
            analysis_json["analyzed_views"] = [
                {
                    "view_index": i + 1,
                    "disease_detected": analysis_json.get("disease_detected"),
                    "confidence": analysis_json.get("confidence", 0.5)
                }
                for i in range(len(valid_images))
            ]
            analysis_json["requires_expert"] = analysis_json.get("urgency") == "high" or analysis_json.get("confidence", 0.5) < 0.6
            analysis_json["model"] = self.model
            analysis_json["vision_provider"] = self.provider_name.lower()
            
            return analysis_json
            
        except Exception as e:
            print(f"❌ Erreur GPT-4o Vision: {e}")
            import traceback
            print(traceback.format_exc())
            return {
                "disease_detected": "Erreur analyse",
                "confidence": 0,
                "analysis": f"Erreur lors de l'analyse: {str(e)}",
                "urgency": "medium",
                "requires_expert": True,
                "photo_count": len(valid_images),
                "error": str(e)
            }


def _normalize_photo_analysis_status(
    analysis: Dict[str, Any], requested_category: Optional[str]
) -> Dict[str, Any]:
    """Normalise le contrôle catégorie et l'état du problème pour tous les moteurs."""
    result = dict(analysis or {})
    requested = (requested_category or "agriculture").strip().lower()
    aliases = {
        "health": "sos_accident",
        "urgence": "sos_accident",
        "animal": "elevage",
        "veterinaire": "elevage",
        "vétérinaire": "elevage",
        "plant": "agriculture",
        "plante": "agriculture",
    }
    requested = aliases.get(requested, requested)

    raw_category = str(result.get("image_category") or "").strip().lower()
    image_category = aliases.get(raw_category, raw_category)
    if not image_category:
        consultation = _normalize_search_text(str(result.get("consultation_type") or ""))
        diagnosis_type = str(result.get("diagnosis_type") or "").strip().lower()
        if result.get("culture_detected") or "agricol" in consultation:
            image_category = "agriculture"
        elif result.get("animal_species") or "veterinaire" in consultation:
            image_category = "elevage"
        elif result.get("situation_type") or "premiers secours" in consultation:
            image_category = "sos_accident"
        elif result.get("threat_type") or "cyber" in consultation:
            image_category = "cybersecurity"
        elif diagnosis_type == "plant_disease_stress":
            image_category = "agriculture"
        elif diagnosis_type == "animal_health_injury":
            image_category = "elevage"
        elif diagnosis_type == "human_first_aid":
            image_category = "sos_accident"
        else:
            # Les moteurs OpenAI/Groq (utilisés par défaut, cf AI_PROVIDER)
            # ne renvoient pas les champs spécifiques à Gemini vérifiés
            # ci-dessus (culture_detected, animal_species, ...) et oublient
            # souvent le champ optionnel "image_category" demandé en fin de
            # prompt. Sans ce filet, une photo correctement diagnostiquée
            # (ex: feuille malade avec disease_detected/analysis renseignés)
            # retombait à tort sur "unclear" simplement parce que ce champ
            # de contrôle manquait — alors que l'analyse elle-même était
            # bonne. On fait donc confiance à la catégorie demandée dès lors
            # que le moteur a produit un vrai contenu de diagnostic.
            has_diagnostic_content = bool(
                str(result.get("disease_detected") or "").strip()
                or result.get("symptoms")
                or result.get("all_symptoms")
                or result.get("visual_observations")
                or str(result.get("analysis") or "").strip()
            ) and not result.get("error")
            image_category = requested if has_diagnostic_content else "unclear"

    image_usable = result.get("image_usable")
    if not isinstance(image_usable, bool):
        image_usable = not bool(result.get("error")) and float(result.get("confidence") or 0) > 0

    category_match = result.get("category_match")
    if not isinstance(category_match, bool):
        category_match = image_category == requested if image_category != "unclear" else False

    diagnosis = str(
        result.get("problem_label")
        or result.get("disease_detected")
        or result.get("threat_type")
        or result.get("situation_type")
        or ""
    ).strip()
    normalized_diagnosis = _normalize_search_text(diagnosis)
    no_problem_markers = (
        "aucun", "aucune", "pas de probleme", "non identifie", "indetermine",
        "image inexploitable", "erreur analyse",
    )
    explicit_problem = result.get("problem_identified")
    if isinstance(explicit_problem, bool):
        problem_identified = explicit_problem
    elif isinstance(result.get("threat_detected"), bool):
        problem_identified = bool(result.get("threat_detected"))
    elif not image_usable or not category_match:
        problem_identified = False
    else:
        problem_identified = bool(diagnosis) and not any(
            marker in normalized_diagnosis for marker in no_problem_markers
        )

    if not image_usable:
        status = "unusable"
    elif not category_match:
        status = "wrong_category"
    elif problem_identified:
        status = "identified"
    elif any(marker in normalized_diagnosis for marker in ("aucun", "aucune", "pas de probleme")):
        status = "not_identified"
    else:
        status = "uncertain"

    messages = {
        "unusable": "La photo n'est pas assez exploitable pour réaliser une analyse fiable.",
        "wrong_category": f"La photo semble appartenir à la catégorie {image_category}, pas à {requested}.",
        "identified": f"Un problème probable a été identifié: {diagnosis}.",
        "not_identified": "Aucun problème visible n'a été identifié sur cette photo.",
        "uncertain": "La photo correspond à la catégorie, mais aucun problème précis ne peut être confirmé.",
    }
    result.update({
        "requested_category": requested,
        "image_category": image_category,
        "category_match": category_match,
        "image_usable": image_usable,
        "problem_identified": problem_identified,
        "problem_status": status,
        "problem_label": diagnosis or "Aucun problème confirmé",
        "validation_message": str(result.get("validation_message") or messages[status]),
    })
    return result


class ResilientVisionEngine:
    """Moteur de vision résilient avec redirection dynamique et bascule automatique.
    Sélectionne la clé préférée selon AI_PROVIDER et bascule sur l'autre en cas d'erreur de facturation/dunning (403/429/etc.).
    """
    def __init__(
        self,
        gemini_key: Optional[str],
        openai_client: Optional[Any],
        groq_client: Optional[Any],
    ):
        self.gemini_engine = GeminiVisionEngine(gemini_key) if (gemini_key and GeminiVisionEngine) else None
        self.groq_engine = GPTVisionEngine(
            groq_client,
            model=os.getenv("GROQ_VISION_MODEL", "qwen/qwen3.6-27b"),
            provider_name="Groq",
        ) if groq_client else None
        self.gpt_engine = GPTVisionEngine(openai_client) if openai_client else None
        self.local_engine = LocalComputerVision()

    def analyze_images(self, images_data: List[bytes], text_description: str = "", category: Optional[str] = None) -> Dict[str, Any]:
        provider = os.getenv("AI_PROVIDER", "openai").lower()
        engines = []
        if provider == "groq":
            if self.groq_engine:
                engines.append(("Groq Vision", self.groq_engine))
        elif provider == "openai":
            if self.gpt_engine:
                engines.append(("OpenAI (GPT-4o)", self.gpt_engine))
            if self.groq_engine:
                engines.append(("Groq Vision", self.groq_engine))
            if self.gemini_engine:
                engines.append(("Gemini", self.gemini_engine))
        else:
            if self.gemini_engine:
                engines.append(("Gemini", self.gemini_engine))
            if self.groq_engine:
                engines.append(("Groq Vision", self.groq_engine))
            if self.gpt_engine:
                engines.append(("OpenAI (GPT-4o)", self.gpt_engine))
        
        engines.append(("Local Computer Vision", self.local_engine))

        last_error = None
        for name, engine in engines:
            try:
                print(f"[VISION-RESILIENT] Tentative d'analyse avec {name}...")
                res = engine.analyze_images(images_data, text_description, category)
                # Si l'analyse retourne un dictionnaire signalant une erreur de clé ou de dunning, on lève une exception pour forcer le fallback
                if isinstance(res, dict):
                    err_msg = str(res.get("analysis", "")).lower() + " " + str(res.get("error", "")).lower()
                    if res.get("disease_detected") == "Erreur analyse" or "dunning" in err_msg or "403" in err_msg or "permission" in err_msg:
                        raise RuntimeError(f"Analyse invalide ou erreur d'API retournée par {name} : {res.get('analysis')}")
                return _normalize_photo_analysis_status(res, category)
            except Exception as e:
                print(f"[VISION-RESILIENT] Échec avec {name} : {e}")
                last_error = e
        
        # En cas d'échec total (très improbable)
        return {
            "disease_detected": "Erreur globale",
            "confidence": 0.0,
            "analysis": f"Toutes les méthodes de vision ont échoué. Dernière erreur : {last_error}",
            "urgency": "medium",
            "requires_expert": True
        }


cv_engine = ResilientVisionEngine(GEMINI_API_KEY, openai_client, groq_client)

# ==========================================
# TRADUCTION LOCALE — LANGUES DU BURKINA FASO
# ==========================================

_LOCAL_LANG_NAMES = {
    "moore":    "Mooré",
    "dioula":   "Dioula",
    "fulfulde": "Fulfuldé",
}

# Champs du diagnostic à traduire (clés JSON retournées par le scanner)
_ANALYSIS_FIELDS_TO_TRANSLATE = [
    "what_i_see",
    "disease_detected",
    "analysis",
    "detailed_analysis",
    "treatment",
    "treatment_local",
    "treatment_chemical",
    "prevention",
    "urgency_message",
    "recommendations",
    "consultation_type",
    "symptoms_observed",
]


def translate_analysis_to_local_lang(analysis: Dict[str, Any], target_lang: str) -> Dict[str, Any]:
    """Traduit les champs textuels d'un diagnostic vers une langue locale burkinabè.

    Args:
        analysis  : dict retourné par cv_engine.analyze_images()
        target_lang : code langue ("moore", "dioula", "fulfulde")

    Returns:
        Copie du dict d'analyse avec un sous-objet 'local_translation' contenant
        chaque champ traduit dans la langue cible, plus les métadonnées phonétiques.
    """
    lang_name = _LOCAL_LANG_NAMES.get(target_lang)
    if not lang_name:
        return analysis

    # Collecter les champs textuels non vides à traduire
    fields_to_translate: Dict[str, str] = {}
    for field in _ANALYSIS_FIELDS_TO_TRANSLATE:
        value = analysis.get(field)
        if isinstance(value, str) and value.strip():
            fields_to_translate[field] = value.strip()
        elif isinstance(value, list):
            # listes de strings (ex: recommendations)
            joined = " | ".join([str(v) for v in value if v])
            if joined:
                fields_to_translate[field] = joined

    if not fields_to_translate:
        return analysis

    # Utiliser le moteur centralisé : dictionnaire validé d'abord, puis le
    # fournisseur configuré (Groq/OpenAI/Gemini). Cela évite que cet ancien
    # endpoint contourne Groq et appelle Gemini directement.
    translated_fields = translate_fields(
        fields_to_translate,
        target_lang,
        GEMINI_API_KEY,
        category=str(analysis.get("type_probleme") or "agriculture"),
    )
    analysis_copy = dict(analysis)
    analysis_copy["local_translation"] = {
        "translations": {
            field: {
                "text": result.get("translation", fields_to_translate.get(field, "")),
                "phonetic": result.get("speech_text", result.get("translation", "")),
                "vocal_writing": result.get("speech_text", result.get("translation", "")),
                "confidence": result.get("confidence", 0.0),
                "source": result.get("source", "unknown"),
            }
            for field, result in translated_fields.items()
        },
        "target_lang": target_lang,
        "lang_name": lang_name,
        "confidence": min(
            [float(item.get("confidence", 0.0)) for item in translated_fields.values()] or [0.0]
        ),
    }
    return analysis_copy

    # Ancienne implémentation Gemini conservée temporairement comme référence
    # mais rendue inaccessible par le retour centralisé ci-dessus.
    fields_json = json.dumps(fields_to_translate, ensure_ascii=False, indent=2)

    prompt = f"""Vous êtes un linguiste expert en langues locales du Burkina Faso.
Votre tâche est de traduire les champs d'un diagnostic médical / agricole / vétérinaire
du Français vers la langue locale : {lang_name} (code: {target_lang}).

NORMES DE TRANSCRIPTION (Commission Nationale des Langues du Burkina Faso) :
- Utilisez les caractères officiels : 'ɛ' et 'ɔ' lorsque requis.
- Nasalisation : insérer 'n' immédiatement après la voyelle nasalisée.
- Longueur vocalique : doublez la voyelle longue (ex: 'ee', 'oo').
- Emprunts : adaptez phonologiquement les termes techniques sans traduction directe.
- Tons : respectez les intonations dans l'écriture phonétique.

CONSIGNES STRICTES :
1. Traduisez chaque champ de manière fluide et naturelle (pas mot-à-mot).
2. Pour les termes médicaux/agricoles sans équivalent local, adaptez phonologiquement
   ou utilisez une périphrase explicative dans la langue cible.
3. Renvoyez UNIQUEMENT du JSON valide, sans aucune explication extérieure.
4. Pour chaque champ traduit, ajoutez aussi sa transcription phonétique (lecture française).

Champs à traduire (JSON) :
{fields_json}

Format de réponse JSON strict :
{{
  "translations": {{
    "<nom_champ>": {{
      "text": "<traduction dans {lang_name}>",
      "phonetic": "<transcription phonétique pour lecture française>",
      "vocal_writing": "<écriture syllabique avec tirets pour TTS>"
    }}
  }},
  "target_lang": "{target_lang}",
  "lang_name": "{lang_name}",
  "confidence": <score 0.0-1.0>
}}"""

    import urllib.request as _urllib_req
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}"
    payload_req = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "responseMimeType": "application/json",
            "temperature": 0.1
        }
    }
    try:
        req = _urllib_req.Request(
            url,
            data=json.dumps(payload_req).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        with _urllib_req.urlopen(req, timeout=15) as resp:
            resp_data = json.loads(resp.read().decode("utf-8"))
            text_res = resp_data["candidates"][0]["content"]["parts"][0]["text"]
            translation_result = json.loads(text_res.strip())

        # Fusionner le résultat de traduction dans l'analyse
        analysis_copy = dict(analysis)
        analysis_copy["local_translation"] = translation_result
        return analysis_copy

    except Exception as e:
        print(f"[TRADUCTION] Échec traduction vers {lang_name}: {e}")
        # En cas d'échec, on renvoie l'analyse originale sans traduction
        analysis_copy = dict(analysis)
        analysis_copy["local_translation"] = {
            "error": f"Traduction vers {lang_name} échouée : {str(e)}",
            "target_lang": target_lang,
            "lang_name": lang_name
        }
        return analysis_copy

# ==========================================
# MODULE IA TEXTE (NLP Local) - RESTAURÉ
# ==========================================

class AITriageEngine:
    def __init__(self):
        self.urgency_keywords = {
            "high": ["urgence", "urgent", "grave", "danger", "sang", "brûlure", "piraté", 
                    "volé", "mort", "mourir", "pourrir", "invasion", "attaque"],
            "medium": ["problème", "aide", "rapidement", "besoin", "important", "malade"],
            "low": ["conseil", "information", "question", "quand", "comment", "préventif"]
        }
        
        self.category_keywords = {
            "agriculture": ["maïs", "sorgho", "mil", "culture", "plante", "champ", "récolte", 
                          "irrigation", "tomate", "oignon", "arachide", "coton",
                          "manioc", "riz", "feuille", "insecte", "parasite", "engrais"],
            # Catégorie élevage : animaux, bétail, poules…
            "elevage": ["bétail", "vache", "boeuf", "chèvre", "mouton", "poules", "volaille",
                        "lapin", "lapereau", "clapier",
                        "agneau", "veau", "animal", "troupeau", "abri", "vermifuge", "parasites"],
            # Catégorie SOS Accident / premiers soins : on garde le domaine health pour le RAG
            "sos_accident": ["blessure", "coupure", "accident", "saigne", "sang", "brûlure",
                              "tomber", "chute", "fracture", "douleur", "secours"],
            "cybersecurity": ["arnaque", "pirate", "mobile money", "code", "mot de passe", 
                            "orange money", "sms suspect", "compte", "fraude", "virus"]
        }
    
    def classify(self, text: str):
        text_lower = text.lower()
        
        # Catégorie
        category_scores = {}
        for cat, keywords in self.category_keywords.items():
            score = sum(1 for kw in keywords if kw in text_lower)
            category_scores[cat] = score
        
        category = max(category_scores, key=category_scores.get) if any(category_scores.values()) else "agriculture"
        confidence = category_scores[category] / (len(self.category_keywords[category]) + 1)
        
        # Urgence
        urgency = "low"
        for level, keywords in self.urgency_keywords.items():
            if any(kw in text_lower for kw in keywords):
                urgency = level
                break
        
        keywords = [word for word in text_lower.split() if len(word) > 3][:5]
        
        return {
            "category": category,
            "urgency": urgency,
            "confidence": float(confidence),
            "keywords": keywords
        }

ai_engine = AITriageEngine()

FOCUS_SUBJECTS: Dict[str, List[Dict[str, Any]]] = {
    "agriculture": [
        {"label": "Maïs", "aliases": ["maïs", "mais"]},
        {"label": "Tomate", "aliases": ["tomate", "tomates"]},
        {"label": "Manioc", "aliases": ["manioc", "bouture de manioc"]},
        {"label": "Sorgho", "aliases": ["sorgho", "sorghos"]},
        {"label": "Oignon", "aliases": ["oignon", "oignons"]},
        {"label": "Arachide", "aliases": ["arachide", "arachides", "cacahuète", "cacahuete"]},
        {"label": "Riz", "aliases": ["riz", "rizière", "riziere"]},
        {"label": "Mil", "aliases": ["mil"]},
        {"label": "Coton", "aliases": ["coton"]},
    ],
    "elevage": [
        {"label": "Lapin", "aliases": ["lapin", "lapins", "lapereau", "lapereaux", "clapier"]},
        {"label": "Volaille", "aliases": ["volaille", "poule", "poules", "poulet", "poulets", "coq", "canard"]},
        {"label": "Vache", "aliases": ["vache", "vaches", "boeuf", "boeufs", "bovin", "bovins", "veau", "veaux"]},
        {"label": "Chèvre", "aliases": ["chèvre", "chevre", "chèvres", "chevres"]},
        {"label": "Mouton", "aliases": ["mouton", "moutons", "brebis", "agneau", "agneaux"]},
        {"label": "Porc", "aliases": ["porc", "porcs", "cochon", "cochons"]},
    ],
    "sos_accident": [
        {"label": "Main", "aliases": ["main", "mains", "doigt", "doigts", "paume"]},
        {"label": "Bras", "aliases": ["bras", "coude"]},
        {"label": "Jambe", "aliases": ["jambe", "jambes", "genou", "cuisse"]},
        {"label": "Pied", "aliases": ["pied", "pieds", "orteil", "orteils"]},
        {"label": "Œil", "aliases": ["oeil", "œil", "yeux"]},
        {"label": "Peau", "aliases": ["peau", "visage", "tête", "tete"]},
    ],
    "cybersecurity": [
        {"label": "Orange Money", "aliases": ["orange money", "orangemoney"]},
        {"label": "Moov Money", "aliases": ["moov money", "moovmoney"]},
        {"label": "Mobile Money", "aliases": ["mobile money", "wallet"]},
        {"label": "WhatsApp", "aliases": ["whatsapp", "whats app"]},
        {"label": "Facebook", "aliases": ["facebook", "fb"]},
        {"label": "Téléphone", "aliases": ["telephone", "téléphone", "portable", "smartphone"]},
        {"label": "Carte SIM", "aliases": ["carte sim", "sim"]},
    ],
}

FOCUS_ISSUES: Dict[str, List[Dict[str, Any]]] = {
    "agriculture": [
        {"label": "Taches jaunes", "aliases": ["taches jaunes", "tache jaune", "jaunissement", "jaune"]},
        {"label": "Rouille", "aliases": ["rouille", "pustule orange", "pustules orange"]},
        {"label": "Mildiou", "aliases": ["mildiou", "pourriture", "feuilles noires"]},
        {"label": "Mosaïque", "aliases": ["mosaïque", "mosaique", "feuilles déformées", "feuilles deformees"]},
        {"label": "Bactériose", "aliases": ["bacteriose", "bactériose", "brûlure bactérienne", "brulure bacterienne"]},
    ],
    "elevage": [
        {"label": "Plaie", "aliases": ["plaie", "blessure", "saigne", "coupure"]},
        {"label": "Boiterie", "aliases": ["boite", "boiterie", "patte", "sabot", "pied"]},
        {"label": "Fièvre", "aliases": ["fièvre", "fievre", "chaud", "abattu"]},
        {"label": "Infection cutanée", "aliases": ["croûte", "croute", "peau", "gale", "plaque"]},
        {"label": "Atteinte oculaire", "aliases": ["oeil", "œil", "écoulement", "ecoulement", "narine"]},
        {"label": "Diarrhée", "aliases": ["diarrhee", "diarrhée", "selles liquides"]},
    ],
    "sos_accident": [
        {"label": "Plaie ouverte", "aliases": ["plaie", "coupure", "saigne", "saignement"]},
        {"label": "Brûlure", "aliases": ["brûlure", "brulure", "huile chaude", "feu"]},
        {"label": "Fracture", "aliases": ["fracture", "cassé", "casse", "déboîté", "deboite"]},
        {"label": "Infection", "aliases": ["pus", "infecté", "infecte", "gonflé", "gonfle"]},
        {"label": "Contusion", "aliases": ["bleu", "choc", "contusion", "hématome", "hematome"]},
    ],
    "cybersecurity": [
        {"label": "Arnaque", "aliases": ["arnaque", "escroquerie", "fraude"]},
        {"label": "Pirâtage", "aliases": ["piraté", "pirate", "hacking", "compte volé", "compte vole"]},
        {"label": "Code OTP", "aliases": ["otp", "code secret", "code de validation"]},
        {"label": "SIM swap", "aliases": ["sim swap", "carte sim", "reseau perdu", "réseau perdu"]},
        {"label": "Virus", "aliases": ["virus", "lien suspect", "application suspecte"]},
    ],
}


# ==========================================
# APPLICATION FASTAPI
# ==========================================

app = FastAPI(
    title="SONGRA API - IA Locale",
    version="5.0",
    description="Plateforme d'assistance avec IA locale pour l'analyse de photos"
)

_RATE_BUCKETS: Dict[str, List[float]] = {}
_RATE_LIMITED_PATHS = {
    "/api/auth/phone/start": (5, 300),
    "/api/auth/phone/verify": (15, 300),
    "/api/auth/pin-login": (20, 300),
    "/api/auth/login": (20, 300),
    "/api/sos/alert": (10, 300),
}


@app.middleware("http")
async def security_rate_limit(request: Request, call_next):
    """Protection locale minimale; Redis sera la source partagee en production."""
    rule = _RATE_LIMITED_PATHS.get(request.url.path)
    if rule:
        limit, window = rule
        now = time.time()
        client = request.client.host if request.client else "unknown"
        key = f"{client}:{request.url.path}"
        recent = [stamp for stamp in _RATE_BUCKETS.get(key, []) if stamp > now - window]
        if len(recent) >= limit:
            return JSONResponse(status_code=429, content={"detail": "Trop de requetes. Reessayez plus tard."})
        recent.append(now)
        _RATE_BUCKETS[key] = recent
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    return response

app.include_router(agri_services.router)
app.include_router(yingr_ai_api.router)

# CORS - origines locales explicites pour le front web et mobile
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "http://localhost:4173",
        "http://127.0.0.1:4173",
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "https://comstratmedia.com",
        "https://www.comstratmedia.com",
        "https://songra.yingr-ai.com",
        "https://www.songra.yingr-ai.com",
        "capacitor://localhost",
        "ionic://localhost",
    ],
    # Flutter Web choisit un port de developpement dynamique. Cette regex
    # reste volontairement limitee a la machine locale et n'ouvre aucune
    # origine distante supplementaire.
    allow_origin_regex=r"^https?://(localhost|127\.0\.0\.1)(:\d+)?$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Servir les fichiers statiques
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
app.mount("/audio", StaticFiles(directory=EXPERT_AUDIO_UPLOAD_DIR), name="audio")

# ==========================================
# FONCTIONS UTILITAIRES
# ==========================================

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    if (hashed or "").startswith("$2"):
        try:
            return bcrypt.checkpw(password.encode(), hashed.encode())
        except ValueError:
            return False
    return secrets.compare_digest(hash_password(password), hashed or "")


def hash_expert_password(password: str) -> str:
    """Hash adaptatif pour les comptes experts; les anciens SHA-256 migrent au login."""
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt(rounds=12)).decode()


def _normalize_bf_phone(phone_number: str) -> str:
    digits = re.sub(r"\D", "", phone_number or "")
    if digits.startswith("00"):
        digits = digits[2:]
    if len(digits) == 8:
        digits = "226" + digits
    if len(digits) < 11 or len(digits) > 15:
        raise HTTPException(status_code=422, detail="Numéro de téléphone invalide")
    return "+" + digits


def _hash_pin(pin: str, salt: Optional[str] = None) -> str:
    clean_pin = (pin or "").strip()
    if not re.fullmatch(r"\d{4}", clean_pin):
        raise HTTPException(status_code=422, detail="Le code PIN doit contenir exactement 4 chiffres")
    resolved_salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", clean_pin.encode(), resolved_salt.encode(), 210000).hex()
    return f"pbkdf2_sha256${resolved_salt}${digest}"


def _verify_pin(pin: str, stored_hash: str) -> bool:
    try:
        algorithm, salt, _ = (stored_hash or "").split("$", 2)
        if algorithm != "pbkdf2_sha256":
            return verify_password(pin, stored_hash)  # compatibilité anciens comptes
        return secrets.compare_digest(_hash_pin(pin, salt), stored_hash)
    except (ValueError, HTTPException):
        return False


_ORANGE_SMS_TOKEN: Optional[str] = None
_ORANGE_SMS_TOKEN_EXPIRES_AT: float = 0.0


def _orange_sms_access_token(force_refresh: bool = False) -> str:
    """Obtient et réutilise le jeton OAuth Orange (valable environ une heure)."""
    global _ORANGE_SMS_TOKEN, _ORANGE_SMS_TOKEN_EXPIRES_AT
    if not force_refresh and _ORANGE_SMS_TOKEN and time.time() < _ORANGE_SMS_TOKEN_EXPIRES_AT:
        return _ORANGE_SMS_TOKEN

    client_id = os.getenv("ORANGE_SMS_CLIENT_ID", "").strip()
    client_secret = os.getenv("ORANGE_SMS_CLIENT_SECRET", "").strip()
    if not client_id or not client_secret:
        raise HTTPException(
            status_code=503,
            detail="Service SMS Orange non configuré. Ajoutez ORANGE_SMS_CLIENT_ID et ORANGE_SMS_CLIENT_SECRET.",
        )
    credentials = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
    request = urllib.request.Request(
        "https://api.orange.com/oauth/v3/token",
        data=b"grant_type=client_credentials",
        method="POST",
        headers={
            "Authorization": f"Basic {credentials}",
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=15) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=503, detail="Impossible de joindre l'authentification SMS Orange") from exc
    token = str(payload.get("access_token") or "").strip()
    if not token:
        raise HTTPException(status_code=503, detail="Orange n'a pas fourni de jeton SMS valide")
    expires_in = max(60, int(payload.get("expires_in") or 3600))
    _ORANGE_SMS_TOKEN = token
    _ORANGE_SMS_TOKEN_EXPIRES_AT = time.time() + expires_in - 60
    return token


def _send_orange_sms(phone_number: str, message: str, retry: bool = True) -> None:
    sender_address = os.getenv("ORANGE_SMS_SENDER_ADDRESS", "tel:+2260000").strip()
    encoded_sender = urllib.parse.quote(sender_address, safe="")
    body = {
        "outboundSMSMessageRequest": {
            "address": f"tel:{phone_number}",
            "senderAddress": sender_address,
            "outboundSMSTextMessage": {"message": message},
        }
    }
    sender_name = os.getenv("ORANGE_SMS_SENDER_NAME", "").strip()
    if sender_name:
        body["outboundSMSMessageRequest"]["senderName"] = sender_name[:11]
    request = urllib.request.Request(
        f"https://api.orange.com/smsmessaging/v1/outbound/{encoded_sender}/requests",
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        method="POST",
        headers={
            "Authorization": f"Bearer {_orange_sms_access_token()}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=15) as response:
            if response.status not in (200, 201, 202):
                raise HTTPException(status_code=503, detail="Orange n'a pas accepté le SMS")
    except urllib.error.HTTPError as exc:
        if exc.code == 401 and retry:
            _orange_sms_access_token(force_refresh=True)
            return _send_orange_sms(phone_number, message, retry=False)
        raise HTTPException(status_code=503, detail="Échec d'envoi du SMS Orange") from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail="Impossible de joindre le service SMS Orange") from exc


def _otp_dev_mode_enabled() -> bool:
    return os.getenv("OTP_DEV_MODE", "false").strip().lower() == "true"


def _send_auth_otp_sms(phone_number: str, code: str) -> None:
    """Envoie réellement l'OTP avec Orange SMS Burkina Faso.

    Si OTP_DEV_MODE=true (ex: crédits Orange épuisés/pas encore rechargés),
    on n'envoie AUCUN vrai SMS et on se contente de logger le code côté
    serveur : /api/auth/phone/start renvoie alors le code directement dans
    sa réponse JSON (champ debug_otp) pour que l'app puisse l'afficher, vu
    que l'utilisateur final n'a pas accès aux logs backend. Il suffit de
    repasser OTP_DEV_MODE=false une fois les crédits Orange rechargés pour
    réactiver l'envoi réel par SMS, sans autre changement de code.
    """
    if _otp_dev_mode_enabled():
        print(f"[OTP-DEV] {phone_number}: {code}")
        return
    provider = os.getenv("SMS_PROVIDER", "orange").strip().lower()
    message = f"SONGRA : votre code de vérification est {code}. Il expire dans 5 minutes. Ne le partagez avec personne."
    if provider == "orange":
        return _send_orange_sms(phone_number, message)
    raise HTTPException(
        status_code=503,
        detail="Service SMS non configuré. Utilisez SMS_PROVIDER=orange et configurez les identifiants Orange.",
    )


def create_access_token(user: User) -> str:
    now = datetime.utcnow()
    payload = {
        "sub": str(user.id),
        "phone_number": user.phone_number,
        "type": "user",
        "iat": int(now.timestamp()),
        "exp": int((now + timedelta(hours=JWT_EXPIRE_HOURS)).timestamp()),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def create_expert_access_token(expert: Expert) -> str:
    now = datetime.utcnow()
    payload = {
        "sub": str(expert.id), "type": "expert", "role": getattr(expert, "role", "expert"),
        "jti": secrets.token_urlsafe(24), "iat": int(now.timestamp()),
        "exp": int((now + timedelta(hours=min(JWT_EXPIRE_HOURS, 24))).timestamp()),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def _extract_bearer_token(authorization: Optional[str]) -> str:
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing authorization header")

    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=401, detail="Invalid authorization header")

    return token


def _get_or_create_user_by_phone(phone_number: str, db: Session) -> User:
    normalized_phone = (phone_number or "").strip()
    if not normalized_phone:
        raise HTTPException(status_code=401, detail="Phone number is required")

    user = db.query(User).filter(User.phone_number == normalized_phone).first()
    if user:
        return user

    user = User(phone_number=normalized_phone)
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def get_current_user(
    authorization: Optional[str] = Header(default=None),
    db: Session = Depends(get_db),
) -> User:
    token = _extract_bearer_token(authorization)

    if token.startswith("phone:"):
        raise HTTPException(
            status_code=401,
            detail="Ancienne session non sécurisée. Vérifiez votre numéro pour vous reconnecter.",
        )

    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError as exc:
        raise HTTPException(status_code=401, detail="Token expired") from exc
    except jwt.InvalidTokenError as exc:
        raise HTTPException(status_code=401, detail="Invalid token") from exc

    if payload.get("type") != "user":
        raise HTTPException(status_code=401, detail="Invalid token scope")

    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token subject")

    user = db.query(User).filter(User.id == int(user_id)).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")

    return user


# ==========================================
# FACTURATION, ABONNEMENTS ET QUOTAS SONGRA
# ==========================================

SONGRA_OFFERS: Dict[str, Dict[str, Any]] = {
    "week": {"label": "Semaine", "price": 500, "days": 7, "analyses": 25, "tickets": 3, "courses": 5},
    "month": {"label": "Mois Essentiel", "price": 1000, "days": 30, "analyses": 100, "tickets": 10, "courses": -1},
    "pro": {"label": "Mois Pro", "price": 2000, "days": 30, "analyses": 200, "tickets": 20, "courses": -1},
}
SONGRA_PRODUCTS: Dict[str, Dict[str, Any]] = {
    "ticket": {"label": "1 ticket expert", "price": 300},
    "course": {"label": "1 cours de l'académie", "price": 100},
    "analysis_pack": {"label": "Pack de 10 analyses", "price": 200, "quantity": 10},
}
YENGAPAY_BASE_URL = os.getenv("YENGAPAY_BASE_URL", "https://api.yengapay.com/api/v1").rstrip("/")
YENGAPAY_API_KEY = os.getenv("YENGAPAY_API_KEY", "").strip()
YENGAPAY_ORGANIZATION_ID = os.getenv("YENGAPAY_ORGANIZATION_ID", "").strip()
YENGAPAY_PROJECT_ID = os.getenv("YENGAPAY_PROJECT_ID", "58165").strip()
YENGAPAY_WEBHOOK_SECRET = os.getenv("YENGAPAY_WEBHOOK_SECRET", "").strip()


class PaymentIntentIn(BaseModel):
    kind: str
    amount: Optional[int] = None
    target_id: Optional[str] = None


class WalletPurchaseIn(BaseModel):
    kind: str
    target_id: Optional[str] = None


def _active_offer(user: User) -> Tuple[Optional[str], Optional[Dict[str, Any]]]:
    now = datetime.utcnow()
    plan = (user.subscription_plan or "").strip().lower()
    if plan in SONGRA_OFFERS and user.subscription_expires_at and user.subscription_expires_at > now:
        return plan, SONGRA_OFFERS[plan]
    return None, None


def _usage_count(db: Session, user_id: int, resource: str, since: datetime) -> int:
    return db.query(UsageEventDB).filter(
        UsageEventDB.user_id == user_id,
        UsageEventDB.resource == resource,
        UsageEventDB.created_at >= since,
    ).count()


def _resource_status(db: Session, user: User, resource: str) -> Dict[str, Any]:
    plan, offer = _active_offer(user)
    now = datetime.utcnow()
    if offer:
        since = user.subscription_started_at or (user.subscription_expires_at - timedelta(days=int(offer["days"])))
        limit = int(offer[resource])
        used = _usage_count(db, user.id, resource, since)
        return {"allowed": used < limit, "used": used, "limit": limit, "remaining": max(0, limit - used), "source": plan}
    if resource == "analyses":
        today = datetime(now.year, now.month, now.day)
        used = _usage_count(db, user.id, resource, today)
        if used < 3:
            return {"allowed": True, "used": used, "limit": 3, "remaining": 3 - used, "source": "free_daily"}
    credits = int((user.analysis_credits if resource == "analyses" else user.ticket_credits) or 0)
    if credits > 0:
        return {"allowed": True, "used": 0, "limit": credits, "remaining": credits, "source": "credits"}
    if resource == "analyses":
        return {"allowed": False, "used": 3, "limit": 3, "remaining": 0, "source": "free_daily"}
    return {"allowed": False, "used": 0, "limit": 0, "remaining": 0, "source": "free"}


def _require_resource(db: Session, user: User, resource: str) -> Dict[str, Any]:
    status = _resource_status(db, user, resource)
    if not status["allowed"]:
        label = "analyses" if resource == "analyses" else "tickets expert"
        raise HTTPException(status_code=402, detail=f"Votre quota de {label} est épuisé. Rechargez votre compte ou choisissez un abonnement.")
    return status


def _consume_resource(db: Session, user: User, resource: str, source: str) -> None:
    status = _require_resource(db, user, resource)
    if status["source"] == "credits":
        if resource == "analyses":
            user.analysis_credits = max(0, int(user.analysis_credits or 0) - 1)
        else:
            user.ticket_credits = max(0, int(user.ticket_credits or 0) - 1)
    db.add(UsageEventDB(user_id=user.id, resource=resource, source=source))
    db.commit()


def _billing_snapshot(db: Session, user: User) -> Dict[str, Any]:
    plan, offer = _active_offer(user)
    course_access = db.query(CourseAccessDB).filter(CourseAccessDB.user_id == user.id).count()
    return {
        "wallet_balance": int(user.wallet_balance or 0),
        "currency": "XOF",
        "subscription": {
            "plan": plan,
            "label": offer["label"] if offer else "Gratuit",
            "expires_at": user.subscription_expires_at.isoformat() if plan and user.subscription_expires_at else None,
        },
        "analyses": _resource_status(db, user, "analyses"),
        "tickets": _resource_status(db, user, "tickets"),
        "analysis_credits": int(user.analysis_credits or 0),
        "ticket_credits": int(user.ticket_credits or 0),
        "course_access_count": course_access,
        "offers": SONGRA_OFFERS,
        "products": SONGRA_PRODUCTS,
        "topup_amounts": [500, 1000, 2000, 5000],
    }


def _price_for(kind: str, amount: Optional[int], target_id: Optional[str]) -> Tuple[int, str]:
    if kind == "wallet_topup":
        value = int(amount or 0)
        if value not in {500, 1000, 2000, 5000}:
            raise HTTPException(status_code=400, detail="Montant de recharge non autorisé")
        return value, f"Recharge portefeuille SONGRA {value} F"
    if kind.startswith("plan_"):
        plan = kind.removeprefix("plan_")
        if plan not in SONGRA_OFFERS:
            raise HTTPException(status_code=400, detail="Abonnement inconnu")
        return int(SONGRA_OFFERS[plan]["price"]), f"Abonnement {SONGRA_OFFERS[plan]['label']}"
    if kind not in SONGRA_PRODUCTS:
        raise HTTPException(status_code=400, detail="Produit inconnu")
    if kind == "course" and not target_id:
        raise HTTPException(status_code=400, detail="Cours requis")
    product = SONGRA_PRODUCTS[kind]
    return int(product["price"]), str(product["label"])


def _activate_purchase(db: Session, tx: BillingTransaction, user: User) -> None:
    if tx.status == "done":
        return
    now = datetime.utcnow()
    kind = tx.kind
    if kind == "wallet_topup":
        user.wallet_balance = int(user.wallet_balance or 0) + tx.amount
    elif kind.startswith("plan_"):
        plan = kind.removeprefix("plan_")
        offer = SONGRA_OFFERS[plan]
        user.subscription_plan = plan
        user.subscription_started_at = now
        user.subscription_expires_at = now + timedelta(days=int(offer["days"]))
        user.is_premium = True
        user.premium_expires_at = user.subscription_expires_at
    elif kind == "ticket":
        user.ticket_credits = int(user.ticket_credits or 0) + 1
    elif kind == "analysis_pack":
        user.analysis_credits = int(user.analysis_credits or 0) + int(SONGRA_PRODUCTS[kind]["quantity"])
    elif kind == "course":
        course_id = int(tx.target_id or 0)
        exists = db.query(CourseAccessDB).filter(CourseAccessDB.user_id == user.id, CourseAccessDB.course_id == course_id, CourseAccessDB.permanent == True).first()
        if not exists:
            db.add(CourseAccessDB(user_id=user.id, course_id=course_id, source="purchase", permanent=True))
    tx.status = "done"
    tx.paid_at = now
    db.commit()


def _yengapay_call(method: str, path: str, payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    if not YENGAPAY_API_KEY or not YENGAPAY_ORGANIZATION_ID:
        raise HTTPException(status_code=503, detail="YengaPay n'est pas encore configuré sur le serveur")
    body = json.dumps(payload).encode("utf-8") if payload is not None else None
    req = urllib.request.Request(f"{YENGAPAY_BASE_URL}{path}", data=body, method=method, headers={"x-api-key": YENGAPAY_API_KEY, "Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=25) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise HTTPException(status_code=502, detail=f"YengaPay a refusé la demande: {detail[:300]}") from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail="YengaPay est momentanément indisponible") from exc


@app.get("/api/billing/status")
async def billing_status(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return _billing_snapshot(db, current_user)


@app.post("/api/billing/payment-intent")
async def create_billing_payment_intent(payload: PaymentIntentIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    import asyncio
    amount, label = _price_for(payload.kind, payload.amount, payload.target_id)
    reference = f"SONGRA-{current_user.id}-{int(time.time())}-{secrets.token_hex(4)}"
    request_payload = {"paymentAmount": amount, "reference": reference, "articles": [{"title": label, "description": "Service numérique SONGRA", "pictures": [], "price": amount}]}
    result = await asyncio.to_thread(_yengapay_call, "POST", f"/groups/{YENGAPAY_ORGANIZATION_ID}/payment-intent/{YENGAPAY_PROJECT_ID}", request_payload)
    checkout_url = result.get("checkoutPageUrlWithPaymentToken")
    if not checkout_url or not result.get("id"):
        raise HTTPException(status_code=502, detail="Réponse YengaPay incomplète")
    tx = BillingTransaction(id=secrets.token_hex(16), user_id=current_user.id, reference=reference, provider_intent_id=str(result["id"]), kind=payload.kind, target_id=payload.target_id, amount=amount, status="pending", checkout_url=checkout_url, provider_payload_json=json.dumps(result, ensure_ascii=False))
    db.add(tx)
    db.commit()
    return {"status": "pending", "transaction_id": tx.id, "reference": reference, "checkout_url": checkout_url, "amount": amount, "currency": "XOF"}


@app.post("/api/billing/wallet-purchase")
async def wallet_purchase(payload: WalletPurchaseIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    amount, _ = _price_for(payload.kind, None, payload.target_id)
    if int(current_user.wallet_balance or 0) < amount:
        raise HTTPException(status_code=402, detail="Solde insuffisant. Rechargez votre compte SONGRA.")
    current_user.wallet_balance = int(current_user.wallet_balance or 0) - amount
    tx = BillingTransaction(id=secrets.token_hex(16), user_id=current_user.id, reference=f"WALLET-{current_user.id}-{int(time.time())}-{secrets.token_hex(3)}", provider="wallet", kind=payload.kind, target_id=payload.target_id, amount=amount, status="pending")
    db.add(tx)
    _activate_purchase(db, tx, current_user)
    return {"status": "done", "billing": _billing_snapshot(db, current_user)}


@app.get("/api/billing/transactions/{transaction_id}")
async def billing_transaction_status(transaction_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    import asyncio
    tx = db.query(BillingTransaction).filter(BillingTransaction.id == transaction_id, BillingTransaction.user_id == current_user.id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction introuvable")
    if tx.status == "pending" and tx.provider_intent_id:
        result = await asyncio.to_thread(_yengapay_call, "GET", f"/groups/{YENGAPAY_ORGANIZATION_ID}/payment-intent/project/{YENGAPAY_PROJECT_ID}/intent/{tx.provider_intent_id}")
        if str(result.get("transactionStatus", "")).upper() == "DONE":
            tx.provider_payload_json = json.dumps(result, ensure_ascii=False)
            _activate_purchase(db, tx, current_user)
        elif str(result.get("transactionStatus", "")).upper() == "FAILED":
            tx.status = "failed"
            db.commit()
    return {"status": tx.status, "billing": _billing_snapshot(db, current_user)}


@app.post("/api/v1/webhooks/yengapay")
async def yengapay_webhook(request: Request, db: Session = Depends(get_db)):
    import asyncio
    payload = await request.json()
    reference = str(payload.get("reference") or "")
    tx = db.query(BillingTransaction).filter(BillingTransaction.reference == reference).first()
    if not tx:
        return {"received": True}
    if tx.status == "done":
        return {"received": True}
    received_amount = int(payload.get("paymentAmount") or 0)
    received_fees = int(payload.get("paymentFees") or 0)
    amount_matches = received_amount == tx.amount or (received_amount + received_fees) == tx.amount
    if str(payload.get("projectId") or "") != YENGAPAY_PROJECT_ID or not amount_matches:
        raise HTTPException(status_code=400, detail="Notification de paiement invalide")
    verified = await asyncio.to_thread(_yengapay_call, "GET", f"/groups/{YENGAPAY_ORGANIZATION_ID}/payment-intent/project/{YENGAPAY_PROJECT_ID}/intent/{tx.provider_intent_id}")
    if str(verified.get("transactionStatus", "")).upper() == "DONE":
        user = db.query(User).filter(User.id == tx.user_id).first()
        if user:
            tx.provider_payload_json = json.dumps(payload, ensure_ascii=False)
            _activate_purchase(db, tx, user)
    return {"received": True}


@app.get("/api/payments/yengapay/return", response_class=HTMLResponse)
async def yengapay_payment_return():
    """Page de retour YengaPay. Le webhook reste la source de vérité du paiement."""
    return HTMLResponse("""<!doctype html>
<html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Paiement SONGRA</title><style>body{font-family:system-ui;background:#f4faf7;color:#123;padding:32px;text-align:center}.card{max-width:430px;margin:12vh auto;background:white;border-radius:24px;padding:30px;box-shadow:0 12px 40px #1232}a{display:block;background:#118c4f;color:white;text-decoration:none;padding:15px;border-radius:13px;font-weight:800;margin-top:22px}</style></head>
<body><div class="card"><h1>Paiement enregistré</h1><p>Retour automatique vers SONGRA…</p><a id="open-app" href="intent://payment-return#Intent;scheme=songra;package=com.songra.mobileapp;end">Revenir dans SONGRA</a></div>
<script>function openSongra(){location.href='intent://payment-return#Intent;scheme=songra;package=com.songra.mobileapp;end'};setTimeout(openSongra,250);setTimeout(function(){location.href='songra://payment-return'},1200);</script></body></html>""")


def get_current_expert(
    authorization: Optional[str] = Header(default=None),
    db: Session = Depends(get_db),
) -> Expert:
    token = _extract_bearer_token(authorization)

    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError as exc:
        raise HTTPException(status_code=401, detail="Expert token expired") from exc
    except jwt.InvalidTokenError as exc:
        raise HTTPException(status_code=401, detail="Invalid expert token") from exc
    if payload.get("type") != "expert" or not payload.get("sub") or not payload.get("jti"):
        raise HTTPException(status_code=401, detail="Invalid expert token scope")
    from voice_alert import is_expert_token_revoked
    if is_expert_token_revoked(SessionLocal, str(payload["jti"])):
        raise HTTPException(status_code=401, detail="Expert token revoked")
    expert = db.query(Expert).filter(Expert.id == int(payload["sub"])).first()
    if not expert or not expert.is_active:
        raise HTTPException(status_code=401, detail="Expert not found")

    return expert


def get_current_admin_expert(
    current_expert: Expert = Depends(get_current_expert),
) -> Expert:
    if (getattr(current_expert, "role", "expert") or "expert").lower() != "admin":
        raise HTTPException(status_code=403, detail="Accès administrateur requis")
    return current_expert


@app.get("/api/admin/billing")
async def admin_billing_dashboard(
    current_admin: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    del current_admin
    transactions = db.query(BillingTransaction).order_by(BillingTransaction.created_at.desc()).limit(300).all()
    paid = [item for item in transactions if item.status == "done"]
    active_subscribers = db.query(User).filter(User.subscription_expires_at > datetime.utcnow()).count()
    return {
        "status": "success",
        "summary": {
            "revenue_xof": sum(item.amount for item in paid if item.provider == "yengapay"),
            "paid_transactions": len(paid),
            "pending_transactions": sum(1 for item in transactions if item.status == "pending"),
            "active_subscribers": active_subscribers,
        },
        "transactions": [{
            "id": item.id,
            "reference": item.reference,
            "user_id": item.user_id,
            "kind": item.kind,
            "amount": item.amount,
            "provider": item.provider,
            "status": item.status,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "paid_at": item.paid_at.isoformat() if item.paid_at else None,
        } for item in transactions],
    }


@app.get("/api/admin/activity-report")
async def admin_activity_report(
    days: int = Query(default=30, ge=1, le=365),
    current_admin: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    del current_admin
    now = datetime.utcnow()
    since = now - timedelta(days=days)
    tickets = db.query(Ticket).all()
    period_tickets = [ticket for ticket in tickets if ticket.created_at and ticket.created_at >= since]
    transactions = db.query(BillingTransaction).filter(BillingTransaction.created_at >= since).order_by(BillingTransaction.created_at.desc()).all()
    paid = [item for item in transactions if item.status == "done"]
    yengapay_paid = [item for item in paid if item.provider == "yengapay"]
    statuses: Dict[str, int] = {}
    categories: Dict[str, int] = {}
    for ticket in period_tickets:
        status = ticket.status or "open"
        category = ticket.category or "agriculture"
        statuses[status] = statuses.get(status, 0) + 1
        categories[category] = categories.get(category, 0) + 1
    payment_kinds: Dict[str, Dict[str, int]] = {}
    for item in paid:
        bucket = payment_kinds.setdefault(item.kind, {"count": 0, "amount": 0})
        bucket["count"] += 1
        bucket["amount"] += item.amount

    experts = db.query(Expert).order_by(Expert.full_name.asc()).all()
    expert_reports: List[Dict[str, Any]] = []
    for expert in experts:
        assigned = [ticket for ticket in tickets if ticket.expert_id == expert.id]
        assigned_period = [ticket for ticket in assigned if ticket.created_at and ticket.created_at >= since]
        resolved = [ticket for ticket in assigned_period if ticket.status == "resolved"]
        active = [ticket for ticket in assigned if ticket.status not in {"resolved", "closed"}]
        replies = db.query(Message).filter(
            Message.sender_type == "expert",
            Message.sender_id == expert.id,
            Message.sent_at >= since,
        ).count()
        voice_replies = db.query(Message).filter(
            Message.sender_type == "expert",
            Message.sender_id == expert.id,
            Message.sent_at >= since,
            Message.audio_url.isnot(None),
        ).count()
        durations = [
            (ticket.resolved_at - ticket.created_at).total_seconds() / 3600
            for ticket in resolved
            if ticket.resolved_at and ticket.created_at and ticket.resolved_at >= ticket.created_at
        ]
        expert_reports.append({
            "id": expert.id,
            "name": expert.full_name,
            "email": expert.email,
            "specialization": expert.specialization,
            "zone": expert.zone,
            "is_active": expert.is_active,
            "assigned": len(assigned_period),
            "assigned_all_time": len(assigned),
            "in_progress": len(active),
            "resolved": len(resolved),
            "resolution_rate": round((len(resolved) / len(assigned_period) * 100), 1) if assigned_period else 0,
            "responses": replies,
            "voice_responses": voice_replies,
            "average_resolution_hours": round(sum(durations) / len(durations), 1) if durations else None,
        })

    return {
        "status": "success",
        "period_days": days,
        "generated_at": now.isoformat(),
        "payments": {
            "revenue_xof": sum(item.amount for item in yengapay_paid),
            "wallet_spending_xof": sum(item.amount for item in paid if item.provider == "wallet"),
            "successful": len(paid),
            "pending": sum(1 for item in transactions if item.status == "pending"),
            "failed": sum(1 for item in transactions if item.status == "failed"),
            "active_subscribers": db.query(User).filter(User.subscription_expires_at > now).count(),
            "by_product": payment_kinds,
            "recent": [{
                "reference": item.reference,
                "user_id": item.user_id,
                "kind": item.kind,
                "amount": item.amount,
                "provider": item.provider,
                "status": item.status,
                "created_at": item.created_at.isoformat() if item.created_at else None,
            } for item in transactions[:50]],
        },
        "tickets": {
            "total": len(period_tickets),
            "all_time": len(tickets),
            "resolved": sum(1 for item in period_tickets if item.status == "resolved"),
            "unassigned": sum(1 for item in tickets if not item.expert_id and item.status not in {"resolved", "closed"}),
            "in_progress": sum(1 for item in tickets if item.status not in {"resolved", "closed"}),
            "by_status": statuses,
            "by_category": categories,
        },
        "experts": expert_reports,
    }


def get_current_user_or_expert(
    authorization: Optional[str] = Header(default=None),
    db: Session = Depends(get_db),
) -> Any:
    """Accepte soit un utilisateur mobile, soit un expert."""
    try:
        return get_current_user(authorization, db)
    except HTTPException:
        try:
            return get_current_expert(authorization, db)
        except HTTPException:
            raise HTTPException(status_code=401, detail="Authentification requise")


def serialize_user(user: User) -> Dict[str, Any]:
    return {
        "id": user.id,
        "phone_number": user.phone_number,
        "name": user.name,
        "location": user.location,
        "organization": getattr(user, "organization", None),
        "organization_id": getattr(user, "organization_id", None),
        "is_premium": user.is_premium,
        "messages_used": user.messages_used,
        "messages_limit": user.messages_limit if user.is_premium else 1,
        "created_at": user.created_at.isoformat() if user.created_at else None,
    }


def _validate_user_credentials(phone_number: str, password: str) -> None:
    if not (phone_number or "").strip():
        raise HTTPException(status_code=422, detail="Phone number is required")
    if len((password or "").strip()) < 4:
        raise HTTPException(status_code=422, detail="Password must contain at least 4 characters")


# ==========================================
# BASE DE CONNAISSANCE (RAG SIMPLE)
# ==========================================

# INITIALISATION AU DÉMARRAGE
# ==========================================

def _run_startup_migrations(db: Session):
    """S'assurer que les colonnes audio_url existent dans les tables messages et chat_messages."""
    from sqlalchemy import text
    # 1. messages
    try:
        db.execute(text("ALTER TABLE messages ADD COLUMN audio_url TEXT;"))
        db.commit()
        print("[OK] Migration startup: audio_url ajoutée à la table messages")
    except Exception:
        db.rollback()

    # 2. chat_messages
    try:
        db.execute(text("ALTER TABLE chat_messages ADD COLUMN audio_url TEXT;"))
        db.commit()
        print("[OK] Migration startup: audio_url ajoutée à la table chat_messages")
    except Exception:
        db.rollback()


def _seed_academy_courses(db: Session) -> int:
    """Crée le catalogue pédagogique initial sans écraser les cours existants."""
    crops = {
        "Maïs": "céréale exigeante qui valorise une bonne fertilité et un semis à temps",
        "Mil": "céréale rustique adaptée aux zones sèches et aux sols légers",
        "Sorgho": "céréale résistante dont les variétés doivent suivre la durée de la saison",
        "Riz": "culture qui exige une bonne maîtrise de l’eau et du nivellement",
        "Niébé": "légumineuse alimentaire qui enrichit le sol et supporte la sécheresse",
        "Arachide": "légumineuse qui préfère un sol meuble, drainant et peu compact",
        "Sésame": "culture de rente sensible à l’excès d’eau et aux récoltes tardives",
        "Coton": "culture de rente nécessitant un calendrier rigoureux et une lutte intégrée",
        "Tomate": "culture maraîchère exigeant eau régulière, tuteurage et surveillance sanitaire",
        "Oignon": "culture maraîchère demandant une pépinière saine et une irrigation maîtrisée",
    }
    crop_modules = [
        ("Installer la culture de {crop}", "Préparer la campagne, choisir la variété et réussir l’implantation.", [
            ("Choisir la parcelle", "Choisissez une parcelle adaptée, accessible et sans stagnation durable d’eau."),
            ("Choisir la semence", "Utilisez une semence saine et une variété adaptée à la zone et à la durée des pluies."),
            ("Préparer le sol", "Nettoyez, ameublissez et apportez de la matière organique bien décomposée."),
            ("Semer ou repiquer", "Intervenez au bon moment, respectez profondeur, espacement et densité recommandés."),
        ]),
        ("Entretenir et fertiliser le {crop}", "Conduire la culture depuis la levée jusqu’à la production.", [
            ("Contrôler la levée", "Repérez rapidement les manques, plants faibles et dégâts après implantation."),
            ("Désherber tôt", "Supprimez les adventices avant qu’elles concurrencent la culture pour l’eau et les éléments nutritifs."),
            ("Nourrir le sol", "Apportez compost et fertilisants selon les besoins, en évitant les doses excessives."),
            ("Suivre la croissance", "Visitez la parcelle chaque semaine et notez les changements importants."),
        ]),
        ("Gérer l’eau pour le {crop}", "Réduire le stress hydrique, l’érosion et les pertes d’eau.", [
            ("Observer l’humidité", "Vérifiez l’humidité près des racines avant de décider d’arroser."),
            ("Conserver l’eau", "Utilisez paillage, zaï, demi-lunes ou travail du sol adapté selon la parcelle."),
            ("Arroser efficacement", "Arrosez tôt le matin ou le soir, directement dans la zone racinaire."),
            ("Évacuer l’excès", "Maintenez rigoles et drainage lorsque les pluies risquent d’asphyxier les racines."),
        ]),
        ("Protéger le {crop}", "Reconnaître tôt ravageurs et maladies et intervenir sans danger.", [
            ("Inspecter régulièrement", "Observez feuilles, tiges, fleurs, fruits et sol sur plusieurs points de la parcelle."),
            ("Identifier avant d’agir", "Comparez les symptômes et demandez confirmation si le problème est inconnu ou se propage vite."),
            ("Privilégier la prévention", "Assainissez la parcelle, alternez les cultures et protégez les auxiliaires utiles."),
            ("Traiter avec prudence", "N’utilisez qu’un produit autorisé, à la dose indiquée, avec protection et sans traiter par vent fort."),
        ]),
        ("Récolter et valoriser le {crop}", "Limiter les pertes et améliorer conservation, qualité et revenu.", [
            ("Reconnaître la maturité", "Récoltez au stade adapté pour préserver rendement, goût et qualité marchande."),
            ("Récolter proprement", "Utilisez du matériel propre et évitez contact prolongé avec le sol humide."),
            ("Sécher et trier", "Triez les produits abîmés et atteignez un séchage suffisant avant stockage."),
            ("Stocker et vendre", "Utilisez un stockage propre et ventilé, puis comparez prix, transport et débouchés."),
        ]),
    ]
    techniques = {
        "Zaï": "concentrer l’eau et la fumure dans des poquets sur sol dégradé",
        "Demi-lunes": "ralentir le ruissellement et restaurer les terres en pente douce",
        "Cordons pierreux": "freiner l’érosion suivant les courbes de niveau",
        "Compostage": "transformer les résidus organiques en amendement mûr et sûr",
        "Irrigation goutte-à-goutte": "apporter peu d’eau directement près des racines",
        "Paillage": "couvrir le sol pour conserver l’humidité et limiter les herbes",
        "Association et rotation": "alterner ou associer les cultures pour fertilité et protection",
        "Lutte intégrée": "prévenir et contrôler les ravageurs avec plusieurs méthodes complémentaires",
    }
    technique_modules = [
        ("Installer la technique : {name}", "Comprendre, préparer et mettre correctement en place la technique.", [
            ("Comprendre l’objectif", "Cette technique sert à {purpose}."),
            ("Choisir l’emplacement", "Observez pente, type de sol, circulation de l’eau et culture prévue avant installation."),
            ("Préparer le matériel", "Rassemblez les outils et matériaux locaux nécessaires avant de commencer."),
            ("Installer progressivement", "Testez d’abord sur une petite zone, contrôlez les dimensions puis étendez la technique."),
        ]),
        ("Entretenir et améliorer : {name}", "Contrôler l’efficacité, corriger les défauts et pérenniser la technique.", [
            ("Contrôler après usage", "Vérifiez l’état du dispositif après pluie, irrigation ou travaux dans le champ."),
            ("Corriger rapidement", "Réparez les ruptures, bouchons, zones érodées ou parties inefficaces."),
            ("Mesurer le résultat", "Comparez humidité, vigueur, rendement, temps de travail et dépenses avec une zone témoin."),
            ("Adapter à la parcelle", "Ajustez dimensions et fréquence d’entretien selon le sol, la pente et la culture."),
        ]),
    ]
    existing_titles = {str(value[0]) for value in db.query(AcademyCourseDB.title).all()}
    created = 0
    for crop, profile in crops.items():
        for title_tpl, summary_tpl, steps_tpl in crop_modules:
            title = title_tpl.format(crop=crop.lower())
            if title in existing_titles:
                continue
            steps = [{"id": f"step-{index + 1}", "title": step_title, "content": step_content, "image_url": None, "audio": {}} for index, (step_title, step_content) in enumerate(steps_tpl)]
            db.add(AcademyCourseDB(title=title, course_type="culture", crop=crop, summary=f"{summary_tpl} Le {crop.lower()} est une {profile}.", steps_json=json.dumps(steps, ensure_ascii=False), audio_json="{}", status="published"))
            existing_titles.add(title)
            created += 1
    for name, purpose in techniques.items():
        for title_tpl, summary, steps_tpl in technique_modules:
            title = title_tpl.format(name=name)
            if title in existing_titles:
                continue
            steps = [{"id": f"step-{index + 1}", "title": step_title, "content": step_content.format(purpose=purpose), "image_url": None, "audio": {}} for index, (step_title, step_content) in enumerate(steps_tpl)]
            db.add(AcademyCourseDB(title=title, course_type="technique", crop=name, summary=summary, steps_json=json.dumps(steps, ensure_ascii=False), audio_json="{}", status="published"))
            existing_titles.add(title)
            created += 1
    if created:
        db.commit()
    return created


@app.on_event("startup")
async def startup_seed_data():
    """Initialiser les données minimales au démarrage, y compris sous Gunicorn."""
    try:
        db = SessionLocal()
        _run_startup_migrations(db)
        default_admin_email = os.getenv("DEFAULT_ADMIN_EMAIL", "superadmin@songra.bf").strip()
        default_admin_password = os.getenv("DEFAULT_ADMIN_PASSWORD", "").strip()
        existing = db.query(Expert).filter(Expert.email == default_admin_email).first()
        if not existing:
            if not default_admin_password:
                if APP_ENV == "production":
                    raise RuntimeError("DEFAULT_ADMIN_PASSWORD requis pour creer le premier administrateur")
                default_admin_password = secrets.token_urlsafe(18)
                print(f"[DEV-ONLY] Mot de passe admin initial genere: {default_admin_password}")
            expert = Expert(
                email=default_admin_email,
                password_hash=hash_expert_password(default_admin_password),
                full_name="Super Administrateur SONGRA",
                specialization="all",
                role="admin",
                is_active=True,
                zone="Burkina Faso",
                language="Français",
            )
            db.add(expert)
            db.commit()
            print(f"[OK] Compte super-admin initialisé: {default_admin_email}")
        else:
            existing.role = "admin"
            existing.is_active = True
            existing.full_name = "Super Administrateur SONGRA"
            existing.specialization = "all"
            existing.zone = existing.zone or "Burkina Faso"
            db.commit()

        default_expert_email = os.getenv("DEFAULT_EXPERT_EMAIL", "expert@songra.bf").strip()
        default_expert_password = os.getenv("DEFAULT_EXPERT_PASSWORD", "").strip()
        default_expert = db.query(Expert).filter(Expert.email == default_expert_email).first()
        if not default_expert:
            if not default_expert_password:
                if APP_ENV == "production":
                    raise RuntimeError("DEFAULT_EXPERT_PASSWORD requis pour creer le premier expert")
                default_expert_password = secrets.token_urlsafe(18)
                print(f"[DEV-ONLY] Mot de passe expert initial genere: {default_expert_password}")
            default_expert = Expert(
                email=default_expert_email,
                password_hash=hash_expert_password(default_expert_password),
                full_name="Expert Terrain Agriculture",
                specialization="agriculture",
                role="expert",
                is_active=True,
                zone="Ouagadougou",
                language="Mooré",
            )
            db.add(default_expert)
        else:
            default_expert.role = "expert"
            default_expert.is_active = True
            default_expert.specialization = default_expert.specialization or "agriculture"
            default_expert.zone = default_expert.zone or "Ouagadougou"
        db.commit()

        try:
            load_knowledge_from_json(db)
            total_items = db.query(KnowledgeItem).count()
            print(f"[OK] Base de connaissances chargée ({total_items} fiches)")
        except Exception as e_load:
            print(f"[WARN] Erreur chargement base de connaissances: {e_load}")

        try:
            academy_created = _seed_academy_courses(db)
            academy_total = db.query(AcademyCourseDB).count()
            print(f"[OK] Académie du paysan chargée ({academy_total} cours, {academy_created} nouveaux)")
        except Exception as academy_error:
            db.rollback()
            print(f"[WARN] Erreur initialisation Académie: {academy_error}")

        try:
            # Seed default emergency numbers if empty
            if db.query(EmergencyNumber).count() == 0:
                defaults = [
                    # MAERAH
                    {"label": "🌾 MAERAH - Numéro vert", "number": "(+226) 51 51 34 04", "description": "Numéro vert officiel du Ministère de l'Agriculture, de l'Eau, des Ressources Animales et Halieutiques.", "display_order": 1},
                    {"label": "🌾 MAERAH - Standard central", "number": "(+226) 25 49 99 00 à 09", "description": "Standard central (secrétariat et directions à Ouaga 2000).", "display_order": 2},
                    # Urgences Sécuritaires
                    {"label": "🛡️ CNVA - Alerte sécurité", "number": "10 10", "description": "Centre National de Veille et d'Alerte, numéro d'urgence gratuit pour signaler attaques, suspects ou coupeurs de route.", "display_order": 3},
                    {"label": "🛡️ Collaboration FDS - Terrorisme", "number": "199", "description": "Ligne verte d'urgence dédiée au signalement d'actes de terrorisme et collaboration avec les FDS.", "display_order": 4},
                    {"label": "🛡️ Anti-corruption", "number": "80 00 11 50", "description": "Numéro vert anti-corruption pour dénoncer anonymement les manquements.", "display_order": 5},
                    # Urgences Sanitaires
                    {"label": "🩺 Urgence Sanitaire - CORUS", "number": "35 35", "description": "Numéro vert du CORUS pour signaler toute menace ou urgence liée à la santé publique.", "display_order": 6},
                    {"label": "🩺 Protection Civile", "number": "112", "description": "Ligne d'urgence d'assistance générale et de protection civile.", "display_order": 7},
                    # Lignes directes
                    {"label": "📞 Police Nationale (Standard)", "number": "(+226) 25 31 68 91", "description": "Ligne d'assistance directe en cas de saturation.", "display_order": 8},
                    {"label": "📞 Sapeurs-Pompiers", "number": "(+226) 25 40 96 96 / 25 40 26", "description": "Sapeurs-Pompiers (Ouagadougou) en cas de saturation.", "display_order": 9},
                ]
                for item in defaults:
                    db.add(EmergencyNumber(**item))
                db.commit()
                print("[OK] Numéros d'urgence par défaut créés")
        except Exception as e_emerg:
            print(f"[WARN] Erreur initialisation numéros d'urgence: {e_emerg}")

        try:
            _ensure_audio_map_store()
            migrated_entries = _upsert_expert_local_knowledge_from_legacy_seed(db)
            if migrated_entries > 0:
                print(f"[OK] Fiches locales expertes migrées ({migrated_entries})")
        except Exception as e_local_knowledge:
            print(f"[WARN] Erreur initialisation studio expert local: {e_local_knowledge}")

        try:
            backfilled = _backfill_resolved_tickets_to_offline_corpus(db)
            print(f"[OK] Corpus offline enrichi depuis {backfilled} tickets resolus")
        except Exception as e_backfill:
            print(f"[WARN] Erreur backfill tickets resolus: {e_backfill}")
        finally:
            db.close()
    except Exception as e:
        print(f"[WARN] Erreur initialisation startup: {e}")


def load_knowledge_from_json(db: Session, file_path: str = "knowledge_base.json") -> None:
    """Charger une base de connaissances simple à partir d'un fichier JSON.

    Le fichier doit contenir une liste d'objets de la forme :
    {
        "domain": "agriculture" | "elevage" | "cybersecurity" | "health",
        "title": "Titre court compréhensible par un agriculteur",
        "question": "Formulation typique de la question",
        "answer": "Réponse détaillée, validée par les experts locaux",
        "tags": ["mais", "taches jaunes", "engrais"],
        "language": "fr",
        "source": "ONG locale",  # optionnel
    }
    """
    # Toujours se baser sur le dossier du fichier main.py pour trouver le JSON,
    # afin que ça fonctionne même si le serveur est lancé depuis la racine.
    base_dir = os.path.dirname(os.path.abspath(__file__))
    if not os.path.isabs(file_path):
        file_path = os.path.join(base_dir, file_path)

    if not os.path.exists(file_path):
        return

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception as e:
        print(f"[WARN] Impossible de charger {file_path}: {e}")
        return

    if not isinstance(raw, list):
        print("[WARN] knowledge_base.json doit contenir une liste d'entrées")
        return

    for item in raw:
        try:
            _upsert_knowledge_item(
                db,
                domain=item.get("domain", "agriculture"),
                title=item.get("title"),
                question=item.get("question"),
                answer=item.get("answer"),
                tags=item.get("tags") or [],
                language=item.get("language", "fr"),
                source=item.get("source"),
                media=item.get("media"),
            )
        except Exception as e:
            print(f"[WARN] Erreur lors de l'import d'une entrée de connaissance: {e}")

    db.commit()


def _upsert_knowledge_item(
    db: Session,
    domain: str,
    title: Optional[str],
    answer: Optional[str],
    question: Optional[str] = None,
    tags: Optional[List[str]] = None,
    language: str = "fr",
    source: Optional[str] = None,
    media: Optional[Any] = None,
) -> None:
    """Mettre à jour la KB en considérant le titre comme clé canonique.

    Cela répare les bases déjà polluées où une même fiche a pu être enregistrée
    sous un mauvais domaine ou en doublon. Au prochain démarrage/import, la
    fiche est réalignée sur le domaine déclaré dans le JSON.
    """
    if not title or not answer:
        return

    normalized_domain = (domain or "agriculture").strip().lower()
    serialized_tags = json.dumps(tags or [], ensure_ascii=False)
    serialized_media = json.dumps(media, ensure_ascii=False) if media is not None else None

    existing_items = (
        db.query(KnowledgeItem)
        .filter(KnowledgeItem.title == title)
        .order_by(KnowledgeItem.id.asc())
        .all()
    )

    primary = next(
        (item for item in existing_items if (item.domain or "").strip().lower() == normalized_domain),
        None,
    )
    if primary is None and existing_items:
        primary = existing_items[0]

    if primary is None:
        db.add(
            KnowledgeItem(
                domain=normalized_domain,
                title=title,
                question=question,
                answer=answer,
                tags=serialized_tags,
                language=language,
                source=source,
                media=serialized_media,
            )
        )
        return

    primary.domain = normalized_domain
    primary.question = question
    primary.answer = answer
    primary.tags = serialized_tags
    primary.language = language
    primary.source = source
    primary.media = serialized_media

    for duplicate in existing_items:
        if duplicate.id != primary.id:
            db.delete(duplicate)


def _normalize_token(token: str) -> str:
    """Normaliser grossièrement un mot français pour améliorer le matching.

    - Passe en minuscules
    - Supprime les accents (maïs -> mais)
    - Supprime un "s" final (pluriel simple : jaunes -> jaune)
    """
    if not token:
        return ""

    # Minuscules + suppression des accents
    token = unicodedata.normalize("NFD", token.lower())
    token = "".join(ch for ch in token if unicodedata.category(ch) != "Mn")

    # Supprimer la ponctuation pour éviter les "jaune," vs "jaune"
    token = "".join(ch for ch in token if ch.isalnum())

    # Pluriel très simple : retirer un "s" final
    if len(token) > 3 and token.endswith("s"):
        token = token[:-1]

    return token


def _tokenize(text: str) -> List[str]:
    if not text:
        return []
    raw_tokens = text.split()
    tokens: List[str] = []
    for w in raw_tokens:
        norm = _normalize_token(w)
        if len(norm) > 2:
            tokens.append(norm)
    return tokens


def _normalize_free_text(text: str) -> str:
    normalized = unicodedata.normalize("NFD", (text or "").lower())
    normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return " ".join(normalized.split())


def _build_focus_terms(label: Optional[str], aliases: Optional[List[str]] = None) -> List[str]:
    values = [label or "", *(aliases or [])]
    terms: List[str] = []
    for value in values:
        normalized_value = _normalize_free_text(value)
        if normalized_value and normalized_value not in terms:
            terms.append(normalized_value)
        for token in _tokenize(value):
            if token not in terms:
                terms.append(token)
    return terms


def _find_best_focus_match(category: str, text: str, focus_map: Dict[str, List[Dict[str, Any]]]) -> Optional[Dict[str, Any]]:
    candidates = focus_map.get(category, [])
    if not candidates:
        return None

    normalized_text = _normalize_free_text(text)
    best: Optional[Dict[str, Any]] = None
    best_score = 0
    best_alias_length = 0

    for candidate in candidates:
        aliases = candidate.get("aliases", [])
        matched_aliases = []
        for alias in aliases:
            normalized_alias = _normalize_free_text(alias)
            if normalized_alias and normalized_alias in normalized_text:
                matched_aliases.append(alias)

        score = len(matched_aliases)
        alias_length = max((len(_normalize_free_text(alias)) for alias in matched_aliases), default=0)
        if score > best_score or (score == best_score and alias_length > best_alias_length):
            best = candidate
            best_score = score
            best_alias_length = alias_length

    if not best:
        return None

    return {
        "label": best["label"],
        "aliases": best.get("aliases", []),
        "terms": _build_focus_terms(best["label"], best.get("aliases", [])),
    }


def extract_focus_context(
    category: str,
    text: str,
    photo_analysis: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    subject = _find_best_focus_match(category, text, FOCUS_SUBJECTS)
    issue = _find_best_focus_match(category, text, FOCUS_ISSUES)

    if photo_analysis:
        detected_subject = photo_analysis.get("detected_subject")
        if detected_subject:
            subject = {
                "label": str(detected_subject),
                "aliases": [str(detected_subject)],
                "terms": _build_focus_terms(str(detected_subject)),
            }

        disease_detected = photo_analysis.get("disease_detected")
        if disease_detected:
            issue = {
                "label": str(disease_detected),
                "aliases": [str(disease_detected)],
                "terms": _build_focus_terms(str(disease_detected)),
            }

    return {
        "subject": subject,
        "issue": issue,
    }


def _focus_label(focus_context: Optional[Dict[str, Any]], key: str) -> Optional[str]:
    focus_entry = (focus_context or {}).get(key)
    if isinstance(focus_entry, dict):
        label = focus_entry.get("label")
        return str(label) if label else None
    return None


def _build_precise_no_match_answer(domain: str, focus_context: Optional[Dict[str, Any]] = None) -> str:
    """Réponse quand aucune fiche RAG n'est trouvée - Songra/Yingr-AI reste humble."""
    focus_subject = _focus_label(focus_context, "subject")
    focus_issue = _focus_label(focus_context, "issue")

    focus_parts = [label for label in [focus_subject, focus_issue] if label]
    if focus_parts:
        target = " / ".join(focus_parts)
        return (
            f"Moi, Songra (assistant de Yingr-AI), je n'ai pas trouvé une fiche assez précise pour ton cas : {target}. "
            "Je préfère rester honnête plutôt que de généraliser vers un autre animal, une autre culture ou un autre problème. \n"
            "Ce que je te conseille : "
            "Ajoute si possible un symptôme clé, une photo plus nette de la situation, "
            "ou rapproche-toi d'un expert local (agent agricole, vétérinaire, service de santé du Burkina Faso) pour une vérification sur place."
        )

    domain_label = {
        "agriculture": "agriculture",
        "elevage": "élevage",
        "health": "premiers secours",
        "cybersecurity": "cybersécurité",
    }.get(domain, domain)
    return (
        f"Je suis Songra, l'assistant IA de Yingr-AI dédié au Burkina Faso. "
        f"Je n'ai pas trouvé de fiche suffisamment précise dans le domaine {domain_label}. "
        "Accumuler des connaissances locales du Burkina Faso est mon mission, "
        "mais pour cette question spécifique, je dois rester prudent. \n"
        "Ajoute des détails concrets sur ton problème (quelle culture, quel animal, quelle région au BF?) "
        "pour que je puisse mieux t'aider. "
        "Ou contacte un expert local du domaine pour une vérification fiable."
    )


def retrieve_knowledge(
    db: Session,
    domain: str,
    query: str,
    limit: int = 5,
    expand_scope: bool = True,
    focus_subject: Optional[Dict[str, Any]] = None,
    focus_issue: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """Récupération améliorée basée sur le recouvrement de mots-clés pondéré.

    - Les correspondances dans le titre et les tags comptent plus que celles
      présentes uniquement dans la réponse longue.
        - La recherche est limitée au domaine demandé. Aucun mélange automatique
            entre agriculture et élevage n'est autorisé.
        - Si ``expand_scope`` est activé et qu'aucune fiche n'est trouvée dans le
            domaine demandé, on fait un second passage sur toutes les fiches pour
            éviter de rater une correspondance évidente.
    """
    query_tokens = set(_tokenize(query))
    if not query_tokens:
        return []

    focus_subject_terms = focus_subject.get("terms", []) if focus_subject else []
    focus_issue_terms = focus_issue.get("terms", []) if focus_issue else []

    def serialize_knowledge_item(it: KnowledgeItem) -> Dict[str, Any]:
        media_data: Optional[Any] = None
        if it.media:
            try:
                media_data = json.loads(it.media)
            except Exception:
                media_data = None

        return {
            "id": it.id,
            "domain": it.domain,
            "title": it.title,
            "question": it.question,
            "answer": it.answer,
            "tags": _load_json_list(it.tags),
            "language": it.language,
            "source": it.source,
            "media": media_data,
        }

    def score_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        scored_local: List[Dict[str, Any]] = []
        for it in items:
            # Séparer les zones de texte pour mieux pondérer
            title_tokens = set(_tokenize(it.get("title") or ""))
            question_tokens = set(_tokenize(it.get("question") or ""))
            answer_tokens = set(_tokenize(it.get("answer") or ""))

            tags_list = [str(tag) for tag in (it.get("tags") or []) if str(tag).strip()]
            tags_tokens = set(_tokenize(" ".join(tags_list))) if tags_list else set()

            overlap_title = len(query_tokens & title_tokens)
            overlap_question = len(query_tokens & question_tokens)
            overlap_answer = len(query_tokens & answer_tokens)
            overlap_tags = len(query_tokens & tags_tokens)

            # Pondération : titre > tags > question > réponse
            # answer reçoit un poids très faible pour éviter que les longs textes
            # (fiches entreprendre, analyses) ne matchent sur des mots communs.
            score = (
                overlap_title * 3.0
                + overlap_tags * 2.5
                + overlap_question * 2.0
                + overlap_answer * 0.3
            )

            combined_text = _normalize_free_text(
                f"{it.get('title') or ''}\n{it.get('question') or ''}\n{it.get('answer') or ''}\n{' '.join(tags_list)}"
            )
            subject_match = any(term in combined_text for term in focus_subject_terms) if focus_subject_terms else False
            issue_match = any(term in combined_text for term in focus_issue_terms) if focus_issue_terms else False

            if focus_subject_terms and subject_match:
                score += 8.0
            if focus_issue_terms and issue_match:
                score += 4.5

            if score <= 0:
                continue

            scored_local.append({
                "item": it,
                "score": score,
                "subject_match": subject_match,
                "issue_match": issue_match,
            })

        return scored_local

    generated_source_kinds = _trusted_shared_rag_source_kinds()

    def fetch_generated_items(target_domain: Optional[str] = None) -> List[Dict[str, Any]]:
        query_builder = db.query(OfflineKnowledgeEntryDB).filter(
            OfflineKnowledgeEntryDB.source_kind.in_(generated_source_kinds)
        )
        if target_domain:
            query_builder = query_builder.filter(OfflineKnowledgeEntryDB.domain == target_domain)
        return [_serialize_offline_entry_for_rag(item) for item in query_builder.all()]

    def fetch_studio_items(target_domain: Optional[str] = None) -> List[Dict[str, Any]]:
        query_builder = db.query(ExpertLocalKnowledgeDB).filter(
            ExpertLocalKnowledgeDB.status.in_(["validated", "resolved", "expert_verified"])
        )
        if target_domain:
            query_builder = query_builder.filter(
                ExpertLocalKnowledgeDB.category
                == _normalize_expert_local_category(target_domain)
            )
        studio_items: List[Dict[str, Any]] = []
        for item in query_builder.all():
            studio_items.append({
                "id": f"studio-{item.id}",
                "domain": item.category,
                "title": item.title,
                "question": item.question_fr,
                "answer": item.resolution_fr,
                "tags": _load_json_list(item.tags_json),
                "language": "fr",
                "source": "studio_connaissances",
                "translations": _load_json_dict(item.translations_json),
                "audio": _load_json_dict(item.audio_json),
            })
        return studio_items

    # 1) Fiches strictement dans le domaine demandé
    primary_items = [
        serialize_knowledge_item(item)
        for item in db.query(KnowledgeItem).filter(KnowledgeItem.domain == domain).all()
    ]
    primary_items.extend(fetch_generated_items(domain))
    primary_items.extend(fetch_studio_items(domain))
    scored = score_items(primary_items)

    # 2) Fallback global optionnel : si rien trouvé, on regarde toutes les fiches
    if not scored and expand_scope:
        all_items = [serialize_knowledge_item(item) for item in db.query(KnowledgeItem).all()]
        all_items.extend(fetch_generated_items())
        all_items.extend(fetch_studio_items())
        scored = score_items(all_items)

    # 3) Dernier recours : recherche par sous-chaîne, soit dans le domaine
    # strict, soit sur toute la base si l'élargissement est autorisé.
    if not scored:
        fallback_items = primary_items
        if expand_scope:
            fallback_items = [serialize_knowledge_item(item) for item in db.query(KnowledgeItem).all()]
            fallback_items.extend(fetch_generated_items())
            fallback_items.extend(fetch_studio_items())

        def normalize_text(text: str) -> str:
            if not text:
                return ""
            tokens = _tokenize(text)
            return " ".join(tokens)

        norm_query_parts = list(query_tokens)
        for it in fallback_items:
            tags_list = [str(tag) for tag in (it.get("tags") or []) if str(tag).strip()]
            big_text = f"{it.get('title') or ''}\n{it.get('question') or ''}\n{it.get('answer') or ''}\n"
            if tags_list:
                big_text += " ".join(tags_list)
            norm_text = normalize_text(big_text)
            if any(part in norm_text for part in norm_query_parts):
                scored.append({"item": it, "score": 1.0})

    if not expand_scope:
        scored = [
            entry
            for entry in scored
            if _normalize_expert_local_category(entry["item"].get("domain"))
            == _normalize_expert_local_category(domain)
        ]

    if focus_subject_terms and any(entry.get("subject_match") for entry in scored):
        scored = [entry for entry in scored if entry.get("subject_match")]

    if focus_issue_terms and any(entry.get("issue_match") for entry in scored):
        scored = [entry for entry in scored if entry.get("issue_match")]

    scored.sort(key=lambda x: x["score"], reverse=True)
    top_items = [s["item"] for s in scored[:limit]]

    results: List[Dict[str, Any]] = []
    for it in top_items:
        results.append(dict(it))

    return results


def generate_llm_answer_with_general_knowledge(
    question: str,
    language: str,
    domain: str,
    conversation_context: Optional[List[Dict[str, str]]] = None,
    photo_analysis: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    """Assistant rural utilisant la connaissance générale de ChatGPT.
    
    Quand il n'y a pas de fiche dans la base de connaissances locale,
    Songra utilise ses connaissances générales pour aider la communauté.
    Elle reste un assistant rural accessible et pratique.
    
    📷 Si une photo a été analysée, son diagnostic guide la réponse.
    """
    if not openai_client or not OPENAI_API_KEY:
        return None

    conversation_text = ""
    if conversation_context:
        serialized_turns = []
        for turn in conversation_context[-6:]:
            role = "Utilisateur" if turn.get("role") == "user" else "Assistant"
            content = (turn.get("content") or "").strip()
            if content:
                serialized_turns.append(f"{role}: {content}")
        if serialized_turns:
            conversation_text = "\n\nContexte de la conversation :\n" + "\n".join(serialized_turns) + "\n"

    # Construire le diagnostic photo s'il existe
    # ⚠️ SI UNE PHOTO, C'EST LE CŒUR DE LA RÉPONSE
    photo_diagnosis_section = ""
    has_photo_diagnosis = False
    if photo_analysis and (photo_analysis.get("disease_detected") or photo_analysis.get("observations")):
        has_photo_diagnosis = True
        diagnosis_parts = []
        
        if photo_analysis.get("disease_detected"):
            diagnosis_parts.append(f"Problème : {photo_analysis.get('disease_detected')}")
        
        if photo_analysis.get("detected_subject"):
            diagnosis_parts.append(f"Sujet : {photo_analysis.get('detected_subject')}")
        
        if photo_analysis.get("observations"):
            diagnosis_parts.append(f"Observations : {photo_analysis.get('observations')}")
        
        if photo_analysis.get("urgency"):
            urgency_fr = {
                "immediate": "IMMÉDIATE 🚨",
                "high": "ÉLEVÉE ⚠️",
                "medium": "MOYENNE",
                "low": "BASSE"
            }.get(photo_analysis.get("urgency"), photo_analysis.get("urgency"))
            diagnosis_parts.append(f"Urgence : {urgency_fr}")
        
        if diagnosis_parts:
            photo_diagnosis_section = (
                "📸 DIAGNOSTIC DE TA PHOTO :\n"
                + " | ".join(diagnosis_parts) + 
                "\n"
            )

    domain_description = {
        "agriculture": "l'agriculture et les cultures",
        "elevage": "l'élevage et l'élevage du bétail",
        "health": "les premiers secours et la sécurité sanitaire",
        "cybersecurity": "la cybersécurité et la sécurité en ligne",
    }.get(domain, domain)

    system_prompt = (
        "Tu es Songra, un assistant rural qui aide les communautés à résoudre leurs difficultés quotidiennes. \n"
        "Tu travailles avec Yingr-AI, une intelligence artificielle pour soutenir les populations. \n"
        f"Ta spécialité actuelle : {domain_description}. \n\n"
        "Tu es pratique, accessible et toujours prêt à aider. \n"
        "Tes réponses doivent être : \n"
        "- SIMPLES et directes (compréhensible par tout le monde) \n"
        "- PRATIQUES avec des conseils qu'on peut appliquer tout de suite \n"
        "- HONNÊTE sur ce que tu sais et ce que tu ne sais pas \n"
        "- ENCOURAGEANTE : tu crois que la communauté peut réussir \n\n"
        "Tu n'as pas de fiche spécialisée exacte pour cette question, donc tu utilises tes connaissances générales. \n"
        "Mais tu donnes toujours des conseils pratiques et adaptés aux situations réelles. \n"
        "Tu n'inventes jamais, tu dis toujours si tu n'es pas sûr. \n"
    )

    # Construire le prompt selon qu'on a une photo ou pas
    if has_photo_diagnosis:
        # PHOTO ANALYSÉE = point de départ obligatoire
        user_prompt = (
            f"Domaine: {domain}. Langue: {language or 'fr'}.\n\n"
            f"🎯 TU DOIS COMMENCER PAR :\n"
            f"{photo_diagnosis_section}\n"
            "Tâche uniquement :\n"
            "1️⃣ Commence par : 'D'après l'analyse de ta photo :'\n"
            "2️⃣ Décris ce problème (2-3 phrases claires)\n"
            "3️⃣ Actions concrètes & pratiques (numérotées)\n"
            "4️⃣ Dis si expert est VRAIMENT nécessaire\n"
            "5️⃣ Max 10-15 phrases. Langage SIMPLE.\n"
            "6️⃣ Ne dévie JAMAIS du diagnostic photo.\n\n"
            f"{conversation_text}\n"
        )
    else:
        # Sans photo = question textuelle normale
        user_prompt = (
            f"Langue : {language or 'fr'}. Domaine : {domain}.\n"
            f"Question : {question}\n"
            f"{conversation_text}\n"
            "Tâche : Aide cette personne de manière pratique et simple. \n"
            "- Explique ce que tu comprends du problème (2-3 phrases). \n"
            "- Donne des conseils concrets qu'on peut faire tout de suite (numérotés). \n"
            "- Dis si tu penses qu'il faut l'aide d'un expert et pourquoi. \n"
            "- Sois honnête si tu n'es pas totalement sûr. \n"
            "- Utilise un langage simple et pratique. \n"
            "- Limite à 10-15 phrases maximum. \n\n"
            "Réponds TOUJOURS de manière pratique pour aider la communauté à résoudre ses difficultés."
        )

    def call_openai():
        if not openai_client or not OPENAI_API_KEY:
            raise RuntimeError("OPENAI_API_KEY non configuree")
        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.3,
            max_tokens=1500,
        )
        return response.choices[0].message.content if response.choices else ""

    def call_gemini():
        if not GEMINI_API_KEY:
            raise RuntimeError("GEMINI_API_KEY non configuree")
        full_prompt = system_prompt + "\n\n" + user_prompt
        model = genai.GenerativeModel('gemini-2.5-flash')
        result = model.generate_content(full_prompt)
        return result.text

    provider_errors = []
    if AI_PROVIDER == "openai":
        try:
            return call_openai()
        except Exception as openai_exc:
            provider_errors.append(f"openai: {openai_exc}")
            print(f"[WARN] OpenAI error in general LLM answer, fallback Gemini: {openai_exc}")
            try:
                return call_gemini()
            except Exception as gemini_exc:
                provider_errors.append(f"gemini: {gemini_exc}")
                print(f"[WARN] Erreur complete (generate_llm_answer_with_general_knowledge): {' | '.join(provider_errors)}")
                return None
    else:
        try:
            return call_gemini()
        except Exception as gemini_exc:
            provider_errors.append(f"gemini: {gemini_exc}")
            print(f"[WARN] Gemini error in general LLM answer, fallback OpenAI: {gemini_exc}")
            try:
                return call_openai()
            except Exception as openai_exc:
                provider_errors.append(f"openai: {openai_exc}")
                print(f"[WARN] Erreur complete (generate_llm_answer_with_general_knowledge): {' | '.join(provider_errors)}")
                return None


def resolve_knowledge_answer(
    db: Session,
    domain: str,
    question: str,
    language: str = "fr",
    conversation_context: Optional[List[Dict[str, str]]] = None,
    limit: int = 5,
    focus_context: Optional[Dict[str, Any]] = None,
    photo_analysis: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Répondre d'abord depuis une fiche validée du Studio de connaissances.

    Stratégie ACTUELLE :
    1. OpenAI / LLM général en PREMIER (analyse directe par l'IA)
    2. Fallback RAG : fiches locales si OpenAI échoue ou est indisponible
    3. Fallback ultime : réponse générique "Je ne sais pas"

    📷 Le diagnostic photo (si disponible) est passé au LLM pour enrichir la réponse.
    """
    studio_match = _find_studio_knowledge_match(
        db,
        category=domain,
        query_text=question,
        photo_analysis=photo_analysis,
    )
    if studio_match:
        return {
            "rag_items": [studio_match],
            "llm_answer": studio_match["resolution_fr"],
            "rag_fallback_answer": None,
            "knowledge_mode": "studio_knowledge",
            "knowledge_fallback_used": False,
            "studio_match": studio_match,
        }

    # Aucune fiche Studio assez proche : analyse générale en repli.
    import asyncio
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as pool:
                general_answer = pool.submit(
                    asyncio.run,
                    v2_services.gemini_llm_general_knowledge(
                        question=question,
                        language=language,
                        domain=domain,
                        conversation_context=conversation_context,
                        photo_analysis=photo_analysis,
                    )
                ).result()
        else:
            general_answer = asyncio.run(
                v2_services.gemini_llm_general_knowledge(
                    question=question,
                    language=language,
                    domain=domain,
                    conversation_context=conversation_context,
                    photo_analysis=photo_analysis,
                )
            )
    except Exception as e:
        print(f"[WARN] OpenAI general knowledge error: {e}")
        general_answer = None

    if general_answer:
        return {
            "rag_items": [],
            "llm_answer": general_answer,
            "rag_fallback_answer": None,
            "knowledge_mode": "llm_general_knowledge",
            "knowledge_fallback_used": False,
        }

    # ÉTAPE 2 : Fallback RAG - fiches locales si OpenAI indisponible
    rag_items = retrieve_knowledge(
        db,
        domain,
        question,
        limit=limit + 3,
        expand_scope=False,
        focus_subject=(focus_context or {}).get("subject"),
        focus_issue=(focus_context or {}).get("issue"),
    )

    if rag_items:
        llm_answer = generate_llm_answer(
            question=question,
            language=language,
            domain=domain,
            knowledge_items=rag_items,
            conversation_context=conversation_context,
            focus_context=focus_context,
            photo_analysis=photo_analysis,
        )
        return {
            "rag_items": rag_items,
            "llm_answer": llm_answer,
            "rag_fallback_answer": None if llm_answer else rag_items[0].get("answer"),
            "knowledge_mode": "rag_strict",
            "knowledge_fallback_used": True,
        }

    # ÉTAPE 3 : Fallback ultime - aucune source d'info disponible
    return {
        "rag_items": [],
        "llm_answer": None,
        "rag_fallback_answer": _build_precise_no_match_answer(domain, focus_context),
        "knowledge_mode": "no_match",
        "knowledge_fallback_used": False,
    }


def generate_llm_answer(
    question: str,
    language: str,
    domain: str,
    knowledge_items: List[Dict[str, Any]],
    conversation_context: Optional[List[Dict[str, str]]] = None,
    focus_context: Optional[Dict[str, Any]] = None,
    photo_analysis: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    """Songra (Yingr-AI) reformule et raisonne à partir de la base RAG validée localement au Burkina Faso.

    IMPORTANT : Songra NE DOIT PAS inventer de faits en dehors des fiches RAG du Burkina Faso.
    - S'il n'y a pas assez de fiches, Songra le dit clairement.
    - Toutes les réponses doivent respecter le contexte socio-climatique du BF.
    - Songra reste humble et recommande toujours un expert local si doute.
    - SI une photo a été analysée, sa diagnosis DOIT être au cœur de la réponse.
    """
    if not knowledge_items:
        # Pas de base de connaissance pertinente, on ne force pas le modèle
        return None

    focus_subject_label = _focus_label(focus_context, "subject")
    focus_issue_label = _focus_label(focus_context, "issue")

    # Petit fallback local : si le LLM n'est pas disponible, on formate au
    # minimum une réponse structurée à partir de la meilleure fiche.
    def build_structured_from_rag() -> str:
        best = knowledge_items[0]
        titre = best.get("title") or "Conseil local"
        reponse = best.get("answer") or ""
        source = best.get("source") or "fiches locales"

        parts = []
        focus_hint = ""
        if focus_subject_label or focus_issue_label:
            focus_parts = [label for label in [focus_subject_label, focus_issue_label] if label]
            focus_hint = f" de façon précise sur {' / '.join(focus_parts)}"
        parts.append(
            f"1) Ce que Songra (moi!) comprends de ton problème :\n"
            f"Tu signales un souci lié à : {titre}. "
            f"Je vais te partager les conseils validés et déjà accumulés au Burkina Faso{focus_hint}, "
            f"car je suis Songra, l'assistant IA local de Yingr-AI dédié aux communautés du BF."
        )
        parts.append(
            "2) Conseils pratiques pour le contexte du Burkina Faso :\n" + reponse
        )
        parts.append(
            "3) Quand ABSOLUMENT contacter un expert local :\n"
            "Si malgré ces conseils la situation ne s'améliore pas, si le problème s'aggrave, ou si tu as un doute, "
            "rapproche-toi d'un agent agricole, d'un vétérinaire, ou d'un service technique local pour vérifier sur place. "
            "C'est important pour ta sécurité et celle de ta famille/communauté."
            f" (Conseil validé localement - Source : {source})."
        )
        return "\n\n".join(parts)

    # Si Gemini n'est pas configuré, on renvoie la version structurée basée sur le RAG.
    if not GEMINI_API_KEY:
        return build_structured_from_rag()

    context_blocks = []
    for idx, item in enumerate(knowledge_items, start=1):
        context_blocks.append(
            f"FICHE {idx} ({item.get('domain', '')}) - {item.get('title', '')}\n"
            f"Question typique: {item.get('question', '')}\n"
            f"Réponse validée: {item.get('answer', '')}\n"
            f"Mots-clés: {', '.join(item.get('tags', []))}\n"
        )

    context_text = "\n\n".join(context_blocks)

    system_prompt = (
        "Tu es Songra, l'agent d'assistance IA de Yingr-AI (Yingr Artificial Intelligence). \n"
        "Yingr-AI est une intelligence artificielle LOCALE et SOUVERAINE basée au Burkina Faso. \n"
        "Tu es le lien entre la connaissance validée et les communautés rurales du Burkina Faso. \n\n"
        "IDENTITÉ ET MISSION : \n"
        "- Tu es Songra, dédié à l'assistance des communautés du Burkina Faso \n"
        "- Tu fournis des conseils en agriculture, élevage, et cybersécurité adaptés au contexte burkinabè \n"
        "- Ton objectif : autonomiser les agriculteurs et éleveurs du BF avec des solutions locales \n\n"
        "CONTEXTE BURKINABÈ - TOUJOURS à l'esprit : \n"
        "- Climat Sahélien avec sécheresses périodiques \n"
        "- Ressources naturelles limitées mais exploitables intelligemment \n"
        "- Cultures principales : mil, sorgho, maïs, arachide, coton, oignons \n"
        "- Élevages : bovins, ovins, caprins, volailles adaptés aux conditions locales \n"
        "- Langues locales : Mooré, Dioula, Fulfuldé dominent les zones rurales \n\n"
        "RÈGLES STRICTES : \n"
        "- Tu dois répondre UNIQUEMENT avec les fiches ci-dessous. \n"
        "- Si les fiches ne suffisent pas, dis-le clairement sans généralisations hasardeuses. \n"
        "- Pas de hors-sujet, zéro conseil médical avancé ou dangereux. \n"
        "- Langage TRÈS simple, phrases courtes, concret, sans jargon - adapté aux populations peu alphabétisées. \n"
        "- Réponds en français clair mais intègre des mots en langue locale si c'est plus approprié \n"
        "(ex: 'zaï' pour les trous en agroforesterie, 'daba' pour la houe). \n"
        "- TOUJOURS recommander de vérifier avec un expert local (agent agricole, vétérinaire, service de santé du BF). \n"
        "- Adapte chaque conseil à la climatologie et aux réalités socio-économiques burkinabè. \n"
    )

    focus_instruction = ""
    if focus_subject_label or focus_issue_label:
        focus_parts = [label for label in [focus_subject_label, focus_issue_label] if label]
        focus_instruction = (
            "\nContrainte de précision : reste strictement centré sur cet objet précis"
            f" : {' / '.join(focus_parts)}. "
            "Ne dérive pas vers des conseils généraux d'une autre espèce, d'une autre culture, d'une autre blessure ou d'un autre service. "
            "Si les fiches ne couvrent pas précisément cet objet, dis-le clairement au lieu de généraliser.\n"
        )

    conversation_text = ""
    if conversation_context:
        serialized_turns = []
        for turn in conversation_context[-6:]:
            role = "Utilisateur" if turn.get("role") == "user" else "Assistant"
            content = (turn.get("content") or "").strip()
            if content:
                serialized_turns.append(f"{role}: {content}")
        if serialized_turns:
            conversation_text = "\n\nContexte de la conversation en cours :\n" + "\n".join(serialized_turns) + "\n"

    # Construire le diagnostic photo détaillé pour enrichir le prompt
    # ⚠️ SI UNE PHOTO EST ANALYSÉE, C'EST LE POINT DE DÉPART OBLIGATOIRE
    photo_diagnosis_section = ""
    has_photo_diagnosis = False
    if photo_analysis and (photo_analysis.get("disease_detected") or photo_analysis.get("observations")):
        has_photo_diagnosis = True
        diagnosis_parts = []
        
        if photo_analysis.get("disease_detected"):
            diagnosis_parts.append(f"Problème détecté : {photo_analysis.get('disease_detected')}")
        
        if photo_analysis.get("detected_subject"):
            diagnosis_parts.append(f"Sujet identifié : {photo_analysis.get('detected_subject')}")
        
        if photo_analysis.get("observations"):
            diagnosis_parts.append(f"Observations : {photo_analysis.get('observations')}")
        
        if photo_analysis.get("urgency"):
            urgency_fr = {
                "immediate": "IMMÉDIATE 🚨",
                "high": "ÉLEVÉE ⚠️",
                "medium": "MOYENNE",
                "low": "BASSE"
            }.get(photo_analysis.get("urgency"), photo_analysis.get("urgency"))
            diagnosis_parts.append(f"Urgence : {urgency_fr}")
        
        if photo_analysis.get("confidence"):
            diagnosis_parts.append(f"Confiance : {photo_analysis.get('confidence')}%")
        
        if diagnosis_parts:
            photo_diagnosis_section = (
                "📸 DIAGNOSTIC PHOTO ANALYSÉE PAR IA :\n"
                + " | ".join(diagnosis_parts) + 
                "\n"
            )

    user_prompt = ""
    
    # SI photo analysée, c'est LA PRIORITÉ - elle guide TOUT
    if has_photo_diagnosis:
        user_prompt = (
            f"Domaine: {domain}. Langue: {language or 'fr'}.\n\n"
            f"🎯 POINT DE DÉPART OBLIGATOIRE - Analyse photo :\n"
            f"{photo_diagnosis_section}\n"
            f"FICHES DE CONNAISSANCE DISPONIBLES :\n{context_text}\n\n"
            "INSTRUCTIONS STRICTES :\n"
            "1️⃣ Commence OBLIGATOIREMENT par : 'D'après l'analyse de ta photo :'\n"
            "2️⃣ Expose d'abord le diagnostic détecté de façon ultra-brève\n"
            "3️⃣ Donne des actions ultra-brèves basées sur ce diagnostic (max 2 actions)\n"
            "4️⃣ Dis clairement en 1 courte phrase si un expert est nécessaire\n"
            "5️⃣ Sois extrêmement bref et concis. Maximum 4 à 5 courtes phrases simples au total pour faciliter la traduction.\n"
            "6️⃣ Ne varie PAS du diagnostic. S'il y a des fiches pertinentes, base-toi dessus.\n"
            "Si tu n'as pas assez d'infos des fiches, dis-le franchement.\n\n"
            f"{focus_instruction}"
            f"{conversation_text}"
        )
    else:
        # Sans photo analysée - CONSULTATION EN LIGNE COMPLÈTE en texte
        user_prompt = (
            f"Domaine: {domain}. Langue: {language or 'fr'}.\n"
            f"Tu es un consultant expert en ligne SONGRA. L'utilisateur a une question spécifique:\n"
            f"Question: {question}\n\n"
            f"FICHES DE CONNAISSANCE DISPONIBLES (base locale BF validée):\n{context_text}\n\n"
            f"{focus_instruction}"
            f"{conversation_text}\n"
            "✅ STRUCTURE OBLIGATOIRE POUR CONSULTATION EN LIGNE:\n\n"
            "1️⃣ **DIAGNOSTIC & ANALYSE** (1 phrase max):\n"
            "   → Ce que tu comprends du problème de l'utilisateur de façon très brève.\n\n"
            "2️⃣ **RECOMMANDATIONS PRATIQUES** (3 étapes max, numérotées 1., 2., 3.):\n"
            "   → Actions concrètes ultra-brèves (max 8 mots par étape) adaptées au Burkina Faso.\n\n"
            "3️⃣ **QUAND CONSULTER UN EXPERT** (1 phrase max):\n"
            "   → Recommandation claire de consulter un expert local (vétérinaire/agronome/santé).\n\n"
            "⚠️ RÈGLES STRICTES:\n"
            "- Langage TRÈS SIMPLE (populations peu alphabétisées Burkina Faso)\n"
            "- Phrases COURTES (max 10 mots par phrase)\n"
            "- BASÉ 100% sur les fiches fournies\n"
            "- Sois extrêmement bref et concis : maximum 4 à 6 phrases au total pour faciliter la traduction.\n"
            "- Utilise français clair + mots locaux si approprié (ex: 'zaï', 'daba', 'vétérinaire')"
        )

    def call_openai():
        if not openai_client or not OPENAI_API_KEY:
            raise RuntimeError("OPENAI_API_KEY non configuree")
        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.3,
            max_tokens=1500,
        )
        return response.choices[0].message.content if response.choices else ""

    def call_gemini():
        if not GEMINI_API_KEY:
            raise RuntimeError("GEMINI_API_KEY non configuree")
        full_prompt = system_prompt + "\n\n" + user_prompt
        model = genai.GenerativeModel('gemini-2.5-flash')
        result = model.generate_content(full_prompt)
        return result.text

    provider_errors = []
    if AI_PROVIDER == "openai":
        try:
            return call_openai()
        except Exception as openai_exc:
            provider_errors.append(f"openai: {openai_exc}")
            print(f"[WARN] OpenAI error in generate_llm_answer, fallback Gemini: {openai_exc}")
            try:
                return call_gemini()
            except Exception as gemini_exc:
                provider_errors.append(f"gemini: {gemini_exc}")
                print(f"[WARN] Erreur complete (generate_llm_answer): {' | '.join(provider_errors)}")
                return build_structured_from_rag()
    else:
        try:
            return call_gemini()
        except Exception as gemini_exc:
            provider_errors.append(f"gemini: {gemini_exc}")
            print(f"[WARN] Gemini error in generate_llm_answer, fallback OpenAI: {gemini_exc}")
            try:
                return call_openai()
            except Exception as openai_exc:
                provider_errors.append(f"openai: {openai_exc}")
                print(f"[WARN] Erreur complete (generate_llm_answer): {' | '.join(provider_errors)}")
                return build_structured_from_rag()

# ==========================================
# ROUTES API (Avec analyse IA restaurée)
# ==========================================

@app.get("/")
async def root():
    return {
        "service": "Songra - Agent Yingr-AI",
        "description": "Assistant IA local et souverain au service des communautés du Burkina Faso",
        "organization": "Yingr-AI (Yingr Artificial Intelligence)",
        "mission": "Autonomiser les agriculteurs et éleveurs du Burkina Faso avec l'IA locale",
        "specializations": ["Agriculture BF", "Élevage BF", "Cybersécurité BF"],
        "version": "5.0",
        "features": [
            "Analyse IA texte (classification)",
            "Analyse IA photo (Computer Vision local)",
            "Base de connaissances RAG validée localement",
            "Raisonnement contextuel au Burkina Faso",
            "Fallback multi-niveaux (RAG strict → RAG élargi → Connaissances générales)"
        ]
    }

@app.get("/health")
async def health_check():
    return {
        "service": "Songra/Yingr-AI",
        "status": "ok",
        "timestamp": datetime.utcnow().isoformat(),
        "ia_status": "active",
        "ai_name": "Songra",
        "ai_role": "Assistant IA dédié aux communautés du Burkina Faso"
    }


@app.post("/api/auth/phone/start")
async def start_phone_auth(data: PhoneAuthStartIn, db: Session = Depends(get_db)):
    phone = _normalize_bf_phone(data.phone_number)
    user = db.query(User).filter(User.phone_number == phone).first()
    pin_configured = bool(user and user.password_hash)
    if pin_configured and not data.force_otp:
        return {
            "sent": False,
            "phone_number": phone,
            "account_exists": True,
            "pin_configured": True,
            "expires_in_seconds": 0,
            "message": "Compte reconnu. Entrez votre code PIN.",
        }
    now = datetime.utcnow()
    recent = (
        db.query(PhoneOtpDB)
        .filter(PhoneOtpDB.phone_number == phone, PhoneOtpDB.created_at >= now - timedelta(minutes=1))
        .count()
    )
    if recent >= 1:
        raise HTTPException(status_code=429, detail="Patientez une minute avant de demander un nouveau code")

    code = f"{secrets.randbelow(1000000):06d}"
    otp = PhoneOtpDB(
        phone_number=phone,
        code_hash=hash_password(code),
        purpose="authentication",
        expires_at=now + timedelta(minutes=5),
    )
    db.add(otp)
    db.commit()
    try:
        _send_auth_otp_sms(phone, code)
    except Exception:
        db.delete(otp)
        db.commit()
        raise

    response_payload: Dict[str, Any] = {
        "sent": True,
        "phone_number": phone,
        "account_exists": bool(user),
        "pin_configured": pin_configured,
        "expires_in_seconds": 300,
        "message": "Un code secret à 6 chiffres vient d'être envoyé par SMS.",
    }

    # Mode temporaire tant que les crédits SMS Orange ne sont pas rechargés :
    # aucun SMS n'est réellement envoyé (cf _send_auth_otp_sms), donc on
    # renvoie le code directement à l'app pour ne pas bloquer l'inscription
    # (l'utilisateur final n'a pas accès aux logs backend). Repasser
    # OTP_DEV_MODE=false (ou le retirer) dès que les crédits sont rechargés
    # pour réactiver l'envoi réel par SMS et arrêter d'exposer ce champ.
    if _otp_dev_mode_enabled():
        response_payload["debug_otp"] = code
        response_payload["message"] = (
            f"Mode test (SMS temporairement désactivé) : votre code est {code}."
        )

    return response_payload


@app.post("/api/auth/phone/verify")
async def verify_phone_otp(data: PhoneOtpVerifyIn, db: Session = Depends(get_db)):
    phone = _normalize_bf_phone(data.phone_number)
    now = datetime.utcnow()
    otp = (
        db.query(PhoneOtpDB)
        .filter(PhoneOtpDB.phone_number == phone, PhoneOtpDB.consumed_at.is_(None))
        .order_by(PhoneOtpDB.created_at.desc())
        .first()
    )
    if not otp or otp.expires_at < now:
        raise HTTPException(status_code=401, detail="Code expiré. Demandez un nouveau SMS")
    if otp.attempts >= 5:
        raise HTTPException(status_code=429, detail="Trop d'essais. Demandez un nouveau code")
    if not secrets.compare_digest(otp.code_hash, hash_password((data.code or "").strip())):
        otp.attempts += 1
        db.commit()
        raise HTTPException(status_code=401, detail="Code incorrect")

    user = db.query(User).filter(User.phone_number == phone).first()
    if not user:
        if not (data.name or "").strip():
            raise HTTPException(status_code=422, detail="Indiquez votre nom pour créer le compte")
        if not data.pin:
            raise HTTPException(status_code=422, detail="Choisissez votre code PIN à 4 chiffres")
        user = User(
            phone_number=phone,
            name=data.name.strip(),
            location=(data.location or "").strip() or None,
            password_hash=_hash_pin(data.pin),
            is_premium=False,
            messages_used=0,
            messages_limit=1,
        )
        db.add(user)
    elif data.pin:
        user.password_hash = _hash_pin(data.pin)
        if (data.name or "").strip():
            user.name = data.name.strip()
        if (data.location or "").strip():
            user.location = data.location.strip()

    otp.consumed_at = now
    db.commit()
    db.refresh(user)
    return {"token": create_access_token(user), "user": serialize_user(user)}


@app.post("/api/auth/pin-login")
async def login_with_pin(data: PhonePinLoginIn, db: Session = Depends(get_db)):
    phone = _normalize_bf_phone(data.phone_number)
    user = db.query(User).filter(User.phone_number == phone).first()
    now = datetime.utcnow()
    if user and user.pin_locked_until and user.pin_locked_until > now:
        raise HTTPException(status_code=429, detail="Compte temporairement bloqué. Utilisez le SMS ou réessayez dans 15 minutes")
    if not user or not user.password_hash or not _verify_pin(data.pin, user.password_hash):
        if user:
            user.failed_pin_attempts = int(user.failed_pin_attempts or 0) + 1
            if user.failed_pin_attempts >= 5:
                user.pin_locked_until = now + timedelta(minutes=15)
                user.failed_pin_attempts = 0
            db.commit()
        raise HTTPException(status_code=401, detail="Numéro ou code PIN incorrect")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Ce compte est désactivé")
    user.failed_pin_attempts = 0
    user.pin_locked_until = None
    db.commit()
    return {"token": create_access_token(user), "user": serialize_user(user)}


@app.post("/api/register")
async def register_user(data: UserRegister, db: Session = Depends(get_db)):
    _validate_user_credentials(data.phone_number, data.password)

    existing = db.query(User).filter(User.phone_number == data.phone_number.strip()).first()
    if existing and existing.password_hash:
        raise HTTPException(status_code=409, detail="User already exists")

    if existing:
        user = existing
        user.name = data.name.strip()
        user.location = (data.location or "").strip() or user.location
        user.password_hash = hash_password(data.password)
    else:
        user = User(
            phone_number=data.phone_number.strip(),
            password_hash=hash_password(data.password),
            name=data.name.strip(),
            location=(data.location or "").strip() or None,
            is_premium=False,
            messages_used=0,
            messages_limit=1,
        )
        db.add(user)

    db.commit()
    db.refresh(user)

    return {
        "token": create_access_token(user),
        "user": serialize_user(user),
    }


@app.post("/api/login")
async def login_user(data: UserLogin, db: Session = Depends(get_db)):
    _validate_user_credentials(data.phone_number, data.password)

    user = db.query(User).filter(User.phone_number == data.phone_number.strip()).first()
    if not user or not user.password_hash or not verify_password(data.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    return {
        "token": create_access_token(user),
        "user": serialize_user(user),
    }


@app.get("/api/user")
async def get_authenticated_user(current_user: User = Depends(get_current_user)):
    return serialize_user(current_user)


# @app.get("/api/dashboard")
# async def get_user_dashboard(
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db),
# ):
#     recent_questions = await get_user_tickets(current_user.phone_number, db)
#     recent_questions = recent_questions[:5]
# 
#     preferred_domain = next(
#         (item.get("category") for item in recent_questions if item.get("category")),
#         "agriculture",
#     )
#     ai_suggestions = retrieve_knowledge(db, preferred_domain, preferred_domain, limit=3)
# 
#     community_messages = (
#         db.query(ChatMessageDB)
#         .filter(ChatMessageDB.is_hidden == False)
#         .order_by(ChatMessageDB.created_at.desc())
#         .limit(6)
#         .all()
#     )
# 
#     return {
#         "user": serialize_user(current_user),
#         "recent_questions": recent_questions,
#         "ai_suggestions": [
#             {
#                 "id": item.get("id"),
#                 "title": item.get("title"),
#                 "answer": item.get("answer"),
#                 "domain": item.get("domain"),
#                 "tags": item.get("tags") or [],
#             }
#             for item in ai_suggestions
#         ],
#         "community_activity": [
#             {
#                 "id": message.id,
#                 "sender": message.sender,
#                 "text": message.text,
#                 "is_bot": message.is_bot,
#                 "created_at": message.created_at.isoformat() if message.created_at else None,
#             }
#             for message in community_messages
#         ],
#     }
# 
# 
@app.post("/api/scanner/analyze")
async def analyze_scanner_photo(
    data: MobileQuestionCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Analyse UNIQUEMENT une photo du scanner - pas de ticket, pas de chat IA.
    
    Si `target_lang` est fourni, le diagnostic reste canonique en français et
    un véritable audio humain est recherché dans la base locale validée.
    """

    # Collecter les photos
    photo_payloads = _collect_photo_payloads(data.photo_base64, data.photo_base64_list)

    if not photo_payloads:
        raise HTTPException(status_code=400, detail="Aucune photo fournie")

    # Valider la langue cible si fournie
    valid_langs = {"moore", "dioula", "fulfulde"}
    target_lang = (data.target_lang or "").strip().lower() or None
    if target_lang and target_lang not in valid_langs:
        raise HTTPException(
            status_code=400,
            detail=f"Langue '{target_lang}' non reconnue. Valeurs acceptées : {sorted(valid_langs)}"
        )

    try:
        # 1. Analyser les photos avec Gemini / GPT-4o
        photo_data_list = [_decode_photo_payload(payload) for payload in photo_payloads]
        photo_analysis = cv_engine.analyze_images(
            photo_data_list,
            data.content or "",
            data.category or "agriculture"
        )

        recorded_case = None
        if target_lang:
            recorded_case = _find_recorded_local_case_audio(
                db,
                category=data.category or "agriculture",
                language=target_lang,
                photo_analysis=photo_analysis,
                french_answer=str(photo_analysis.get("analysis") or ""),
            )

        return {
            "status": "success",
            "analysis": photo_analysis,
            "category": data.category,
            "model": photo_analysis.get("model", "gemini-2.5-flash"),
            "translated": False,
            "target_lang": target_lang,
            "lang_name": _LOCAL_LANG_NAMES.get(target_lang) if target_lang else None,
            "local_audio_available": recorded_case is not None,
            "audio_url": recorded_case.get("audio_url") if recorded_case else None,
            "audio_mime_type": recorded_case.get("audio_mime_type") if recorded_case else None,
            "local_knowledge_match": recorded_case,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Analyse scanner photo: {e}")
        return {
            "status": "error",
            "error": str(e),
            "analysis": {"error": str(e), "requires_expert": True}
        }


@app.get("/api/questions")
async def list_mobile_questions(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return await get_user_tickets(current_user.phone_number, db)


@app.post("/api/questions")
async def create_mobile_question(
    data: MobileQuestionCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_resource(db, current_user, "tickets")
    message_data = MessageCreate(
        content=data.content,
        phone_number=current_user.phone_number,
        channel="app",
        category=data.category,
        photo_base64=data.photo_base64,
        photo_base64_list=data.photo_base64_list,
        conversation_context=data.conversation_context,
        target_lang=data.target_lang,
    )
    result = await incoming_sms(message_data, db)
    _consume_resource(db, current_user, "tickets", "mobile_question")
    return result


@app.post("/api/expert-escalations")
async def create_fast_expert_escalation(
    data: MessageCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Crée immédiatement un ticket depuis une fiche déjà analysée.

    Aucun second passage Vision/RAG n'est nécessaire : le résumé de la fiche
    est transmis tel quel à l'expert, ce qui rend l'envoi quasi immédiat.
    """
    _require_resource(db, current_user, "tickets")
    user = current_user

    photo_path = None
    photo_paths: List[str] = []
    photo_payloads = _collect_photo_payloads(data.photo_base64, data.photo_base64_list)
    if photo_payloads:
        photo_data_list = [_decode_photo_payload(payload) for payload in photo_payloads]
        photo_paths = _store_photo_payloads(user.id, photo_data_list, prefix="expert")
        photo_path = photo_paths[0] if photo_paths else None

    ticket = Ticket(
        user_id=user.id,
        category=_normalize_category(data.category),
        urgency="medium",
        status="open",
        preferred_language=_normalize_expert_local_language(data.target_lang),
        photo_path=photo_path,
        photo_paths_json=json.dumps(photo_paths, ensure_ascii=False),
        ai_photo_analysis=json.dumps({
            "source": "studio_result_escalation",
            "submitted_question_and_ai_response": data.content,
            "photo_attached": bool(photo_paths),
            "user": {
                "name": user.name,
                "phone_number": user.phone_number,
                "location": user.location,
            },
        }, ensure_ascii=False),
        internal_notes=json.dumps({
            "source": "studio_result_escalation",
            "user_name": user.name,
            "user_phone": user.phone_number,
            "user_location": user.location,
        }, ensure_ascii=False),
    )
    db.add(ticket)
    db.flush()
    db.add(Message(
        ticket_id=ticket.id,
        sender_type="user",
        sender_id=user.id,
        content=data.content,
        channel="app",
    ))
    db.commit()
    db.refresh(ticket)
    _consume_resource(db, current_user, "tickets", "expert_escalation")
    return {
        "status": "success",
        "ticket_id": ticket.id,
        "category": ticket.category,
        "photo_analysis": None,
    }


@app.get("/api/questions/{question_id}")
async def get_mobile_question_detail(
    question_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ticket = db.query(Ticket).filter(Ticket.id == question_id, Ticket.user_id == current_user.id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Question not found")

    # Appel interne : passer explicitement la session DB. L'ancien appel
    # positionnel plaçait la session dans current_expert et laissait `db`
    # avec l'objet Depends, empêchant la récupération des nouveaux messages.
    ticket_detail = await get_ticket_detail(question_id, current_expert=None, db=db)
    ai_summary_data = await get_ticket_ai_summary(question_id, db)

    return {
        "question": ticket_detail["ticket"],
        "user": ticket_detail["user"],
        "messages": ticket_detail["messages"],
        "ai_summary": ai_summary_data.get("ai_summary"),
        "rag_items": ai_summary_data.get("rag_items", []),
        "latest_expert_answer": next(
            (message["content"] for message in reversed(ticket_detail["messages"]) if message.get("sender_type") == "expert"),
            None,
        ),
    }

@app.post("/api/auth/login")
async def login(data: ExpertLogin, db: Session = Depends(get_db)):
    expert = db.query(Expert).filter(Expert.email == data.email).first()
    if not expert or not expert.is_active or not verify_password(data.password, expert.password_hash):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Migration transparente des anciens SHA-256 vers bcrypt sans casser les comptes.
    if not expert.password_hash.startswith("$2"):
        expert.password_hash = hash_expert_password(data.password)
        db.commit()
    return {
        "token": create_expert_access_token(expert),
        "expert": {
            "id": expert.id,
            "name": expert.full_name,
            "email": expert.email,
            "specialization": expert.specialization,
            "role": getattr(expert, "role", "expert"),
            "organization": expert.institution,
            "organization_id": getattr(expert, "organization_id", None),
            "zone": expert.zone,
            "language": expert.language,
            "project": expert.project,
        }
    }


def _normalized_scope_value(value: Optional[str]) -> str:
    normalized = (value or "").strip().lower()
    aliases = {
        "élevage": "elevage",
        "agricole": "agriculture",
        "cybersécurité": "cybersecurity",
        "cybersecurite": "cybersecurity",
        "cyber sécurité": "cybersecurity",
        "santé": "sante",
    }
    return aliases.get(normalized, normalized)


def _scope_ticket_query_for_expert(query: Any, expert: Expert) -> Any:
    """Applique le cloisonnement metier au niveau SQL, jamais seulement dans l'UI."""
    domain = _normalized_scope_value(expert.specialization)
    if domain and domain not in {"all", "admin", "general", "général", "tous"}:
        query = query.filter(func.lower(func.coalesce(Ticket.category, "")) == domain)

    organization_id = getattr(expert, "organization_id", None)
    organization = _normalized_scope_value(expert.institution)
    if organization_id:
        query = query.join(User, User.id == Ticket.user_id).filter(
            User.organization_id == organization_id
        )
    elif organization:
        query = query.join(User, User.id == Ticket.user_id).filter(
            func.lower(func.coalesce(User.organization, "")) == organization
        )
    return query


def _expert_ticket_or_404(db: Session, expert: Expert, ticket_id: int) -> Ticket:
    ticket = _scope_ticket_query_for_expert(
        db.query(Ticket).filter(Ticket.id == ticket_id), expert
    ).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Dossier introuvable dans votre périmètre")
    return ticket


def _serialize_expert_ticket(db: Session, ticket: Ticket) -> Dict[str, Any]:
    user = db.query(User).filter(User.id == ticket.user_id).first()
    last_msg = db.query(Message).filter(Message.ticket_id == ticket.id).order_by(Message.sent_at.desc()).first()
    photo_paths = _load_json_list(ticket.photo_paths_json)
    return {
        "id": ticket.id,
        "category": ticket.category or "agriculture",
        "urgency": ticket.urgency or "low",
        "status": ticket.status or "open",
        "preferred_language": _normalize_expert_local_language(ticket.preferred_language),
        "created_at": ticket.created_at.isoformat() if ticket.created_at else None,
        "last_message": last_msg.content if last_msg else "Aucun message",
        "ai_confidence": ticket.ai_confidence_score,
        "photo_url": _build_upload_url(ticket.photo_path),
        "photo_urls": [_build_upload_url(path) for path in photo_paths if path],
        "user": {
            "id": user.id if user else None,
            "name": user.name if user else None,
            "phone": user.phone_number if user else None,
            "location": user.location if user else None,
            "organization": getattr(user, "organization", None) if user else None,
        },
    }


@app.get("/api/expert/dashboard")
async def expert_dashboard(
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    scoped = _scope_ticket_query_for_expert(db.query(Ticket), current_expert)
    tickets = scoped.all()
    return {
        "expert": {
            "id": current_expert.id,
            "name": current_expert.full_name,
            "email": current_expert.email,
            "specialization": current_expert.specialization,
            "organization": current_expert.institution,
            "organization_id": getattr(current_expert, "organization_id", None),
            "role": getattr(current_expert, "role", "expert"),
            "zone": current_expert.zone,
            "language": current_expert.language,
            "project": current_expert.project,
        },
        "stats": {
            "total": len(tickets),
            "open": sum(1 for item in tickets if item.status in {None, "open", "awaiting_expert"}),
            "assigned": sum(1 for item in tickets if item.status == "assigned"),
            "resolved": sum(1 for item in tickets if item.status == "resolved"),
            "urgent": sum(1 for item in tickets if _normalized_scope_value(item.urgency) in {"high", "urgent", "critical"}),
        },
    }


@app.get("/api/expert/tickets")
async def expert_tickets(
    status: Optional[str] = None,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    query = _scope_ticket_query_for_expert(db.query(Ticket), current_expert)
    if status:
        query = query.filter(Ticket.status == status)
    tickets = query.order_by(Ticket.created_at.desc()).all()
    return [_serialize_expert_ticket(db, ticket) for ticket in tickets]


@app.get("/api/expert/users")
async def expert_users(
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    scoped_ticket_ids = _scope_ticket_query_for_expert(db.query(Ticket.user_id), current_expert).subquery()
    users = db.query(User).filter(User.id.in_(scoped_ticket_ids)).order_by(User.created_at.desc()).all()
    return [
        {
            **serialize_user(user),
            "ticket_count": _scope_ticket_query_for_expert(
                db.query(Ticket).filter(Ticket.user_id == user.id), current_expert
            ).count(),
        }
        for user in users
    ]


@app.get("/api/expert/tickets/{ticket_id}")
async def expert_ticket_detail(
    ticket_id: int,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    ticket = _expert_ticket_or_404(db, current_expert, ticket_id)
    result = _serialize_expert_ticket(db, ticket)
    messages = db.query(Message).filter(Message.ticket_id == ticket.id).order_by(Message.sent_at).all()
    result["messages"] = [
        {
            "id": message.id,
            "content": message.content,
            "sender_type": message.sender_type,
            "sent_at": message.sent_at.isoformat() if message.sent_at else None,
            "audio_url": _build_upload_url(message.audio_url) if message.audio_url else None,
            "language": message.language,
        }
        for message in messages
    ]
    result["photo_analysis"] = ticket.ai_photo_analysis
    # Retrouver également la fiche Studio associée afin que l'expert puisse
    # consulter sa consigne et écouter sa voix locale depuis le ticket.
    user_text = "\n".join(
        (message.content or "").strip()
        for message in messages
        if message.sender_type == "user" and (message.content or "").strip()
    )
    parsed_photo_analysis: Dict[str, Any] = {}
    if ticket.ai_photo_analysis:
        try:
            raw_analysis = json.loads(ticket.ai_photo_analysis)
            if isinstance(raw_analysis, dict):
                parsed_photo_analysis = raw_analysis
        except Exception:
            parsed_photo_analysis = {"analysis": ticket.ai_photo_analysis}
    studio_match = _find_studio_knowledge_match(
        db,
        category=ticket.category or "agriculture",
        query_text=user_text,
        photo_analysis=parsed_photo_analysis,
    )
    if studio_match:
        language = _normalize_expert_local_language(ticket.preferred_language)
        audio_map = _normalize_expert_local_audio(studio_match.get("audio") or {})
        selected_audio = audio_map.get(language) if isinstance(audio_map.get(language), dict) else {}
        result["knowledge_card"] = {
            "id": studio_match.get("id"),
            "title": studio_match.get("title"),
            "resolution_fr": studio_match.get("resolution_fr"),
            "language": language,
            "audio_url": selected_audio.get("url"),
            "audio_mime_type": selected_audio.get("mime_type"),
            "audio": audio_map,
        }
    return result


@app.post("/api/expert/tickets/{ticket_id}/reply")
async def expert_reply(
    ticket_id: int,
    content: ReplyMessage,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    ticket = _expert_ticket_or_404(db, current_expert, ticket_id)
    message_text = content.message.strip()
    if not message_text:
        raise HTTPException(status_code=422, detail="La réponse est vide")
    reply_language = _normalize_expert_local_language(content.language or ticket.preferred_language)
    db.add(Message(ticket_id=ticket.id, sender_type="expert", sender_id=current_expert.id, content=message_text, channel="mobile_expert", language=reply_language))
    ticket.expert_id = current_expert.id
    if ticket.status in {None, "open", "awaiting_expert"}:
        ticket.status = "assigned"
    db.commit()
    return {"status": "success", "ticket_status": ticket.status}


@app.post("/api/expert/tickets/{ticket_id}/reply-audio")
async def expert_reply_audio(
    ticket_id: int,
    audio: UploadFile = File(...),
    language: Optional[str] = Form(None),
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    """Ajoute une réponse vocale au ticket dans le périmètre de l'expert."""
    ticket = _expert_ticket_or_404(db, current_expert, ticket_id)
    content = await audio.read()
    if not content:
        raise HTTPException(status_code=400, detail="Le message vocal est vide")
    extension = os.path.splitext(audio.filename or "message.m4a")[1].lower()
    if extension not in {".webm", ".ogg", ".mp3", ".wav", ".m4a", ".mp4", ".mpeg", ".aac"}:
        extension = ".m4a"
    relative_path = os.path.join(
        "uploads", "audio", "expert-replies",
        f"expert_{current_expert.id}_ticket_{ticket.id}_{int(time.time() * 1000)}{extension}",
    ).replace("\\", "/")
    absolute_path = os.path.abspath(relative_path)
    _ensure_parent_dir(absolute_path)
    with open(absolute_path, "wb") as handle:
        handle.write(content)
    reply_language = _normalize_expert_local_language(language or ticket.preferred_language)
    db.add(Message(
        ticket_id=ticket.id,
        sender_type="expert",
        sender_id=current_expert.id,
        content="Réponse vocale de l’expert",
        channel="mobile_expert",
        audio_url=relative_path,
        language=reply_language,
    ))
    ticket.expert_id = current_expert.id
    if ticket.status in {None, "open", "awaiting_expert"}:
        ticket.status = "assigned"
    db.commit()
    return {
        "status": "success",
        "ticket_status": ticket.status,
        "audio_url": _build_upload_url(relative_path),
        "language": reply_language,
    }


@app.put("/api/expert/tickets/{ticket_id}/status")
async def expert_ticket_status(
    ticket_id: int,
    payload: Dict[str, Any],
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    ticket = _expert_ticket_or_404(db, current_expert, ticket_id)
    requested_status = str(payload.get("status") or "").strip()
    if requested_status not in {"open", "assigned", "awaiting_expert", "resolved"}:
        raise HTTPException(status_code=422, detail="Statut invalide")
    ticket.status = requested_status
    ticket.expert_id = current_expert.id
    ticket.resolved_at = datetime.utcnow() if requested_status == "resolved" else None
    db.commit()
    return {"status": "success", "new_status": ticket.status}

@app.get("/api/tickets")
async def get_tickets(
    status: Optional[str] = None,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db)
):
    del current_expert
    query = db.query(Ticket)
    if status:
        query = query.filter(Ticket.status == status)
    
    tickets = query.order_by(Ticket.created_at.desc()).all()
    
    result = []
    for ticket in tickets:
        user = db.query(User).filter(User.id == ticket.user_id).first()
        last_msg = db.query(Message).filter(
            Message.ticket_id == ticket.id
        ).order_by(Message.sent_at.desc()).first()
        latest_expert_msg = db.query(Message).filter(
            Message.ticket_id == ticket.id,
            Message.sender_type == "expert",
        ).order_by(Message.sent_at.desc(), Message.id.desc()).first()
        
        # Construire l'URL de la photo
        photo_url = _build_upload_url(ticket.photo_path)
        photo_paths = _load_json_list(ticket.photo_paths_json)
        photo_urls = [_build_upload_url(path) for path in photo_paths if path]
        
        result.append({
            "id": ticket.id,
            "user_phone": user.phone_number if user else "Inconnu",
            "category": ticket.category or "agriculture",
            "urgency": ticket.urgency or "low",
            "status": ticket.status or "open",
            "preferred_language": _normalize_expert_local_language(ticket.preferred_language),
            "created_at": ticket.created_at,
            "last_message": last_msg.content if last_msg else "Aucun message",
            "ai_confidence": ticket.ai_confidence_score,
            "has_photo": ticket.photo_path is not None,
            "photo_url": photo_url,
            "photo_urls": photo_urls,
            "photo_path": ticket.photo_path,
            "has_photo_analysis": ticket.ai_photo_analysis is not None
        })
    
    return result

@app.get("/api/stats")
async def get_stats(
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    del current_expert
    total_tickets = db.query(Ticket).count()
    open_tickets = db.query(Ticket).filter(Ticket.status == "open").count()
    assigned_tickets = db.query(Ticket).filter(Ticket.status == "assigned").count()
    
    # Tickets résolus aujourd'hui
    today = datetime.utcnow().date()
    resolved_today = db.query(Ticket).filter(
        Ticket.status == "resolved",
        func.date(Ticket.resolved_at) == today
    ).count()
    
    tickets_with_photos = db.query(Ticket).filter(
        Ticket.photo_path.isnot(None)
    ).count()
    
    return {
        "total_tickets": total_tickets,
        "open_tickets": open_tickets,
        "assigned_tickets": assigned_tickets,
        "resolved_today": resolved_today,
        "tickets_with_photos": tickets_with_photos
    }

@app.get("/api/tickets/{ticket_id}")
async def get_ticket_detail(
    ticket_id: int,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    del current_expert
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    user = db.query(User).filter(User.id == ticket.user_id).first()
    messages = db.query(Message).filter(
        Message.ticket_id == ticket_id
    ).order_by(Message.sent_at).all()
    
    # Gérer l'analyse photo
    photo_analysis = None
    if ticket.ai_photo_analysis:
        try:
            if isinstance(ticket.ai_photo_analysis, str):
                photo_analysis = json.loads(ticket.ai_photo_analysis)
            else:
                photo_analysis = ticket.ai_photo_analysis
        except:
            photo_analysis = {"error": "Failed to parse analysis"}
    
    # Construire l'URL de la photo
    photo_url = _build_upload_url(ticket.photo_path)
    photo_paths = _load_json_list(ticket.photo_paths_json)
    photo_urls = [_build_upload_url(path) for path in photo_paths if path]
    
    # Extraire les mots-clés
    keywords = []
    if ticket.ai_extracted_keywords:
        try:
            keywords = json.loads(ticket.ai_extracted_keywords)
        except:
            keywords = []
    
    return {
        "ticket": {
            "id": ticket.id,
            "category": ticket.category or "agriculture",
            "urgency": ticket.urgency or "low",
            "status": ticket.status or "open",
            "preferred_language": _normalize_expert_local_language(ticket.preferred_language),
            "keywords": keywords,
            "confidence": ticket.ai_confidence_score or 0.5,
            "photo_url": photo_url,
            "photo_urls": photo_urls,
            "photo_path": ticket.photo_path,
            "photo_filename": ticket.photo_path,  # Alias pour le frontend expert
            "photo_analysis": photo_analysis,
            "created_at": ticket.created_at,
            "resolved_at": ticket.resolved_at
        },
        "user": {
            "phone": user.phone_number if user else None,
            "name": user.name if user else None,
            "location": user.location if user else None
        },
        "messages": [{
            "id": msg.id,
            "content": msg.content,
            "sender_type": msg.sender_type,
            "sent_at": msg.sent_at,
            "audio_url": _build_upload_url(msg.audio_url) if msg.audio_url else None
            , "language": msg.language
        } for msg in messages]
    }

@app.post("/api/webhooks/incoming-sms")
async def incoming_sms(data: MessageCreate, db: Session = Depends(get_db)):
    # 1. Trouver ou créer l'utilisateur
    user = db.query(User).filter(User.phone_number == data.phone_number).first()
    if not user:
        user = User(phone_number=data.phone_number)
        db.add(user)
        db.commit()
        db.refresh(user)
    
    # 2. Analyse IA texte (RESTAURÉ)
    ai_result = ai_engine.classify(data.content)

    # 2.bis Déterminer la catégorie finale et le domaine RAG
    # On combine la catégorie choisie dans l'app (data.category) et
    # la catégorie NLP locale.
    nlp_category = ai_result.get("category", "agriculture")
    chosen_category = data.category or nlp_category
    ai_result["classifier_category"] = nlp_category
    ai_result["category"] = chosen_category

    # Mapping catégorie -> domaine de la base de connaissances
    if chosen_category == "agriculture":
        kb_domain = "agriculture"
    elif chosen_category == "elevage":
        kb_domain = "elevage"
    elif chosen_category == "sos_accident":
        kb_domain = "health"  # les fiches premiers soins sont dans le domaine health
    elif chosen_category == "cybersecurity":
        kb_domain = "cybersecurity"
    else:
        kb_domain = "agriculture"
    
    # ✅ DÉTECTION D'URGENCE - CRITÈRE AUTOMATIQUE AVANT RAG
    print("\n[EMERGENCY-CHECK] Vérification d'urgence critique...")
    emergency_info = detect_emergency(data.content, chosen_category)
    
    if emergency_info["is_emergency"]:
        print(f"🚨 [EMERGENCY-DETECTED] Type: {emergency_info['emergency_type']}, Sévérité: {emergency_info['severity']}")
        print(f"   Actions: {len(emergency_info['immediate_actions'])} étapes numérotées")
        print(f"   Numéro secours: {emergency_info['call_emergency_number']}")
        
        # Créer la réponse d'urgence avec protocoles de premiers secours
        emergency_response = f"""🚨🚨🚨 SITUATION D'URGENCE CRITIQUE 🚨🚨🚨

{emergency_info['warning_alert']}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

**ACTIONS IMMÉDIATES À FAIRE MAINTENANT:**

"""
        for idx, action in enumerate(emergency_info['immediate_actions'], 1):
            emergency_response += f"{action}\n"
        
        emergency_response += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

**APPELEZ IMMÉDIATEMENT LES SECOURS:**
🚑 {emergency_info['call_emergency_number']}

⏱️ CHAQUE MINUTE EST CRUCIALE
Ne perdez pas de temps à discuter - commencez les gestes de premiers secours MAINTENANT
Appelez les secours pendant que vous effectuez ces gestes

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        
        # Créer le ticket normalement mais marquer comme URGENCE CRITIQUE
        ticket = Ticket(
            user_id=user.id,
            category=chosen_category,
            urgency="critical",
            ai_confidence_score=100,
            ai_extracted_keywords=json.dumps(["urgence", "SOS", emergency_info["emergency_type"]], ensure_ascii=False),
            ai_photo_analysis=None,
            photo_path=None,
            status="emergency",
            preferred_language=_normalize_expert_local_language(data.target_lang),
            internal_notes=json.dumps({
                "emergency_type": emergency_info["emergency_type"],
                "severity": emergency_info["severity"],
                "protocol": emergency_info["emergency_type"],
                "auto_detected": True
            }, ensure_ascii=False)
        )
        db.add(ticket)
        db.commit()
        db.refresh(ticket)
        
        # Créer le message utilisateur
        message = Message(
            ticket_id=ticket.id,
            sender_type="user",
            sender_id=user.id,
            content=data.content,
            channel=data.channel
        )
        db.add(message)
        db.commit()
        
        # Créer le message de réponse d'urgence
        ai_message = Message(
            ticket_id=ticket.id,
            sender_type="assistant",
            sender_id=None,
            content=emergency_response,
            channel=data.channel,
            is_ai_generated=True
        )
        db.add(ai_message)
        db.commit()
        
        # Retourner la réponse d'urgence
        return {
            "status": "success",
            "ticket_id": ticket.id,
            "is_emergency": True,
            "severity": emergency_info["severity"],
            "emergency_type": emergency_info["emergency_type"],
            "llm_answer": emergency_response,
            "ai_analysis": {
                "category": chosen_category,
                "urgency": "critical",
                "confidence": 100,
                "emergency_detected": True,
                "emergency_protocol": emergency_info["emergency_type"]
            }
        }
    
    # 3. Analyse photo si présente (RESTAURÉ)
    photo_analysis = None
    photo_path = None
    photo_paths: List[str] = []
    
    photo_payloads = _collect_photo_payloads(data.photo_base64, data.photo_base64_list)
    if photo_payloads:
        try:
            photo_data_list = [_decode_photo_payload(payload) for payload in photo_payloads]
            photo_analysis_result = cv_engine.analyze_images(photo_data_list, data.content, data.category)
            photo_analysis = json.dumps(photo_analysis_result, ensure_ascii=False)
            
            photo_paths = _store_photo_payloads(user.id, photo_data_list, prefix="ticket")
            best_index = max(0, photo_analysis_result.get("best_view_index", 1) - 1)
            photo_path = photo_paths[best_index] if best_index < len(photo_paths) else photo_paths[0]
            
            # Ajuster urgence si maladie grave détectée
            if photo_analysis_result.get("urgency") == "high":
                ai_result["urgency"] = "high"
                
        except Exception as e:
            print(f"Erreur analyse photo: {e}")
            photo_analysis = json.dumps({"error": str(e), "requires_expert": True})
    
    # 3.bis RAG strict d'abord, base élargie seulement si le RAG ne répond pas.
    photo_analysis_payload = None
    if photo_analysis:
        try:
            photo_analysis_payload = json.loads(photo_analysis)
        except Exception:
            photo_analysis_payload = None

    focus_context = extract_focus_context(
        chosen_category,
        data.content,
        photo_analysis_payload,
    )

    knowledge_result = resolve_knowledge_answer(
        db=db,
        domain=kb_domain,
        question=data.content,
        language="fr",
        focus_context=focus_context,
        photo_analysis=photo_analysis_payload,
    )
    rag_items = knowledge_result["rag_items"]
    llm_answer = knowledge_result["llm_answer"]

    # LOG DEBUG : affichage des fiches RAG et de la réponse LLM (si disponible)
    try:
        print("[RAG] Domaine:", kb_domain)
        knowledge_mode = knowledge_result['knowledge_mode']
        fallback_used = knowledge_result['knowledge_fallback_used']
        
        if knowledge_mode == "rag_strict":
            print(f"[OK] [RAG-STRICT] {len(rag_items)} fiche(s) trouvée(s) dans le domaine exact")
        elif knowledge_mode == "rag_expanded":
            print(f"⚠ [RAG-EXPANDED] {len(rag_items)} fiche(s) trouvée(s) (recherche élargie dans d'autres domaines)")
        elif knowledge_mode == "llm_general_knowledge":
            print(f"[INFO] [LLM-GENERAL] Pas de fiche RAG, utilisation des connaissances générales de Lia")
        elif knowledge_mode == "no_match":
            print(f"[ERROR] [NO-MATCH] Aucune connaissance trouvée, réponse générique")
        
        for idx, item in enumerate(rag_items, start=1):
            print(f"  - FICHE {idx}: {item.get('title')} (domaine: {item.get('domain')})")

        studio_match_debug = knowledge_result.get("studio_match")
        if studio_match_debug:
            print(
                f"[FICHE] Reponse basee sur la fiche Studio #{studio_match_debug.get('id')} "
                f"'{studio_match_debug.get('title')}' (score={studio_match_debug.get('match_score')})"
            )

        if llm_answer:
            if knowledge_mode == "llm_general_knowledge":
                print("[LLM-GENERAL] Réponse générée avec connaissances générales (début):", llm_answer[:250].replace("\n", " "))
            else:
                print("[LLM] Réponse générée à partir de RAG (début):", llm_answer[:250].replace("\n", " "))
        else:
            print("[LLM] Aucune réponse LLM générée")
    except Exception as e_log:
        print(f"[DEBUG] Erreur lors du log RAG/LLM: {e_log}")

    # 4. Créer ticket avec analyse IA
    ticket = Ticket(
        user_id=user.id,
        category=chosen_category,
        urgency=ai_result["urgency"],
        ai_confidence_score=ai_result["confidence"],
        ai_extracted_keywords=json.dumps(ai_result["keywords"], ensure_ascii=False),
        ai_photo_analysis=photo_analysis,
        photo_path=photo_path,
        photo_paths_json=json.dumps(photo_paths, ensure_ascii=False) if photo_paths else None,
        status="open",
        preferred_language=_normalize_expert_local_language(data.target_lang),
    )
    db.add(ticket)
    db.commit()
    db.refresh(ticket)
    
    # 5. Créer message - si photo_analysis existe, sauvegarder l'analyse formatée comme premier message
    # Car l'analyse de l'image EST la première demande, pas le texte générique "Question 1"
    message_content = data.content
    if photo_analysis_payload:
        # Formater l'analyse de l'image comme premier message utilisateur
        message_content = f"""📸 ANALYSE DE L'IMAGE - {photo_analysis_payload.get('disease_detected', 'Analyse détectée')}

**Diagnostic:** {photo_analysis_payload.get('disease_detected', 'Non identifié')}
**Confiance:** {(photo_analysis_payload.get('confidence', 0) * 100):.0f}%
**Urgence:** {photo_analysis_payload.get('urgency', 'Normal')}

{photo_analysis_payload.get('analysis', photo_analysis_payload.get('recommendations', 'Voir analyse complète ci-dessous'))}"""
    
    message = Message(
        ticket_id=ticket.id,
        sender_type="user",
        sender_id=user.id,
        content=message_content,
        channel=data.channel
    )
    db.add(message)
    db.commit()
    
    # 6. Retourner résultat avec analyse IA COMPLÈTE
    response = {
        "status": "success",
        "ticket_id": ticket.id,
        "ai_analysis": ai_result,
        "knowledge_mode": knowledge_result["knowledge_mode"],
        "knowledge_fallback_used": knowledge_result["knowledge_fallback_used"],
    }
    
    # Ajouter l'analyse photo si disponible
    if photo_analysis:
        try:
            response["photo_analysis"] = photo_analysis_payload or json.loads(photo_analysis)
        except:
            response["photo_analysis"] = {"analysis": "Analyse photo en cours"}
    
    # Ajouter l'URL de la photo
    if photo_path:
        response["photo_url"] = _build_upload_url(photo_path)
    if photo_paths:
        response["photo_urls"] = [_build_upload_url(path) for path in photo_paths]

    # Toujours retourner les fiches RAG utilisées (pour debug et fallback côté frontend)
    if rag_items:
        response["rag_items"] = rag_items

    # Identifier explicitement la fiche du Studio utilisée (si une fiche a
    # été retenue), pour permettre à l'app/aux logs de savoir précisément
    # quelle fiche a servi de base à la réponse.
    studio_match_info = knowledge_result.get("studio_match")
    response["knowledge_fiche_id"] = studio_match_info.get("id") if studio_match_info else None
    response["knowledge_fiche_title"] = studio_match_info.get("title") if studio_match_info else None

    # Ajouter la réponse principale générée par le LLM si disponible
    if llm_answer:
        response["llm_answer"] = llm_answer
    elif knowledge_result["rag_fallback_answer"]:
        # Fallback déterministe : utiliser la réponse de la meilleure fiche
        # pour que l'utilisateur ait au moins la réponse validée locale,
        # même si la clé OpenAI n'est pas configurée.
        response["rag_fallback_answer"] = knowledge_result["rag_fallback_answer"]

    return response


def _find_recorded_local_case_audio(
    db: Session,
    *,
    category: str,
    language: str,
    photo_analysis: Optional[Dict[str, Any]],
    french_answer: str,
) -> Optional[Dict[str, Any]]:
    """Trouve une fiche française validée possédant un vrai audio humain local."""
    if language not in _TRANSLATOR_VALID_LANGS:
        return None
    if photo_analysis and photo_analysis.get("problem_status") != "identified":
        return None

    normalized_category = _normalize_expert_local_category(category)
    candidates = (
        db.query(ExpertLocalKnowledgeDB)
        .filter(
            ExpertLocalKnowledgeDB.category == normalized_category,
            ExpertLocalKnowledgeDB.status.in_(["validated", "resolved", "expert_verified"]),
        )
        .order_by(ExpertLocalKnowledgeDB.updated_at.desc())
        .limit(500)
        .all()
    )

    analysis = photo_analysis or {}
    diagnostic_parts: List[str] = [
        str(analysis.get("problem_label") or ""),
        str(analysis.get("disease_detected") or ""),
        str(analysis.get("diagnosis") or ""),
        str(analysis.get("situation_type") or ""),
        str(analysis.get("threat_type") or ""),
        str(analysis.get("analysis") or ""),
        french_answer or "",
    ]
    for key in ("all_symptoms", "visible_symptoms", "symptoms", "red_flags"):
        value = analysis.get(key)
        if isinstance(value, list):
            diagnostic_parts.extend(str(item) for item in value)

    diagnostic_text = " ".join(part for part in diagnostic_parts if part).strip()
    diagnostic_tokens = set(_tokenize(diagnostic_text))
    if not diagnostic_tokens:
        return None

    best: Optional[Tuple[float, ExpertLocalKnowledgeDB, Dict[str, Any]]] = None
    for item in candidates:
        audio_map = _load_json_dict(item.audio_json)
        audio = audio_map.get(language)
        if not isinstance(audio, dict) or not str(audio.get("url") or "").strip():
            continue

        tags = [str(tag) for tag in _load_json_list(item.tags_json)]
        title_tokens = set(_tokenize(item.title or ""))
        question_tokens = set(_tokenize(item.question_fr or ""))
        resolution_tokens = set(_tokenize(item.resolution_fr or ""))
        tag_tokens = set(_tokenize(" ".join(tags)))
        score = (
            4.0 * len(diagnostic_tokens & tag_tokens)
            + 3.0 * len(diagnostic_tokens & title_tokens)
            + 1.5 * len(diagnostic_tokens & question_tokens)
            + 0.5 * len(diagnostic_tokens & resolution_tokens)
        )
        normalized_title = _normalize_search_text(item.title or "")
        normalized_problem = _normalize_search_text(str(analysis.get("problem_label") or ""))
        if normalized_title and normalized_problem and (
            normalized_title in normalized_problem or normalized_problem in normalized_title
        ):
            score += 12.0

        if best is None or score > best[0]:
            best = (score, item, audio)

    if best is None or best[0] < 4.0:
        return None

    score, item, audio = best
    translations = _load_json_dict(item.translations_json)
    local_text = translations.get(language) if isinstance(translations.get(language), dict) else {}
    return {
        "knowledge_id": item.id,
        "title": item.title,
        "question_fr": item.question_fr,
        "resolution_fr": item.resolution_fr,
        "match_score": round(score, 2),
        "language": language,
        "audio_url": str(audio.get("url") or "").strip(),
        "audio_mime_type": str(audio.get("mime_type") or "audio/webm"),
        "local_text": str((local_text or {}).get("text") or "").strip() or None,
        "source": "recorded_expert_local_knowledge",
    }


@app.post("/api/assistant/query")
async def assistant_query(data: MessageCreate, db: Session = Depends(get_db)):
    """Endpoint conversation IA seule (RAG + GPT) sans création de ticket.

    Utilisé par l'application pour discuter avec l'IA et affiner le problème.
    Aucun Ticket/Message n'est créé ici, uniquement une réponse IA.
    """
    # ── TRADUCTION DE LA REQUETE DE LANGUE LOCALE VERS LE FRANCAIS ──────────────
    original_query = (data.content or "").strip()
    search_query_fr = original_query
    reconstructed_query_local = original_query
    query_interpretation_confidence = 1.0
    target_lang = (data.target_lang or "").strip().lower() or None
    # Aucune traduction : Groq analyse la photo/le texte, puis la langue cible
    # sert seulement à sélectionner l'audio humain de la fiche Studio.

    # 1. Analyse IA texte
    ai_result = ai_engine.classify(search_query_fr)


    # 2. Déterminer la catégorie finale et le domaine RAG
    nlp_category = ai_result.get("category", "agriculture")
    allowed_categories = {"agriculture", "elevage", "sos_accident", "cybersecurity"}
    chosen_category = data.category if data.category in allowed_categories else nlp_category
    ai_result["classifier_category"] = nlp_category
    ai_result["category"] = chosen_category

    if chosen_category == "agriculture":
        kb_domain = "agriculture"
    elif chosen_category == "elevage":
        kb_domain = "elevage"
    elif chosen_category == "sos_accident":
        kb_domain = "health"
    elif chosen_category == "cybersecurity":
        kb_domain = "cybersecurity"
    else:
        kb_domain = "agriculture"

    # 3. Analyse photo si présente (sans sauvegarder de fichier)
    photo_analysis = None
    photo_payloads = _collect_photo_payloads(data.photo_base64, data.photo_base64_list)
    if photo_payloads:
        try:
            photo_data_list = [_decode_photo_payload(payload) for payload in photo_payloads]
            photo_analysis_result = cv_engine.analyze_images(photo_data_list, search_query_fr, chosen_category)
            photo_analysis = photo_analysis_result

            if photo_analysis_result.get("urgency") == "high":
                ai_result["urgency"] = "high"
        except Exception as e:
            print(f"Erreur analyse photo (assistant_query): {e}")
            photo_analysis = {"error": str(e), "requires_expert": True}

    # 4. Construire le contexte conversationnel et enrichir la recherche RAG
    conversation_context = []
    if data.conversation_context:
        conversation_context = [
            {"role": turn.role, "content": turn.content}
            for turn in data.conversation_context
            if (turn.content or "").strip()
        ]

    contextual_query_parts = [
        turn["content"]
        for turn in conversation_context[-4:]
    ]
    contextual_query_parts.append(search_query_fr)
    contextual_query = "\n".join(part for part in contextual_query_parts if part)

    # 5. RAG strict d'abord, puis base élargie seulement si besoin
    focus_context = extract_focus_context(
        chosen_category,
        contextual_query,
        photo_analysis,
    )

    # Les recherches utilisateur consultent le Studio, pas l'historique des tickets.
    reusable_entry = None
    knowledge_result = resolve_knowledge_answer(
        db=db,
        domain=kb_domain,
        question=contextual_query,
        language="fr",
        conversation_context=conversation_context,
        focus_context=focus_context,
        photo_analysis=photo_analysis,
    )
    rag_items = knowledge_result["rag_items"]
    llm_answer = _clean_assistant_text(knowledge_result["llm_answer"])
    if knowledge_result.get("rag_fallback_answer"):
        knowledge_result["rag_fallback_answer"] = _clean_assistant_text(
            knowledge_result["rag_fallback_answer"]
        )

    image_result = None
    video_result = None

    if data.generate_media:
        try:
            media_category = _normalize_category(chosen_category)
            media_images = _collect_images_b64(data.photo_base64, data.photo_base64_list)
            media_seed_parts = [
                search_query_fr,
                llm_answer,
                knowledge_result.get("rag_fallback_answer"),
                photo_analysis.get("analysis") if isinstance(photo_analysis, dict) else None,
                photo_analysis.get("diagnosis") if isinstance(photo_analysis, dict) else None,
                photo_analysis.get("observations") if isinstance(photo_analysis, dict) else None,
            ]
            media_seed_text = "\n".join(part.strip() for part in media_seed_parts if isinstance(part, str) and part.strip())

            media_analysis = await v2_services.gemini_analyze(
                text=media_seed_text or search_query_fr or "Diagnostic Songra",
                images_b64=media_images,
                category=media_category,
            )
            media_decision = v2_services.decide(media_analysis)

            if media_decision.get("generer_image") and media_decision.get("prompt_image"):
                style = "schema" if media_decision.get("mode_urgence") else "illustration"
                image_result = await v2_services.generate_image(
                    media_decision["prompt_image"],
                    style=style,
                    category=media_category,
                )

            if media_decision.get("generer_video") and media_decision.get("prompt_video"):
                video_result = await v2_services.generate_video(
                    media_decision["prompt_video"],
                    gemini_api_key=GEMINI_API_KEY,
                    duration_sec=5 if media_decision.get("mode_urgence") else 8,
                    is_urgency=media_decision.get("mode_urgence", False),
                    category=media_category,
                )
        except Exception as e:
            print(f"[ASSISTANT] Erreur génération média: {e}")

    # Log the knowledge source
    try:
        knowledge_mode = knowledge_result['knowledge_mode']
        if knowledge_mode == "rag_strict":
            print(f"[OK] [ASSISTANT] RAG-STRICT | {len(rag_items)} fiches trouvées")
        elif knowledge_mode == "rag_expanded":
            print(f"⚠ [ASSISTANT] RAG-EXPANDED | {len(rag_items)} fiches (recherche élargie)")
        elif knowledge_mode == "llm_general_knowledge":
            print(f"[INFO] [ASSISTANT] LLM-GENERAL | Connaissances générales sans RAG")
        elif knowledge_mode == "no_match":
            print(f"[ERROR] [ASSISTANT] NO-MATCH | Aucune source disponible")

        studio_match_debug = knowledge_result.get("studio_match")
        if studio_match_debug:
            print(
                f"[FICHE] [ASSISTANT] Reponse basee sur la fiche Studio #{studio_match_debug.get('id')} "
                f"'{studio_match_debug.get('title')}' (score={studio_match_debug.get('match_score')})"
            )
    except Exception as e:
        print(f"[ASSISTANT] Log error: {e}")

    # 6. Construction de la réponse (sans ticket)
    studio_match_info = knowledge_result.get("studio_match")
    response: Dict[str, Any] = {
        "status": "success",
        "ai_analysis": ai_result,
        "category": chosen_category,
        "knowledge_mode": knowledge_result["knowledge_mode"],
        "knowledge_fallback_used": knowledge_result["knowledge_fallback_used"],
        "original_query": original_query,
        "normalized_query_fr": search_query_fr,
        "reconstructed_query_local": reconstructed_query_local,
        "query_interpretation_confidence": query_interpretation_confidence,
        # Identifie explicitement la fiche du Studio utilisée pour répondre
        # (null si aucune fiche ne correspondait et que la réponse vient du
        # LLM général ou du RAG élargi).
        "knowledge_fiche_id": studio_match_info.get("id") if studio_match_info else None,
        "knowledge_fiche_title": studio_match_info.get("title") if studio_match_info else None,
    }

    if photo_analysis is not None:
        response["photo_analysis"] = photo_analysis

    if rag_items:
        response["rag_items"] = rag_items

    if llm_answer:
        response["llm_answer"] = llm_answer
    elif knowledge_result["rag_fallback_answer"]:
        response["rag_fallback_answer"] = knowledge_result["rag_fallback_answer"]

    if isinstance(photo_analysis, dict):
        photo_status = photo_analysis.get("problem_status")
        validation_message = _clean_assistant_text(
            photo_analysis.get("validation_message")
        )
        if photo_status in {"wrong_category", "unusable", "not_identified", "uncertain"}:
            # Ne pas laisser le RAG inventer un problème quand la vision ne l'a
            # pas confirmé ou lorsque la photo ne correspond pas à la catégorie.
            response["llm_answer"] = validation_message
            response.pop("rag_fallback_answer", None)
        elif photo_status == "identified" and validation_message:
            current_answer = _clean_assistant_text(
                response.get("llm_answer") or response.get("rag_fallback_answer")
            )
            if current_answer and validation_message not in current_answer:
                response["llm_answer"] = f"{validation_message}\n\n{current_answer}"

    if image_result and image_result.get("success"):
        response["image_base64"] = image_result.get("image_base64")
        response["image_mime_type"] = image_result.get("mime_type", "image/png")
    elif image_result and image_result.get("fallback_description"):
        response["image_description"] = image_result.get("fallback_description")

    if video_result and video_result.get("success"):
        response["video_base64"] = video_result.get("video_base64")
        response["video_url"] = video_result.get("video_url")
        response["video_mime_type"] = video_result.get("mime_type", "video/mp4")
        response["video_duration"] = video_result.get("duration_sec")
    elif video_result and video_result.get("fallback"):
        response["video_description"] = video_result.get("video_description")
        response["video_steps"] = video_result.get("steps_visuelles")

    # Conserver la version française canonique dans le cache. Elle sera
    # retraduite selon la langue choisie lors de chaque future consultation.
    cache_answer_fr = _clean_assistant_text(
        response.get("llm_answer") or response.get("rag_fallback_answer")
    )

    # Le endpoint historique renvoie aussi le contrat V2 afin que les sessions
    # phone-only ouvrent exactement la même fiche résultat que les comptes JWT.
    if studio_match_info:
        photo_requires_expert = bool(
            isinstance(photo_analysis, dict)
            and photo_analysis.get("requires_expert")
        )
        v2_payload = {
            "message": cache_answer_fr or studio_match_info.get("resolution_fr") or "",
            "diagnostic": {
                "type": chosen_category,
                "description": studio_match_info.get("title") or "Fiche SONGRA",
                "gravite": (
                    str((photo_analysis or {}).get("severity") or "moyenne")
                    if isinstance(photo_analysis, dict) else "moyenne"
                ),
                "confiance": (
                    (photo_analysis or {}).get("confidence", 0.8)
                    if isinstance(photo_analysis, dict) else 0.8
                ),
                "causes": [],
            },
            "actions": [],
            "urgence": bool(
                isinstance(photo_analysis, dict)
                and (photo_analysis.get("urgency") == "high")
            ),
            "priorite": 1 if photo_requires_expert else 3,
            "consulter_expert": photo_requires_expert,
        }
        response.update(
            _apply_studio_match_to_v2_response(
                v2_payload, studio_match_info, target_lang
            )
        )

    # ── AUDIO HUMAIN DE LA BASE LOCALE ───────────────────────────────────────
    # Le diagnostic et la recherche restent en français. Pour une langue locale,
    # on ne traduit plus par LLM et on ne génère plus de TTS: on cherche une
    # fiche validée correspondant au cas et possédant un enregistrement humain.
    if target_lang and target_lang in _TRANSLATOR_VALID_LANGS:
        try:
            studio_case = knowledge_result.get("studio_match")
            recorded_case = None
            if studio_case:
                audio_entry = (studio_case.get("audio") or {}).get(target_lang) or {}
                translation_entry = (studio_case.get("translations") or {}).get(target_lang) or {}
                if str(audio_entry.get("url") or "").strip():
                    recorded_case = {
                        **studio_case,
                        "audio_url": audio_entry.get("url"),
                        "audio_mime_type": audio_entry.get("mime_type") or "audio/webm",
                        "local_text": translation_entry.get("text"),
                    }
                response["local_knowledge_match"] = studio_case
                response["local_text"] = translation_entry.get("text")
                response["french_text"] = studio_case.get("resolution_fr")
            else:
                recorded_case = _find_recorded_local_case_audio(
                    db,
                    category=chosen_category,
                    language=target_lang,
                    photo_analysis=photo_analysis,
                    french_answer=cache_answer_fr,
                )
            response["translated"] = False
            response["target_lang"] = target_lang
            response["lang_name"] = _TRANSLATOR_LANG_NAMES.get(target_lang)
            response["local_audio_available"] = recorded_case is not None
            response["local_audio_source"] = (
                "recorded_expert_local_knowledge" if recorded_case else None
            )
            if recorded_case:
                response["audio_url"] = recorded_case["audio_url"]
                response["audio_mime_type"] = recorded_case["audio_mime_type"]
                response["voice_summary"] = recorded_case.get("local_text")
                response["local_knowledge_match"] = recorded_case
            else:
                response["local_audio_message"] = (
                    "La voix dédiée à cette fiche est indisponible dans cette langue."
                )
                response["french_fallback_available"] = True
        except Exception as local_audio_exc:
            response["local_audio_available"] = False
            response["local_audio_message"] = "Recherche d'audio local indisponible."
            print(f"[LOCAL-AUDIO] Recherche impossible: {local_audio_exc}")

    # Mémoriser la réponse finale pour une future question strictement
    # identique. Une réponse déjà réutilisée n'est pas enregistrée à nouveau.
    if reusable_entry is None:
        if cache_answer_fr:
            try:
                _persist_offline_knowledge_entry(
                    db=db,
                    user_id=None,
                    source_kind="assistant_query",
                    category=chosen_category,
                    question_text=contextual_query,
                    response_payload={
                        "message": cache_answer_fr,
                        "category": chosen_category,
                        "target_lang": target_lang,
                    },
                )
            except Exception as cache_exc:
                print(f"[ASSISTANT] Mise en cache réponse impossible: {cache_exc}")

    return response

@app.post("/api/tickets/{ticket_id}/reply")
async def reply_to_ticket(
    ticket_id: int, 
    content: ReplyMessage,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db)
):
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    message = Message(
        ticket_id=ticket_id,
        sender_type="expert",
        sender_id=current_expert.id,
        content=content.message,
        channel="web",
        language=_normalize_expert_local_language(content.language or ticket.preferred_language),
    )
    db.add(message)
    
    if not ticket.expert_id:
        ticket.expert_id = current_expert.id
        ticket.status = "assigned"
    
    db.commit()
    return {"status": "success"}


@app.post("/api/tickets/{ticket_id}/reply-audio")
async def reply_to_ticket_with_audio(
    ticket_id: int,
    audio_file: UploadFile = File(...),
    language: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Permet à l'expert de répondre par message vocal à un ticket."""
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Sauvegarder le fichier audio
    audio_dir = "uploads/audio/expert-replies"
    os.makedirs(audio_dir, exist_ok=True)
    
    ext = os.path.splitext(audio_file.filename or "audio.webm")[1].lower() or ".webm"
    allowed_audio_exts = [".webm", ".ogg", ".mp3", ".wav", ".m4a", ".mp4", ".mpeg"]
    if ext not in allowed_audio_exts:
        ext = ".webm"
    
    filename = f"expert_reply_{ticket_id}_{int(time.time())}{ext}"
    filepath = os.path.join(audio_dir, filename)
    
    content_bytes = await audio_file.read()
    with open(filepath, "wb") as buffer:
        buffer.write(content_bytes)
    
    audio_relative_path = f"uploads/audio/expert-replies/{filename}"
    
    # Créer le message avec audio_url
    message = Message(
        ticket_id=ticket_id,
        sender_type="expert",
        sender_id=1,
        content="🔊 Réponse vocale de l'expert",
        channel="web",
        audio_url=audio_relative_path,
        language=_normalize_expert_local_language(language or ticket.preferred_language),
    )
    db.add(message)
    
    if not ticket.expert_id:
        ticket.expert_id = 1
        ticket.status = "assigned"
    
    db.commit()
    return {"status": "success", "audio_url": f"/{audio_relative_path}"}

@app.post("/api/tickets/{ticket_id}/send-to-expert")
async def send_ticket_to_expert(
    ticket_id: int,
    db: Session = Depends(get_db)
):
    """Envoie un ticket à un expert humain (enregistre la demande)."""
    
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Assigner à un expert (ID 1 par défaut, ou le premier expert disponible)
    if not ticket.expert_id:
        expert = db.query(Expert).filter(Expert.is_active == True).first()
        if expert:
            ticket.expert_id = expert.id
        else:
            ticket.expert_id = 1  # Fallback
    
    # Marquer comme "awaiting_expert"
    ticket.status = "awaiting_expert"
    
    # Créer un message système pour tracer
    message = Message(
        ticket_id=ticket_id,
        sender_type="system",
        sender_id=None,
        content="[SYSTEME] Aide d'expert demandee par l'utilisateur",
        channel="app"
    )
    db.add(message)
    db.commit()
    db.refresh(ticket)
    
    return {
        "status": "success",
        "ticket_id": ticket.id,
        "expert_assigned": ticket.expert_id,
        "message": "Votre demande a ete envoyee a un expert. Un expert humain vous contactera bientot."
    }

@app.post("/api/tickets/{ticket_id}/resolve")
async def resolve_ticket(ticket_id: int, db: Session = Depends(get_db)):
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    ticket.status = "resolved"
    ticket.resolved_at = datetime.utcnow()
    
    message = Message(
        ticket_id=ticket_id,
        sender_type="system",
        sender_id=None,
        content="Ticket marqué comme résolu par l'expert",
        channel="system"
    )
    db.add(message)
    
    db.commit()

    try:
        payload = _build_offline_payload_from_resolved_ticket(ticket, db)
        if payload:
            user_messages = (
                db.query(Message)
                .filter(Message.ticket_id == ticket.id, Message.sender_type == "user")
                .order_by(Message.sent_at.asc(), Message.id.asc())
                .all()
            )
            question_text = "\n".join(
                msg.content.strip()
                for msg in user_messages
                if (msg.content or "").strip()
            ).strip() or "Ticket resolu Songra"
            _persist_offline_knowledge_entry(
                db=db,
                user_id=ticket.user_id,
                source_kind="resolved_ticket",
                category=ticket.category or "agriculture",
                question_text=question_text,
                response_payload=payload,
            )
    except Exception as e:
        print(f"[OFFLINE-CORPUS] Erreur persistance ticket resolu {ticket.id}: {e}")

    return {"status": "success"}


@app.get("/api/tickets/{ticket_id}/ai-summary")
async def get_ticket_ai_summary(ticket_id: int, db: Session = Depends(get_db)):
    """Retourne un résumé IA (RAG + GPT ou fallback) pour un ticket donné.

    L'IA se base sur le dernier message utilisateur du ticket et sur
    la catégorie du ticket pour choisir le bon domaine de connaissance.
    """

    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    # Récupérer le dernier message utilisateur lié au ticket
    last_user_msg = (
        db.query(Message)
        .filter(Message.ticket_id == ticket_id, Message.sender_type == "user")
        .order_by(Message.sent_at.desc())
        .first()
    )

    if not last_user_msg:
        return {"status": "success", "ai_summary": None, "detail": "Aucun message utilisateur pour ce ticket."}

    content = last_user_msg.content or ""

    # Mapper la catégorie du ticket vers le domaine RAG
    chosen_category = ticket.category or "agriculture"
    if chosen_category == "agriculture":
        kb_domain = "agriculture"
    elif chosen_category == "elevage":
        kb_domain = "elevage"
    elif chosen_category == "sos_accident":
        kb_domain = "health"
    elif chosen_category == "cybersecurity":
        kb_domain = "cybersecurity"
    else:
        kb_domain = "agriculture"

    focus_context = extract_focus_context(
        chosen_category,
        content,
    )

    knowledge_result = resolve_knowledge_answer(
        db=db,
        domain=kb_domain,
        question=content,
        language="fr",
        focus_context=focus_context,
    )
    rag_items = knowledge_result["rag_items"]
    llm_answer = knowledge_result["llm_answer"] or knowledge_result["rag_fallback_answer"]
    studio_match_info = knowledge_result.get("studio_match")

    return {
        "status": "success",
        "ai_summary": llm_answer,
        "category": chosen_category,
        "rag_items": rag_items,
        "knowledge_mode": knowledge_result["knowledge_mode"],
        "knowledge_fallback_used": knowledge_result["knowledge_fallback_used"],
        "knowledge_fiche_id": studio_match_info.get("id") if studio_match_info else None,
        "knowledge_fiche_title": studio_match_info.get("title") if studio_match_info else None,
    }


def _extract_text_from_photo_analysis(raw_analysis: Optional[str]) -> str:
    if not raw_analysis:
        return ""
    try:
        parsed = json.loads(raw_analysis)
    except Exception:
        return raw_analysis.strip()

    if not isinstance(parsed, dict):
        return str(parsed).strip()

    parts: List[str] = []
    for key in (
        "disease_detected",
        "analysis",
        "observations",
        "treatment",
        "recommendations",
        "when_to_call_expert",
    ):
        value = parsed.get(key)
        if isinstance(value, str) and value.strip():
            parts.append(value.strip())
        elif isinstance(value, list) and value:
            parts.append("; ".join(str(item).strip() for item in value if str(item).strip()))
    return "\n\n".join(part for part in parts if part).strip()


def _build_offline_payload_from_resolved_ticket(ticket: Ticket, db: Session) -> Optional[Dict[str, Any]]:
    user_messages = (
        db.query(Message)
        .filter(Message.ticket_id == ticket.id, Message.sender_type == "user")
        .order_by(Message.sent_at.asc(), Message.id.asc())
        .all()
    )
    if not user_messages:
        return None

    latest_expert_message = (
        db.query(Message)
        .filter(Message.ticket_id == ticket.id, Message.sender_type == "expert")
        .order_by(Message.sent_at.desc(), Message.id.desc())
        .first()
    )

    question_text = "\n".join(
        message.content.strip()
        for message in user_messages
        if (message.content or "").strip()
    ).strip()
    if not question_text:
        return None

    answer_text = ""
    if latest_expert_message and (latest_expert_message.content or "").strip():
        answer_text = latest_expert_message.content.strip()
    elif (ticket.resolution_notes or "").strip():
        answer_text = ticket.resolution_notes.strip()
    else:
        answer_text = _extract_text_from_photo_analysis(ticket.ai_photo_analysis)

    if not answer_text:
        return None

    return {
        "message": answer_text,
        "diagnostic": {
            "type": _normalize_category(ticket.category),
            "description": answer_text[:240],
            "gravite": "moyenne",
            "confiance": 0.95 if latest_expert_message else 0.75,
            "causes": _extract_offline_keywords(question_text)[:5],
        },
        "actions": [],
        "consulter_expert": False,
        "ticket_id": ticket.id,
        "resolved_at": ticket.resolved_at.isoformat() if ticket.resolved_at else None,
    }


def _backfill_resolved_tickets_to_offline_corpus(db: Session) -> int:
    resolved_tickets = (
        db.query(Ticket)
        .filter(Ticket.status == "resolved")
        .order_by(Ticket.resolved_at.desc().nullslast(), Ticket.id.desc())
        .all()
    )

    persisted = 0
    for ticket in resolved_tickets:
        payload = _build_offline_payload_from_resolved_ticket(ticket, db)
        if not payload:
            continue
        try:
            _persist_offline_knowledge_entry(
                db=db,
                user_id=ticket.user_id,
                source_kind="resolved_ticket",
                category=ticket.category or "agriculture",
                question_text=payload.get("diagnostic", {}).get("description") and "\n".join(
                    message.content.strip()
                    for message in db.query(Message)
                    .filter(Message.ticket_id == ticket.id, Message.sender_type == "user")
                    .order_by(Message.sent_at.asc(), Message.id.asc())
                    .all()
                    if (message.content or "").strip()
                ) or "Ticket resolu Songra",
                response_payload=payload,
            )
            persisted += 1
        except Exception as e:
            print(f"[OFFLINE-CORPUS] Ticket resolu {ticket.id} ignore: {e}")
    return persisted


@app.post("/api/admin/offline-corpus/backfill-resolved-tickets")
async def backfill_resolved_tickets_offline_corpus(
    db: Session = Depends(get_db),
):
    count = _backfill_resolved_tickets_to_offline_corpus(db)
    return {
        "status": "success",
        "backfilled": count,
        "scope": "resolved_tickets_only",
    }

@app.get("/api/user-tickets")
async def get_user_tickets(phone: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.phone_number == phone).first()
    if not user:
        return []
    
    # Une analyse V2 reste dans l'historique des consultations. Elle ne devient
    # une "demande" que lorsque l'utilisateur appuie explicitement sur
    # "Contacter un expert". Les anciens tickets créés par la synchronisation
    # automatique sont donc exclus de cette liste.
    automatic_v2_ticket_ids = db.query(Message.ticket_id).filter(
        Message.channel == "v2_mobile"
    )
    tickets = db.query(Ticket).filter(
        Ticket.user_id == user.id,
        ~Ticket.id.in_(automatic_v2_ticket_ids),
    ).order_by(Ticket.created_at.desc()).all()
    
    result = []
    for ticket in tickets:
        # Récupérer le PREMIER message utilisateur (la demande initiale, y compris analyse d'image)
        first_user_msg = db.query(Message).filter(
            Message.ticket_id == ticket.id,
            Message.sender_type == 'user'
        ).order_by(Message.sent_at.asc()).first()
        
        # Récupérer le DERNIER message (pour savoir le statut de la conversation)
        last_msg = db.query(Message).filter(
            Message.ticket_id == ticket.id
        ).order_by(Message.sent_at.desc()).first()
        latest_expert_msg = db.query(Message).filter(
            Message.ticket_id == ticket.id,
            Message.sender_type == "expert",
        ).order_by(Message.sent_at.desc(), Message.id.desc()).first()
        
        # Construire l'URL de la photo
        photo_url = _build_upload_url(ticket.photo_path)
        photo_paths = _load_json_list(ticket.photo_paths_json)
        photo_urls = [_build_upload_url(path) for path in photo_paths if path]
        
        # Gérer l'analyse photo
        photo_analysis = None
        if ticket.ai_photo_analysis:
            try:
                if isinstance(ticket.ai_photo_analysis, str):
                    photo_analysis = json.loads(ticket.ai_photo_analysis)
                else:
                    photo_analysis = ticket.ai_photo_analysis
            except:
                photo_analysis = None
        
        result.append({
            "id": ticket.id,
            "category": ticket.category or "agriculture",
            "urgency": ticket.urgency or "low",
            "status": ticket.status or "open",
            "created_at": ticket.created_at.isoformat() if ticket.created_at else None,
            "last_message": (first_user_msg.content if first_user_msg else None) or (last_msg.content if last_msg else "Aucun message"),
            "has_photo": ticket.photo_path is not None,
            "photo_url": photo_url,
            "photo_urls": photo_urls,
            "photo_analysis": photo_analysis,
            "latest_expert_message_at": latest_expert_msg.sent_at.isoformat() if latest_expert_msg and latest_expert_msg.sent_at else None,
            "latest_expert_has_audio": bool(latest_expert_msg and latest_expert_msg.audio_url),
        })
    
    return result


@app.get("/api/user-tickets/{ticket_id}")
async def get_user_ticket_detail(
    ticket_id: int,
    phone: str,
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.phone_number == phone).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    ticket = db.query(Ticket).filter(
        Ticket.id == ticket_id,
        Ticket.user_id == user.id,
    ).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Demande introuvable")
    return await get_ticket_detail(ticket_id, current_expert=None, db=db)


@app.get("/api/user-history")
async def get_user_history(phone: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.phone_number == phone).first()
    if not user:
        return []

    tickets = await get_user_tickets(phone, db)

    entreprendre_entries = (
        db.query(OfflineKnowledgeEntryDB)
        .filter(
            OfflineKnowledgeEntryDB.user_id == user.id,
            OfflineKnowledgeEntryDB.source_kind == "entreprendre",
        )
        .order_by(OfflineKnowledgeEntryDB.updated_at.desc())
        .limit(50)
        .all()
    )

    consultation_entries = (
        db.query(OfflineKnowledgeEntryDB)
        .filter(
            OfflineKnowledgeEntryDB.user_id == user.id,
            OfflineKnowledgeEntryDB.source_kind.in_(["v2_analyze", "v2_scanner", "v2_assistant_query"]),
        )
        .order_by(OfflineKnowledgeEntryDB.updated_at.desc())
        .limit(100)
        .all()
    )

    media_entries = (
        db.query(OfflineKnowledgeEntryDB)
        .filter(
            OfflineKnowledgeEntryDB.user_id == user.id,
            OfflineKnowledgeEntryDB.source_kind.in_(["generated_image_illustration", "generated_video_illustration"]),
        )
        .order_by(OfflineKnowledgeEntryDB.updated_at.desc())
        .limit(100)
        .all()
    )

    merged = [
        *tickets,
        *[_serialize_entreprendre_history_entry(entry) for entry in entreprendre_entries],
        *[_serialize_v2_history_entry(entry) for entry in consultation_entries],
        *[_serialize_generated_media_history_entry(entry) for entry in media_entries],
    ]
    merged.sort(
        key=lambda item: _history_sort_value(item.get("updated_at") or item.get("created_at")),
        reverse=True,
    )
    return merged


@app.get("/api/photo-analyses")
async def get_photo_analyses(phone: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.phone_number == phone).first()
    if not user:
        return []

    records = (
        db.query(PhotoAnalysisHistoryDB)
        .filter(PhotoAnalysisHistoryDB.user_id == user.id)
        .order_by(PhotoAnalysisHistoryDB.created_at.desc())
        .limit(50)
        .all()
    )
    return [_serialize_photo_history_record(record) for record in records]


@app.post("/api/photo-analyses")
async def save_photo_analysis_history(
    data: PhotoAnalysisHistoryIn,
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.phone_number == data.phone_number.strip()).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    record = None
    if data.client_record_id:
        record = (
            db.query(PhotoAnalysisHistoryDB)
            .filter(
                PhotoAnalysisHistoryDB.user_id == user.id,
                PhotoAnalysisHistoryDB.client_record_id == data.client_record_id,
            )
            .first()
        )

    photo_data_list = [
        _decode_photo_payload(payload)
        for payload in data.photo_base64_list[:3]
        if payload
    ]
    stored_paths = (
        _store_photo_payloads(user.id, photo_data_list, prefix="analysis")
        if photo_data_list
        else []
    )

    if record is None:
        record = PhotoAnalysisHistoryDB(
            user_id=user.id,
            client_record_id=data.client_record_id,
            category=data.category,
            prompt=data.prompt,
            analysis_json=json.dumps(data.analysis, ensure_ascii=False),
            photo_paths_json=json.dumps(stored_paths, ensure_ascii=False),
            photo_labels_json=json.dumps(
                data.photo_labels[: len(stored_paths)],
                ensure_ascii=False,
            ),
        )
        db.add(record)
    else:
        record.category = data.category
        record.prompt = data.prompt
        record.analysis_json = json.dumps(data.analysis, ensure_ascii=False)
        if stored_paths:
            record.photo_paths_json = json.dumps(stored_paths, ensure_ascii=False)
            record.photo_labels_json = json.dumps(
                data.photo_labels[: len(stored_paths)],
                ensure_ascii=False,
            )

    db.commit()
    db.refresh(record)
    return _serialize_photo_history_record(record)


@app.get("/api/knowledge/offline-cache")
async def knowledge_offline_cache(
    domain: Optional[str] = None,
    language: str = "fr",
    limit: int = Query(default=100, le=500),
    db: Session = Depends(get_db),
):
    return _build_offline_cache_payload(
        db,
        domain=domain,
        language=language,
        limit=limit,
    )

# ==========================================
# GESTION AUDIO (EXPERT)
# ==========================================

@app.get("/api/admin/audio-map")
async def get_audio_map():
    """Récupérer la carte des correspondances audio."""
    if not os.path.exists(EXPERT_AUDIO_MAP_PATH):
        return {}
    try:
        with open(EXPERT_AUDIO_MAP_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lecture audio_map: {e}")

@app.post("/api/admin/audio-map/{key}/{language}")
async def upload_audio_file(
    key: str,
    language: str,
    file: UploadFile = File(...),
    current_expert: Expert = Depends(get_current_admin_expert),
):
    """Uploader un fichier audio pour une clé et une langue spécifique."""
    # Dossier spécifique par langue (ex: uploads/audio/moore)
    lang_dir = os.path.join(EXPERT_AUDIO_UPLOAD_DIR, language)
    os.makedirs(lang_dir, exist_ok=True)
    
    # Nom du fichier (ex: menu_entreprendre.mp3)
    filename = f"{key}.mp3"
    file_path = os.path.join(lang_dir, filename)
    
    # Sauvegarder le fichier sur le disque
    try:
        content = await file.read()
        with open(file_path, "wb") as buffer:
            buffer.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur écriture fichier: {e}")
        
    # Mettre à jour la carte audio (JSON)
    audio_map = {}
    if os.path.exists(EXPERT_AUDIO_MAP_PATH):
        try:
            with open(EXPERT_AUDIO_MAP_PATH, "r", encoding="utf-8") as f:
                audio_map = json.load(f)
        except:
            audio_map = {}
            
    # Initialiser la structure si inexistante
    if key not in audio_map:
        audio_map[key] = {"label": key.replace("_", " ").capitalize(), "voices": {}}
    
    if isinstance(audio_map[key], dict) and "voices" not in audio_map[key]:
        # Migration si ancien format
        old_audio = audio_map[key].get("audio")
        audio_map[key]["voices"] = {"fr": old_audio} if old_audio else {}
        
    # URL relative pour le téléchargement
    audio_map[key]["voices"][language] = f"/audio/{language}/{filename}"
    
    try:
        with open(EXPERT_AUDIO_MAP_PATH, "w", encoding="utf-8") as f:
            json.dump(audio_map, f, indent=2, ensure_ascii=False)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur mise à jour audio_map: {e}")
        
    return {"status": "success", "url": audio_map[key]["voices"][language], "key": key, "language": language}

@app.delete("/api/admin/audio-map/{key}/{language}")
async def delete_audio_mapping(
    key: str, language: str,
    current_expert: Expert = Depends(get_current_admin_expert),
):
    """Supprimer une association audio."""
    if not os.path.exists(EXPERT_AUDIO_MAP_PATH):
        return {"status": "skipped"}
        
    with open(EXPERT_AUDIO_MAP_PATH, "r", encoding="utf-8") as f:
        audio_map = json.load(f)
        
    if key in audio_map and "voices" in audio_map[key]:
        if language in audio_map[key]["voices"]:
            del audio_map[key]["voices"][language]
            
            with open(EXPERT_AUDIO_MAP_PATH, "w", encoding="utf-8") as f:
                json.dump(audio_map, f, indent=2, ensure_ascii=False)
                
    return {"status": "success"}


@app.get("/api/admin/settings")
async def get_system_settings(current_identity: Any = Depends(get_current_user_or_expert)):
    """Récupère les réglages système (Voice ID, etc)."""
    settings_path = os.path.join(BACKEND_DIR, "system_settings.json")
    default_settings = {
        "elevenlabs_voice_id": "EXAVITQu4vr4xnSDxMaL", # Bella par défaut
        "ai_model": "eleven_multilingual_v2",
        "whatsapp_webhook_url": "",
        "alert_target_phone": "+22670000017"
    }
    
    if not os.path.exists(settings_path):
        return default_settings
        
    try:
        with open(settings_path, "r", encoding="utf-8") as f:
            return {**default_settings, **json.load(f)}
    except:
        return default_settings

@app.post("/api/admin/settings")
async def update_system_settings(
    payload: Dict[str, Any],
    current_expert: Expert = Depends(get_current_admin_expert),
):
    """Met à jour les réglages système."""
    settings_path = os.path.join(BACKEND_DIR, "system_settings.json")
    existing = {}
    if os.path.exists(settings_path):
        try:
            with open(settings_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except: pass
        
    updated = {**existing, **payload}
    with open(settings_path, "w", encoding="utf-8") as f:
        json.dump(updated, f, ensure_ascii=False, indent=2)
        
    return {"status": "success", "settings": updated}


@app.post("/api/admin/broadcast")
async def create_broadcast(
    file: UploadFile = File(...),
    title: str = Form(...),
    category: str = Form("general"),
    language: str = Form("Français"),
    region: str = Form("Toutes"),
    description: str = Form(""),
    offline_allowed: bool = Form(True),
    current_expert: Expert = Depends(get_current_expert),
):
    """Crée une nouvelle diffusion communautaire (Journal Vocal)."""
    language_aliases = {
        "fr": "fr", "francais": "fr", "français": "fr",
        "moore": "moore", "mooré": "moore",
        "dioula": "dioula",
        "fulfulde": "fulfulde", "fulfuldé": "fulfulde",
    }
    normalized_language = language_aliases.get(language.strip().lower())
    if not normalized_language:
        raise HTTPException(status_code=400, detail="Langue non supportee pour la radio.")
    safe_original_name = os.path.basename(file.filename or "message.mp3")
    extension = os.path.splitext(safe_original_name)[1].lower()
    if extension not in {".mp3", ".m4a", ".wav", ".ogg", ".aac"}:
        raise HTTPException(status_code=400, detail="Format audio non supporte.")
    os.makedirs("uploads/broadcasts", exist_ok=True)
    filename = f"broadcast_{int(time.time())}_{safe_original_name}"
    filepath = f"uploads/broadcasts/{filename}"

    total_size = 0
    try:
        with open(filepath, "wb") as buffer:
            while chunk := await file.read(1024 * 1024):
                total_size += len(chunk)
                if total_size > 25 * 1024 * 1024:
                    raise HTTPException(status_code=413, detail="Le fichier audio depasse 25 Mo.")
                buffer.write(chunk)
    except Exception:
        if os.path.exists(filepath):
            os.remove(filepath)
        raise
    
    # Enregistrer dans broadcasts.json
    db_path = os.path.join(BACKEND_DIR, "broadcasts.json")
    broadcasts = []
    if os.path.exists(db_path):
        try:
            with open(db_path, "r", encoding="utf-8") as f:
                broadcasts = json.load(f)
        except: pass
    
    new_entry = {
        "id": int(time.time()),
        "title": title,
        "category": category,
        "language": normalized_language,
        "region": region,
        "description": description,
        "offline_allowed": offline_allowed,
        "audio_url": f"/uploads/broadcasts/{filename}",
        "timestamp": datetime.now().isoformat(),
        "listeners": 0
        ,"created_by_expert_id": current_expert.id
        ,"created_by_expert": current_expert.full_name
        ,"organization_id": getattr(current_expert, "organization_id", None)
        ,"organization": current_expert.institution
    }
    broadcasts.insert(0, new_entry)
    
    with open(db_path, "w", encoding="utf-8") as f:
        json.dump(broadcasts[:50], f, ensure_ascii=False, indent=2) # Garder les 50 derniers
        
    return {"status": "success", "broadcast": new_entry}

@app.get("/api/community/broadcasts")
async def get_broadcasts():
    """Récupère les dernières diffusions pour l'app mobile."""
    db_path = os.path.join(BACKEND_DIR, "broadcasts.json")
    if not os.path.exists(db_path):
        return []
    try:
        with open(db_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return []
# 
@app.post("/api/admin/upload")
async def upload_general_file(
    file: UploadFile = File(...),
    current_expert: Expert = Depends(get_current_admin_expert),
):
    """Upload un fichier générique (image, audio, etc) vers le dossier uploads."""
    os.makedirs("uploads", exist_ok=True)
    
    # Sécuriser le nom du fichier
    ext = os.path.splitext(file.filename)[1].lower()
    # On autorise images et audios
    allowed_exts = ['.jpg', '.jpeg', '.png', '.gif', '.mp3', '.wav', '.m4a', '.ogg']
    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail=f"Extension {ext} non autorisée")
        
    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{hashlib.md5(file.filename.encode()).hexdigest()[:8]}{ext}"
    file_path = os.path.join("uploads", filename)
    
    try:
        content = await file.read()
        with open(file_path, "wb") as buffer:
            buffer.write(content)
        
        return {"status": "success", "url": f"/uploads/{filename}", "filename": filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur upload: {e}")


@app.get("/api/admin/knowledge")
async def list_knowledge_items(
    domain: Optional[str] = None,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    """Lister les fiches de la base de connaissances.

    Pour l'instant, cette route n'est pas protégée (usage local / démo). Pour
    un déploiement réel, il faudra ajouter une authentification admin.
    """
    query = db.query(KnowledgeItem)
    if domain:
        query = query.filter(KnowledgeItem.domain == domain)

    items = query.order_by(KnowledgeItem.created_at.desc()).all()

    result: List[Dict[str, Any]] = []
    for it in items:
        media_data = None
        if it.media:
            try:
                media_data = json.loads(it.media)
            except Exception:
                media_data = None

        result.append(
            {
                "id": it.id,
                "domain": it.domain,
                "title": it.title,
                "question": it.question,
                "answer": it.answer,
                "tags": _load_json_list(it.tags),
                "language": it.language,
                "source": it.source,
                "media": media_data,
                "created_at": it.created_at,
                "updated_at": it.updated_at,
            }
        )

    return result


@app.get("/api/emergency-numbers")
async def public_emergency_numbers(db: Session = Depends(get_db)):
    """Liste publique des numéros d'urgence / numéros utiles.

    Utilisée par l'application utilisateur pour afficher les numéros utiles.
    """
    items = db.query(EmergencyNumber).filter(EmergencyNumber.is_active == True).order_by(
        EmergencyNumber.display_order.asc(), EmergencyNumber.id.asc()
    ).all()

    return [
        {
            "id": it.id,
            "label": it.label,
            "number": it.number,
            "description": it.description,
            "display_order": it.display_order,
        }
        for it in items
    ]


@app.post("/api/admin/reload-knowledge")
async def reload_knowledge_endpoint(
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    """Recharger la base de connaissances depuis le fichier JSON.

    End-point simple pour éviter de redémarrer le serveur quand tu mets à jour
    knowledge_base.json. Dans un vrai déploiement, il faudra protéger cette route
    (token admin, VPN, etc.).
    """
    load_knowledge_from_json(db)
    total_items = db.query(KnowledgeItem).count()
    return {
        "status": "success",
        "total_items": total_items,
    }


@app.post("/api/admin/knowledge")
async def create_knowledge_item(
    payload: KnowledgeItemIn,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    """Créer une nouvelle fiche de connaissance (usage panneau expert)."""

    media_json = None
    if payload.media:
        media_json = json.dumps(
            [m.dict() for m in payload.media], ensure_ascii=False
        )

    item = KnowledgeItem(
        domain=payload.domain,
        title=payload.title,
        question=payload.question,
        answer=payload.answer,
        tags=json.dumps(payload.tags or [], ensure_ascii=False),
        language=payload.language,
        source=payload.source,
        media=media_json,
    )
    db.add(item)
    db.commit()
    db.refresh(item)

    media_data = None
    if item.media:
        try:
            media_data = json.loads(item.media)
        except Exception:
            media_data = None

    return {
        "id": item.id,
        "domain": item.domain,
        "title": item.title,
        "question": item.question,
        "answer": item.answer,
        "tags": json.loads(item.tags) if item.tags else [],
        "language": item.language,
        "source": item.source,
        "media": media_data,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


@app.get("/api/admin/emergency-numbers")
async def list_emergency_numbers_admin(db: Session = Depends(get_db)):
    """Lister tous les numéros utiles (admin panneau expert)."""

    items = db.query(EmergencyNumber).order_by(
        EmergencyNumber.display_order.asc(), EmergencyNumber.id.asc()
    ).all()

    return [
        {
            "id": it.id,
            "label": it.label,
            "number": it.number,
            "description": it.description,
            "display_order": it.display_order,
            "is_active": it.is_active,
        }
        for it in items
    ]


@app.post("/api/admin/emergency-numbers")
async def create_emergency_number(
    payload: EmergencyNumberIn,
    db: Session = Depends(get_db),
):
    """Créer un nouveau numéro utile (pompier, police, clinique...)."""

    item = EmergencyNumber(
        label=payload.label,
        number=payload.number,
        description=payload.description,
        display_order=payload.display_order,
        is_active=True,
    )
    db.add(item)
    db.commit()
    db.refresh(item)

    return {
        "id": item.id,
        "label": item.label,
        "number": item.number,
        "description": item.description,
        "display_order": item.display_order,
        "is_active": item.is_active,
    }


@app.put("/api/admin/knowledge/{item_id}")
async def update_knowledge_item(
    item_id: int,
    payload: KnowledgeItemIn,
    db: Session = Depends(get_db),
):
    """Mettre à jour une fiche de connaissance existante ou la créer si elle n'existe plus.

    Cela évite une erreur 404 côté panneau expert si, pour une raison quelconque,
    l'ID stocké dans le frontend ne correspond plus à une ligne en base.
    """

    item = db.query(KnowledgeItem).filter(KnowledgeItem.id == item_id).first()

    media_json = None
    if payload.media:
        media_json = json.dumps(
            [m.dict() for m in payload.media], ensure_ascii=False
        )

    if not item:
        # Comportement "upsert" : si l'ID n'existe plus, on crée une nouvelle fiche
        item = KnowledgeItem(
            domain=payload.domain,
            title=payload.title,
            question=payload.question,
            answer=payload.answer,
            tags=json.dumps(payload.tags or [], ensure_ascii=False),
            language=payload.language,
            source=payload.source,
            media=media_json,
        )
        db.add(item)
        db.commit()
        db.refresh(item)
    else:
        item.domain = payload.domain
        item.title = payload.title
        item.question = payload.question
        item.answer = payload.answer
        item.tags = json.dumps(payload.tags or [], ensure_ascii=False)
        item.language = payload.language
        item.source = payload.source
        item.media = media_json

        db.commit()
        db.refresh(item)

    media_data = None
    if item.media:
        try:
            media_data = json.loads(item.media)
        except Exception:
            media_data = None

    return {
        "id": item.id,
        "domain": item.domain,
        "title": item.title,
        "question": item.question,
        "answer": item.answer,
        "tags": json.loads(item.tags) if item.tags else [],
        "language": item.language,
        "source": item.source,
        "media": media_data,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


@app.delete("/api/admin/knowledge/{item_id}")
async def delete_knowledge_item(
    item_id: int,
    db: Session = Depends(get_db),
):
    """Supprimer une fiche de connaissance."""

    item = db.query(KnowledgeItem).filter(KnowledgeItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Knowledge item not found")

    db.delete(item)
    db.commit()
    return {"status": "success"}


@app.put("/api/admin/emergency-numbers/{item_id}")
async def update_emergency_number(
    item_id: int,
    payload: EmergencyNumberIn,
    db: Session = Depends(get_db),
):
    """Mettre à jour un numéro utile existant."""

    item = db.query(EmergencyNumber).filter(EmergencyNumber.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Emergency number not found")

    item.label = payload.label
    item.number = payload.number
    item.description = payload.description
    item.display_order = payload.display_order

    db.commit()
    db.refresh(item)

    return {
        "id": item.id,
        "label": item.label,
        "number": item.number,
        "description": item.description,
        "display_order": item.display_order,
        "is_active": item.is_active,
    }


@app.delete("/api/admin/emergency-numbers/{item_id}")
async def delete_emergency_number(
    item_id: int,
    db: Session = Depends(get_db),
):
    """Supprimer un numéro utile."""

    item = db.query(EmergencyNumber).filter(EmergencyNumber.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Emergency number not found")

    db.delete(item)
    db.commit()
    return {"status": "success"}


@app.post("/api/admin/knowledge/import-json")
async def import_knowledge_from_json(
    payload: KnowledgeBulkImport,
    db: Session = Depends(get_db),
):
    """Importer/mettre à jour des fiches à partir d'un JSON envoyé par le panneau expert.

    Les fiches sont identifiées par (domain, title). Si une fiche existe déjà,
    elle est mise à jour; sinon elle est créée.
    """

    created = 0
    updated = 0

    for entry in payload.items:
        before = db.query(KnowledgeItem).filter(KnowledgeItem.title == entry.title).count()
        _upsert_knowledge_item(
            db,
            domain=entry.domain,
            title=entry.title,
            question=entry.question,
            answer=entry.answer,
            tags=entry.tags or [],
            language=entry.language,
            source=entry.source,
            media=[m.dict() for m in entry.media] if entry.media else None,
        )
        after = db.query(KnowledgeItem).filter(KnowledgeItem.title == entry.title).count()
        if before == 0 and after == 1:
            created += 1
        else:
            updated += 1

    db.commit()

    total_items = db.query(KnowledgeItem).count()
    return {
        "status": "success",
        "created": created,
        "updated": updated,
        "total_items": total_items,
    }


@app.post("/api/localization/translate")
async def translate_localization_payload(
    payload: LocalizationTranslateIn,
    background_tasks: BackgroundTasks,
    current_identity: Any = Depends(get_current_user_or_expert),
):
    del current_identity
    translations = _generate_local_translations(
        payload.question_fr,
        payload.resolution_fr,
        _normalize_expert_local_category(payload.category),
        background_tasks,
        payload.actions_fr,
    )
    return {
        "status": "success",
        "question_fr": payload.question_fr,
        "resolution_fr": payload.resolution_fr,
        "translations": translations,
    }


def _serialize_academy_course(course: AcademyCourseDB, include_content: bool = True, access: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    result = {
        "id": course.id,
        "title": course.title,
        "course_type": course.course_type or "culture",
        "crop": course.crop,
        "summary": course.summary,
        "cover_url": _build_upload_url(course.cover_url) if course.cover_url and not str(course.cover_url).startswith("http") else course.cover_url,
        "lesson_count": len(_load_json_list(course.steps_json)),
        "status": course.status or "published",
        "created_at": course.created_at.isoformat() if course.created_at else None,
        "updated_at": course.updated_at.isoformat() if course.updated_at else None,
    }
    if include_content:
        result["steps"] = _load_json_list(course.steps_json)
        result["audio"] = _normalize_expert_local_audio(_load_json_dict(course.audio_json))
    if access:
        result["access"] = access
    return result


def _course_access_status(db: Session, user: User, course_id: int) -> Dict[str, Any]:
    existing = db.query(CourseAccessDB).filter(CourseAccessDB.user_id == user.id, CourseAccessDB.course_id == course_id).first()
    if existing and (existing.permanent or existing.period_key == (user.subscription_started_at.isoformat() if user.subscription_started_at else None)):
        return {"allowed": True, "source": existing.source}
    plan, offer = _active_offer(user)
    if offer and int(offer["courses"]) == -1:
        return {"allowed": True, "source": plan}
    if offer:
        period_key = user.subscription_started_at.isoformat() if user.subscription_started_at else plan
        used = db.query(CourseAccessDB).filter(CourseAccessDB.user_id == user.id, CourseAccessDB.period_key == period_key).count()
        return {"allowed": used < int(offer["courses"]), "source": plan, "remaining": max(0, int(offer["courses"]) - used)}
    free_used = db.query(CourseAccessDB).filter(CourseAccessDB.user_id == user.id, CourseAccessDB.source == "free").count()
    return {"allowed": free_used < 1, "source": "free", "remaining": max(0, 1 - free_used)}


def _normalize_academy_steps(raw_steps: Any, previous: Optional[List[Dict[str, Any]]] = None) -> List[Dict[str, Any]]:
    previous_by_id = {str(item.get("id")): item for item in (previous or []) if isinstance(item, dict)}
    result: List[Dict[str, Any]] = []
    for index, raw in enumerate(raw_steps if isinstance(raw_steps, list) else []):
        if not isinstance(raw, dict):
            continue
        step_id = re.sub(r"[^a-zA-Z0-9_-]", "", str(raw.get("id") or f"step-{index + 1}")) or f"step-{index + 1}"
        old = previous_by_id.get(step_id, {})
        title = str(raw.get("title") or "").strip()
        content = str(raw.get("content") or "").strip()
        if not title or not content:
            continue
        result.append({
            "id": step_id,
            "title": title,
            "content": content,
            "image_url": raw.get("image_url") or old.get("image_url"),
            "audio": _normalize_expert_local_audio({**_load_json_dict(old.get("audio")), **_load_json_dict(raw.get("audio"))}),
        })
    return result


@app.get("/api/academy/courses")
async def list_public_academy_courses(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    courses = db.query(AcademyCourseDB).filter(AcademyCourseDB.status == "published").order_by(AcademyCourseDB.updated_at.desc()).all()
    return {"status": "success", "courses": [_serialize_academy_course(course, include_content=False, access=_course_access_status(db, current_user, course.id)) for course in courses]}


@app.post("/api/academy/courses/{course_id}/access")
async def access_academy_course(course_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    course = db.query(AcademyCourseDB).filter(AcademyCourseDB.id == course_id, AcademyCourseDB.status == "published").first()
    if not course:
        raise HTTPException(status_code=404, detail="Cours introuvable")
    access = _course_access_status(db, current_user, course_id)
    if not access.get("allowed"):
        raise HTTPException(status_code=402, detail="Votre accès aux cours est épuisé. Achetez ce cours à 100 F ou choisissez un abonnement.")
    existing = db.query(CourseAccessDB).filter(CourseAccessDB.user_id == current_user.id, CourseAccessDB.course_id == course_id).first()
    plan, offer = _active_offer(current_user)
    if not existing and not (offer and int(offer["courses"]) == -1):
        source = plan or "free"
        period_key = current_user.subscription_started_at.isoformat() if plan and current_user.subscription_started_at else None
        db.add(CourseAccessDB(user_id=current_user.id, course_id=course_id, source=source, period_key=period_key, permanent=(source == "free")))
        db.commit()
    return {"status": "success", "course": _serialize_academy_course(course, include_content=True, access={"allowed": True, "source": access.get("source")})}


@app.get("/api/admin/academy/courses")
async def list_admin_academy_courses(
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    del current_expert
    courses = db.query(AcademyCourseDB).order_by(AcademyCourseDB.updated_at.desc()).all()
    return {"status": "success", "courses": [_serialize_academy_course(course) for course in courses]}


@app.post("/api/admin/academy/courses")
async def create_academy_course(
    payload: AcademyCourseIn,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    if not payload.title.strip() or not payload.summary.strip():
        raise HTTPException(status_code=422, detail="Titre et présentation du cours obligatoires")
    course = AcademyCourseDB(
        title=payload.title.strip(),
        course_type="technique" if payload.course_type == "technique" else "culture",
        crop=(payload.crop or "").strip() or None,
        summary=payload.summary.strip(),
        steps_json=json.dumps(_normalize_academy_steps(payload.steps), ensure_ascii=False),
        audio_json="{}",
        status="draft" if payload.status == "draft" else "published",
        created_by=current_expert.id,
    )
    db.add(course)
    db.commit()
    db.refresh(course)
    return {"status": "success", "course": _serialize_academy_course(course)}


@app.put("/api/admin/academy/courses/{course_id}")
async def update_academy_course(
    course_id: int,
    payload: AcademyCourseIn,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    del current_expert
    course = db.query(AcademyCourseDB).filter(AcademyCourseDB.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Cours introuvable")
    course.title = payload.title.strip()
    course.course_type = "technique" if payload.course_type == "technique" else "culture"
    course.crop = (payload.crop or "").strip() or None
    course.summary = payload.summary.strip()
    course.steps_json = json.dumps(_normalize_academy_steps(payload.steps, _load_json_list(course.steps_json)), ensure_ascii=False)
    course.status = "draft" if payload.status == "draft" else "published"
    db.commit()
    db.refresh(course)
    return {"status": "success", "course": _serialize_academy_course(course)}


@app.post("/api/admin/academy/courses/{course_id}/media")
async def upload_academy_course_media(
    course_id: int,
    media_kind: str = Form(...),
    file: UploadFile = File(...),
    step_id: Optional[str] = Form(None),
    language: Optional[str] = Form(None),
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    del current_expert
    course = db.query(AcademyCourseDB).filter(AcademyCourseDB.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Cours introuvable")
    kind = media_kind.strip().lower()
    if kind not in {"image", "audio"}:
        raise HTTPException(status_code=422, detail="Type média invalide")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Le fichier envoyé est vide")
    extension = os.path.splitext(file.filename or "")[1].lower() or (".jpg" if kind == "image" else ".m4a")
    safe_step = re.sub(r"[^a-zA-Z0-9_-]", "", step_id or "course") or "course"
    normalized_language = _normalize_expert_local_language(language) if kind == "audio" else "image"
    relative_path = os.path.join("uploads", "academy", str(course.id), f"{safe_step}-{normalized_language}-{int(time.time() * 1000)}{extension}").replace("\\", "/")
    absolute_path = os.path.abspath(relative_path)
    _ensure_parent_dir(absolute_path)
    with open(absolute_path, "wb") as handle:
        handle.write(content)
    public_url = _build_upload_url(relative_path)
    if step_id:
        steps = _load_json_list(course.steps_json)
        target = next((item for item in steps if isinstance(item, dict) and str(item.get("id")) == str(step_id)), None)
        if target is None:
            raise HTTPException(status_code=404, detail="Étape introuvable")
        if kind == "image":
            target["image_url"] = public_url
        else:
            audio_map = _normalize_expert_local_audio(target.get("audio") or {})
            audio_map[normalized_language] = {"url": public_url, "mime_type": file.content_type, "uploaded_at": datetime.utcnow().isoformat()}
            target["audio"] = audio_map
        course.steps_json = json.dumps(steps, ensure_ascii=False)
    elif kind == "image":
        course.cover_url = relative_path
    else:
        audio_map = _normalize_expert_local_audio(_load_json_dict(course.audio_json))
        audio_map[normalized_language] = {"url": public_url, "mime_type": file.content_type, "uploaded_at": datetime.utcnow().isoformat()}
        course.audio_json = json.dumps(audio_map, ensure_ascii=False)
    db.commit()
    db.refresh(course)
    return {"status": "success", "course": _serialize_academy_course(course)}


@app.get("/api/expert/local-knowledge")
async def list_expert_local_knowledge(
    category: Optional[str] = None,
    language: Optional[str] = None,
    limit: int = Query(default=200, le=10000),
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    query = db.query(ExpertLocalKnowledgeDB)
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    if not is_admin:
        query = query.filter(ExpertLocalKnowledgeDB.expert_id == current_expert.id)
    if category:
        query = query.filter(
            ExpertLocalKnowledgeDB.category == _normalize_expert_local_category(category)
        )

    items = query.order_by(ExpertLocalKnowledgeDB.updated_at.desc()).limit(limit).all()
    serialized = [_serialize_expert_local_knowledge_item(item) for item in items]

    requested_language = _normalize_expert_local_language(language)
    if requested_language != "fr":
        serialized = [
            item
            for item in serialized
            if item.get("translations", {}).get(requested_language, {}).get("text")
            or item.get("audio", {}).get(requested_language, {}).get("url")
        ]

    return {"status": "success", "items": serialized}


@app.post("/api/expert/local-knowledge/import")
async def import_expert_local_knowledge(
    file: UploadFile = File(...),
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    """Importe des fiches et leurs vrais audios depuis JSON ou ZIP.

    Un ZIP doit contenir `manifest.json` et les fichiers audio référencés par
    `audio: {"moore": "audio/cas-moore.mp3", ...}` dans chaque fiche.
    """
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Fichier d'import vide")
    if len(raw) > 60 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Import limité à 60 Mo")

    archive: Optional[zipfile.ZipFile] = None
    try:
        if (file.filename or "").lower().endswith(".zip"):
            archive = zipfile.ZipFile(BytesIO(raw))
            manifest_name = next(
                (name for name in archive.namelist() if name.replace("\\", "/").lower().endswith("manifest.json")),
                None,
            )
            if not manifest_name:
                raise HTTPException(status_code=400, detail="manifest.json absent du ZIP")
            manifest = json.loads(archive.read(manifest_name).decode("utf-8"))
        else:
            manifest = json.loads(raw.decode("utf-8-sig"))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Import JSON/ZIP invalide: {exc}") from exc

    items = manifest.get("items") if isinstance(manifest, dict) else manifest
    if not isinstance(items, list) or not items:
        raise HTTPException(status_code=422, detail="Le manifeste doit contenir une liste `items`")
    if len(items) > 500:
        raise HTTPException(status_code=422, detail="Maximum 500 fiches par import")

    created = 0
    updated = 0
    audio_count = 0
    errors: List[Dict[str, Any]] = []
    for index, payload in enumerate(items, start=1):
        if not isinstance(payload, dict):
            errors.append({"index": index, "error": "Fiche non JSON"})
            continue
        title = str(payload.get("title") or "").strip()
        question_fr = str(payload.get("question_fr") or payload.get("question") or "").strip()
        resolution_fr = str(payload.get("resolution_fr") or payload.get("answer") or "").strip()
        if not title:
            title = question_fr[:120]
        if not title or not question_fr or not resolution_fr:
            errors.append({"index": index, "title": title, "error": "title, question_fr et resolution_fr requis"})
            continue

        category = _infer_expert_local_category(
            payload.get("category"), title, question_fr, resolution_fr, payload.get("tags") or []
        )
        item = db.query(ExpertLocalKnowledgeDB).filter(
            ExpertLocalKnowledgeDB.title == title,
            ExpertLocalKnowledgeDB.category == category,
        ).first()
        if item is None:
            item = ExpertLocalKnowledgeDB(title=title, category=category, question_fr=question_fr, resolution_fr=resolution_fr)
            db.add(item)
            db.flush()
            created += 1
        else:
            updated += 1

        item.question_fr = question_fr
        item.resolution_fr = resolution_fr
        item.tags_json = json.dumps(payload.get("tags") or [], ensure_ascii=False)
        item.status = _normalize_expert_local_status(payload.get("status"))
        item.origin = str(payload.get("origin") or "bulk_import")
        item.expert_id = item.expert_id or current_expert.id
        item.translations_json = json.dumps(
            _normalize_expert_local_translations(payload.get("translations")), ensure_ascii=False
        )

        next_audio = _normalize_expert_local_audio(payload.get("audio"))
        audio_refs = payload.get("audio") or payload.get("audios") or {}
        if archive and isinstance(audio_refs, dict):
            archive_names = {name.replace("\\", "/"): name for name in archive.namelist()}
            for language, audio_ref in audio_refs.items():
                normalized_language = _normalize_expert_local_language(language)
                ref_name = audio_ref.get("file") if isinstance(audio_ref, dict) else audio_ref
                ref_name = str(ref_name or "").replace("\\", "/").lstrip("/")
                member = archive_names.get(ref_name)
                if not member or member.endswith("/"):
                    continue
                audio_bytes = archive.read(member)
                if not audio_bytes or len(audio_bytes) > 20 * 1024 * 1024:
                    continue
                extension = os.path.splitext(ref_name)[1].lower()
                if extension not in {".mp3", ".wav", ".m4a", ".ogg", ".webm", ".aac"}:
                    extension = ".mp3"
                filename = f"expert-local-{item.id}-{normalized_language}{extension}"
                relative_path = os.path.join(EXPERT_AUDIO_UPLOAD_DIR, filename).replace("\\", "/")
                with open(os.path.abspath(relative_path), "wb") as handle:
                    handle.write(audio_bytes)
                next_audio[normalized_language] = {
                    "url": _build_upload_url(relative_path),
                    "mime_type": mimetypes.guess_type(filename)[0] or "audio/mpeg",
                    "uploaded_at": datetime.utcnow().isoformat(),
                }
                audio_count += 1
        item.audio_json = json.dumps(next_audio, ensure_ascii=False)

    db.commit()
    return {
        "status": "success",
        "created": created,
        "updated": updated,
        "audio_imported": audio_count,
        "errors": errors,
        "total_received": len(items),
    }


def _read_knowledge_source_document(filename: str, content: bytes) -> Tuple[str, List[Dict[str, Any]]]:
    extension = os.path.splitext(filename or "")[1].lower()
    rows: List[Dict[str, Any]] = []
    if extension == ".txt":
        return content.decode("utf-8-sig", errors="replace")[:120000], rows
    if extension == ".csv":
        decoded = content.decode("utf-8-sig", errors="replace")
        sample = decoded[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        rows = [dict(row) for row in csv.DictReader(decoded.splitlines(), dialect=dialect)][:500]
    elif extension == ".json":
        try:
            parsed = json.loads(content.decode("utf-8-sig"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise HTTPException(status_code=400, detail=f"Fichier JSON invalide : {exc}") from exc
        raw_items = parsed.get("items") if isinstance(parsed, dict) else parsed
        if not isinstance(raw_items, list):
            raise HTTPException(
                status_code=400,
                detail="Le JSON doit être une liste de fiches ou un objet contenant une clé 'items'.",
            )
        rows = [dict(item) for item in raw_items[:500] if isinstance(item, dict)]
    elif extension in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=503, detail="Le support Excel requiert openpyxl") from exc
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(values, [])]
            for values_row in values:
                row = {headers[index]: value for index, value in enumerate(values_row) if index < len(headers) and headers[index]}
                if any(value not in {None, ""} for value in row.values()):
                    row["_feuille"] = sheet.title
                    rows.append(row)
                if len(rows) >= 500:
                    break
            if len(rows) >= 500:
                break
    else:
        raise HTTPException(status_code=400, detail="Formats acceptés : TXT, CSV, JSON, XLSX")
    document = "\n".join(json.dumps(row, ensure_ascii=False, default=str) for row in rows)
    return document[:120000], rows


def _structured_knowledge_fallback(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def raw_value(row: Dict[str, Any], aliases: List[str]) -> Any:
        normalized = {_normalize_search_text(str(key)): val for key, val in row.items()}
        for alias in aliases:
            for key, raw in normalized.items():
                if alias in key and raw is not None and raw != "":
                    return raw
        return None

    def value(row: Dict[str, Any], aliases: List[str]) -> str:
        raw = raw_value(row, aliases)
        return str(raw).strip() if raw is not None else ""
    items = []
    for row in rows:
        title = value(row, ["maladie", "probleme", "titre", "title", "diagnostic", "nom"])
        question = value(row, ["symptome", "question", "description", "observation"])
        solution = value(row, ["solution", "traitement", "recommandation", "resolution", "conseil"])
        if not title or not solution:
            continue
        tags_raw = raw_value(row, ["tags", "mots cles", "mot cle", "culture", "espece"])
        if isinstance(tags_raw, list):
            tags = [str(part).strip() for part in tags_raw if str(part).strip()][:15]
        else:
            tags = [
                part.strip()
                for part in re.split(r"[,;|]", str(tags_raw or ""))
                if part.strip()
            ][:15]
        category = _infer_expert_local_category(
            value(row, ["categorie", "category", "domaine", "type"]),
            title, question, solution, tags,
        )
        items.append({
            "title": title, "category": category, "question_fr": question or title,
            "resolution_fr": solution,
            "tags": tags,
        })
    return items[:200]


@app.post("/api/expert/local-knowledge/ai-bulk-import")
async def ai_bulk_import_expert_local_knowledge(
    file: UploadFile = File(...),
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    """Transforme un document TXT, CSV, JSON ou Excel en fiches françaises."""
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Document vide")
    if len(raw) > 15 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Document limité à 15 Mo")
    document, structured_rows = _read_knowledge_source_document(file.filename or "document", raw)
    if not document.strip():
        raise HTTPException(status_code=422, detail="Aucune information exploitable dans le document")

    prompt = f"""Tu es l'agent documentaliste de Songra au Burkina Faso.
Transforme le document fourni en fiches de connaissances terrain distinctes. N'invente aucune maladie ni aucun dosage absent du document.
Pour chaque fiche, produis : title (nom court), category (agriculture, elevage, urgence ou cybersecurity), question_fr (symptômes/problème observable), resolution_fr (solution complète et prudente), tags (5 à 12 mots utiles normalisés).
Fusionne les doublons. Conserve les précautions et les conditions de consultation d'un agent agricole, vétérinaire ou soignant.
Retourne uniquement un JSON strict : {{"items":[{{"title":"...","category":"agriculture","question_fr":"...","resolution_fr":"...","tags":["..."]}}]}}.
Maximum 200 fiches.

DOCUMENT :
{document}"""
    extension = os.path.splitext(file.filename or "")[1].lower()
    parsed_items: List[Dict[str, Any]] = (
        _structured_knowledge_fallback(structured_rows) if extension == ".json" else []
    )
    provider_used = "json_structured" if parsed_items else "structured_fallback"
    provider_errors: List[str] = []
    providers = []
    if groq_client:
        providers.append(("groq", groq_client, os.getenv("GROQ_MODEL", "qwen/qwen3.6-27b")))
    if openai_client:
        providers.append(("openai", openai_client, os.getenv("OPENAI_TEXT_MODEL", "gpt-4o-mini")))
    for provider_name, client, model_name in ([] if parsed_items else providers):
        try:
            response = client.chat.completions.create(
                model=model_name, messages=[{"role": "user", "content": prompt}],
                temperature=0.1, max_tokens=6000,
            )
            parsed = _parse_json_object_from_text(response.choices[0].message.content or "")
            if isinstance(parsed.get("items"), list):
                parsed_items, provider_used = parsed["items"][:200], provider_name
                break
        except Exception as exc:
            provider_errors.append(f"{provider_name}: {exc}")
    if not parsed_items:
        parsed_items = _structured_knowledge_fallback(structured_rows)
    if not parsed_items:
        raise HTTPException(status_code=502, detail="L'agent IA n'a pas pu structurer ce document. " + " | ".join(provider_errors))

    created = updated = rejected = 0
    imported_items = []
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    for payload in parsed_items:
        title = str(payload.get("title") or "").strip()[:180]
        question = str(payload.get("question_fr") or "").strip()
        solution = str(payload.get("resolution_fr") or "").strip()
        if not title or not question or not solution:
            rejected += 1
            continue
        category = _infer_expert_local_category(
            payload.get("category"), title, question, solution, payload.get("tags") or []
        )
        item_query = db.query(ExpertLocalKnowledgeDB).filter(
            ExpertLocalKnowledgeDB.title == title,
            ExpertLocalKnowledgeDB.expert_id == current_expert.id,
        )
        # Pour un JSON structuré, le titre identifie la fiche même si une
        # ancienne version l'avait rangée dans une mauvaise catégorie.
        if extension != ".json":
            item_query = item_query.filter(ExpertLocalKnowledgeDB.category == category)
        item = item_query.first()
        if item:
            updated += 1
        else:
            item = ExpertLocalKnowledgeDB(title=title, category=category, expert_id=current_expert.id)
            db.add(item)
            created += 1
        item.category = category
        item.question_fr, item.resolution_fr = question, solution
        item.tags_json = json.dumps([str(tag).strip() for tag in (payload.get("tags") or []) if str(tag).strip()][:15], ensure_ascii=False)
        item.status = "validated" if is_admin else "pending_review"
        item.origin = f"ai_bulk_import:{provider_used}"
        item.translations_json = item.translations_json or "{}"
        item.audio_json = item.audio_json or "{}"
        imported_items.append(item)
    db.commit()
    return {
        "status": "success", "provider": provider_used, "created": created,
        "updated": updated, "rejected": rejected, "total_detected": len(parsed_items),
        "message": "Fiches françaises créées. Ajoutez maintenant les voix locales dans le Studio.",
    }


@app.get("/api/expert/local-knowledge/{item_id}")
async def get_expert_local_knowledge_item(
    item_id: int,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    item = db.query(ExpertLocalKnowledgeDB).filter(ExpertLocalKnowledgeDB.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Fiche locale introuvable")
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    if not is_admin and item.expert_id != current_expert.id:
        raise HTTPException(status_code=403, detail="Cette fiche appartient à un autre expert")
    return {"status": "success", "item": _serialize_expert_local_knowledge_item(item)}


@app.post("/api/expert/local-knowledge")
async def create_expert_local_knowledge_item(
    payload: ExpertLocalKnowledgeIn,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    item = ExpertLocalKnowledgeDB(
        title=payload.title.strip() or payload.question_fr[:120],
        category=_normalize_expert_local_category(payload.category),
        question_fr=payload.question_fr.strip(),
        resolution_fr=payload.resolution_fr.strip(),
        tags_json=json.dumps(payload.tags or [], ensure_ascii=False),
        status=_normalize_expert_local_status(payload.status) if is_admin else "pending_review",
        origin=str(payload.origin or "expert_manual"),
        expert_id=current_expert.id,
        translations_json=json.dumps(
            _normalize_expert_local_translations(payload.translations),
            ensure_ascii=False,
        ),
        audio_json=json.dumps(
            _normalize_expert_local_audio(payload.audio),
            ensure_ascii=False,
        ),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"status": "success", "item": _serialize_expert_local_knowledge_item(item)}


@app.patch("/api/expert/local-knowledge/{item_id}")
async def update_expert_local_knowledge_item(
    item_id: int,
    payload: ExpertLocalKnowledgeIn,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    item = db.query(ExpertLocalKnowledgeDB).filter(ExpertLocalKnowledgeDB.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Fiche locale introuvable")
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    if not is_admin and item.expert_id != current_expert.id:
        raise HTTPException(status_code=403, detail="Cette fiche appartient à un autre expert")

    next_title = payload.title.strip() or payload.question_fr[:120]
    next_category = _normalize_expert_local_category(payload.category)
    next_question = payload.question_fr.strip()
    next_resolution = payload.resolution_fr.strip()
    next_tags = [str(tag).strip() for tag in (payload.tags or []) if str(tag).strip()]
    content_changed = any((
        next_title != (item.title or ""),
        next_category != (item.category or ""),
        next_question != (item.question_fr or ""),
        next_resolution != (item.resolution_fr or ""),
        next_tags != [str(tag).strip() for tag in _load_json_list(item.tags_json)],
    ))
    previous_status = item.status
    item.title = next_title
    item.category = next_category
    item.question_fr = next_question
    item.resolution_fr = next_resolution
    item.tags_json = json.dumps(next_tags, ensure_ascii=False)
    if is_admin:
        requested_status = _normalize_expert_local_status(payload.status)
        # Le formulaire envoie historiquement pending_review même lorsqu'un admin
        # ajoute seulement une voix à une fiche déjà publiée.
        item.status = previous_status if previous_status in {"validated", "resolved", "expert_verified"} and requested_status == "pending_review" else requested_status
    else:
        item.status = previous_status if not content_changed and previous_status in {"validated", "resolved", "expert_verified"} else "pending_review"
    item.expert_id = item.expert_id or current_expert.id
    item.origin = str(payload.origin or item.origin or "expert_manual")
    item.translations_json = json.dumps(
        {
            **_load_json_dict(item.translations_json),
            **_normalize_expert_local_translations(payload.translations),
        },
        ensure_ascii=False,
    )
    item.audio_json = json.dumps(
        {
            **_load_json_dict(item.audio_json),
            **_normalize_expert_local_audio(payload.audio),
        },
        ensure_ascii=False,
    )
    db.commit()
    db.refresh(item)
    return {"status": "success", "item": _serialize_expert_local_knowledge_item(item)}


@app.post("/api/expert/local-knowledge/{item_id}/translate")
async def translate_expert_local_knowledge_item(
    item_id: int,
    background_tasks: BackgroundTasks,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    item = db.query(ExpertLocalKnowledgeDB).filter(ExpertLocalKnowledgeDB.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Fiche locale introuvable")
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    if not is_admin and item.expert_id != current_expert.id:
        raise HTTPException(status_code=403, detail="Cette fiche appartient à un autre expert")
    translations = _generate_local_translations(
        item.question_fr,
        item.resolution_fr,
        item.category,
        background_tasks,
    )
    item.translations_json = json.dumps(
        {**_load_json_dict(item.translations_json), **translations},
        ensure_ascii=False,
    )
    db.commit()
    db.refresh(item)
    return {"status": "success", "item": _serialize_expert_local_knowledge_item(item)}


@app.post("/api/expert/local-knowledge/{item_id}/audio-upload")
async def upload_expert_local_knowledge_audio(
    item_id: int,
    language: str = Form(default="fr"),
    audio: UploadFile = File(...),
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    item = db.query(ExpertLocalKnowledgeDB).filter(ExpertLocalKnowledgeDB.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Fiche locale introuvable")
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    if not is_admin and item.expert_id != current_expert.id:
        raise HTTPException(status_code=403, detail="Cette fiche appartient à un autre expert")

    normalized_language = _normalize_expert_local_language(language)
    extension = os.path.splitext(audio.filename or "")[1] or ".webm"
    file_name = f"expert-local-{item_id}-{normalized_language}{extension}"
    relative_path = os.path.join(EXPERT_AUDIO_UPLOAD_DIR, file_name).replace("\\", "/")
    absolute_path = os.path.abspath(relative_path)
    content = await audio.read()
    if not content:
        raise HTTPException(status_code=400, detail="Le fichier audio envoyé est vide")
    _ensure_parent_dir(absolute_path)
    with open(absolute_path, "wb") as handle:
        handle.write(content)

    next_audio = _normalize_expert_local_audio(_load_json_dict(item.audio_json))
    next_audio[normalized_language] = {
        "url": _build_upload_url(relative_path),
        "mime_type": audio.content_type or "audio/webm",
        "uploaded_at": datetime.utcnow().isoformat(),
    }
    item.audio_json = json.dumps(next_audio, ensure_ascii=False)
    db.commit()
    db.refresh(item)
    print(
        f"[STUDIO-AUDIO-UPLOAD] fiche=#{item.id} langue={normalized_language} "
        f"octets={len(content)} url={next_audio[normalized_language]['url']}"
    )
    return {"status": "success", "item": _serialize_expert_local_knowledge_item(item)}


@app.post("/api/expert/local-knowledge/{item_id}/review")
async def review_expert_local_knowledge_item(
    item_id: int,
    payload: ExpertLocalKnowledgeReviewIn,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    item = db.query(ExpertLocalKnowledgeDB).filter(ExpertLocalKnowledgeDB.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Fiche locale introuvable")
    status = str(payload.status or "").strip().lower()
    if status not in {"validated", "pending_review", "rejected", "archived"}:
        raise HTTPException(status_code=422, detail="Statut de validation invalide")
    item.status = status
    item.reviewer_id = current_expert.id
    item.review_notes = str(payload.review_notes or "").strip() or None
    item.reviewed_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    return {"status": "success", "item": _serialize_expert_local_knowledge_item(item)}


@app.get("/api/expert/local-knowledge-dashboard")
async def get_expert_local_knowledge_dashboard(
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    query = db.query(ExpertLocalKnowledgeDB)
    if not is_admin:
        query = query.filter(ExpertLocalKnowledgeDB.expert_id == current_expert.id)
    items = query.all()
    languages = ("moore", "dioula", "fulfulde")
    return {
        "total": len(items),
        "pending": sum(1 for item in items if item.status == "pending_review"),
        "validated": sum(1 for item in items if item.status in {"validated", "expert_verified", "resolved"}),
        "rejected": sum(1 for item in items if item.status == "rejected"),
        "complete_audio": sum(
            1 for item in items
            if all(_load_json_dict(item.audio_json).get(language, {}).get("url") for language in languages)
        ),
        "by_category": {
            category: sum(1 for item in items if item.category == category)
            for category in ("agriculture", "elevage", "urgence")
        },
    }


@app.get("/api/audio-map")
async def get_audio_map():
    return {"status": "success", "items": _load_audio_map_store()}


@app.put("/api/audio-map/{audio_key}")
async def update_audio_map_entry(
    audio_key: str,
    body: Dict[str, Any],
    current_expert: Expert = Depends(get_current_expert),
):
    del current_expert
    key = str(audio_key or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail="Clé audio requise")
    items = _load_audio_map_store()
    items[key] = _sanitize_audio_map_entry(body, items.get(key))
    _save_audio_map_store(items)
    return {"status": "success", "key": key, "item": items[key]}


@app.post("/api/audio-map/{audio_key}/upload")
async def upload_audio_map_entry(
    audio_key: str,
    language: str = Form(default="fr"),
    audio: UploadFile = File(...),
    current_expert: Expert = Depends(get_current_expert),
):
    del current_expert
    key = str(audio_key or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail="Clé audio requise")

    normalized_language = _normalize_expert_local_language(language)
    extension = os.path.splitext(audio.filename or "")[1] or ".webm"
    file_name = f"{key}-{normalized_language}{extension}".replace("/", "-")
    relative_path = os.path.join(EXPERT_AUDIO_UPLOAD_DIR, file_name).replace("\\", "/")
    absolute_path = os.path.abspath(relative_path)
    content = await audio.read()
    with open(absolute_path, "wb") as handle:
        handle.write(content)

    items = _load_audio_map_store()
    current = items.get(key, {})
    next_entry = _sanitize_audio_map_entry(current, current)
    next_audio_meta = {
        "url": _build_upload_url(relative_path),
        "mime_type": audio.content_type or "audio/webm",
        "updated_at": datetime.utcnow().isoformat(),
    }
    next_entry.setdefault("audios", {})
    next_entry["audios"][normalized_language] = next_audio_meta
    if normalized_language == "fr":
        next_entry["audio"] = next_audio_meta["url"]
        next_entry["mime_type"] = next_audio_meta["mime_type"]
    next_entry["updated_at"] = datetime.utcnow().isoformat()
    items[key] = next_entry
    _save_audio_map_store(items)
    return {"status": "success", "key": key, "item": items[key]}

@app.post("/api/create-test-expert")
async def create_test_expert(db: Session = Depends(get_db)):
    if os.getenv("ENABLE_TEST_EXPERT", "false").lower() != "true":
        raise HTTPException(status_code=404, detail="Route indisponible")
    existing = db.query(Expert).filter(Expert.email == "test@resolvehub.bf").first()
    if existing:
        return {
            "message": "Expert déjà existant", 
            "email": "test@resolvehub.bf", 
            "status": "already_exists"
        }
    
    expert = Expert(
        email="test@resolvehub.bf",
        password_hash=hash_password("test123"),
        full_name="Expert Test IA",
        specialization="agriculture",
        is_active=True
    )
    db.add(expert)
    db.commit()
    
    return {
        "message": "Expert créé avec succès", 
        "email": "test@resolvehub.bf", 
        "status": "created"
    }

# ==========================================
# NOUVEAUX ENDPOINTS POUR FONCTIONNALITÉS PRIORITAIRES
# ==========================================

class SOSAlert(BaseModel):
    phoneNumber: str
    type: str  # accident, attack, fire, animal_sick, community
    description: Optional[str] = None
    location: Dict[str, Any]  # {latitude, longitude, accuracy, note}
    timestamp: int
    urgent: bool = True

class SOSAlertDB(Base):
    __tablename__ = "sos_alerts"
    id = Column(Integer, primary_key=True, index=True)
    phone_number = Column(String, nullable=False)
    alert_type = Column(String, nullable=False)
    description = Column(Text, nullable=True)
    latitude = Column(Float, nullable=True)
    longitude = Column(Float, nullable=True)
    accuracy = Column(Float, nullable=True)
    location_note = Column(String, nullable=True)
    timestamp = Column(DateTime, default=datetime.utcnow)
    status = Column(String, default="pending")  # pending, acknowledged, resolved
    notified_authorities = Column(Boolean, default=False)
    assigned_expert_id = Column(Integer, nullable=True, index=True)
    passed_expert_ids_json = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class ChatMessageDB(Base):
    __tablename__ = "chat_messages"
    id = Column(Integer, primary_key=True, index=True)
    sender = Column(String, nullable=False)  # numéro ou nom affiché
    text = Column(Text, nullable=False)
    is_bot = Column(Boolean, default=False)
    room = Column(String, default="general")
    sender_role = Column(String, default="member")
    report_count = Column(Integer, default=0)
    is_hidden = Column(Boolean, default=False)
    is_pinned = Column(Boolean, default=False)
    pinned_at = Column(DateTime, nullable=True)
    audio_url = Column(String, nullable=True)  # URL du fichier audio communautaire
    created_at = Column(DateTime, default=datetime.utcnow)


class CommunityFieldCaseDB(Base):
    __tablename__ = "community_field_cases"
    id = Column(Integer, primary_key=True, index=True)
    room = Column(String, default="general")
    category = Column(String, default="agriculture")
    title = Column(String, nullable=False)
    description = Column(Text, nullable=False)
    reporter_name = Column(String, nullable=False)
    reporter_phone = Column(String, nullable=True)
    contributor_role = Column(String, default="member")
    severity = Column(String, default="medium")
    status = Column(String, default="new")
    location_label = Column(String, nullable=True)
    latitude = Column(Float, nullable=True)
    longitude = Column(Float, nullable=True)
    crop_or_livestock = Column(String, nullable=True)
    tags_json = Column(Text, nullable=True)
    before_photo_paths_json = Column(Text, nullable=True)
    before_audio_paths_json = Column(Text, nullable=True)
    after_photo_paths_json = Column(Text, nullable=True)
    after_audio_paths_json = Column(Text, nullable=True)
    promoted_to_offline = Column(Boolean, default=False)
    resolved_at = Column(DateTime, nullable=True)
    last_follow_up_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class CommunitySolutionDB(Base):
    __tablename__ = "community_case_solutions"
    id = Column(Integer, primary_key=True, index=True)
    case_id = Column(Integer, nullable=False, index=True)
    author_name = Column(String, nullable=False)
    author_phone = Column(String, nullable=True)
    contributor_role = Column(String, default="member")
    text = Column(Text, nullable=False)
    action_taken = Column(Text, nullable=True)
    cost_note = Column(String, nullable=True)
    delay_note = Column(String, nullable=True)
    result_status = Column(String, default="proposed")
    photo_paths_json = Column(Text, nullable=True)
    is_expert = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class CommunitySolutionFeedbackDB(Base):
    __tablename__ = "community_solution_feedback"
    id = Column(Integer, primary_key=True, index=True)
    solution_id = Column(Integer, nullable=False, index=True)
    voter_name = Column(String, nullable=False)
    voter_phone = Column(String, nullable=True)
    feedback_type = Column(String, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)


class CommunityCaseConfirmationDB(Base):
    __tablename__ = "community_case_confirmations"
    id = Column(Integer, primary_key=True, index=True)
    case_id = Column(Integer, nullable=False, index=True)
    confirmer_name = Column(String, nullable=False)
    confirmer_phone = Column(String, nullable=True)
    contributor_role = Column(String, default="member")
    note = Column(Text, nullable=True)
    location_label = Column(String, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class CommunityCaseFollowUpDB(Base):
    __tablename__ = "community_case_followups"
    id = Column(Integer, primary_key=True, index=True)
    case_id = Column(Integer, nullable=False, index=True)
    author_name = Column(String, nullable=False)
    author_phone = Column(String, nullable=True)
    contributor_role = Column(String, default="member")
    note = Column(Text, nullable=False)
    status_after = Column(String, nullable=True)
    outcome_label = Column(String, nullable=True)
    photo_paths_json = Column(Text, nullable=True)
    audio_paths_json = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

Base.metadata.create_all(bind=engine)
# 
# 
def _ensure_sos_assignment_columns() -> None:
    try:
        with engine.connect() as conn:
            _add_column_if_missing(conn, "sos_alerts", "assigned_expert_id", "INTEGER", "INTEGER")
            _add_column_if_missing(conn, "sos_alerts", "passed_expert_ids_json", "TEXT", "TEXT")
            conn.commit()
    except Exception as exc:
        print(f"[WARN] Migration affectation SOS impossible: {exc}")


_ensure_sos_assignment_columns()


def _ensure_community_case_media_columns() -> None:
    try:
        with engine.connect() as conn:
            if str(engine.url).startswith("sqlite"):
                case_columns = [row[1] for row in conn.exec_driver_sql("PRAGMA table_info(community_field_cases)")]
                if "before_audio_paths_json" not in case_columns:
                    conn.exec_driver_sql(
                        "ALTER TABLE community_field_cases ADD COLUMN before_audio_paths_json TEXT"
                    )
                if "after_audio_paths_json" not in case_columns:
                    conn.exec_driver_sql(
                        "ALTER TABLE community_field_cases ADD COLUMN after_audio_paths_json TEXT"
                    )

                follow_up_columns = [row[1] for row in conn.exec_driver_sql("PRAGMA table_info(community_case_followups)")]
                if "audio_paths_json" not in follow_up_columns:
                    conn.exec_driver_sql(
                        "ALTER TABLE community_case_followups ADD COLUMN audio_paths_json TEXT"
                    )
            else:
                conn.exec_driver_sql(
                    "ALTER TABLE community_field_cases ADD COLUMN IF NOT EXISTS before_audio_paths_json TEXT"
                )
                conn.exec_driver_sql(
                    "ALTER TABLE community_field_cases ADD COLUMN IF NOT EXISTS after_audio_paths_json TEXT"
                )
                conn.exec_driver_sql(
                    "ALTER TABLE community_case_followups ADD COLUMN IF NOT EXISTS audio_paths_json TEXT"
                )
    except Exception as e:
        print(f"[WARN] Impossible d'ajouter les colonnes audio des cas terrain: {e}")


_ensure_community_case_media_columns()


def _ensure_sos_description_column() -> None:
    """Migration: ajouter la colonne 'description' à sos_alerts si absente."""
    try:
        with engine.connect() as conn:
            _add_column_if_missing(conn, "sos_alerts", "description", "TEXT", "TEXT")
            conn.commit()
    except Exception as e:
        print(f"[WARN] Impossible d'ajouter la colonne 'description' à sos_alerts: {e}")

_ensure_sos_description_column()


def _ensure_chat_message_columns() -> None:
    """Migration légère pour enrichir le module communauté."""
    try:
        with engine.connect() as conn:
            _add_column_if_missing(conn, "chat_messages", "room", "TEXT DEFAULT 'general'", "TEXT DEFAULT 'general'")
            _add_column_if_missing(conn, "chat_messages", "sender_role", "TEXT DEFAULT 'member'", "TEXT DEFAULT 'member'")
            _add_column_if_missing(conn, "chat_messages", "report_count", "INTEGER DEFAULT 0", "INTEGER DEFAULT 0")
            _add_column_if_missing(conn, "chat_messages", "is_hidden", "BOOLEAN DEFAULT 0", "BOOLEAN DEFAULT FALSE")
            _add_column_if_missing(conn, "chat_messages", "is_pinned", "BOOLEAN DEFAULT 0", "BOOLEAN DEFAULT FALSE")
            _add_column_if_missing(conn, "chat_messages", "pinned_at", "DATETIME", "TIMESTAMP")
            conn.commit()
    except Exception as e:
        print(f"[WARN] Impossible d'ajouter les colonnes de communauté: {e}")


_ensure_chat_message_columns()

@app.post("/api/sos/alert")
async def create_sos_alert(alert: SOSAlert, db: Session = Depends(get_db)):
    """
    🔑 ENDPOINT SOS - PRIORITÉ CRITIQUE
    Enregistrer une alerte SOS avec géolocalisation
    """
    try:
        # Créer l'alerte SOS
        sos_alert = SOSAlertDB(
            phone_number=alert.phoneNumber or "Anonyme",
            alert_type=alert.type,
            description=alert.description,
            latitude=alert.location.get("latitude"),
            longitude=alert.location.get("longitude"),
            accuracy=alert.location.get("accuracy"),
            location_note=alert.location.get("note"),
            status="pending",
            notified_authorities=False
        )
        
        db.add(sos_alert)
        db.commit()
        db.refresh(sos_alert)
        
        # Notification automatique vers WhatsApp (via Webhook) et Africa's Talking
        try:
            import httpx
            sys_settings = {}
            settings_path = os.path.join(BACKEND_DIR, "system_settings.json")
            if os.path.exists(settings_path):
                try:
                    with open(settings_path, "r", encoding="utf-8") as f:
                        sys_settings = json.load(f)
                except:
                    pass

            whatsapp_webhook = sys_settings.get("whatsapp_webhook_url") or os.getenv("WHATSAPP_WEBHOOK_URL")
            alert_phone = sys_settings.get("alert_target_phone") or os.getenv("ALERT_TARGET_PHONE", "+22670000017")
            
            # Nettoyage automatique pour WhatsApp (pas de +, pas d'espaces)
            clean_phone = alert_phone.replace("+", "").replace(" ", "").strip()
            if clean_phone.startswith("00"):
                clean_phone = clean_phone[2:]
            
            message_payload = f"🚨 *NOUVEAU SIGNALEMENT SONGRA ({alert.type.upper()})* 🚨\n\n"
            message_payload += f"👤 *Auteur* : {alert.phoneNumber or 'Anonyme'}\n"
            message_payload += f"📝 *Détails* : {alert.description}\n"
            
            if alert.location:
                lat = alert.location.get("latitude")
                lng = alert.location.get("longitude")
                if lat and lng:
                    message_payload += f"📍 *Position* : https://maps.google.com/?q={lat},{lng}\n"
                
            print(f"[WHATSAPP OUTGOING] Destinataire: {clean_phone} | Message: {message_payload}")
            
            # 1. Si Webhook configuré
            if whatsapp_webhook and whatsapp_webhook != "votre_webhook_url":
                async with httpx.AsyncClient() as client:
                    await client.post(whatsapp_webhook, json={
                        "to": clean_phone,
                        "message": message_payload,
                        "type": alert.type
                    }, timeout=5.0)
            
            # 2. Si Africa's Talking configuré
            at_username = os.getenv("AFRICAS_TALKING_USERNAME")
            at_api_key = os.getenv("AFRICAS_TALKING_API_KEY")
            if at_username and at_api_key and at_api_key != "votre_cle_api":
                url = "https://api.africastalking.com/version1/messaging"
                headers = {
                    "ApiKey": at_api_key,
                    "Content-Type": "application/x-www-form-urlencoded",
                    "Accept": "application/json"
                }
                data = {
                    "username": at_username,
                    "to": alert_phone,
                    "message": message_payload
                }
                async with httpx.AsyncClient() as client:
                    await client.post(url, headers=headers, data=data, timeout=5.0)
        except Exception as alert_err:
            print(f"[WARN] Erreur lors de l'envoi de la notification externe: {alert_err}")

        
        print(f"🚨 ALERTE SOS REÇUE - Type: {alert.type}, Tel: {alert.phoneNumber}")
        if alert.location.get("latitude"):
            print(f"   📍 Position: {alert.location.get('latitude')}, {alert.location.get('longitude')}")
        
        return {
            "success": True,
            "alert_id": sos_alert.id,
            "message": "Alerte SOS enregistrée. Secours notifiés.",
            "emergency_numbers": [
                {"label": "Police", "number": "17"},
                {"label": "Pompiers", "number": "18"},
                {"label": "SAMU", "number": "15"}
            ]
        }
    
    except Exception as e:
        print(f"❌ Erreur SOS: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur enregistrement SOS: {str(e)}")

@app.patch("/api/sos/alerts/{alert_id}/status")
async def update_sos_alert_status(
    alert_id: int,
    body: Dict[str, Any],
    current_identity: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db)
):
    """
    Mettre à jour le statut d'une alerte SOS (pour panel expert).
    Body JSON: { "status": "acknowledged" | "resolved" }
    """
    alert = db.query(SOSAlertDB).filter(SOSAlertDB.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte SOS introuvable")
    new_status = body.get("status")
    if new_status not in ("pending", "acknowledged", "resolved"):
        raise HTTPException(status_code=422, detail="Statut invalide")
    alert.status = new_status
    db.commit()
    return {"success": True, "alert_id": alert_id, "status": new_status}

@app.get("/api/sos/alerts")
async def get_sos_alerts(
    status: Optional[str] = None,
    limit: int = Query(default=50, le=200),
    current_identity: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db)
):
    """
    Récupérer les alertes SOS (pour dashboard admin/expert)
    """
    query = db.query(SOSAlertDB).order_by(SOSAlertDB.created_at.desc())
    
    if status:
        query = query.filter(SOSAlertDB.status == status)
    
    alerts = query.limit(limit).all()
    
    return {
        "alerts": [
            {
                "id": alert.id,
                "phoneNumber": alert.phone_number,
                "type": alert.alert_type,
                "description": alert.description,
                "location": {
                    "latitude": alert.latitude,
                    "longitude": alert.longitude,
                    "accuracy": alert.accuracy,
                    "note": alert.location_note
                },
                "status": alert.status,
                "timestamp": alert.created_at.isoformat() if alert.created_at else None
            }
            for alert in alerts
        ],
        "total": len(alerts)
    }


def _expert_has_passed(raw_ids: Optional[str], expert_id: int) -> bool:
    return expert_id in [int(value) for value in _load_json_list(raw_ids) if str(value).isdigit()]


def _append_passed_expert(raw_ids: Optional[str], expert_id: int) -> str:
    ids = [int(value) for value in _load_json_list(raw_ids) if str(value).isdigit()]
    if expert_id not in ids:
        ids.append(expert_id)
    return json.dumps(ids)


@app.get("/api/expert/nearby-work")
async def get_expert_nearby_work(
    latitude: Optional[float] = None,
    longitude: Optional[float] = None,
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    """File terrain triée par zone : SOS et tickets libres/affectés à l'expert."""
    expert_zone = _normalize_search_text(current_expert.zone or current_expert.project or "").strip()
    specialization = _normalized_scope_value(current_expert.specialization)
    work: List[Dict[str, Any]] = []

    def distance_km(target_lat: Optional[float], target_lng: Optional[float]) -> Optional[float]:
        if latitude is None or longitude is None or target_lat is None or target_lng is None:
            return None
        radius = 6371.0
        lat1, lat2 = math.radians(latitude), math.radians(target_lat)
        delta_lat = math.radians(target_lat - latitude)
        delta_lng = math.radians(target_lng - longitude)
        value = math.sin(delta_lat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(delta_lng / 2) ** 2
        return radius * 2 * math.atan2(math.sqrt(value), math.sqrt(1 - value))

    alerts = db.query(SOSAlertDB).filter(
        SOSAlertDB.status != "resolved",
        or_(SOSAlertDB.assigned_expert_id.is_(None), SOSAlertDB.assigned_expert_id == current_expert.id),
    ).order_by(SOSAlertDB.created_at.desc()).limit(100).all()
    for alert in alerts:
        if _expert_has_passed(alert.passed_expert_ids_json, current_expert.id):
            continue
        location = str(alert.location_note or "").strip()
        zone_match = bool(expert_zone and expert_zone in _normalize_search_text(location))
        distance = distance_km(alert.latitude, alert.longitude)
        work.append({
            "work_type": "sos", "id": alert.id,
            "title": f"Alerte {alert.alert_type}", "description": alert.description,
            "category": "urgence", "urgency": "critical", "location": location or "Position GPS disponible",
            "latitude": alert.latitude, "longitude": alert.longitude, "status": alert.status,
            "assigned_to_me": alert.assigned_expert_id == current_expert.id,
            "distance_km": round(distance, 1) if distance is not None else None,
            "proximity_score": max(1, 1000 - distance) if distance is not None else (100 if zone_match else (60 if alert.latitude is not None else 20)),
            "created_at": alert.created_at.isoformat() if alert.created_at else None,
        })

    ticket_query = db.query(Ticket, User).join(User, User.id == Ticket.user_id).filter(
        Ticket.status != "resolved",
        or_(Ticket.expert_id.is_(None), Ticket.expert_id == current_expert.id),
    )
    if specialization and specialization not in {"all", "admin", "general", "tous"}:
        ticket_query = ticket_query.filter(func.lower(func.coalesce(Ticket.category, "")) == specialization)
    for ticket, user in ticket_query.order_by(Ticket.created_at.desc()).limit(100).all():
        if _expert_has_passed(ticket.passed_expert_ids_json, current_expert.id):
            continue
        location = str(user.location or "").strip()
        zone_match = bool(expert_zone and expert_zone in _normalize_search_text(location))
        work.append({
            "work_type": "ticket", "id": ticket.id,
            "title": f"Ticket {ticket.category or 'terrain'} #{ticket.id}",
            "description": (db.query(Message).filter(Message.ticket_id == ticket.id).order_by(Message.sent_at.asc()).first() or Message(content="Aucune description", ticket_id=ticket.id, sender_type="user", channel="app")).content,
            "category": ticket.category, "urgency": ticket.urgency, "location": location or "Localisation non précisée",
            "latitude": None, "longitude": None, "status": ticket.status,
            "assigned_to_me": ticket.expert_id == current_expert.id,
            "proximity_score": 100 if zone_match else 10,
            "created_at": ticket.created_at.isoformat() if ticket.created_at else None,
        })
    work.sort(key=lambda item: (item["assigned_to_me"], item["proximity_score"], item["urgency"] in {"critical", "high"}, item["created_at"] or ""), reverse=True)
    return {"items": work[:100], "total": len(work), "expert_zone": current_expert.zone}


@app.post("/api/expert/nearby-work/{work_type}/{work_id}/action")
async def act_on_expert_nearby_work(
    work_type: str,
    work_id: int,
    body: Dict[str, Any],
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db),
):
    action = str(body.get("action") or "").strip().lower()
    if action not in {"claim", "pass", "resolve"}:
        raise HTTPException(status_code=422, detail="Action invalide")
    if work_type == "sos":
        item = db.query(SOSAlertDB).filter(SOSAlertDB.id == work_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Alerte introuvable")
        if action == "claim":
            if item.assigned_expert_id not in {None, current_expert.id}:
                raise HTTPException(status_code=409, detail="Alerte déjà prise par un autre expert")
            item.assigned_expert_id, item.status = current_expert.id, "acknowledged"
        elif action == "pass":
            if item.assigned_expert_id == current_expert.id:
                item.assigned_expert_id, item.status = None, "pending"
            item.passed_expert_ids_json = _append_passed_expert(item.passed_expert_ids_json, current_expert.id)
        else:
            if item.assigned_expert_id != current_expert.id:
                raise HTTPException(status_code=403, detail="Prenez d'abord cette alerte en charge")
            item.status = "resolved"
    elif work_type == "ticket":
        item = db.query(Ticket).filter(Ticket.id == work_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Ticket introuvable")
        if action == "claim":
            if item.expert_id not in {None, current_expert.id}:
                raise HTTPException(status_code=409, detail="Ticket déjà pris par un autre expert")
            item.expert_id, item.status = current_expert.id, "assigned"
        elif action == "pass":
            if item.expert_id == current_expert.id:
                item.expert_id, item.status = None, "open"
            item.passed_expert_ids_json = _append_passed_expert(item.passed_expert_ids_json, current_expert.id)
        else:
            if item.expert_id != current_expert.id:
                raise HTTPException(status_code=403, detail="Prenez d'abord ce ticket en charge")
            item.status, item.resolved_at = "resolved", datetime.utcnow()
    else:
        raise HTTPException(status_code=422, detail="Type de travail invalide")
    db.commit()
    return {"status": "success", "action": action, "work_type": work_type, "id": work_id}


# ==========================================
# COMMUNITY CHAT
# ==========================================

BOT_REPLIES = [
    "Merci pour votre partage ! Un expert vous répondra bientôt.",
    "Bonne question ! D'autres membres ont eu le même problème.",
    "Consultez aussi la section Agriculture pour plus de conseils.",
    "Vous n'êtes pas seul ! La communauté SONGRA est là pour vous aider.",
    "Votre question a bien été notée. Restez connecté !",
]

COMMUNITY_ROOM_LABELS = {
    "general": "Général",
    "agriculture": "Agriculture",
    "elevage": "Élevage",
    "securite": "Sécurité",
    "marche": "Marché",
}

COMMUNITY_ROOM_TO_DOMAIN = {
    "general": "agriculture",
    "agriculture": "agriculture",
    "elevage": "elevage",
    "securite": "cybersecurity",
    "marche": "agriculture",
}

COMMUNITY_DUPLICATE_WINDOW_SECONDS = 120
COMMUNITY_REPORT_HIDE_THRESHOLD = 3
COMMUNITY_BLOCKED_TERMS = {
    "arnaque",
    "escroc",
    "haine",
    "insulte",
}

COMMUNITY_CASE_STATUS_LABELS = {
    "new": "Nouveau",
    "in_progress": "En cours",
    "improved": "Ameliore",
    "resolved": "Resolu",
    "watch": "A surveiller",
}

COMMUNITY_CASE_SEVERITY_LABELS = {
    "low": "Faible",
    "medium": "Moyenne",
    "high": "Elevee",
    "critical": "Critique",
}

COMMUNITY_SOLUTION_FEEDBACK_TYPES = {"useful", "tested", "worked", "failed"}

COMMUNITY_CONTRIBUTOR_ROLE_LABELS = {
    "member": "Acteur terrain",
    "observer": "Observateur terrain",
    "solver": "Solutionneur local",
    "referent": "Producteur referent",
    "expert": "Expert communaute",
}


def _normalize_community_room(raw_room: Optional[str]) -> str:
    normalized = (raw_room or "general").strip().lower()
    return normalized if normalized in COMMUNITY_ROOM_LABELS else "general"


def _normalize_community_case_status(raw_status: Optional[str]) -> str:
    normalized = (raw_status or "new").strip().lower()
    return normalized if normalized in COMMUNITY_CASE_STATUS_LABELS else "new"


def _normalize_community_case_severity(raw_severity: Optional[str]) -> str:
    normalized = (raw_severity or "medium").strip().lower()
    return normalized if normalized in COMMUNITY_CASE_SEVERITY_LABELS else "medium"


def _normalize_community_feedback_type(raw_feedback_type: Optional[str]) -> str:
    normalized = (raw_feedback_type or "useful").strip().lower()
    return normalized if normalized in COMMUNITY_SOLUTION_FEEDBACK_TYPES else "useful"


def _normalize_community_category(raw_category: Optional[str], room: str) -> str:
    normalized = (raw_category or "").strip().lower()
    if normalized in {"agriculture", "elevage", "cybersecurity", "sos_accident"}:
        return normalized
    if room == "elevage":
        return "elevage"
    if room == "securite":
        return "cybersecurity"
    return "agriculture"


def _serialize_community_photo_urls(raw_paths_json: Optional[str]) -> List[Dict[str, Any]]:
    paths = _load_json_list(raw_paths_json)
    return [
        {
            "path": path,
            "url": _build_upload_url(path),
        }
        for path in paths
        if path
    ]


def _community_person_filter(phone_field: Any, name_field: Any, phone_number: Optional[str], name: Optional[str]):
    if phone_number:
        return phone_field == phone_number
    return name_field == (name or "")


# def _build_community_contributor_profile(
#     db: Session,
#     *,
#     name: Optional[str],
#     phone_number: Optional[str],
#     preferred_role: Optional[str] = None,
#     is_expert: bool = False,
# ) -> Dict[str, Any]:
#     display_name = (name or "Membre SONGRA").strip() or "Membre SONGRA"
#     if is_expert or preferred_role == "expert":
#         role_key = "expert"
#         return {
#             "name": display_name,
#             "role_key": role_key,
#             "role_label": COMMUNITY_CONTRIBUTOR_ROLE_LABELS[role_key],
#             "case_count": 0,
#             "solution_count": 0,
#             "confirmation_count": 0,
#             "follow_up_count": 0,
#         }
# 
#     case_count = db.query(CommunityFieldCaseDB).filter(
#         _community_person_filter(
#             CommunityFieldCaseDB.reporter_phone,
#             CommunityFieldCaseDB.reporter_name,
#             phone_number,
#             display_name,
#         )
#     ).count()
#     solution_count = db.query(CommunitySolutionDB).filter(
#         _community_person_filter(
#             CommunitySolutionDB.author_phone,
#             CommunitySolutionDB.author_name,
#             phone_number,
#             display_name,
#         )
#     ).count()
#     confirmation_count = db.query(CommunityCaseConfirmationDB).filter(
#         _community_person_filter(
#             CommunityCaseConfirmationDB.confirmer_phone,
#             CommunityCaseConfirmationDB.confirmer_name,
#             phone_number,
#             display_name,
#         )
#     ).count()
#     follow_up_count = db.query(CommunityCaseFollowUpDB).filter(
#         _community_person_filter(
#             CommunityCaseFollowUpDB.author_phone,
#             CommunityCaseFollowUpDB.author_name,
#             phone_number,
#             display_name,
#         )
#     ).count()
# 
#     if preferred_role in {"observer", "solver", "referent"}:
#         role_key = preferred_role
#     elif solution_count >= 3 and follow_up_count >= 1:
#         role_key = "referent"
#     elif solution_count >= 2:
#         role_key = "solver"
#     elif case_count + confirmation_count + follow_up_count >= 3:
#         role_key = "observer"
#     else:
#         role_key = "member"
# 
#     return {
#         "name": display_name,
#         "role_key": role_key,
#         "role_label": COMMUNITY_CONTRIBUTOR_ROLE_LABELS[role_key],
#         "case_count": case_count,
#         "solution_count": solution_count,
#         "confirmation_count": confirmation_count,
#         "follow_up_count": follow_up_count,
#     }
# 
# 
# def _serialize_community_solution_feedback_counts(db: Session, solution_id: int) -> Dict[str, int]:
#     rows = (
#         db.query(
#             CommunitySolutionFeedbackDB.feedback_type,
#             func.count(CommunitySolutionFeedbackDB.id),
#         )
#         .filter(CommunitySolutionFeedbackDB.solution_id == solution_id)
#         .group_by(CommunitySolutionFeedbackDB.feedback_type)
#         .all()
#     )
#     counts = {"useful": 0, "tested": 0, "worked": 0, "failed": 0}
#     for feedback_type, total in rows:
#         if feedback_type in counts:
#             counts[feedback_type] = int(total or 0)
#     return counts
# 
# 
# def _serialize_community_solution(db: Session, solution: CommunitySolutionDB) -> Dict[str, Any]:
#     feedback = _serialize_community_solution_feedback_counts(db, solution.id)
#     contributor = _build_community_contributor_profile(
#         db,
#         name=solution.author_name,
#         phone_number=solution.author_phone,
#         preferred_role=solution.contributor_role,
#         is_expert=bool(solution.is_expert),
#     )
#     return {
#         "id": solution.id,
#         "case_id": solution.case_id,
#         "author": contributor,
#         "text": solution.text,
#         "action_taken": solution.action_taken,
#         "cost_note": solution.cost_note,
#         "delay_note": solution.delay_note,
#         "result_status": solution.result_status,
#         "photos": _serialize_community_photo_urls(solution.photo_paths_json),
#         "feedback": feedback,
#         "created_at": solution.created_at.isoformat() if solution.created_at else None,
#         "updated_at": solution.updated_at.isoformat() if solution.updated_at else None,
#     }
# 
# 
# def _serialize_community_confirmation(db: Session, confirmation: CommunityCaseConfirmationDB) -> Dict[str, Any]:
#     return {
#         "id": confirmation.id,
#         "case_id": confirmation.case_id,
#         "contributor": _build_community_contributor_profile(
#             db,
#             name=confirmation.confirmer_name,
#             phone_number=confirmation.confirmer_phone,
#             preferred_role=confirmation.contributor_role,
#         ),
#         "note": confirmation.note,
#         "location_label": confirmation.location_label,
#         "created_at": confirmation.created_at.isoformat() if confirmation.created_at else None,
#     }
# 
# 
# def _serialize_community_follow_up(db: Session, follow_up: CommunityCaseFollowUpDB) -> Dict[str, Any]:
#     return {
#         "id": follow_up.id,
#         "case_id": follow_up.case_id,
#         "contributor": _build_community_contributor_profile(
#             db,
#             name=follow_up.author_name,
#             phone_number=follow_up.author_phone,
#             preferred_role=follow_up.contributor_role,
#         ),
#         "note": follow_up.note,
#         "status_after": _normalize_community_case_status(follow_up.status_after),
#         "status_after_label": COMMUNITY_CASE_STATUS_LABELS.get(
#             _normalize_community_case_status(follow_up.status_after),
#             COMMUNITY_CASE_STATUS_LABELS["new"],
#         ),
#         "outcome_label": follow_up.outcome_label,
#         "photos": _serialize_community_photo_urls(follow_up.photo_paths_json),
#         "audios": _serialize_community_audio_urls(follow_up.audio_paths_json),
#         "created_at": follow_up.created_at.isoformat() if follow_up.created_at else None,
#     }
# 
# 
# def _serialize_community_field_case_summary(db: Session, case: CommunityFieldCaseDB) -> Dict[str, Any]:
#     solution_count = db.query(CommunitySolutionDB).filter(CommunitySolutionDB.case_id == case.id).count()
#     confirmation_count = db.query(CommunityCaseConfirmationDB).filter(CommunityCaseConfirmationDB.case_id == case.id).count()
#     follow_up_count = db.query(CommunityCaseFollowUpDB).filter(CommunityCaseFollowUpDB.case_id == case.id).count()
#     contributor = _build_community_contributor_profile(
#         db,
#         name=case.reporter_name,
#         phone_number=case.reporter_phone,
#         preferred_role=case.contributor_role,
#     )
#     normalized_status = _normalize_community_case_status(case.status)
#     normalized_severity = _normalize_community_case_severity(case.severity)
#     tags = _load_json_list(case.tags_json)
#     return {
#         "id": case.id,
#         "room": _normalize_community_room(case.room),
#         "room_label": COMMUNITY_ROOM_LABELS.get(_normalize_community_room(case.room), "General"),
#         "category": case.category,
#         "title": case.title,
#         "description": case.description,
#         "reporter": contributor,
#         "severity": normalized_severity,
#         "severity_label": COMMUNITY_CASE_SEVERITY_LABELS.get(normalized_severity, "Moyenne"),
#         "status": normalized_status,
#         "status_label": COMMUNITY_CASE_STATUS_LABELS.get(normalized_status, "Nouveau"),
#         "location_label": case.location_label,
#         "latitude": case.latitude,
#         "longitude": case.longitude,
#         "crop_or_livestock": case.crop_or_livestock,
#         "tags": tags,
#         "before_photos": _serialize_community_photo_urls(case.before_photo_paths_json),
#         "before_audios": _serialize_community_audio_urls(case.before_audio_paths_json),
#         "after_photos": _serialize_community_photo_urls(case.after_photo_paths_json),
#         "after_audios": _serialize_community_audio_urls(case.after_audio_paths_json),
#         "solution_count": solution_count,
#         "confirmation_count": confirmation_count,
#         "follow_up_count": follow_up_count,
#         "promoted_to_offline": bool(case.promoted_to_offline),
#         "created_at": case.created_at.isoformat() if case.created_at else None,
#         "updated_at": case.updated_at.isoformat() if case.updated_at else None,
#         "resolved_at": case.resolved_at.isoformat() if case.resolved_at else None,
#         "last_follow_up_at": case.last_follow_up_at.isoformat() if case.last_follow_up_at else None,
#     }
# 
# 
# def _serialize_community_field_case_detail(db: Session, case: CommunityFieldCaseDB) -> Dict[str, Any]:
#     summary = _serialize_community_field_case_summary(db, case)
#     solutions = (
#         db.query(CommunitySolutionDB)
#         .filter(CommunitySolutionDB.case_id == case.id)
#         .order_by(CommunitySolutionDB.created_at.desc(), CommunitySolutionDB.id.desc())
#         .all()
#     )
#     confirmations = (
#         db.query(CommunityCaseConfirmationDB)
#         .filter(CommunityCaseConfirmationDB.case_id == case.id)
#         .order_by(CommunityCaseConfirmationDB.created_at.desc(), CommunityCaseConfirmationDB.id.desc())
#         .all()
#     )
#     follow_ups = (
#         db.query(CommunityCaseFollowUpDB)
#         .filter(CommunityCaseFollowUpDB.case_id == case.id)
#         .order_by(CommunityCaseFollowUpDB.created_at.desc(), CommunityCaseFollowUpDB.id.desc())
#         .all()
#     )
#     summary["solutions"] = [_serialize_community_solution(db, item) for item in solutions]
#     summary["confirmations"] = [_serialize_community_confirmation(db, item) for item in confirmations]
#     summary["follow_ups"] = [_serialize_community_follow_up(db, item) for item in follow_ups]
#     return summary
# 
# 
# def _build_community_case_offline_payload(case: CommunityFieldCaseDB, solution: Optional[CommunitySolutionDB]) -> Dict[str, Any]:
#     summary_parts: List[str] = []
#     if case.crop_or_livestock:
#         summary_parts.append(f"Sujet: {case.crop_or_livestock}")
#     if case.location_label:
#         summary_parts.append(f"Zone: {case.location_label}")
#     summary_parts.append(case.description)
#     if solution and solution.action_taken:
#         summary_parts.append(f"Action testee: {solution.action_taken}")
#     if solution and solution.text:
#         summary_parts.append(f"Retour terrain: {solution.text}")
#     if solution and solution.delay_note:
#         summary_parts.append(f"Delai observe: {solution.delay_note}")
#     if solution and solution.cost_note:
#         summary_parts.append(f"Cout: {solution.cost_note}")
# 
#     diagnostic = {
#         "description": case.title,
#         "type": case.category,
#         "causes": _load_json_list(case.tags_json),
#     }
#     actions = []
#     if solution and solution.action_taken:
#         actions.append({"texte": solution.action_taken})
#     if solution and solution.text:
#         actions.append({"texte": solution.text})
# 
#     return {
#         "message": "\n\n".join(part for part in summary_parts if part).strip(),
#         "diagnostic": diagnostic,
#         "actions": actions,
#     }
# 
# 
# def _promote_community_case_to_offline(db: Session, case: CommunityFieldCaseDB) -> None:
#     if not case or _normalize_community_case_status(case.status) != "resolved":
#         return
#     best_solution = (
#         db.query(CommunitySolutionDB)
#         .filter(CommunitySolutionDB.case_id == case.id)
#         .order_by(CommunitySolutionDB.created_at.desc(), CommunitySolutionDB.id.desc())
#         .first()
#     )
#     payload = _build_community_case_offline_payload(case, best_solution)
#     _persist_offline_knowledge_entry(
#         db=db,
#         user_id=None,
#         source_kind="community_case",
#         category=case.category,
#         question_text=case.description,
#         response_payload=payload,
#     )
#     case.promoted_to_offline = True
#     db.commit()
# 
# 
# def _serialize_community_message(message: ChatMessageDB) -> Dict[str, Any]:
#     normalized_room = _normalize_community_room(message.room)
#     sender_role = message.sender_role or ("assistant" if message.is_bot else "member")
#     return {
#         "id": message.id,
#         "sender": message.sender,
#         "text": message.text,
#         "is_bot": message.is_bot,
#         "created_at": message.created_at.isoformat() if message.created_at else None,
#         "room": normalized_room,
#         "room_label": COMMUNITY_ROOM_LABELS.get(normalized_room, "Général"),
#         "sender_role": sender_role,
#         "is_expert": sender_role == "expert",
#         "report_count": message.report_count or 0,
#         "is_pinned": bool(message.is_pinned),
#         "pinned_at": message.pinned_at.isoformat() if message.pinned_at else None,
#         "audio_url": _build_upload_url(message.audio_url) if message.audio_url else None,
#     }
# 
# 
# def _get_pinned_community_message(db: Session, room: str) -> Optional[ChatMessageDB]:
#     return (
#         db.query(ChatMessageDB)
#         .filter(ChatMessageDB.room == room)
#         .filter(ChatMessageDB.is_hidden == False)
#         .filter(ChatMessageDB.is_pinned == True)
#         .order_by(ChatMessageDB.pinned_at.desc(), ChatMessageDB.id.desc())
#         .first()
#     )
# 
# 
# def _set_pinned_community_message(
#     db: Session,
#     message: ChatMessageDB,
#     pinned: bool,
# ) -> ChatMessageDB:
#     db.query(ChatMessageDB).filter(ChatMessageDB.room == message.room).update(
#         {ChatMessageDB.is_pinned: False, ChatMessageDB.pinned_at: None},
#         synchronize_session=False,
#     )
# 
#     if pinned:
#         message.is_pinned = True
#         message.pinned_at = datetime.utcnow()
#     else:
#         message.is_pinned = False
#         message.pinned_at = None
# 
#     db.commit()
#     db.refresh(message)
#     return message
# 
# 
def _contains_blocked_community_text(text: str) -> bool:
    lowered = text.lower()
    return any(term in lowered for term in COMMUNITY_BLOCKED_TERMS)


def _build_contextual_bot_reply(db: Session, room: str, text: str) -> str:
    domain = COMMUNITY_ROOM_TO_DOMAIN.get(room, "agriculture")
    items = retrieve_knowledge(db, domain, text, limit=2)
    room_label = COMMUNITY_ROOM_LABELS.get(room, "Général")

    if items:
        best = items[0]
        title = best.get("title") or room_label
        answer = (best.get("answer") or "").replace("\n", " ").strip()
        if len(answer) > 220:
            answer = answer[:220].rsplit(" ", 1)[0] + "..."

        if room == "marche":
            return (
                f"Repère marché: {title}. {answer} "
                "Compare aussi les prix locaux avant de vendre ou d'acheter."
            )

        if room == "securite":
            return (
                f"Point sécurité: {title}. {answer} "
                "Ne partage jamais ton code ou ton mot de passe par message."
            )

        return f"Conseil {room_label.lower()}: {title}. {answer}"

    lowered = text.lower()
    if any(keyword in lowered for keyword in ["urgent", "grave", "mort", "attaque"]):
        return (
            "La situation semble sérieuse. Décris le lieu, depuis quand le problème a commencé, "
            "et contacte aussi un expert ou le module SOS si nécessaire."
        )

    if room == "marche":
        return (
            "Précise le produit, la quantité, le prix observé et la commune. "
            "La communauté pourra comparer plus facilement."
        )

    return BOT_REPLIES[hash(f"{room}:{text[:40]}") % len(BOT_REPLIES)]

# @app.get("/api/community/messages")
# async def get_community_messages(
#     room: Optional[str] = None,
#     limit: int = Query(default=50, le=200),
#     db: Session = Depends(get_db)
# ):
#     """Récupérer les derniers messages du chat communautaire."""
#     query = db.query(ChatMessageDB).filter(ChatMessageDB.is_hidden == False)
#     if room is not None:
#         query = query.filter(ChatMessageDB.room == _normalize_community_room(room))
# 
#     messages = query.order_by(
#         ChatMessageDB.created_at.desc(), ChatMessageDB.id.desc()
#     ).limit(limit).all()
#     messages.reverse()
#     return [_serialize_community_message(m) for m in messages]
# 
# 
# @app.get("/api/community/pinned")
# async def get_pinned_community_message(
#     room: Optional[str] = None,
#     db: Session = Depends(get_db)
# ):
#     """Récupérer l'annonce experte épinglée pour un salon."""
#     normalized_room = _normalize_community_room(room)
#     message = _get_pinned_community_message(db, normalized_room)
#     return {
#         "room": normalized_room,
#         "message": _serialize_community_message(message) if message else None,
#     }
# 
# @app.post("/api/community/messages")
# async def post_community_message(
#     body: Dict[str, Any],
#     db: Session = Depends(get_db)
# ):
#     """Poster un message dans le chat communautaire et générer une réponse bot."""
#     sender = (body.get("sender") or "Anonyme")[:80]
#     text = (body.get("text") or "").strip()
#     room = _normalize_community_room(body.get("room"))
#     sender_role = (body.get("sender_role") or "member").strip().lower()
#     if sender_role not in {"member", "expert", "assistant"}:
#         sender_role = "member"
#     if not text:
#         raise HTTPException(status_code=422, detail="Le message ne peut pas être vide")
#     if len(text) > 1000:
#         raise HTTPException(status_code=422, detail="Message trop long (max 1000 caractères)")
#     if _contains_blocked_community_text(text):
#         raise HTTPException(status_code=422, detail="Message bloqué par la modération légère de la communauté")
# 
#     duplicate_since = datetime.utcnow() - timedelta(seconds=COMMUNITY_DUPLICATE_WINDOW_SECONDS)
#     duplicate = (
#         db.query(ChatMessageDB)
#         .filter(ChatMessageDB.sender == sender)
#         .filter(ChatMessageDB.text == text)
#         .filter(ChatMessageDB.room == room)
#         .filter(ChatMessageDB.created_at >= duplicate_since)
#         .filter(ChatMessageDB.is_hidden == False)
#         .first()
#     )
#     if duplicate:
#         raise HTTPException(status_code=409, detail="Message déjà publié récemment dans ce salon")
# 
#     user_msg = ChatMessageDB(
#         sender=sender,
#         text=text,
#         is_bot=False,
#         room=room,
#         sender_role=sender_role,
#     )
#     db.add(user_msg)
#     db.commit()
#     db.refresh(user_msg)
# 
#     bot_text = _build_contextual_bot_reply(db, room, text)
#     bot_msg = ChatMessageDB(
#         sender="Assistant SONGRA",
#         text=bot_text,
#         is_bot=True,
#         room=room,
#         sender_role="assistant",
#     )
#     db.add(bot_msg)
#     db.commit()
#     db.refresh(bot_msg)
# 
#     return {
#         "user": _serialize_community_message(user_msg),
#         "bot": _serialize_community_message(bot_msg),
#     }
# 
# 
# @app.post("/api/community/messages/audio")
# async def post_community_audio_message(
#     sender: str = Form("Anonyme"),
#     room: str = Form("general"),
#     sender_role: str = Form("member"),
#     audio_file: UploadFile = File(...),
#     db: Session = Depends(get_db)
# ):
#     """Poster un message vocal dans le chat communautaire."""
#     os.makedirs("uploads/community_audio", exist_ok=True)
#     ext = os.path.splitext(audio_file.filename or "audio.webm")[1].lower() or ".webm"
#     filename = f"comm_{int(time.time())}_{hashlib.md5((audio_file.filename or '').encode()).hexdigest()[:8]}{ext}"
#     filepath = os.path.join("uploads/community_audio", filename)
#     
#     content = await audio_file.read()
#     with open(filepath, "wb") as buffer:
#         buffer.write(content)
#         
#     audio_relative_path = f"uploads/community_audio/{filename}"
#     
#     user_msg = ChatMessageDB(
#         sender=sender,
#         text="🔊 Message vocal",
#         is_bot=False,
#         room=room,
#         sender_role=sender_role,
#         audio_url=audio_relative_path
#     )
#     db.add(user_msg)
#     db.commit()
#     db.refresh(user_msg)
#     
#     return _serialize_community_message(user_msg)
# 
# 
# @app.post("/api/community/messages/{message_id}/report")
# async def report_community_message(
#     message_id: int,
#     db: Session = Depends(get_db)
# ):
#     """Signaler un message communautaire. Auto-masquage après plusieurs signalements."""
#     message = db.query(ChatMessageDB).filter(ChatMessageDB.id == message_id).first()
#     if not message:
#         raise HTTPException(status_code=404, detail="Message introuvable")
# 
#     if message.is_hidden:
#         return {
#             "success": True,
#             "message_id": message_id,
#             "report_count": message.report_count or 0,
#             "status": "hidden",
#         }
# 
#     message.report_count = (message.report_count or 0) + 1
#     if message.report_count >= COMMUNITY_REPORT_HIDE_THRESHOLD:
#         message.is_hidden = True
#     db.commit()
#     db.refresh(message)
# 
#     return {
#         "success": True,
#         "message_id": message.id,
#         "report_count": message.report_count,
#         "status": "hidden" if message.is_hidden else "reported",
#     }
# 
# 
# @app.post("/api/expert/community/messages")
# async def post_expert_community_message(
#     body: Dict[str, Any],
#     current_expert: Expert = Depends(get_current_expert),
#     db: Session = Depends(get_db),
# ):
#     """Permet à un expert connecté de publier dans un salon communautaire."""
#     text = (body.get("text") or "").strip()
#     room = _normalize_community_room(body.get("room"))
#     pin_message = bool(body.get("pin"))
#     sender = (current_expert.full_name or current_expert.email or "Expert SONGRA").strip()[:80]
# 
#     if not text:
#         raise HTTPException(status_code=422, detail="Le message ne peut pas être vide")
#     if len(text) > 1000:
#         raise HTTPException(status_code=422, detail="Message trop long (max 1000 caractères)")
#     if _contains_blocked_community_text(text):
#         raise HTTPException(status_code=422, detail="Message bloqué par la modération légère de la communauté")
# 
#     duplicate_since = datetime.utcnow() - timedelta(seconds=COMMUNITY_DUPLICATE_WINDOW_SECONDS)
#     duplicate = (
#         db.query(ChatMessageDB)
#         .filter(ChatMessageDB.sender == sender)
#         .filter(ChatMessageDB.text == text)
#         .filter(ChatMessageDB.room == room)
#         .filter(ChatMessageDB.created_at >= duplicate_since)
#         .filter(ChatMessageDB.is_hidden == False)
#         .first()
#     )
#     if duplicate:
#         raise HTTPException(status_code=409, detail="Message déjà publié récemment dans ce salon")
# 
#     expert_message = ChatMessageDB(
#         sender=sender,
#         text=text,
#         is_bot=False,
#         room=room,
#         sender_role="expert",
#     )
#     db.add(expert_message)
#     db.commit()
#     db.refresh(expert_message)
# 
#     if pin_message:
#         expert_message = _set_pinned_community_message(db, expert_message, True)
# 
#     return {
#         "success": True,
#         "message": _serialize_community_message(expert_message),
#     }
# 
# 
# @app.patch("/api/expert/community/messages/{message_id}/pin")
# async def pin_expert_community_message(
#     message_id: int,
#     body: Dict[str, Any],
#     current_expert: Expert = Depends(get_current_expert),
#     db: Session = Depends(get_db),
# ):
#     """Épingler ou retirer une annonce experte dans un salon communautaire."""
#     del current_expert
#     pinned = bool(body.get("pinned", True))
#     message = db.query(ChatMessageDB).filter(ChatMessageDB.id == message_id).first()
#     if not message:
#         raise HTTPException(status_code=404, detail="Message introuvable")
#     if message.is_hidden:
#         raise HTTPException(status_code=422, detail="Impossible d'épingler un message masqué")
#     if (message.sender_role or "") != "expert":
#         raise HTTPException(status_code=422, detail="Seuls les messages experts peuvent être épinglés")
# 
#     updated_message = _set_pinned_community_message(db, message, pinned)
#     return {
#         "success": True,
#         "message": _serialize_community_message(updated_message),
#         "pinned": bool(updated_message.is_pinned),
#     }
# 
# 
# @app.patch("/api/expert/community/messages/{message_id}")
# async def update_expert_community_message(
#     message_id: int,
#     body: Dict[str, Any],
#     current_expert: Expert = Depends(get_current_expert),
#     db: Session = Depends(get_db),
# ):
#     """Modifier directement un message expert communautaire, notamment une annonce épinglée."""
#     del current_expert
#     text = (body.get("text") or "").strip()
#     if not text:
#         raise HTTPException(status_code=422, detail="Le message ne peut pas être vide")
#     if len(text) > 1000:
#         raise HTTPException(status_code=422, detail="Message trop long (max 1000 caractères)")
#     if _contains_blocked_community_text(text):
#         raise HTTPException(status_code=422, detail="Message bloqué par la modération légère de la communauté")
# 
#     message = db.query(ChatMessageDB).filter(ChatMessageDB.id == message_id).first()
#     if not message:
#         raise HTTPException(status_code=404, detail="Message introuvable")
#     if message.is_hidden:
#         raise HTTPException(status_code=422, detail="Impossible de modifier un message masqué")
#     if (message.sender_role or "") != "expert":
#         raise HTTPException(status_code=422, detail="Seuls les messages experts peuvent être modifiés")
# 
#     message.text = text
#     db.commit()
#     db.refresh(message)
# 
#     return {
#         "success": True,
#         "message": _serialize_community_message(message),
#     }
# 
# 
# @app.get("/api/community/field-cases")
# async def get_community_field_cases(
#     room: Optional[str] = None,
#     status: Optional[str] = None,
#     limit: int = Query(default=60, le=200),
#     db: Session = Depends(get_db),
# ):
#     query = db.query(CommunityFieldCaseDB)
#     if room is not None:
#         query = query.filter(CommunityFieldCaseDB.room == _normalize_community_room(room))
#     if status is not None:
#         query = query.filter(CommunityFieldCaseDB.status == _normalize_community_case_status(status))
# 
#     cases = (
#         query.order_by(CommunityFieldCaseDB.updated_at.desc(), CommunityFieldCaseDB.id.desc())
#         .limit(limit)
#         .all()
#     )
#     return [_serialize_community_field_case_summary(db, item) for item in cases]
# 
# 
# @app.get("/api/community/field-cases/{case_id}")
# async def get_community_field_case_detail(case_id: int, db: Session = Depends(get_db)):
#     case = db.query(CommunityFieldCaseDB).filter(CommunityFieldCaseDB.id == case_id).first()
#     if not case:
#         raise HTTPException(status_code=404, detail="Cas terrain introuvable")
#     return _serialize_community_field_case_detail(db, case)
# 
# 
# @app.post("/api/community/field-cases")
# async def create_community_field_case(body: Dict[str, Any], db: Session = Depends(get_db)):
#     reporter_name = (body.get("reporter_name") or body.get("sender") or "Acteur terrain").strip()[:80]
#     reporter_phone = (body.get("reporter_phone") or body.get("phone_number") or "").strip()[:40] or None
#     room = _normalize_community_room(body.get("room"))
#     category = _normalize_community_category(body.get("category"), room)
#     description = (body.get("description") or "").strip()
#     location_label = (body.get("location_label") or "").strip()[:140] or None
#     crop_or_livestock = (body.get("crop_or_livestock") or "").strip()[:140] or None
#     severity = _normalize_community_case_severity(body.get("severity"))
#     tags = body.get("tags") or []
#     if not isinstance(tags, list):
#         tags = []
# 
#     photo_payloads = _collect_photo_payloads(body.get("photo_base64"), body.get("photo_base64_list"))
#     photo_data_list = [_decode_photo_payload(payload) for payload in photo_payloads[:4] if payload]
#     audio_payloads = _collect_audio_payloads(body.get("audio_base64"), body.get("audio_base64_list"))
#     owner_id = int(datetime.utcnow().timestamp())
#     stored_paths = _store_photo_payloads(owner_id, photo_data_list, prefix="community_case") if photo_data_list else []
#     stored_audio_paths = _store_audio_payloads(owner_id, audio_payloads, prefix="community_case_audio") if audio_payloads else []
# 
#     if not description and not stored_paths and not stored_audio_paths:
#         raise HTTPException(status_code=422, detail="Ajoutez un texte, une photo ou un vocal pour ce signalement")
# 
#     if not description:
#         description = "Signalement terrain envoye avec media sans texte."
# 
#     title = _build_community_case_title(body.get("title") or "", description, room)
# 
#     case = CommunityFieldCaseDB(
#         room=room,
#         category=category,
#         title=title,
#         description=description,
#         reporter_name=reporter_name,
#         reporter_phone=reporter_phone,
#         contributor_role="observer",
#         severity=severity,
#         status="new",
#         location_label=location_label,
#         latitude=body.get("latitude"),
#         longitude=body.get("longitude"),
#         crop_or_livestock=crop_or_livestock,
#         tags_json=json.dumps(_extract_offline_keywords(tags, title, description, crop_or_livestock), ensure_ascii=False),
#         before_photo_paths_json=json.dumps(stored_paths, ensure_ascii=False) if stored_paths else None,
#         before_audio_paths_json=json.dumps(stored_audio_paths, ensure_ascii=False) if stored_audio_paths else None,
#     )
#     db.add(case)
#     db.commit()
#     db.refresh(case)
# 
#     return {
#         "success": True,
#         "case": _serialize_community_field_case_detail(db, case),
#     }
# 
# 
# @app.post("/api/community/field-cases/{case_id}/solutions")
# async def add_solution_to_community_field_case(case_id: int, body: Dict[str, Any], db: Session = Depends(get_db)):
#     case = db.query(CommunityFieldCaseDB).filter(CommunityFieldCaseDB.id == case_id).first()
#     if not case:
#         raise HTTPException(status_code=404, detail="Cas terrain introuvable")
# 
#     author_name = (body.get("author_name") or body.get("sender") or "Acteur terrain").strip()[:80]
#     author_phone = (body.get("author_phone") or body.get("phone_number") or "").strip()[:40] or None
#     text = (body.get("text") or "").strip()
#     action_taken = (body.get("action_taken") or "").strip()
#     if not text:
#         raise HTTPException(status_code=422, detail="La solution proposee est obligatoire")
# 
#     photo_payloads = _collect_photo_payloads(body.get("photo_base64"), body.get("photo_base64_list"))
#     photo_data_list = [_decode_photo_payload(payload) for payload in photo_payloads[:4] if payload]
#     owner_id = case_id * 1000 + int(datetime.utcnow().timestamp())
#     stored_paths = _store_photo_payloads(owner_id, photo_data_list, prefix="community_solution") if photo_data_list else []
# 
#     solution = CommunitySolutionDB(
#         case_id=case_id,
#         author_name=author_name,
#         author_phone=author_phone,
#         contributor_role="solver",
#         text=text,
#         action_taken=action_taken or None,
#         cost_note=(body.get("cost_note") or "").strip()[:120] or None,
#         delay_note=(body.get("delay_note") or "").strip()[:120] or None,
#         result_status=(body.get("result_status") or "proposed").strip()[:80] or "proposed",
#         photo_paths_json=json.dumps(stored_paths, ensure_ascii=False) if stored_paths else None,
#         is_expert=bool(body.get("is_expert", False)),
#     )
#     case.updated_at = datetime.utcnow()
#     db.add(solution)
#     db.commit()
#     db.refresh(solution)
# 
#     return {
#         "success": True,
#         "solution": _serialize_community_solution(db, solution),
#         "case": _serialize_community_field_case_summary(db, case),
#     }
# 
# 
# @app.post("/api/community/solutions/{solution_id}/feedback")
# async def add_feedback_to_community_solution(solution_id: int, body: Dict[str, Any], db: Session = Depends(get_db)):
#     solution = db.query(CommunitySolutionDB).filter(CommunitySolutionDB.id == solution_id).first()
#     if not solution:
#         raise HTTPException(status_code=404, detail="Solution introuvable")
# 
#     voter_name = (body.get("voter_name") or body.get("sender") or "Acteur terrain").strip()[:80]
#     voter_phone = (body.get("voter_phone") or body.get("phone_number") or "").strip()[:40] or None
#     feedback_type = _normalize_community_feedback_type(body.get("feedback_type"))
# 
#     existing = (
#         db.query(CommunitySolutionFeedbackDB)
#         .filter(CommunitySolutionFeedbackDB.solution_id == solution_id)
#         .filter(
#             _community_person_filter(
#                 CommunitySolutionFeedbackDB.voter_phone,
#                 CommunitySolutionFeedbackDB.voter_name,
#                 voter_phone,
#                 voter_name,
#             )
#         )
#         .first()
#     )
#     if existing is None:
#         existing = CommunitySolutionFeedbackDB(
#             solution_id=solution_id,
#             voter_name=voter_name,
#             voter_phone=voter_phone,
#             feedback_type=feedback_type,
#         )
#         db.add(existing)
#     else:
#         existing.feedback_type = feedback_type
#         existing.created_at = datetime.utcnow()
# 
#     db.commit()
#     db.refresh(solution)
#     return {
#         "success": True,
#         "solution": _serialize_community_solution(db, solution),
#     }
# 
# 
# @app.post("/api/community/field-cases/{case_id}/confirm")
# async def confirm_community_field_case(case_id: int, body: Dict[str, Any], db: Session = Depends(get_db)):
#     case = db.query(CommunityFieldCaseDB).filter(CommunityFieldCaseDB.id == case_id).first()
#     if not case:
#         raise HTTPException(status_code=404, detail="Cas terrain introuvable")
# 
#     confirmer_name = (body.get("confirmer_name") or body.get("sender") or "Acteur terrain").strip()[:80]
#     confirmer_phone = (body.get("confirmer_phone") or body.get("phone_number") or "").strip()[:40] or None
# 
#     existing = (
#         db.query(CommunityCaseConfirmationDB)
#         .filter(CommunityCaseConfirmationDB.case_id == case_id)
#         .filter(
#             _community_person_filter(
#                 CommunityCaseConfirmationDB.confirmer_phone,
#                 CommunityCaseConfirmationDB.confirmer_name,
#                 confirmer_phone,
#                 confirmer_name,
#             )
#         )
#         .first()
#     )
#     if existing:
#         raise HTTPException(status_code=409, detail="Confirmation deja enregistree pour ce cas")
# 
#     confirmation = CommunityCaseConfirmationDB(
#         case_id=case_id,
#         confirmer_name=confirmer_name,
#         confirmer_phone=confirmer_phone,
#         contributor_role="observer",
#         note=(body.get("note") or "").strip() or None,
#         location_label=(body.get("location_label") or "").strip()[:140] or None,
#     )
#     case.updated_at = datetime.utcnow()
#     db.add(confirmation)
#     db.commit()
# 
#     return {
#         "success": True,
#         "case": _serialize_community_field_case_detail(db, case),
#     }
# 
# 
# @app.post("/api/community/field-cases/{case_id}/follow-ups")
# async def add_follow_up_to_community_field_case(case_id: int, body: Dict[str, Any], db: Session = Depends(get_db)):
#     case = db.query(CommunityFieldCaseDB).filter(CommunityFieldCaseDB.id == case_id).first()
#     if not case:
#         raise HTTPException(status_code=404, detail="Cas terrain introuvable")
# 
#     author_name = (body.get("author_name") or body.get("sender") or "Acteur terrain").strip()[:80]
#     author_phone = (body.get("author_phone") or body.get("phone_number") or "").strip()[:40] or None
#     note = (body.get("note") or "").strip()
# 
#     photo_payloads = _collect_photo_payloads(body.get("photo_base64"), body.get("photo_base64_list"))
#     photo_data_list = [_decode_photo_payload(payload) for payload in photo_payloads[:4] if payload]
#     audio_payloads = _collect_audio_payloads(body.get("audio_base64"), body.get("audio_base64_list"))
#     owner_id = case_id * 1000 + int(datetime.utcnow().timestamp())
#     stored_paths = _store_photo_payloads(owner_id, photo_data_list, prefix="community_followup") if photo_data_list else []
#     stored_audio_paths = _store_audio_payloads(owner_id, audio_payloads, prefix="community_followup_audio") if audio_payloads else []
# 
#     if not note and not stored_paths and not stored_audio_paths:
#         raise HTTPException(status_code=422, detail="Ajoutez un texte, une photo ou un vocal pour ce suivi")
#     if not note:
#         note = "Retour terrain envoye avec media sans texte."
# 
#     normalized_status = _normalize_community_case_status(body.get("status_after") or case.status)
#     follow_up = CommunityCaseFollowUpDB(
#         case_id=case_id,
#         author_name=author_name,
#         author_phone=author_phone,
#         contributor_role="observer",
#         note=note,
#         status_after=normalized_status,
#         outcome_label=(body.get("outcome_label") or "").strip()[:140] or None,
#         photo_paths_json=json.dumps(stored_paths, ensure_ascii=False) if stored_paths else None,
#         audio_paths_json=json.dumps(stored_audio_paths, ensure_ascii=False) if stored_audio_paths else None,
#     )
#     db.add(follow_up)
# 
#     if stored_paths:
#         existing_after = _load_json_list(case.after_photo_paths_json)
#         case.after_photo_paths_json = json.dumps(existing_after + stored_paths, ensure_ascii=False)
#     if stored_audio_paths:
#         existing_after_audio = _load_json_list(case.after_audio_paths_json)
#         case.after_audio_paths_json = json.dumps(existing_after_audio + stored_audio_paths, ensure_ascii=False)
#     case.status = normalized_status
#     case.last_follow_up_at = datetime.utcnow()
#     case.updated_at = datetime.utcnow()
#     if normalized_status == "resolved":
#         case.resolved_at = datetime.utcnow()
# 
#     db.commit()
#     db.refresh(case)
#     if normalized_status == "resolved":
#         _promote_community_case_to_offline(db, case)
# 
#     return {
#         "success": True,
#         "case": _serialize_community_field_case_detail(db, case),
#     }
# 
# 
# @app.patch("/api/community/follow-ups/{follow_up_id}")
# async def update_community_case_follow_up(follow_up_id: int, body: Dict[str, Any], db: Session = Depends(get_db)):
#     follow_up = db.query(CommunityCaseFollowUpDB).filter(CommunityCaseFollowUpDB.id == follow_up_id).first()
#     if not follow_up:
#         raise HTTPException(status_code=404, detail="Suivi terrain introuvable")
# 
#     case = db.query(CommunityFieldCaseDB).filter(CommunityFieldCaseDB.id == follow_up.case_id).first()
#     if not case:
#         raise HTTPException(status_code=404, detail="Cas terrain introuvable")
# 
#     author_name = (body.get("author_name") or follow_up.author_name or "Acteur terrain").strip()[:80]
#     author_phone = (body.get("author_phone") or body.get("phone_number") or follow_up.author_phone or "").strip()[:40] or None
#     note = (body.get("note") or follow_up.note or "").strip()
# 
#     normalized_status = _normalize_community_case_status(body.get("status_after") or follow_up.status_after or case.status)
#     photo_payloads = _collect_photo_payloads(body.get("photo_base64"), body.get("photo_base64_list"))
#     photo_data_list = [_decode_photo_payload(payload) for payload in photo_payloads[:4] if payload]
#     audio_payloads = _collect_audio_payloads(body.get("audio_base64"), body.get("audio_base64_list"))
#     owner_id = case.id * 1000 + follow_up_id
#     stored_paths = _store_photo_payloads(owner_id, photo_data_list, prefix="community_followup") if photo_data_list else []
#     stored_audio_paths = _store_audio_payloads(owner_id, audio_payloads, prefix="community_followup_audio") if audio_payloads else []
# 
#     if not note and not stored_paths and not stored_audio_paths:
#         raise HTTPException(status_code=422, detail="Ajoutez un texte, une photo ou un vocal pour ce suivi")
#     if not note:
#         note = "Retour terrain envoye avec media sans texte."
# 
#     existing_follow_up_paths = _load_json_list(follow_up.photo_paths_json)
#     merged_follow_up_paths = existing_follow_up_paths + stored_paths if stored_paths else existing_follow_up_paths
#     existing_follow_up_audio_paths = _load_json_list(follow_up.audio_paths_json)
#     merged_follow_up_audio_paths = existing_follow_up_audio_paths + stored_audio_paths if stored_audio_paths else existing_follow_up_audio_paths
# 
#     follow_up.author_name = author_name
#     follow_up.author_phone = author_phone
#     follow_up.note = note
#     follow_up.status_after = normalized_status
#     follow_up.outcome_label = (body.get("outcome_label") or "").strip()[:140] or None
#     follow_up.photo_paths_json = json.dumps(merged_follow_up_paths, ensure_ascii=False) if merged_follow_up_paths else None
#     follow_up.audio_paths_json = json.dumps(merged_follow_up_audio_paths, ensure_ascii=False) if merged_follow_up_audio_paths else None
# 
#     if stored_paths:
#         case_after_paths = _load_json_list(case.after_photo_paths_json)
#         case.after_photo_paths_json = json.dumps(case_after_paths + stored_paths, ensure_ascii=False)
#     if stored_audio_paths:
#         case_after_audio_paths = _load_json_list(case.after_audio_paths_json)
#         case.after_audio_paths_json = json.dumps(case_after_audio_paths + stored_audio_paths, ensure_ascii=False)
# 
#     case.status = normalized_status
#     case.last_follow_up_at = datetime.utcnow()
#     case.updated_at = datetime.utcnow()
#     if normalized_status == "resolved":
#         case.resolved_at = case.resolved_at or datetime.utcnow()
# 
#     db.commit()
#     db.refresh(case)
#     if normalized_status == "resolved":
#         _promote_community_case_to_offline(db, case)
# 
#     return {
#         "success": True,
#         "case": _serialize_community_field_case_detail(db, case),
#     }
# 
# 
# @app.patch("/api/community/field-cases/{case_id}/status")
# async def update_community_field_case_status(case_id: int, body: Dict[str, Any], db: Session = Depends(get_db)):
#     case = db.query(CommunityFieldCaseDB).filter(CommunityFieldCaseDB.id == case_id).first()
#     if not case:
#         raise HTTPException(status_code=404, detail="Cas terrain introuvable")
# 
#     normalized_status = _normalize_community_case_status(body.get("status"))
#     case.status = normalized_status
#     case.updated_at = datetime.utcnow()
#     if normalized_status == "resolved":
#         case.resolved_at = datetime.utcnow()
#     db.commit()
#     db.refresh(case)
#     if normalized_status == "resolved":
#         _promote_community_case_to_offline(db, case)
# 
#     return {
#         "success": True,
#         "case": _serialize_community_field_case_detail(db, case),
#     }
# 
# 
# @app.post("/api/community/field-cases/{case_id}/promote-offline")
# async def promote_community_field_case_offline(case_id: int, db: Session = Depends(get_db)):
#     case = db.query(CommunityFieldCaseDB).filter(CommunityFieldCaseDB.id == case_id).first()
#     if not case:
#         raise HTTPException(status_code=404, detail="Cas terrain introuvable")
#     if _normalize_community_case_status(case.status) != "resolved":
#         raise HTTPException(status_code=422, detail="Seuls les cas resolus peuvent alimenter la base hors ligne")
# 
#     _promote_community_case_to_offline(db, case)
#     db.refresh(case)
#     return {
#         "success": True,
#         "case": _serialize_community_field_case_detail(db, case),
#     }
# 
# 
# @app.get("/api/community/trends")
# async def get_community_trends(
#     room: Optional[str] = None,
#     limit: int = Query(default=5, le=20),
#     db: Session = Depends(get_db),
# ):
#     query = db.query(CommunityFieldCaseDB)
#     if room is not None:
#         query = query.filter(CommunityFieldCaseDB.room == _normalize_community_room(room))
#     cases = query.order_by(CommunityFieldCaseDB.updated_at.desc()).limit(300).all()
# 
#     status_counts: Dict[str, int] = {}
#     category_counts: Dict[str, int] = {}
#     location_counts: Dict[str, int] = {}
#     signal_counts: Dict[str, int] = {}
#     active_contributors = set()
# 
#     for case in cases:
#         status_key = _normalize_community_case_status(case.status)
#         status_counts[status_key] = status_counts.get(status_key, 0) + 1
#         category_counts[case.category or "agriculture"] = category_counts.get(case.category or "agriculture", 0) + 1
#         if case.location_label:
#             location_counts[case.location_label] = location_counts.get(case.location_label, 0) + 1
#         trend_key = (case.crop_or_livestock or case.title or case.category or "terrain").strip()
#         signal_counts[trend_key] = signal_counts.get(trend_key, 0) + 1
#         active_contributors.add(case.reporter_phone or case.reporter_name)
# 
#     return {
#         "overview": {
#             "total_cases": len(cases),
#             "active_contributors": len([item for item in active_contributors if item]),
#             "resolved_cases": status_counts.get("resolved", 0),
#             "watch_cases": status_counts.get("watch", 0),
#         },
#         "statuses": [
#             {
#                 "key": key,
#                 "label": COMMUNITY_CASE_STATUS_LABELS.get(key, key),
#                 "count": count,
#             }
#             for key, count in sorted(status_counts.items(), key=lambda item: item[1], reverse=True)
#         ],
#         "categories": [
#             {"key": key, "count": count}
#             for key, count in sorted(category_counts.items(), key=lambda item: item[1], reverse=True)[:limit]
#         ],
#         "hotspots": [
#             {"label": key, "count": count}
#             for key, count in sorted(location_counts.items(), key=lambda item: item[1], reverse=True)[:limit]
#         ],
#         "signals": [
#             {"label": key, "count": count}
#             for key, count in sorted(signal_counts.items(), key=lambda item: item[1], reverse=True)[:limit]
#         ],
#     }
# 
class OfflineSyncData(BaseModel):
    tickets: List[Dict[str, Any]] = []
    messages: List[Dict[str, Any]] = []
    photos: List[Dict[str, Any]] = []

@app.post("/api/sync/offline")
async def sync_offline_data(sync_data: OfflineSyncData, db: Session = Depends(get_db)):
    """
    🔑 ENDPOINT SYNCHRONISATION OFFLINE
    Synchroniser les données sauvegardées localement quand connexion revient
    """
    synced_tickets = []
    synced_messages = []
    synced_photos = []
    errors = []
    
    try:
        # Synchroniser les tickets
        for ticket_data in sync_data.tickets:
            try:
                # Vérifier si l'utilisateur existe
                phone = ticket_data.get("phoneNumber", "")
                user = db.query(User).filter(User.phone_number == phone).first()
                
                if not user:
                    user = User(phone_number=phone, name=ticket_data.get("userName"))
                    db.add(user)
                    db.commit()
                    db.refresh(user)
                
                # Créer le ticket
                ticket = Ticket(
                    user_id=user.id,
                    category=ticket_data.get("category"),
                    urgency=ticket_data.get("urgency", "medium"),
                    status="open"
                )
                
                db.add(ticket)
                db.commit()
                db.refresh(ticket)
                
                synced_tickets.append({
                    "localId": ticket_data.get("localId"),
                    "serverId": ticket.id
                })
                
            except Exception as e:
                errors.append({"type": "ticket", "error": str(e), "data": ticket_data.get("localId")})
        
        # Synchroniser les messages
        for msg_data in sync_data.messages:
            try:
                message = Message(
                    ticket_id=msg_data.get("ticketId"),
                    sender_type="user",
                    sender_id=msg_data.get("userId"),
                    content=msg_data.get("content"),
                    channel="app"
                )
                
                db.add(message)
                db.commit()
                
                synced_messages.append({
                    "localId": msg_data.get("localId"),
                    "serverId": message.id
                })
                
            except Exception as e:
                errors.append({"type": "message", "error": str(e)})
        
        return {
            "success": True,
            "synced": {
                "tickets": len(synced_tickets),
                "messages": len(synced_messages),
                "photos": len(synced_photos)
            },
            "mapping": {
                "tickets": synced_tickets,
                "messages": synced_messages
            },
            "errors": errors
        }
    
    except Exception as e:
        print(f"❌ Erreur sync offline: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur synchronisation: {str(e)}")

@app.get("/api/knowledge/offline-cache")
async def get_knowledge_for_offline_cache(
    domain: Optional[str] = None,
    language: str = "fr",
    limit: int = Query(default=100, le=500),
    db: Session = Depends(get_db)
):
    return _build_offline_cache_payload(
        db,
        domain=domain,
        language=language,
        limit=limit,
    )

# ==========================================
# ENDPOINTS FREEMIUM & DIALOGUE
# ==========================================

class SendMessageRequest(BaseModel):
    ticket_id: int
    sender_type: str  # 'user' ou 'expert'
    content: str

@app.get("/api/user-status")
async def get_user_status(phone: str, db: Session = Depends(get_db)):
    """Retourne le statut premium et les limites de messages d'un utilisateur"""
    user = db.query(User).filter(User.phone_number == phone).first()
    
    if not user:
        # Créer l'utilisateur s'il n'existe pas
        user = User(
            phone_number=phone,
            is_premium=False,
            messages_used=0,
            messages_limit=1
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    
    # Vérifier si l'abonnement premium est expiré
    if user.is_premium and user.premium_expires_at:
        if user.premium_expires_at < datetime.utcnow():
            user.is_premium = False
            user.messages_limit = 1
            db.commit()
    
    return {
        "is_premium": user.is_premium,
        "messages_used": user.messages_used,
        "messages_limit": user.messages_limit if user.is_premium else 1,
        "premium_expires_at": user.premium_expires_at.isoformat() if user.premium_expires_at else None
    }

@app.post("/api/send-message")
async def send_message(request: SendMessageRequest, db: Session = Depends(get_db)):
    """Envoie un message dans un ticket (user ou expert)"""
    
    # Récupérer le ticket
    ticket = db.query(Ticket).filter(Ticket.id == request.ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket non trouvé")
    
    # Récupérer l'utilisateur
    user = db.query(User).filter(User.id == ticket.user_id).first()
    
    if request.sender_type == 'user':
        # Vérifier les limites de messages
        user_messages = db.query(Message).filter(
            Message.ticket_id == request.ticket_id,
            Message.sender_type == 'user'
        ).count()
        
        limit = user.messages_limit if user.is_premium else 1
        
        if user_messages >= limit:
            raise HTTPException(
                status_code=403, 
                detail=f"Limite de messages atteinte. Version gratuite : 1 message. Premium : 10 messages."
            )
    
    # Créer le message
    message = Message(
        ticket_id=request.ticket_id,
        sender_type=request.sender_type,
        sender_id=user.id if request.sender_type == 'user' else None,
        content=request.content,
        channel='web'
    )
    db.add(message)
    db.commit()
    
    return {"success": True, "message": "Message envoyé"}


# ==========================================
# ROUTES API V2 (Pipeline unifié Gemini)
# ==========================================

class V2AnalyzeRequest(BaseModel):
    text: Optional[str] = ""
    content: Optional[str] = ""
    category: Optional[str] = "agriculture"
    photo_base64: Optional[str] = None
    photo_base64_list: Optional[List[str]] = None
    generate_media: Optional[bool] = True
    target_lang: Optional[str] = None  # Langue locale : "moore", "dioula", "fulfulde"

class V2EntreprendreRequest(BaseModel):
    text: Optional[str] = ""
    content: Optional[str] = ""
    category: Optional[str] = "agriculture"
    photo_base64: Optional[str] = None
    photo_base64_list: Optional[List[str]] = None
    generate_media: Optional[bool] = True
    target_lang: Optional[str] = None  # Langue locale cible

class V2VideoIllustrationRequest(BaseModel):
    diagnostic: str
    steps: Optional[List[str]] = []
    category: Optional[str] = "agriculture"

class V2ImageIllustrationRequest(BaseModel):
    diagnostic: str
    steps: Optional[List[str]] = []
    category: Optional[str] = "agriculture"

def _normalize_category(cat: Optional[str]) -> str:
    if not cat:
        return "agriculture"
    cat = cat.lower().strip()
    if cat in ("elevage", "élevage"):
        return "elevage"
    if cat in ("urgence", "sos_accident", "sos", "health"):
        return "urgence"
    if "cyber" in cat or cat in ("securite_numerique", "sécurité numérique"):
        return "cybersecurity"
    return "agriculture"

def _collect_images_b64(photo_base64: Optional[str], photo_base64_list: Optional[List[str]]) -> List[str]:
    """Collecte et nettoie les images base64"""
    images = []
    for payload in ([photo_base64] if photo_base64 else []) + (photo_base64_list or []):
        if not payload:
            continue
        # Retirer le prefix data:image/...;base64, si présent
        if "," in payload:
            payload = payload.split(",", 1)[1]
        if payload not in images:
            images.append(payload)
    return images[:3]


@app.post("/api/v2/sync-consultation")
async def sync_v2_consultation(
    payload: Dict[str, Any],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Synchronise une consultation faite sur le mobile (V2) vers le serveur.
    Cree un ticket et stocke le resultat de l'analyse.
    """
    try:
        category = payload.get("category", "agriculture")
        query = payload.get("user_query", "Analyse V2")
        response_data = payload.get("response", {})
        
        # Créer un nouveau ticket
        ticket = Ticket(
            user_id=current_user.id,
            category=category,
            status="resolved",
            ai_photo_analysis=json.dumps(response_data, ensure_ascii=False),
        )
        db.add(ticket)
        db.commit()
        db.refresh(ticket)
        
        # Créer le message initial pour le contenu de la requête
        msg = Message(
            ticket_id=ticket.id,
            sender_type="user",
            sender_id=current_user.id,
            content=query,
            channel="v2_mobile",
        )
        db.add(msg)
        
        # Ajouter la réponse IA comme message expert/system
        ai_msg = Message(
            ticket_id=ticket.id,
            sender_type="expert",
            sender_id=None,
            content=response_data.get("message", "Analyse terminée"),
            channel="v2_mobile",
        )
        db.add(ai_msg)
        
        db.commit()
        
        return {"status": "success", "ticket_id": ticket.id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v2/analyze")
async def v2_analyze(
    data: V2AnalyzeRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Pipeline v2 complet : texte + photos → Gemini → décision → médias → réponse unique"""
    import time as _time
    start_time = _time.time()

    text = data.text or data.content or ""
    category = _normalize_category(data.category)
    generate_media = data.generate_media is not False
    images_b64 = _collect_images_b64(data.photo_base64, data.photo_base64_list)

    if not text.strip() and not images_b64:
        raise HTTPException(status_code=400, detail="Envoyez au moins du texte ou une photo pour obtenir un diagnostic.")

    # ÉTAPE 1 : Analyse Gemini unifiée
    analysis = await v2_services.gemini_analyze(text=text, images_b64=images_b64, category=category)

    # ÉTAPE 2 : Décision (image? vidéo? urgence?)
    decision = v2_services.decide(analysis)

    # ÉTAPE 3 : Génération médias (si demandé)
    image_result = None
    video_result = None

    if generate_media:
        import asyncio

        async def _run_media_task(label: str, coro, timeout_seconds: int):
            try:
                result = await asyncio.wait_for(coro, timeout=timeout_seconds)
                return label, result
            except Exception as e:
                print(f"[MEDIA] Erreur {label}: {e}")
                return label, None

        tasks = []

        if decision["generer_image"] and decision["prompt_image"]:
            style = "schema" if decision["mode_urgence"] else "illustration"
            tasks.append(
                _run_media_task(
                    "image",
                    v2_services.generate_image(decision["prompt_image"], style=style, category=category),
                    20,
                )
            )

        if decision["generer_video"] and decision["prompt_video"]:
            tasks.append(
                _run_media_task(
                    "video",
                    v2_services.generate_video(
                        decision["prompt_video"],
                        gemini_api_key=GEMINI_API_KEY,
                        duration_sec=5 if decision["mode_urgence"] else 8,
                        is_urgency=decision["mode_urgence"],
                        category=category,
                    ),
                    25,
                )
            )

        if tasks:
            for label, result in await asyncio.gather(*tasks):
                if label == "image":
                    image_result = result
                elif label == "video":
                    video_result = result

    # ÉTAPE 4 : Réponse unique
    final_response = v2_services.build_response(
        analysis=analysis,
        decision=decision,
        image_result=image_result,
        video_result=video_result,
    )
    target_lang = (data.target_lang or "").strip().lower() or None
    studio_match = _find_studio_knowledge_match(
        db,
        category=category,
        query_text=text,
        photo_analysis=analysis,
    )
    if studio_match:
        print(
            f"[STUDIO] [V2-ANALYZE] Fiche #{studio_match.get('id')} "
            f"'{studio_match.get('title')}' utilisee "
            f"(score={studio_match.get('match_score')}, langue={target_lang or 'fr'})"
        )
        final_response = _apply_studio_match_to_v2_response(
            final_response, studio_match, target_lang
        )
        selected_audio = ((studio_match.get("audio") or {}).get(target_lang) or {}) if target_lang else {}
        if target_lang and not str(selected_audio.get("url") or "").strip():
            print(
                f"[STUDIO] [V2-ANALYZE] Voix {target_lang} indisponible; "
                "le mobile doit proposer la lecture francaise"
            )
    else:
        print(
            f"[STUDIO] [V2-ANALYZE] Aucune fiche validee correspondante "
            f"(categorie={category}, langue={target_lang or 'fr'}); "
            "le mobile doit proposer la lecture francaise"
        )

    offline_payload = {
        **final_response,
        "input_photo_base64": images_b64[0] if images_b64 else None,
    }

    try:
        _persist_offline_knowledge_entry(
            db=db,
            user_id=current_user.id,
            source_kind="v2_analyze",
            category=category,
            question_text=text or "Analyse photo Songra",
            response_payload=offline_payload,
        )
    except Exception as e:
        print(f"[OFFLINE-CORPUS] Erreur persistance v2/analyze: {e}")

    duration = int((_time.time() - start_time) * 1000)

    # La langue locale sélectionne uniquement une voix de fiche Studio.
    # Aucune traduction ni synthèse vocale n'est générée à la volée.
    voice_payload = None

    return {
        "status": "success",
        **final_response,
        "voice_summary": voice_payload.get("voice_summary") if voice_payload else None,
        "audio_url": voice_payload.get("audio_url") if voice_payload else final_response.get("audio_url"),
        "audio_mime_type": voice_payload.get("audio_mime_type") if voice_payload else final_response.get("audio_mime_type"),
        "_meta": {
            "duration_ms": duration,
            "provider": v2_services.AI_PROVIDER,
            "model": (
                v2_services.GROQ_MODEL if v2_services.AI_PROVIDER == "groq"
                else v2_services.OPENAI_MODEL if v2_services.AI_PROVIDER == "openai"
                else v2_services.GEMINI_MODEL
            ),
            "from_cache": analysis.get("from_cache", False),
            "fallback_used": analysis.get("from_fallback", False),
            "translated": False,
            "target_lang": target_lang,
            "lang_name": _TRANSLATOR_LANG_NAMES.get(target_lang) if target_lang else None,
        },
    }


@app.post("/api/v2/scanner/analyze")
async def v2_scanner_analyze(
    data: V2AnalyzeRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Scanner v2 : au moins 1 photo requise"""
    _require_resource(db, current_user, "analyses")
    text = data.text or data.content or ""
    category = _normalize_category(data.category)
    images_b64 = _collect_images_b64(data.photo_base64, data.photo_base64_list)

    # target_lang sert uniquement à choisir la voix enregistrée dans le Studio.
    target_lang = (data.target_lang or "").strip().lower() or None


    if not images_b64:
        raise HTTPException(status_code=400, detail="Le scanner nécessite au moins une photo.")

    analysis = await v2_services.gemini_analyze(text=text, images_b64=images_b64, category=category)
    decision = v2_services.decide(analysis)
    final_response = v2_services.build_response(analysis=analysis, decision=decision)
    studio_match = _find_studio_knowledge_match(
        db,
        category=category,
        query_text=text,
        photo_analysis=analysis,
    )
    if studio_match:
        print(
            f"[STUDIO] [V2-SCANNER] Fiche #{studio_match.get('id')} "
            f"'{studio_match.get('title')}' utilisee "
            f"(score={studio_match.get('match_score')}, langue={target_lang or 'fr'})"
        )
        final_response = _apply_studio_match_to_v2_response(
            final_response, studio_match, target_lang
        )
        selected_audio = ((studio_match.get("audio") or {}).get(target_lang) or {}) if target_lang else {}
        if target_lang and not str(selected_audio.get("url") or "").strip():
            print(
                f"[STUDIO] [V2-SCANNER] Voix {target_lang} indisponible; "
                "proposition de lecture francaise envoyee au mobile"
            )
    else:
        print(
            f"[STUDIO] [V2-SCANNER] Aucune fiche validee correspondante "
            f"(categorie={category}, langue={target_lang or 'fr'}); "
            "proposition de lecture francaise envoyee au mobile"
        )

    offline_payload = {
        **final_response,
        "input_photo_base64": images_b64[0] if images_b64 else None,
    }

    try:
        _persist_offline_knowledge_entry(
            db=db,
            user_id=current_user.id,
            source_kind="v2_scanner",
            category=category,
            question_text=text or "Scan photo Songra",
            response_payload=offline_payload,
        )
    except Exception as e:
        print(f"[OFFLINE-CORPUS] Erreur persistance v2/scanner: {e}")

    # Pas de traduction : voix Studio si disponible, sinon proposition FR.
    voice_payload = None

    _consume_resource(db, current_user, "analyses", "v2_scanner")
    return {
        "status": "success",
        **final_response,
        "translated": False,
        "target_lang": target_lang,
        "lang_name": _TRANSLATOR_LANG_NAMES.get(target_lang) if target_lang else None,
        "voice_summary": voice_payload.get("voice_summary") if voice_payload else None,
        "audio_url": voice_payload.get("audio_url") if voice_payload else final_response.get("audio_url"),
        "audio_mime_type": voice_payload.get("audio_mime_type") if voice_payload else final_response.get("audio_mime_type"),
    }


@app.post("/api/v2/assistant/query")
async def v2_assistant_query(
    data: V2AnalyzeRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Assistant conversationnel v2 : texte +/- image, pas de génération média par défaut"""
    _require_resource(db, current_user, "analyses")
    text = data.text or data.content or ""
    category = _normalize_category(data.category)
    images_b64 = _collect_images_b64(data.photo_base64, data.photo_base64_list)

    # target_lang sert uniquement à choisir la voix enregistrée dans le Studio.
    target_lang = (data.target_lang or "").strip().lower() or None

    if not text.strip() and not images_b64:
        raise HTTPException(status_code=400, detail="Posez une question ou envoyez une photo.")

    analysis = await v2_services.gemini_analyze(text=text, images_b64=images_b64, category=category)
    decision = v2_services.decide(analysis)
    final_response = v2_services.build_response(analysis=analysis, decision=decision)
    studio_match = _find_studio_knowledge_match(
        db,
        category=category,
        query_text=text,
        photo_analysis=analysis,
    )
    if studio_match:
        final_response = _apply_studio_match_to_v2_response(
            final_response, studio_match, target_lang
        )

    offline_payload = {
        **final_response,
        "input_photo_base64": images_b64[0] if images_b64 else None,
    }

    try:
        _persist_offline_knowledge_entry(
            db=db,
            user_id=current_user.id,
            source_kind="v2_assistant_query",
            category=category,
            question_text=text or "Question Songra",
            response_payload=offline_payload,
        )
    except Exception as e:
        print(f"[OFFLINE-CORPUS] Erreur persistance v2/assistant: {e}")

    # Pas de traduction : voix Studio si disponible, sinon proposition FR.
    voice_payload = None

    _consume_resource(db, current_user, "analyses", "v2_assistant")
    return {
        "status": "success",
        **final_response,
        "translated": False,
        "target_lang": target_lang,
        "lang_name": _TRANSLATOR_LANG_NAMES.get(target_lang) if target_lang else None,
        "voice_summary": voice_payload.get("voice_summary") if voice_payload else None,
        "audio_url": voice_payload.get("audio_url") if voice_payload else final_response.get("audio_url"),
        "audio_mime_type": voice_payload.get("audio_mime_type") if voice_payload else final_response.get("audio_mime_type"),
    }


# @app.post("/api/v2/entreprendre")
# async def v2_entreprendre(
#     data: V2EntreprendreRequest,
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db),
# ):
#     """Analyse entrepreneuriale d'un terrain → propositions business"""
#     import time as _time
#     start_time = _time.time()
#     try:
#         text = data.text or data.content or ""
#         category = _normalize_category(data.category)
#         generate_media = data.generate_media is not False
#         images_b64 = _collect_images_b64(data.photo_base64, data.photo_base64_list)
# 
#         # Analyse entrepreneuriale via Gemini
#         cache_hit = False
#         entrepreneurship = await v2_services.gemini_analyze_entrepreneurship(
#             text=text, images_b64=images_b64, category=category
#         )
# 
#         # Génération image/vidéo du plan de découpage
#         image_result = None
#         video_result = None
# 
#         if generate_media:
#             import asyncio
# 
#             async def _run_media_task(label: str, coro, timeout_seconds: int):
#                 try:
#                     result = await asyncio.wait_for(coro, timeout=timeout_seconds)
#                     return label, result
#                 except Exception as e:
#                     print(f"[ENTREPRENDRE] Erreur {label}: {e}")
#                     return label, None
# 
#             tasks = []
#             cat_label = "d'élevage" if category == "elevage" else "agricole"
# 
#             if entrepreneurship.get("besoin_image") and not (image_result and image_result.get("success")):
#                 img_prompt = (
#                     f"Plan d'aménagement de terrain {cat_label} au Burkina Faso vu du dessus. "
#                     f"Montrer le découpage en zones : {entrepreneurship.get('decoupage_terrain', '')}. "
#                     f"Propositions : {', '.join(p.get('titre', '') for p in entrepreneurship.get('propositions', []))}. "
#                     "Style plan/carte colorée, simple, avec des icônes pour chaque zone. Pas de texte."
#                 )
#                 tasks.append(
#                     _run_media_task(
#                         "image",
#                         v2_services.generate_image(img_prompt, style="schema", category=category),
#                         12,
#                     )
#                 )
# 
#             if entrepreneurship.get("besoin_video") and not (video_result and (video_result.get("success") or video_result.get("fallback"))):
#                 propositions = ". ".join(
#                     p.get("titre", "") for p in entrepreneurship.get("propositions", [])[:3] if p.get("titre")
#                 )
#                 calendrier = ". ".join(
#                     f"{item.get('mois', '')}: {item.get('activite', '')}"
#                     for item in entrepreneurship.get("calendrier_cultural", [])[:3]
#                 )
#                 video_prompt = (
#                     f"Vidéo pédagogique courte montrant un plan d'aménagement de terrain {cat_label} au Burkina Faso. "
#                     f"Montrer l'organisation de l'espace selon ce découpage : {entrepreneurship.get('decoupage_terrain', '')}. "
#                     f"Montrer aussi les projets proposés : {propositions}. "
#                     f"Calendrier de mise en oeuvre : {calendrier}. "
#                     "Style clair, vue du dessus puis gestes simples sur le terrain, sans texte à l'écran."
#                 )
#                 tasks.append(
#                     _run_media_task(
#                         "video",
#                         v2_services.generate_video(
#                             video_prompt,
#                             gemini_api_key=GEMINI_API_KEY,
#                             duration_sec=8,
#                             is_urgency=False,
#                             category=category,
#                         ),
#                         18,
#                     )
#                 )
# 
#             if tasks:
#                 results = await asyncio.gather(*tasks)
#                 for label, res in results:
#                     if label == "image":
#                         image_result = res
#                     elif label == "video":
#                         video_result = res
# 
#         # Support multilingue pour Entreprendre
#         localizations = {}
#         try:
#             trans_prompt = _build_entreprendre_translation_prompt(entrepreneurship)
#             # Priorité absolue à Gemini pour la qualité des langues burkinabè et éviter les boucles de répétition
#             if GEMINI_API_KEY:
#                 try:
#                     import asyncio
#                     model = genai.GenerativeModel("gemini-2.5-flash")
#                     response = await asyncio.to_thread(
#                         model.generate_content,
#                         trans_prompt,
#                         generation_config=genai.types.GenerationConfig(response_mime_type="application/json")
#                     )
#                     trans_response = response.text
#                 except Exception as gemini_exc:
#                     print(f"[WARN] Gemini echec traduction entreprendre: {gemini_exc}, fallback OpenAI")
#                     trans_response = await v2_services._openai_llm_answer(
#                         "Tu es un traducteur expert en langues burkinabè.", 
#                         trans_prompt,
#                         json_mode=True
#                     )
#             else:
#                 trans_response = await v2_services._openai_llm_answer(
#                     "Tu es un traducteur expert en langues burkinabè.", 
#                     trans_prompt,
#                     json_mode=True
#                 )
#             if trans_response:
#                 localizations = v2_services._parse_gemini_json(trans_response)
#         except Exception as e:
#             print(f"[ENTREPRENDRE] Erreur traduction: {e}")
# 
#         duration = int((_time.time() - start_time) * 1000)
# 
#         offline_payload = {
#             **entrepreneurship,
#             "localizations": localizations,
#             "image_base64": image_result["image_base64"] if image_result and image_result.get("success") else None,
#             "image_mime_type": image_result.get("mime_type") if image_result and image_result.get("success") else None,
#             "video_url": video_result.get("video_url") if video_result and video_result.get("success") else None,
#             "video_mime_type": video_result.get("mime_type") if video_result and video_result.get("success") else None,
#             "image_description": image_result.get("fallback_description") if image_result else None,
#             "video_description": video_result.get("video_description") if video_result and video_result.get("fallback") else None,
#         }
#         try:
#             _persist_offline_knowledge_entry(
#                 db=db,
#                 user_id=current_user.id,
#                 source_kind="entreprendre",
#                 category=category,
#                 question_text=text or "Analyse terrain Songra",
#                 response_payload=offline_payload,
#             )
#         except Exception as e:
#             print(f"[OFFLINE-CORPUS] Erreur persistance v2/entreprendre: {e}")
# 
#         return {
#             "status": "success",
#             **entrepreneurship,
#             "localizations": localizations,
#             "image_base64": image_result["image_base64"] if image_result and image_result.get("success") else None,
#             "image_mime_type": image_result.get("mime_type") if image_result and image_result.get("success") else None,
#             "image_description": image_result.get("fallback_description") if image_result else None,
#             "video_base64": video_result.get("video_base64") if video_result and video_result.get("success") else None,
#             "video_url": video_result.get("video_url") if video_result and video_result.get("success") else None,
#             "video_mime_type": video_result.get("mime_type") if video_result and video_result.get("success") else None,
#             "video_duration": video_result.get("duration_sec") if video_result and video_result.get("success") else None,
#             "video_description": video_result.get("video_description") if video_result and video_result.get("fallback") else None,
#             "video_steps": video_result.get("steps_visuelles") if video_result and video_result.get("fallback") else None,
#             "_meta": {
#                 "duration_ms": duration,
#                 "model": "gemini-1.5-flash",
#                 "shared_cache_hit": cache_hit,
#                 "shared_source": None,
#             },
#         }
#     except Exception as e:
#         print(f"[ENTREPRENDRE] Erreur globale: {e}")
#         return {"status": "error", "message": f"Erreur serveur: {str(e)}"}
@app.get("/api/v2/health")
async def v2_health():
    """Health check v2"""
    return {
        "status": "ok",
        "service": "songra-v2",
        "gemini_configured": bool(GEMINI_API_KEY),
        "timestamp": datetime.utcnow().isoformat(),
    }


# @app.post("/api/v2/generate-video-illustration")
# async def v2_generate_video_illustration(
#     data: V2VideoIllustrationRequest,
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db),
# ):
#     """Génération on-demand d'une vidéo illustrée"""
#     import time as _time
#     start_time = _time.time()
# 
#     diagnostic = data.diagnostic
#     steps_list = (data.steps or [])[:5]
#     cat = _normalize_category(data.category)
#     cache_question = _build_media_cache_question(diagnostic, steps_list)
#     reusable_entry = _find_reusable_offline_entry(
#         db,
#         domain=_normalize_offline_domain(cat),
#         source_kinds=["generated_video_illustration"],
#         question_text=cache_question,
#     )
#     reusable_payload = _parse_offline_response_json(reusable_entry) if reusable_entry else {}
#     if reusable_payload:
#         return {
#             "status": "success",
#             "type": "video" if reusable_payload.get("video_url") or reusable_payload.get("video_base64") else ("image_steps" if reusable_payload.get("steps") else "text_steps"),
#             "video_base64": reusable_payload.get("video_base64"),
#             "video_url": reusable_payload.get("video_url"),
#             "video_mime_type": reusable_payload.get("video_mime_type"),
#             "duration_sec": reusable_payload.get("video_duration"),
#             "steps": reusable_payload.get("steps") or ([{"description": step} for step in (reusable_payload.get("video_steps") or [])]),
#             "_meta": {
#                 "duration_ms": int((_time.time() - start_time) * 1000),
#                 "shared_cache_hit": True,
#                 "shared_source": reusable_entry.source_kind,
#             },
#         }
# 
#     # Tentative 1 : Veo (vraie vidéo)
#     video_result = None
#     try:
#         if cat == "agriculture":
#             video_prompt = f"Vidéo pédagogique courte (5-10 secondes) montrant les gestes techniques pour traiter : {diagnostic}. Étapes : {'. '.join(steps_list)}. Contexte : champ en Afrique sahélienne."
#         elif cat == "elevage":
#             video_prompt = f"Vidéo pédagogique courte (5-10 secondes) montrant les soins pour : {diagnostic}. Étapes : {'. '.join(steps_list)}. Contexte : élevage rural Burkina Faso."
#         else:
#             video_prompt = f"Vidéo courte (5-8 secondes) montrant les gestes d'urgence pour : {diagnostic}. Étapes : {'. '.join(steps_list)}. Contexte : village africain."
# 
#         video_result = await v2_services.generate_video(
#             video_prompt,
#             gemini_api_key=GEMINI_API_KEY,
#             duration_sec=5 if cat == "urgence" else 8,
#             is_urgency=cat == "urgence",
#             category=cat,
#         )
#     except Exception as e:
#         print(f"[VIDEO-ILLUS] Veo échoué: {e}")
# 
#     # Si Veo a réussi
#     if video_result and video_result.get("success"):
#         try:
#             _persist_offline_knowledge_entry(
#                 db=db,
#                 user_id=current_user.id,
#                 source_kind="generated_video_illustration",
#                 category=cat,
#                 question_text=cache_question,
#                 response_payload=_build_media_offline_payload(
#                     diagnostic=diagnostic,
#                     steps=steps_list,
#                     video_result=video_result,
#                 ),
#             )
#         except Exception as e:
#             print(f"[OFFLINE-CORPUS] Erreur persistance video illustration: {e}")
#         return {
#             "status": "success",
#             "type": "video",
#             "video_base64": video_result.get("video_base64"),
#             "video_url": video_result.get("video_url"),
#             "video_mime_type": video_result.get("mime_type", "video/mp4"),
#             "duration_sec": video_result.get("duration_sec"),
#         }
# 
#     # Fallback : image infographique étape par étape
#     steps_text = ". ".join(f"Étape {i+1}: {s}" for i, s in enumerate(steps_list)) if steps_list else diagnostic
# 
#     if cat == "agriculture":
#         infographic_prompt = f"Infographie pédagogique agricole en 3-4 vignettes montrant les étapes pour traiter : {diagnostic}. {steps_text}. Style bande-dessinée simple, personnages africains. Pas de texte."
#     elif cat == "elevage":
#         infographic_prompt = f"Infographie pédagogique vétérinaire en 3-4 vignettes montrant les soins pour : {diagnostic}. {steps_text}. Style illustration simple. Pas de texte."
#     else:
#         infographic_prompt = f"Infographie premiers secours en 3-4 vignettes montrant les gestes pour : {diagnostic}. {steps_text}. Style schématique clair. Pas de texte."
# 
#     try:
#         img_result = await v2_services.generate_image(infographic_prompt, style="schema", category=cat)
#         if img_result and img_result.get("success"):
#             steps_payload = [{"image_base64": img_result["image_base64"], "mime_type": img_result.get("mime_type", "image/png"), "description": diagnostic}]
#             try:
#                 _persist_offline_knowledge_entry(
#                     db=db,
#                     user_id=current_user.id,
#                     source_kind="generated_video_illustration",
#                     category=cat,
#                     question_text=cache_question,
#                     response_payload=_build_media_offline_payload(
#                         diagnostic=diagnostic,
#                         steps=steps_list,
#                         image_result={
#                             "image_base64": img_result.get("image_base64"),
#                             "image_mime_type": img_result.get("mime_type", "image/png"),
#                             "image_description": diagnostic,
#                         },
#                         video_result={
#                             "type": "image_steps",
#                             "steps": steps_payload,
#                             "video_steps": steps_list,
#                             "video_description": diagnostic,
#                         },
#                     ),
#                 )
#             except Exception as e:
#                 print(f"[OFFLINE-CORPUS] Erreur persistance fallback illustration video: {e}")
#             return {
#                 "status": "success",
#                 "type": "image_steps",
#                 "steps": steps_payload,
#             }
#     except Exception as e:
#         print(f"[VIDEO-ILLUS] Image fallback échoué: {e}")
# 
#     # Fallback ultime : texte
#     text_fallback = {
#         "status": "success",
#         "type": "text_steps",
#         "steps": [{"description": s} for s in steps_list] if steps_list else [{"description": diagnostic}],
#     }
#     try:
#         _persist_offline_knowledge_entry(
#             db=db,
#             user_id=current_user.id,
#             source_kind="generated_video_illustration",
#             category=cat,
#             question_text=cache_question,
#             response_payload=_build_media_offline_payload(
#                 diagnostic=diagnostic,
#                 steps=steps_list,
#                 video_result={
#                     "type": "text_steps",
#                     "video_steps": steps_list,
#                     "video_description": diagnostic,
#                 },
#             ),
#         )
#     except Exception as e:
#         print(f"[OFFLINE-CORPUS] Erreur persistance text steps video: {e}")
#     return text_fallback
# 
# 
# @app.post("/api/v2/generate-image-illustration")
# async def v2_generate_image_illustration(
#     data: V2ImageIllustrationRequest,
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db),
# ):
#     """Génération on-demand d'une illustration pédagogique"""
#     import time as _time
#     start_time = _time.time()
# 
#     diagnostic = data.diagnostic
#     steps_list = (data.steps or [])[:5]
#     cat = _normalize_category(data.category)
#     cache_question = _build_media_cache_question(diagnostic, steps_list)
#     reusable_entry = _find_reusable_offline_entry(
#         db,
#         domain=_normalize_offline_domain(cat),
#         source_kinds=["generated_image_illustration"],
#         question_text=cache_question,
#     )
#     reusable_payload = _parse_offline_response_json(reusable_entry) if reusable_entry else {}
#     # Réutiliser uniquement si 80% des mots-clés correspondent (dans les deux sens)
#     def _cache_similarity(t1: str, t2: str) -> float:
#         s1 = set(_tokenize(t1))
#         s2 = set(_tokenize(t2))
#         if not s1 or not s2:
#             return 0.0
#         overlap = len(s1 & s2)
#         return min(overlap / len(s1), overlap / len(s2))
# 
#     good_cache_match = (
#         reusable_entry is not None
#         and _cache_similarity(cache_question, reusable_entry.question or "") >= 0.8
#     )
#     if good_cache_match and reusable_payload.get("image_base64"):
#         return {
#             "status": "success",
#             "type": "image",
#             "image_base64": reusable_payload.get("image_base64"),
#             "image_mime_type": reusable_payload.get("image_mime_type"),
#             "image_description": reusable_payload.get("image_description") or diagnostic,
#             "_meta": {
#                 "duration_ms": int((_time.time() - start_time) * 1000),
#                 "shared_cache_hit": True,
#                 "shared_source": reusable_entry.source_kind,
#             },
#         }
#     steps_text = ". ".join(f"Étape {i+1}: {s}" for i, s in enumerate(steps_list)) if steps_list else diagnostic
# 
#     if cat == "agriculture":
#         image_prompt = (
#             f"Illustration pédagogique agricole montrant : {diagnostic}. "
#             f"Montrer aussi les gestes recommandés : {steps_text}. "
#             "Contexte : champ, outils et personnes du Burkina Faso. Pas de texte."
#         )
#         style = "illustration"
#     elif cat == "elevage":
#         image_prompt = (
#             f"Illustration vétérinaire simple montrant : {diagnostic}. "
#             f"Montrer les soins ou vérifications utiles : {steps_text}. "
#             "Contexte : élevage rural au Burkina Faso. Pas de texte."
#         )
#         style = "illustration"
#     else:
#         image_prompt = (
#             f"Illustration schématique claire montrant les gestes pour : {diagnostic}. "
#             f"Étapes : {steps_text}. Contexte local Burkina Faso, style premiers secours non choquant. Pas de texte."
#         )
#         style = "schema"
# 
#     image_result = await v2_services.generate_image(image_prompt, style=style, category=cat)
# 
#     if image_result and image_result.get("success"):
#         try:
#             _persist_offline_knowledge_entry(
#                 db=db,
#                 user_id=current_user.id,
#                 source_kind="generated_image_illustration",
#                 category=cat,
#                 question_text=cache_question,
#                 response_payload=_build_media_offline_payload(
#                     diagnostic=diagnostic,
#                     steps=steps_list,
#                     image_result=image_result,
#                 ),
#             )
#         except Exception as e:
#             print(f"[OFFLINE-CORPUS] Erreur persistance image illustration: {e}")
#         return {
#             "status": "success",
#             "type": "image",
#             "image_base64": image_result.get("image_base64"),
#             "image_mime_type": image_result.get("mime_type", "image/png"),
#             "image_description": diagnostic,
#             "_meta": {
#                 "duration_ms": int((_time.time() - start_time) * 1000),
#             },
#         }
# 
#     return {
#         "status": "success",
#         "type": "image_fallback",
#         "image_description": image_result.get("fallback_description") if image_result else diagnostic,
#         "steps": [{"description": s} for s in steps_list] if steps_list else [{"description": diagnostic}],
#         "_meta": {
#             "duration_ms": int((_time.time() - start_time) * 1000),
#         },
#     }
# 
# 
@app.post("/api/contacts/sync")
async def sync_rural_contacts(payload: RuralSyncPayload, current_user: User = Depends(get_current_user_or_expert), db: Session = Depends(get_db)):
    # 1. Update/Insert incoming contacts
    for c in payload.contacts:
        existing = db.query(RuralContactDB).filter(RuralContactDB.id == c.id).first()
        try:
            updated_at_dt = datetime.fromisoformat(c.updated_at.replace("Z", "+00:00"))
        except:
            updated_at_dt = datetime.utcnow()
        
        if not existing:
            new_contact = RuralContactDB(
                id=c.id,
                user_id=current_user.id,
                name=c.name,
                actor_type=c.actor_type,
                phone_number=c.phone_number,
                location_label=c.location_label,
                organization=c.organization,
                market_name=c.market_name,
                notes=c.notes,
                tags_json=json.dumps(c.tags),
                crop_labels_json=json.dumps(c.crop_labels),
                updated_at=updated_at_dt
            )
            db.add(new_contact)
        else:
            if updated_at_dt > existing.updated_at:
                existing.name = c.name
                existing.actor_type = c.actor_type
                existing.phone_number = c.phone_number
                existing.location_label = c.location_label
                existing.organization = c.organization
                existing.market_name = c.market_name
                existing.notes = c.notes
                existing.tags_json = json.dumps(c.tags)
                existing.crop_labels_json = json.dumps(c.crop_labels)
                existing.updated_at = updated_at_dt

    db.commit()

    # 2. Return all contacts to "share with everyone" as requested
    all_contacts = db.query(RuralContactDB).all()
    return {
        "status": "success",
        "contacts": [
            {
                "id": c.id,
                "name": c.name,
                "actor_type": c.actor_type,
                "phone_number": c.phone_number,
                "location_label": c.location_label,
                "organization": c.organization,
                "market_name": c.market_name,
                "notes": c.notes,
                "tags": json.loads(c.tags_json or "[]"),
                "crop_labels": json.loads(c.crop_labels_json or "[]"),
                "updated_at": c.updated_at.isoformat()
            } for c in all_contacts
        ]
    }

class TranslateRequest(BaseModel):
    text: str
    target_language: str
    category: Optional[str] = None


@app.post("/api/translate")
async def translate_text_endpoint(payload: TranslateRequest):
    """Traduit un texte du Français vers une langue locale (Mooré, Dioula, etc.)"""
    try:
        from burkina_translator import translate_text
        res = translate_text(
            text=payload.text,
            target_lang=payload.target_language.strip().lower(),
            gemini_api_key=GEMINI_API_KEY,
            category=payload.category
        )
        return {
            "status": "success",
            "translated_text": res.get("translation", payload.text),
            "speech_text": res.get("speech_text", res.get("translation", payload.text)),
            "confidence": res.get("confidence", 0.0),
            "source": res.get("source", "unknown")
        }
    except Exception as e:
        print(f"[TRANSLATOR] Erreur /api/translate: {e}")
        return {
            "status": "error",
            "message": str(e),
            "translated_text": payload.text,
            "speech_text": payload.text
        }


@app.get("/api/admin/dictionaries")
async def get_dictionary_stats(current_user: Any = Depends(get_current_user_or_expert)):
    return {"languages": dictionary_stats(), "supported": ["fr", "moore", "dioula", "fulfulde"]}


@app.post("/api/admin/dictionaries/import")
async def import_agricultural_dictionary(
    language: str = Form(...),
    replace: bool = Form(False),
    file: UploadFile = File(...),
    current_user: Any = Depends(get_current_user_or_expert),
):
    try:
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Le fichier depasse 10 Mo.")
        result = import_dictionary_file(content, file.filename or "dictionary.csv", language, replace)
        return {"status": "success", **result}
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

# @app.post("/api/entreprendre/sync")
# async def sync_entreprendre(payload: EntreprendreSyncPayload, current_user: User = Depends(get_current_user_or_expert), db: Session = Depends(get_db)):
#     for r in payload.records:
#         existing = db.query(EntreprendreHistoryDB).filter(EntreprendreHistoryDB.id == r.id).first()
#         try:
#             updated_at_dt = datetime.fromisoformat(r.updated_at.replace("Z", "+00:00"))
#         except:
#             updated_at_dt = datetime.utcnow()
#         
#         if not existing:
#             new_record = EntreprendreHistoryDB(
#                 id=r.id,
#                 user_id=current_user.id,
#                 category=r.category,
#                 user_query=r.user_query,
#                 response_json=r.response_json,
#                 photo_path=r.photo_path,
#                 plot_id=r.plot_id,
#                 translations_json=r.translations_json,
#                 updated_at=updated_at_dt
#             )
#             db.add(new_record)
#         else:
#             if updated_at_dt > existing.updated_at:
#                 existing.category = r.category
#                 existing.user_query = r.user_query
#                 existing.response_json = r.response_json
#                 existing.photo_path = r.photo_path
#                 existing.plot_id = r.plot_id
#                 existing.translations_json = r.translations_json
#                 existing.updated_at = updated_at_dt
#     
#     db.commit()
#     
#     my_records = db.query(EntreprendreHistoryDB).filter(EntreprendreHistoryDB.user_id == current_user.id).all()
#     return {
#         "status": "success",
#         "records": [
#             {
#                 "id": rec.id,
#                 "category": rec.category,
#                 "user_query": rec.user_query,
#                 "response_json": rec.response_json,
#                 "photo_path": rec.photo_path,
#                 "plot_id": rec.plot_id,
#                 "translations_json": rec.translations_json,
#                 "updated_at": rec.updated_at.isoformat()
#             } for rec in my_records
#         ]
#     }
# 
# 
class LocalTranslateIn(BaseModel):
    text: str
    target_lang: str


# @app.post("/api/translate/local")
# async def local_dictionary_translate(payload: LocalTranslateIn):
#     # Traduction avec LLM (priorité Gemini, fallback OpenAI) pour garantir une traduction naturelle et fluide,
#     # sans calque littéral mot-à-mot ni jargon incompréhensible, en attendant la complétude de Burkina Dict.
#     target_lang_label = payload.target_lang.strip().lower()
#     
#     # Résoudre les étiquettes de langue
#     lang_names = {
#         "moore": "Mooré",
#         "dioula": "Dioula",
#         "fulfulde": "Fulfuldé",
#         "gourounsi": "Gourounsi",
#         "bissa": "Bissa"
#     }
#     target_lang_name = lang_names.get(target_lang_label, target_lang_label.capitalize())
# 
#     prompt = (
#         f"Tu es un traducteur expert natif en langue {target_lang_name} (Burkina Faso).\n"
#         f"Traduis le texte français suivant en {target_lang_name}.\n"
#         f"CRITÈRES DE HAUTE QUALITÉ :\n"
#         f"1. Ne traduis SURTOUT PAS mot-à-mot (pas de traduction littérale). Adapte le sens en utilisant les expressions et termes les plus naturels possibles en langue locale sans altérer le sens original.\n"
#         f"2. Si un mot spécifique (technique ou moderne) n'a pas d'équivalent direct, utilise des synonymes ou des paraphrases naturelles en langue locale plutôt que de le traduire littéralement.\n"
#         f"3. Le ton doit être oral, clair et adapté à des producteurs ruraux.\n"
#         f"4. Retourne UNIQUEMENT la traduction finale brute en {target_lang_name}. N'écris pas d'introduction, d'explication, ni de bloc de code markdown.\n\n"
#         f"Texte français à traduire : {payload.text}"
#     )
# 
#     try:
#         # Priorité absolue à Gemini pour la qualité et fluidité des langues africaines
#         if GEMINI_API_KEY:
#             model = genai.GenerativeModel("gemini-2.5-flash")
#             result = model.generate_content(prompt)
#             translation = result.text.strip()
#             # Nettoyer les éventuels restes de format markdown de bloc de code
#             translation = re.sub(r"^```[a-zA-Z]*\n", "", translation)
#             translation = re.sub(r"\n```$", "", translation)
#             translation = translation.strip()
#             if translation:
#                 print(f"[INFO] Traduction locale /api/translate/local generee via Gemini pour {target_lang_name}")
#                 return {"translation": translation}
#         
#         # Fallback OpenAI
#         if openai_client and OPENAI_API_KEY:
#             response = openai_client.chat.completions.create(
#                 model="gpt-4o",
#                 messages=[{"role": "user", "content": prompt}],
#                 temperature=0.2,
#                 max_tokens=2500,
#             )
#             content = response.choices[0].message.content if response.choices else ""
#             translation = (content or "").strip()
#             # Nettoyer markdown
#             translation = re.sub(r"^```[a-zA-Z]*\n", "", translation)
#             translation = re.sub(r"\n```$", "", translation)
#             translation = translation.strip()
#             if translation:
#                 print(f"[INFO] Traduction locale /api/translate/local generee via OpenAI (fallback) pour {target_lang_name}")
#                 return {"translation": translation}
# 
#     except Exception as exc:
#         print(f"[WARN] Echec de traduction locale LLM /api/translate/local: {exc}")
# 
#     # Fallback ultime sur le texte brut
#     return {"translation": payload.text}
# 
# 
# # ==========================================
# # LANCEMENT
# # ==========================================
# 
# if __name__ == "__main__":
#     import uvicorn
#     port = int(os.getenv("PORT", "3000"))
#     
#     provider_name = os.getenv("AI_PROVIDER", "openai").upper()
#     print("=" * 50)
#     print(f"SONGRA - Backend unifié v2 ({provider_name})")
#     print("Version 6.0 - Pipeline v2 + RAG + Expert + Entreprendre")
#     print("=" * 50)
#     print(f"Serveur démarré sur http://localhost:{port}")
#     print("API v2: /api/v2/analyze, /api/v2/entreprendre, /api/v2/scanner/analyze")
#     print("Test expert: test@resolvehub.bf / test123")
#     print("=" * 50)
#     
#     # Créer l'expert test au démarrage
#     try:
#         db = SessionLocal()
#         existing = db.query(Expert).filter(Expert.email == "test@resolvehub.bf").first()
#         if not existing:
#             expert = Expert(
#                 email="test@resolvehub.bf",
#                 password_hash=hash_password("test123"),
#                 full_name="Expert Test IA",
#                 specialization="agriculture",
#                 is_active=True
#             )
#             db.add(expert)
#             db.commit()
#             print("[OK] Expert test créé: test@resolvehub.bf / test123")
#         # Charger la base de connaissances locale depuis le JSON
#         try:
#             load_knowledge_from_json(db)
#             total_items = db.query(KnowledgeItem).count()
#             print(f"[OK] Base de connaissances chargée ({total_items} fiches)")
#         except Exception as e_load:
#             print(f"[ERROR] Chargement base de connaissances: {e_load}")
#         db.close()
#     except Exception as e:
#         print(f"[ERROR] Création expert: {e}")
#     
#     # Le mode reload est plutôt à utiliser avec la commande uvicorn en ligne
#     # de commande (ex: `uvicorn main:app --reload`). Ici on garde un run simple.
#     uvicorn.run(app, host="0.0.0.0", port=port, reload=False)
# ── CAREMA PORTAL ADMIN ENDPOINTS ──────────────────────────────────────

def _serialize_organization(organization: Organization, db: Session) -> Dict[str, Any]:
    return {
        "id": organization.id,
        "name": organization.name,
        "code": organization.code,
        "description": organization.description,
        "region": organization.region,
        "phone_number": organization.phone_number,
        "email": organization.email,
        "is_active": organization.is_active,
        "experts_count": db.query(Expert).filter(Expert.organization_id == organization.id).count(),
        "users_count": db.query(User).filter(User.organization_id == organization.id).count(),
        "created_at": organization.created_at.isoformat() if organization.created_at else None,
    }


@app.get("/api/admin/organizations")
async def get_admin_organizations(
    current_admin: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    organizations = db.query(Organization).order_by(Organization.name).all()
    return [_serialize_organization(item, db) for item in organizations]


@app.post("/api/admin/organizations")
async def create_admin_organization(
    payload: Dict[str, Any],
    current_admin: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="Le nom de l'ONG est obligatoire")
    existing = db.query(Organization).filter(func.lower(Organization.name) == name.lower()).first()
    if existing:
        raise HTTPException(status_code=409, detail="Cette ONG existe déjà")
    organization = Organization(
        name=name,
        code=(str(payload.get("code") or "").strip() or None),
        description=(str(payload.get("description") or "").strip() or None),
        region=(str(payload.get("region") or "").strip() or None),
        phone_number=(str(payload.get("phone_number") or "").strip() or None),
        email=(str(payload.get("email") or "").strip() or None),
        is_active=True,
    )
    db.add(organization)
    try:
        db.commit()
        db.refresh(organization)
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="Code ou nom d'ONG déjà utilisé") from exc
    return _serialize_organization(organization, db)


@app.put("/api/admin/organizations/{organization_id}")
async def update_admin_organization(
    organization_id: int,
    payload: Dict[str, Any],
    current_admin: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    organization = db.query(Organization).filter(Organization.id == organization_id).first()
    if not organization:
        raise HTTPException(status_code=404, detail="ONG introuvable")
    for field in ("name", "code", "description", "region", "phone_number", "email"):
        if field in payload:
            value = str(payload.get(field) or "").strip()
            setattr(organization, field, value or None)
    if "is_active" in payload:
        organization.is_active = bool(payload["is_active"])
    db.commit()
    db.refresh(organization)
    return _serialize_organization(organization, db)

@app.get("/api/admin/users")
async def get_admin_users(
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    """Lister les utilisateurs pour CAREMA avec leur historique"""
    users = db.query(User).all()
    result = []
    for u in users:
        ticket_count = db.query(Ticket).filter(Ticket.user_id == u.id).count()
        result.append({
            "id": u.id,
            "phone_number": u.phone_number,
            "name": u.name,
            "location": u.location,
            "organization": getattr(u, "organization", None),
            "organization_id": getattr(u, "organization_id", None),
            "created_at": u.created_at.isoformat() if u.created_at else None,
            "is_active": getattr(u, "is_active", True),
            "role": getattr(u, "role", "utilisateur"),
            "ticket_count": ticket_count
        })
    return result


@app.post("/api/admin/users")
async def create_admin_user(
    payload: Dict[str, Any],
    current_admin: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    phone_number = str(payload.get("phone_number") or "").strip()
    password = str(payload.get("password") or "").strip()
    _validate_user_credentials(phone_number, password)
    if db.query(User).filter(User.phone_number == phone_number).first():
        raise HTTPException(status_code=409, detail="Ce numéro est déjà enregistré")
    organization = None
    organization_id = payload.get("organization_id")
    if organization_id not in (None, ""):
        organization = db.query(Organization).filter(
            Organization.id == int(organization_id), Organization.is_active == True
        ).first()
        if not organization:
            raise HTTPException(status_code=404, detail="ONG introuvable")
    user = User(
        phone_number=phone_number,
        password_hash=hash_password(password),
        name=(str(payload.get("name") or "").strip() or None),
        location=(str(payload.get("location") or "").strip() or None),
        role="utilisateur",
        is_active=True,
        organization_id=organization.id if organization else None,
        organization=organization.name if organization else None,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"status": "success", "user": serialize_user(user)}

@app.put("/api/admin/users/{user_id}/status")
async def update_user_status(
    user_id: int,
    payload: dict,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    """Activer/désactiver ou changer le rôle d'un utilisateur"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if "is_active" in payload:
        user.is_active = payload["is_active"]
    if "role" in payload:
        user.role = payload["role"]
    if "organization" in payload:
        user.organization = (payload.get("organization") or "").strip() or None
    if "organization_id" in payload:
        organization_id = payload.get("organization_id")
        organization = None
        if organization_id not in (None, ""):
            organization = db.query(Organization).filter(Organization.id == int(organization_id), Organization.is_active == True).first()
            if not organization:
                raise HTTPException(status_code=404, detail="ONG introuvable")
        user.organization_id = organization.id if organization else None
        user.organization = organization.name if organization else None
    db.commit()
    return {"status": "success", "user": {
        "id": user.id,
        "is_active": getattr(user, "is_active", True),
        "role": getattr(user, "role", "utilisateur")
        ,"organization": getattr(user, "organization", None)
    }}

@app.get("/api/admin/experts")
async def get_admin_experts(
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    """Lister le réseau d'experts"""
    experts = db.query(Expert).all()
    result = []
    for e in experts:
        result.append({
            "id": e.id,
            "email": e.email,
            "full_name": e.full_name,
            "specialization": e.specialization,
            "is_active": e.is_active,
            "zone": getattr(e, "zone", ""),
            "project": getattr(e, "project", ""),
            "language": getattr(e, "language", ""),
            "institution": getattr(e, "institution", "")
            ,"organization_id": getattr(e, "organization_id", None)
            ,"role": getattr(e, "role", "expert")
        })
    return result

@app.post("/api/admin/experts")
async def create_admin_expert(
    payload: dict,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db),
):
    """Créer ou modifier un expert dans le réseau"""
    email = payload.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")

    organization = None
    organization_id = payload.get("organization_id")
    if organization_id not in (None, ""):
        organization = db.query(Organization).filter(
            Organization.id == int(organization_id),
            Organization.is_active == True,
        ).first()
        if not organization:
            raise HTTPException(status_code=404, detail="ONG introuvable")
    requested_role = str(payload.get("role") or "expert").strip().lower()
    if requested_role not in {"expert", "admin"}:
        raise HTTPException(status_code=422, detail="Rôle expert invalide")
        
    expert = db.query(Expert).filter(Expert.email == email).first()
    if not expert:
        # Create new (meme schema de hash que /api/auth/login : hash_password/verify_password)
        password_hash = hash_password(payload.get("password", "expert123"))
        expert = Expert(
            email=email,
            password_hash=password_hash,
            full_name=payload.get("full_name", ""),
            specialization=payload.get("specialization", ""),
            role=requested_role,
            is_active=payload.get("is_active", True),
            zone=payload.get("zone", ""),
            project=payload.get("project", ""),
            language=payload.get("language", ""),
            institution=organization.name if organization else payload.get("institution", ""),
            organization_id=organization.id if organization else None,
        )
        db.add(expert)
    else:
        # Update existing
        expert.full_name = payload.get("full_name", expert.full_name)
        expert.specialization = payload.get("specialization", expert.specialization)
        expert.role = requested_role
        expert.is_active = payload.get("is_active", expert.is_active)
        expert.zone = payload.get("zone", expert.zone)
        expert.project = payload.get("project", expert.project)
        expert.language = payload.get("language", expert.language)
        if "organization_id" in payload:
            expert.organization_id = organization.id if organization else None
            expert.institution = organization.name if organization else None
        else:
            expert.institution = payload.get("institution", expert.institution)
        
    db.commit()
    return {"status": "success", "expert_id": expert.id}

@app.post("/api/tickets/{ticket_id}/reply-voice")
async def reply_ticket_voice(
    ticket_id: int,
    file: UploadFile = File(...),
    language: Optional[str] = Form(None),
    current_expert: Expert = Depends(get_current_expert),
    db: Session = Depends(get_db)
):
    """Associer une réponse audio locale (traduction) à un ticket"""
    is_admin = (getattr(current_expert, "role", "expert") or "expert").lower() == "admin"
    ticket = (
        db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if is_admin else _expert_ticket_or_404(db, current_expert, ticket_id)
    )
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
        
    os.makedirs("uploads/replies", exist_ok=True)
    filename = f"reply_{ticket_id}_{int(time.time())}_{file.filename}"
    filepath = f"uploads/replies/{filename}"
    
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    # Ajouter un message vocal au ticket
    new_msg = Message(
        ticket_id=ticket_id,
        sender_type="expert",
        sender_id=current_expert.id,
        content="[Message Vocal Traduit en Langue Locale]",
        channel="web",
        audio_url=f"/uploads/replies/{filename}",
        language=_normalize_expert_local_language(language or ticket.preferred_language),
    )
    db.add(new_msg)
    
    # Mettre à jour le ticket
    ticket.status = "resolved"
    db.commit()
    
    return {"status": "success", "audio_url": f"/uploads/replies/{filename}"}

class TicketStatusUpdate(BaseModel):
    status: str

@app.put("/api/tickets/{ticket_id}/status")
async def update_ticket_status(
    ticket_id: int,
    payload: TicketStatusUpdate,
    current_expert: Expert = Depends(get_current_admin_expert),
    db: Session = Depends(get_db)
):
    """Mettre à jour le statut d'un ticket (Workflow CAREMA)"""
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    ticket.status = payload.status
    if payload.status == "resolved":
        ticket.resolved_at = datetime.utcnow()
    db.commit()
    return {"status": "success", "new_status": ticket.status}


@app.post("/api/auth/expert/logout")
async def logout_expert(
    authorization: Optional[str] = Header(default=None),
    current_expert: Expert = Depends(get_current_expert),
):
    token = _extract_bearer_token(authorization)
    payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    from voice_alert import revoke_expert_token
    revoke_expert_token(
        SessionLocal, current_expert.id, str(payload["jti"]),
        datetime.utcfromtimestamp(int(payload["exp"])),
    )
    return {"success": True}


# Domaine agricole distinct des SOS. Le fournisseur inclus est exclusivement mock.
from voice_alert import AlertBase, create_router as create_voice_alert_router

AlertBase.metadata.create_all(bind=engine)
app.include_router(create_voice_alert_router(
    get_db=get_db, get_user=get_current_user, get_expert=get_current_expert,
    get_admin=get_current_admin_expert, session_factory=SessionLocal,
))


if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.getenv("PORT", 3000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=False)


