"""
burkina_translator.py
=====================
Moteur de traduction universel — Langues locales du Burkina Faso.

Priorites :
  1. Dictionnaire local CSV/JSON (confiance max)
  2. Recherche de synonymes dans le dictionnaire local
  3. Gemini AI avec regles grammaticales officielles + connaissances externes

Philosophie :
  - JAMAIS de traduction mot-a-mot (traduction du SENS, pas des mots)
  - Les mots absents du dict sont traites en silence — l'utilisateur ne voit rien
  - Chercher un synonyme avant de passer a l'IA
  - Respecter les normes de la Commission Nationale des Langues du Burkina Faso
"""

import os
import json
import re
import csv
import unicodedata
from difflib import SequenceMatcher
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional

LANG_NAMES: Dict[str, str] = {
    "moore":     "Moore",
    "dioula":    "Dioula",
    "fulfulde":  "Fulfulde",
}

VALID_LANGS = set(LANG_NAMES.keys())

_BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
_TRANSLATE_DIR = os.path.join(_BACKEND_DIR, "TRANSLATE")

_DICT_FILES: Dict[str, str] = {
    "moore":     "dictionnaire_moore_1000.csv",
    "dioula":    "dictionnaire_dioula_1000.csv",
    "fulfulde":  "dictionnaire_fulfulde_1000.csv",
}

_FR_SYNONYMS: Dict[str, List[str]] = {
    "maladie":      ["pathologie", "affection", "trouble"],
    "medicament":   ["remede", "traitement", "therapie"],
    "douleur":      ["mal", "souffrance", "gene"],
    "eau":          ["liquide"],
    "enfant":       ["bebe", "nourrisson", "jeune"],
    "medecin":      ["docteur", "praticien"],
    "champ":        ["terre", "parcelle", "culture"],
    "plante":       ["vegetal", "culture", "herbe"],
    "animal":       ["bete"],
    "urgent":       ["immediat", "critique", "grave"],
    "prevention":   ["protection", "securite"],
    "traitement":   ["soin", "remede", "therapie"],
    "diagnostic":   ["evaluation", "analyse", "bilan"],
    "symptome":     ["signe", "manifestation"],
    "infection":    ["contamination"],
    "consultation": ["visite", "examen"],
    "bonjour":      ["salut", "bonsoir", "salutation"],
    "merci":        ["remerciement", "gratitude"],
    "aide":         ["assistance", "soutien", "secours"],
}

# Le lookup dictionnaire/synonyme ne fait du sens que pour un mot ou une courte
# expression (cle du dictionnaire = 1 mot). Au-dela, "k in key.split()" matcherait
# un simple mot du dictionnaire present n'importe ou dans une phrase/paragraphe et
# retournerait CE SEUL MOT comme "traduction" de tout le texte (reponse tronquee/fausse).
# Pour les textes plus longs on saute directement a la traduction Gemini (phrase complete).
_SHORT_TEXT_MAX_WORDS = 4

_dictionaries: Dict[str, Dict[str, Any]] = {}

# Corrections vérifiées de transcriptions produites par le moteur vocal
# francophone d'Android. Ces formes ne sont pas du français saisi par
# l'utilisateur : ce sont des sons mooré que le STT force vers des mots
# français proches. Elles doivent être corrigées avant toute recherche RAG.
_STT_VERIFIED_ALIASES: Dict[str, Dict[str, Dict[str, Any]]] = {
    "moore": {
        "mon cours": {
            "reconstructed_local": "mam koo-da",
            "french_query": "je cultive",
            "confidence": 0.95,
        },
        "mon cour": {
            "reconstructed_local": "mam koo-da",
            "french_query": "je cultive",
            "confidence": 0.95,
        },
    },
}


def _load_dict_from_path(path: str) -> Dict[str, Any]:
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read().strip()
        if content.startswith("{") or content.startswith("["):
            return json.loads(content)
        result: Dict[str, Any] = {}
        reader = csv.reader(content.splitlines())
        for row in reader:
            if len(row) >= 2:
                key = row[0].strip().lower()
                val = row[1].strip()
                if key and val:
                    result[key] = {"translation": val}
        return result
    except Exception as e:
        print(f"[TRANSLATOR] Erreur chargement dict {path}: {e}")
        return {}


def _init_dictionaries() -> None:
    for lang, filename in _DICT_FILES.items():
        path = os.path.join(_TRANSLATE_DIR, filename)
        if not os.path.exists(path):
            path = os.path.join(_BACKEND_DIR, filename)
        loaded = _load_dict_from_path(path)
        _dictionaries[lang] = loaded
        count = len(loaded)
        status = f"{count} entrees" if count else "vide ou introuvable"
        print(f"[TRANSLATOR] Dict '{lang}' : {status} ({path})")


_init_dictionaries()


def dictionary_stats() -> Dict[str, int]:
    """Nombre d'entrees actives pour les trois langues locales supportees."""
    return {lang: len(_dictionaries.get(lang, {})) for lang in sorted(VALID_LANGS)}


def import_dictionary_file(content: bytes, filename: str, language: str, replace: bool = False) -> Dict[str, Any]:
    """Importe un lexique français -> langue locale depuis TXT, CSV ou XLSX.

    Les deux premières colonnes sont utilisées. Une ligne d'en-tête courante
    (francais, traduction) est ignorée. L'écriture est atomique et le nouveau
    dictionnaire est immédiatement disponible, sans redémarrer le serveur.
    """
    language = (language or "").strip().lower()
    if language not in VALID_LANGS:
        raise ValueError("Langue invalide. Utilisez moore, dioula ou fulfulde.")
    extension = os.path.splitext(filename or "")[1].lower()
    rows: List[List[Any]] = []
    if extension == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Le support Excel requiert openpyxl.") from exc
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = [[cell for cell in row[:2]] for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    elif extension in {".csv", ".txt"}:
        decoded = content.decode("utf-8-sig")
        sample = decoded[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            rows = list(csv.reader(StringIO(decoded), dialect))
        except csv.Error:
            rows = list(csv.reader(StringIO(decoded), delimiter=";" if ";" in sample else ","))
    else:
        raise ValueError("Format non supporte. Utilisez .txt, .csv ou .xlsx.")

    imported: Dict[str, Any] = {}
    header_words = {"fr", "francais", "français", "french", "mot", "terme"}
    for row in rows:
        if len(row) < 2:
            continue
        french = str(row[0] or "").strip().lower()
        translation = str(row[1] or "").strip()
        if not french or not translation or french in header_words:
            continue
        imported[french] = {"translation": translation}
    if not imported:
        raise ValueError("Aucune paire francais/traduction valide trouvee.")

    merged = {} if replace else dict(_dictionaries.get(language, {}))
    merged.update(imported)
    os.makedirs(_TRANSLATE_DIR, exist_ok=True)
    target = os.path.join(_TRANSLATE_DIR, _DICT_FILES[language])
    temp_target = target + ".tmp"
    with open(temp_target, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for french, entry in sorted(merged.items()):
            writer.writerow([french, entry.get("translation", "")])
    os.replace(temp_target, target)
    _dictionaries[language] = merged
    return {"language": language, "imported": len(imported), "total": len(merged), "replaced": replace}


def _dict_lookup(word: str, lang: str) -> Optional[str]:
    d = _dictionaries.get(lang, {})
    key = word.lower().strip()
    entry = d.get(key)
    if entry:
        return entry.get("translation", "") if isinstance(entry, dict) else str(entry)
    for k, v in d.items():
        if key in k.split() or k in key.split():
            t = v.get("translation", "") if isinstance(v, dict) else str(v)
            if t:
                return t
    return None


def _synonym_lookup(word: str, lang: str) -> Optional[str]:
    import unicodedata
    def normalize(s):
        return unicodedata.normalize("NFD", s.lower()).encode("ascii", "ignore").decode()
    key = normalize(word.strip())
    synonyms = _FR_SYNONYMS.get(key, [])
    for syn in synonyms:
        result = _dict_lookup(syn, lang)
        if result:
            return result
    for base_word, syns in _FR_SYNONYMS.items():
        if key in [normalize(s) for s in syns]:
            result = _dict_lookup(base_word, lang)
            if result:
                return result
    return None


def _build_dict_context(text: str, lang: str, max_entries: int = 25) -> Dict[str, Any]:
    d = _dictionaries.get(lang, {})
    words = re.sub(r"[.,!?;:()'\"\\/@]", " ", text.lower()).split()
    relevant: Dict[str, Any] = {}
    for w in words:
        if len(w) < 2:
            continue
        for k, v in d.items():
            if w in k or k in w:
                if len(relevant) < max_entries:
                    t = v.get("translation", "") if isinstance(v, dict) else str(v)
                    if t:
                        relevant[k] = t
    return relevant


_TRANSCRIPTION_RULES = (
    "NORMES DE TRANSCRIPTION (Commission Nationale des Langues du Burkina Faso) :\n"
    "- Alphabet officiel : utilisez les caracteres speciaux (epsilon, o ouvert) quand requis.\n"
    "- Nasalisation : inserez 'n' apres la voyelle nasalisee (ex: 'an', 'en', 'on').\n"
    "- Longueur vocalique : doublez la voyelle longue (ex: 'ee', 'oo', 'aa').\n"
    "- Tons : respectez les tons (haut/moyen/bas) dans le speech_text.\n"
    "- Emprunts : adaptez phonologiquement les termes modernes (ex: 'mobili' pour vehicule).\n"
    "- Ne jamais ecrire le mot francais brut pour un terme sans equivalent local.\n"
    "- Lecture vocale (speech_text) : utilisez impérativement un découpage phonétique francophone et syllabique en séparant les syllabes complexes par des tirets '-' ou des espaces légers (ex: 'ko-no-ko', 'ou-ain-dé' au lieu de 'ouaindé') pour forcer une élocution très lente, claire et décomposée par le moteur TTS."
)

_TRANSLATION_PRINCIPLES = (
    "PRINCIPES DE TRADUCTION (non negociables) :\n"
    "1. JAMAIS de traduction litterale mot-a-mot. Traduire le SENS, pas les mots.\n"
    "2. Si un mot exact n'existe pas, chercher un equivalent semantique ou une periphrase "
    "naturelle. L'utilisateur ne doit jamais voir de mots manquants ni de termes francais bruts.\n"
    "3. Utiliser les expressions idiomatiques locales du Burkina Faso.\n"
    "4. Adapter le registre au public rural burkinabe (langage oral, chaleureux, rassurant).\n"
    "5. Pour les termes techniques : utiliser l'equivalent culturel local si disponible, "
    "sinon adapter phonologiquement avec une periphrase explicative.\n"
    "6. Preferer les expressions culturellement equivalentes aux traductions approchantes."
)


def _call_gemini_translation(prompt: str, gemini_api_key: str) -> Optional[Dict[str, Any]]:
    import urllib.request
    url = (
        "https://generativelanguage.googleapis.com/v1beta/models/"
        f"gemini-2.5-flash:generateContent?key={gemini_api_key}"
    )
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"responseMimeType": "application/json", "temperature": 0.1},
    }
    try:
        req = urllib.request.Request(
            url, data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"}, method="POST",
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            raw = data["candidates"][0]["content"]["parts"][0]["text"]
            return json.loads(raw.strip())
    except Exception as e:
        print(f"[TRANSLATOR] Gemini echec: {e}")
        return None


def _call_openai_translation(prompt: str) -> Optional[Dict[str, Any]]:
    """Repli OpenAI quand Gemini est indisponible (ex: probleme de billing/quota
    Gemini, deja rencontre sur ce projet — cf. AI_PROVIDER dans v2_services.py,
    qui bascule tout le reste de l'app sur OpenAI par defaut pour cette raison).
    Sans ce repli, la traduction echouait silencieusement et gardait le texte
    francais original des que Gemini etait en panne.
    """
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": prompt + "\n\nReponds UNIQUEMENT avec un objet JSON valide, sans markdown ni explication.",
                }
            ],
            temperature=0.2,
            max_tokens=2000,
            response_format={"type": "json_object"},
        )
        content = response.choices[0].message.content if response.choices else ""
        return json.loads(content) if content else None
    except Exception as e:
        print(f"[TRANSLATOR] OpenAI (repli) echec: {e}")
        return None


def _call_groq_translation(prompt: str, json_mode: bool = True) -> Optional[Any]:
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        return None
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key, base_url="https://api.groq.com/openai/v1")
        kwargs = {
            "model": os.getenv("GROQ_MODEL", "qwen/qwen3.6-27b"),
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
            "max_tokens": 2000 if json_mode else 500,
            "extra_body": {"reasoning_effort": "none", "reasoning_format": "hidden"},
        }
        if json_mode:
            kwargs["response_format"] = {"type": "json_object"}
        try:
            response = client.chat.completions.create(**kwargs)
        except Exception as first_error:
            if not json_mode or "json" not in str(first_error).lower():
                raise
            kwargs.pop("response_format", None)
            kwargs["messages"][0]["content"] += (
                "\n\nRetourne uniquement un objet JSON valide, sans markdown ni commentaire."
            )
            response = client.chat.completions.create(**kwargs)
        content = response.choices[0].message.content if response.choices else ""
        if not json_mode:
            return content.strip() if content else None
        cleaned = re.sub(r"```(?:json)?\s*|```", "", content or "").strip()
        start, end = cleaned.find("{"), cleaned.rfind("}")
        if start >= 0 and end > start:
            cleaned = cleaned[start:end + 1]
        return json.loads(cleaned) if cleaned else None
    except Exception as e:
        print(f"[TRANSLATOR] Groq echec: {e}")
        return None


def _call_openai_plain_text(prompt: str) -> Optional[str]:
    """Repli OpenAI pour un prompt attendant une reponse texte brut (pas de JSON)."""
    if os.getenv("AI_PROVIDER", "openai").lower() == "groq":
        groq_result = _call_groq_translation(prompt, json_mode=False)
        return str(groq_result) if groq_result else None
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=500,
        )
        content = response.choices[0].message.content if response.choices else ""
        return content.strip() if content else None
    except Exception as e:
        print(f"[TRANSLATOR] OpenAI (repli texte) echec: {e}")
        return None


def _call_translation_llm(prompt: str, gemini_api_key: Optional[str]) -> Optional[Dict[str, Any]]:
    """Traduit via Gemini en priorite (meilleure qualite pour les langues locales
    du Burkina Faso), puis bascule automatiquement sur OpenAI si Gemini echoue."""
    if os.getenv("AI_PROVIDER", "openai").lower() == "groq":
        # Groq est le fournisseur exclusif quand AI_PROVIDER=groq. Une erreur
        # Groq remonte en réponse indisponible sans consommer Gemini/OpenAI.
        return _call_groq_translation(prompt)
    if gemini_api_key:
        result = _call_gemini_translation(prompt, gemini_api_key)
        if result:
            return result
        print("[TRANSLATOR] Gemini indisponible pour la traduction, repli OpenAI")
    return _call_openai_translation(prompt)


def translate_text(
    text: str,
    target_lang: str,
    gemini_api_key: Optional[str] = None,
    category: Optional[str] = None,
) -> Dict[str, Any]:
    text = (text or "").strip()
    if not text:
        return {"translation": "", "speech_text": "", "confidence": 1.0, "source": "empty"}
    if target_lang not in VALID_LANGS:
        return {"translation": text, "speech_text": text, "confidence": 0.0, "source": "invalid_lang"}

    lang_name = LANG_NAMES[target_lang]

    is_short_text = len(text.split()) <= _SHORT_TEXT_MAX_WORDS
    if is_short_text:
        d = _dict_lookup(text, target_lang)
        if d:
            return {"translation": d, "speech_text": d, "confidence": 1.0, "source": "local_dictionary"}

        s = _synonym_lookup(text, target_lang)
        if s:
            return {"translation": s, "speech_text": s, "confidence": 0.85, "source": "local_synonym"}

    dict_context = _build_dict_context(text, target_lang)
    dict_str = ""
    if dict_context:
        dict_str = (
            "\nExtrait du dictionnaire local (a respecter en priorite) :\n"
            + json.dumps(dict_context, ensure_ascii=False, indent=2) + "\n"
        )
    cat_ctx = f"Contexte : {category}. " if category else ""
    prompt = (
        f"Tu es un linguiste expert natif en {lang_name} (Burkina Faso).\n\n"
        f"{cat_ctx}Traduis le texte suivant du Francais vers le {lang_name} (code: {target_lang}).\n\n"
        f"{_TRANSCRIPTION_RULES}\n\n{_TRANSLATION_PRINCIPLES}\n{dict_str}\n"
        f"TEXTE A TRADUIRE :\n\"\"\"{text}\"\"\"\n\n"
        f"REPONSE JSON STRICTE :\n"
        "{{\n"
        f'  "translation": "<traduction naturelle en {lang_name}>",\n'
        '  "speech_text": "<version phonetique francisee pour TTS>",\n'
        '  "confidence": <0.0-1.0>\n'
        "}}"
    )
    ai = _call_translation_llm(prompt, gemini_api_key)
    if ai and ai.get("translation"):
        return {
            "translation": ai.get("translation", text),
            "speech_text": ai.get("speech_text", ai.get("translation", text)),
            "confidence": float(ai.get("confidence", 0.75)),
            "source": "ai_translation",
        }

    return {"translation": text, "speech_text": text, "confidence": 0.0, "source": "fallback_original"}


def translate_and_summarize_for_speech(
    text: str,
    target_lang: str,
    gemini_api_key: Optional[str] = None,
    category: Optional[str] = None,
    max_sentences: int = 3,
) -> Dict[str, Any]:
    """Résume un texte français long puis le traduit en langue locale pour la VOIX.

    Contrairement à translate_text() (traduction intégrale), cette fonction est
    destinée à la lecture audio : la réponse peut être longue (plusieurs
    paragraphes issus du LLM), mais une voix qui lit tout mot-à-mot en langue
    locale est lente et fatigante à l'oral. On demande donc à Gemini un résumé
    court ORAL (2-3 phrases), déjà en langue locale, plus sa transcription
    phonétique syllabique (speech_text) destinée à être lue par une voix TTS
    française (cf. _TRANSCRIPTION_RULES) pour un rendu sonore réaliste.
    """
    text = (text or "").strip()
    if not text:
        return {"summary": "", "speech_text": "", "confidence": 1.0, "source": "empty"}
    if target_lang not in VALID_LANGS:
        return {"summary": text, "speech_text": text, "confidence": 0.0, "source": "invalid_lang"}

    lang_name = LANG_NAMES[target_lang]
    dict_context = _build_dict_context(text, target_lang)
    dict_str = ""
    if dict_context:
        dict_str = (
            "\nExtrait du dictionnaire local (a respecter en priorite) :\n"
            + json.dumps(dict_context, ensure_ascii=False, indent=2) + "\n"
        )
    cat_ctx = f"Contexte : {category}. " if category else ""

    prompt = (
        f"Tu es un linguiste expert natif en {lang_name} (Burkina Faso).\n\n"
        f"{cat_ctx}Voici une reponse ecrite en francais, potentiellement longue. "
        f"Ta tache : 1) resumer son message essentiel en {max_sentences} phrases courtes MAXIMUM, "
        f"adaptees a une lecture ORALE (ton chaleureux, direct, actionnable) ; "
        f"2) traduire ce resume du Francais vers le {lang_name} (code: {target_lang}).\n\n"
        f"{_TRANSCRIPTION_RULES}\n\n{_TRANSLATION_PRINCIPLES}\n{dict_str}\n"
        f"TEXTE COMPLET A RESUMER PUIS TRADUIRE :\n\"\"\"{text}\"\"\"\n\n"
        f"REPONSE JSON STRICTE :\n"
        "{\n"
        f'  "summary": "<resume en {max_sentences} phrases max, traduit en {lang_name}>",\n'
        '  "speech_text": "<version phonetique francisee et syllabique du resume, pour lecture TTS>",\n'
        '  "confidence": <0.0-1.0>\n'
        "}"
    )
    ai = _call_translation_llm(prompt, gemini_api_key)
    if ai and ai.get("summary"):
        return {
            "summary": ai.get("summary", text),
            "speech_text": ai.get("speech_text", ai.get("summary", text)),
            "confidence": float(ai.get("confidence", 0.75)),
            "source": "ai_summary",
        }

    return {"summary": text, "speech_text": text, "confidence": 0.0, "source": "fallback_original"}


SONGRA_TEXT_FIELDS = [
    "what_i_see", "disease_detected", "analysis", "detailed_analysis",
    "treatment", "treatment_local", "treatment_chemical", "prevention",
    "urgency_message", "recommendations", "consultation_type",
    "symptoms_observed", "diagnosis", "advice",
    "response", "answer", "message", "content", "summary",
    "resolution", "question", "description",
    "description_terrain", "decoupage_terrain", "gestion_eau",
    "title", "text", "body",
]
_SONGRA_TEXT_FIELDS = SONGRA_TEXT_FIELDS  # alias retro-compatible (usage interne)


def translate_fields(
    fields: Dict[str, str],
    target_lang: str,
    gemini_api_key: Optional[str] = None,
    category: Optional[str] = None,
) -> Dict[str, Any]:
    if not fields or target_lang not in VALID_LANGS:
        return {}

    lang_name = LANG_NAMES[target_lang]
    results: Dict[str, Any] = {}
    needs_ai: Dict[str, str] = {}

    for field, text in fields.items():
        if not isinstance(text, str) or not text.strip():
            continue
        if len(text.split()) <= _SHORT_TEXT_MAX_WORDS:
            d = _dict_lookup(text, target_lang)
            if d:
                results[field] = {"translation": d, "speech_text": d, "confidence": 1.0, "source": "local_dictionary"}
                continue
            s = _synonym_lookup(text, target_lang)
            if s:
                results[field] = {"translation": s, "speech_text": s, "confidence": 0.85, "source": "local_synonym"}
                continue
        needs_ai[field] = text.strip()

    if needs_ai:
        dict_context = _build_dict_context(" ".join(needs_ai.values()), target_lang)
        dict_str = ""
        if dict_context:
            dict_str = (
                "\nExtrait du dictionnaire local (a respecter en priorite) :\n"
                + json.dumps(dict_context, ensure_ascii=False, indent=2) + "\n"
            )
        cat_ctx = f"Contexte : {category}. " if category else ""
        fields_json = json.dumps(needs_ai, ensure_ascii=False, indent=2)

        prompt = (
            f"Tu es un linguiste expert natif en {lang_name} (Burkina Faso).\n\n"
            f"{cat_ctx}Traduis chacun des champs suivants du Francais vers le {lang_name}.\n\n"
            f"{_TRANSCRIPTION_RULES}\n\n{_TRANSLATION_PRINCIPLES}\n{dict_str}\n"
            f"CHAMPS A TRADUIRE :\n{fields_json}\n\n"
            "REPONSE JSON STRICTE :\n"
            "{\n"
            '  "translations": {\n'
            f'    "<nom_champ>": {{"translation": "<texte en {lang_name}>", "speech_text": "<phonetique TTS>", "confidence": 0.0}}\n'
            "  }\n"
            "}"
        )
        ai_result = _call_translation_llm(prompt, gemini_api_key)
        if ai_result and "translations" in ai_result:
            for field, res in ai_result["translations"].items():
                if isinstance(res, dict) and res.get("translation"):
                    results[field] = {
                        "translation": res.get("translation", needs_ai.get(field, "")),
                        "speech_text": res.get("speech_text", res.get("translation", "")),
                        "confidence": float(res.get("confidence", 0.75)),
                        "source": "ai_batch",
                    }
                    needs_ai.pop(field, None)

    for field, text in needs_ai.items():
        if field not in results:
            results[field] = {"translation": text, "speech_text": text, "confidence": 0.0, "source": "fallback_original"}

    return results


def translate_fields_and_voice_summary(
    fields: Dict[str, str],
    target_lang: str,
    gemini_api_key: Optional[str] = None,
    category: Optional[str] = None,
    voice_source_field: Optional[str] = None,
    max_sentences: int = 3,
) -> Dict[str, Any]:
    """Traduit des champs ET produit un resume vocal en UNE SEULE requete LLM.

    translate_fields() (traduction complete pour l'affichage) et
    translate_and_summarize_for_speech() (resume court pour la voix) faisaient
    chacune leur propre appel Gemini/OpenAI, donc CHAQUE reponse vocale
    declenchait 2 appels — ce qui doublait le risque de tomber sur le quota
    gratuit Gemini (429 Too Many Requests) et faisait echouer la traduction en
    silence. On fusionne les deux besoins dans un seul prompt/une seule reponse
    JSON pour diviser par deux le nombre d'appels.

    Retourne {"translations": {champ: {...}}, "voice_summary": {...} | None}.
    """
    if not fields or target_lang not in VALID_LANGS:
        return {"translations": {}, "voice_summary": None}

    clean_fields = {
        field: text.strip()
        for field, text in fields.items()
        if isinstance(text, str) and text.strip()
    }
    if not clean_fields:
        return {"translations": {}, "voice_summary": None}

    voice_source_text = (
        clean_fields.get(voice_source_field) if voice_source_field else None
    ) or next(iter(clean_fields.values()))

    lang_name = LANG_NAMES[target_lang]
    dict_context = _build_dict_context(" ".join(clean_fields.values()), target_lang)
    dict_str = ""
    if dict_context:
        dict_str = (
            "\nExtrait du dictionnaire local (a respecter en priorite) :\n"
            + json.dumps(dict_context, ensure_ascii=False, indent=2) + "\n"
        )
    cat_ctx = f"Contexte : {category}. " if category else ""
    fields_json = json.dumps(clean_fields, ensure_ascii=False, indent=2)

    prompt = (
        f"Tu es un linguiste expert natif en {lang_name} (Burkina Faso).\n\n"
        f"{cat_ctx}Tu dois faire DEUX taches EN UNE SEULE reponse JSON :\n"
        f"1) Traduire integralement chacun des champs ci-dessous du Francais vers le {lang_name} "
        f"(pour l'affichage texte).\n"
        f"2) Resumer le message principal en {max_sentences} phrases ORALES courtes MAXIMUM, "
        f"traduites en {lang_name}, adaptees a une lecture a voix haute (ton chaleureux, direct, actionnable).\n\n"
        f"{_TRANSCRIPTION_RULES}\n\n{_TRANSLATION_PRINCIPLES}\n{dict_str}\n"
        f"CHAMPS A TRADUIRE INTEGRALEMENT :\n{fields_json}\n\n"
        f"MESSAGE PRINCIPAL A RESUMER POUR LA VOIX :\n\"\"\"{voice_source_text}\"\"\"\n\n"
        "REPONSE JSON STRICTE (un seul objet, rien d'autre) :\n"
        "{\n"
        '  "translations": {\n'
        f'    "<nom_champ>": {{"translation": "<texte en {lang_name}>", "speech_text": "<phonetique TTS>", "confidence": 0.0}}\n'
        "  },\n"
        '  "voice_summary": {\n'
        f'    "summary": "<resume {max_sentences} phrases max, traduit en {lang_name}>",\n'
        '    "speech_text": "<version phonetique francisee et syllabique du resume, pour lecture TTS>",\n'
        "    \"confidence\": 0.0\n"
        "  }\n"
        "}"
    )

    ai_result = _call_translation_llm(prompt, gemini_api_key)

    translations: Dict[str, Any] = {}
    if ai_result:
        for field, res in (ai_result.get("translations") or {}).items():
            if isinstance(res, dict) and res.get("translation"):
                translations[field] = {
                    "translation": res.get("translation", clean_fields.get(field, "")),
                    "speech_text": res.get("speech_text", res.get("translation", "")),
                    "confidence": float(res.get("confidence", 0.75)),
                    "source": "combined_ai",
                }
    for field, text in clean_fields.items():
        if field not in translations:
            translations[field] = {"translation": text, "speech_text": text, "confidence": 0.0, "source": "fallback_original"}

    voice_summary: Optional[Dict[str, Any]] = None
    raw_voice = ai_result.get("voice_summary") if ai_result else None
    if isinstance(raw_voice, dict) and raw_voice.get("summary"):
        voice_summary = {
            "summary": raw_voice.get("summary"),
            "speech_text": raw_voice.get("speech_text", raw_voice.get("summary")),
            "confidence": float(raw_voice.get("confidence", 0.75)),
            "source": "combined_ai",
        }
    else:
        voice_summary = {
            "summary": voice_source_text,
            "speech_text": voice_source_text,
            "confidence": 0.0,
            "source": "fallback_original",
        }

    return {"translations": translations, "voice_summary": voice_summary}


def translate_module_response(
    response: Dict[str, Any],
    target_lang: str,
    gemini_api_key: Optional[str] = None,
    category: Optional[str] = None,
) -> Dict[str, Any]:
    """Traduit automatiquement tous les champs textuels d'une reponse de module Songra.

    Applicable a TOUS les modules : scanner, assistant, teleconsultation,
    agriculture, veterinaire, SOS, communaute, entreprendre, etc.

    Ajoute un objet 'local_translation' a la reponse sans modifier l'original.
    Operation silencieuse — aucun crash possible, aucun mot manquant visible.
    """
    if not response or target_lang not in VALID_LANGS:
        return response

    to_translate: Dict[str, str] = {}
    for field in _SONGRA_TEXT_FIELDS:
        val = response.get(field)
        if isinstance(val, str) and val.strip():
            to_translate[field] = val.strip()
        elif isinstance(val, list):
            joined = " | ".join(str(v) for v in val if v)
            if joined:
                to_translate[field] = joined

    translated = translate_fields(to_translate, target_lang, gemini_api_key, category)

    result = dict(response)
    result["local_translation"] = {
        "target_lang": target_lang,
        "lang_name": LANG_NAMES[target_lang],
        "fields": translated,
    }
    return result


_SYLLABLE_VOWELS = "aeiouyàâäéèêëïîôöùûüɛɔ"


def _syllabify(word: str) -> List[str]:
    """Découpe approximative d'un mot en syllabes (groupes consonnes+voyelle).

    Sert de repli grossier mais utile : un moteur de reconnaissance vocale
    francophone transcrit une langue locale de façon phonétique et découpe
    souvent les mots au mauvais endroit. Chercher des correspondances au
    niveau syllabe (plutôt qu'au niveau mot entier issu du STT) retrouve des
    correspondances dans le dictionnaire local même quand le mot transcrit
    est faux.
    """
    word = word.lower().strip()
    if len(word) <= 2:
        return [word] if word else []

    syllables: List[str] = []
    current = ""
    for char in word:
        current += char
        if char in _SYLLABLE_VOWELS:
            syllables.append(current)
            current = ""
    if current:
        if syllables:
            syllables[-1] += current
        else:
            syllables.append(current)
    return syllables or [word]


def _syllable_fragments(text: str) -> List[str]:
    """Fragments syllabiques d'un texte : syllabes seules + paires de syllabes adjacentes."""
    words = re.sub(r"[.,!?;:()'\"\\/@]", " ", text.lower()).split()
    fragments: List[str] = []
    for word in words:
        syllables = _syllabify(word)
        fragments.extend(s for s in syllables if len(s) >= 2)
        for i in range(len(syllables) - 1):
            pair = syllables[i] + syllables[i + 1]
            if len(pair) >= 3:
                fragments.append(pair)
    return fragments


def translate_query_to_french(
    query: str,
    source_lang: str,
    gemini_api_key: Optional[str] = None,
    return_details: bool = False,
) -> Any:
    """Traduit une requête écrite ou vocale de l'utilisateur (en langue locale) vers le Français.

    1. Découpe en mots ET en syllabes, cherche des correspondances dans le dictionnaire
       local inversé à chaque niveau (mot entier, sous-chaîne, syllabe) — nécessaire car
       le moteur de reconnaissance vocale (francophone) transcrit la langue locale de façon
       phonétique et découpe souvent les mots au mauvais endroit.
    2. Passe TOUJOURS par Gemini (si la clé est disponible) en lui fournissant les
       correspondances locales trouvées ET un extrait plus large du dictionnaire local,
       pour qu'il reconstruise intelligemment le sens réel de la phrase plutôt que de
       deviner une traduction mot-à-mot en français.
    """
    query = (query or "").strip()
    if not query or source_lang not in VALID_LANGS:
        return query

    normalized_stt = unicodedata.normalize("NFD", query.lower())
    normalized_stt = "".join(
        char for char in normalized_stt
        if unicodedata.category(char) != "Mn"
    )
    normalized_stt = re.sub(r"[^a-z0-9ɛɔ]+", " ", normalized_stt).strip()
    verified_alias = _STT_VERIFIED_ALIASES.get(source_lang, {}).get(normalized_stt)
    if verified_alias:
        details = {
            "original_transcript": query,
            **verified_alias,
            "source": "verified_stt_alias",
        }
        return details if return_details else str(details["french_query"])

    # Inverser le dictionnaire local de cette langue (local_word -> french_word)
    d_local = _dictionaries.get(source_lang, {})
    inverted_dict: Dict[str, str] = {}
    for fr_word, entry in d_local.items():
        local_val = entry.get("translation", "") if isinstance(entry, dict) else str(entry)
        local_val = local_val.lower().strip()
        if local_val and fr_word:
            inverted_dict[local_val] = fr_word

    def _phonetic_key(value: str) -> str:
        value = unicodedata.normalize("NFD", value.lower())
        value = "".join(
            char for char in value if unicodedata.category(char) != "Mn"
        )
        # Le STT francais ajoute souvent espaces et tirets au milieu d'un mot local.
        return re.sub(r"[^a-z0-9ɛɔ]", "", value)

    query_lower = query.lower()
    query_phonetic = _phonetic_key(query)
    matched_french_words: List[str] = []

    def _add_match(fr_word: str) -> None:
        if fr_word and fr_word not in matched_french_words:
            matched_french_words.append(fr_word)

    # 1. Correspondance exacte de groupes de mots inversés dans la requête
    for local_phrase, fr_word in inverted_dict.items():
        if len(_phonetic_key(local_phrase)) >= 3 and re.search(
            rf"(?<!\w){re.escape(local_phrase)}(?!\w)", query_lower
        ):
            _add_match(fr_word)

    # 2. Correspondance mot à mot / sous-chaîne (tolère les frontières de mots
    #    mal placées par le STT)
    words_in_query = re.sub(r"[.,!?;:()'\"\\/@]", " ", query_lower).split()
    for word in words_in_query:
        if len(_phonetic_key(word)) < 4:
            continue
        for local_phrase, fr_word in inverted_dict.items():
            word_key = _phonetic_key(word)
            local_key = _phonetic_key(local_phrase)
            if word_key == local_key or (
                len(local_key) >= 4 and local_key in word_key
            ):
                _add_match(fr_word)

    # 2b. Correspondance phonétique approchée. Elle rapproche par exemple un
    # mot local découpé en plusieurs syllabes par Android de l'entrée continue
    # du dictionnaire, sans accepter les rapprochements faibles et ambigus.
    query_word_keys = [_phonetic_key(word) for word in words_in_query]
    for candidate in [query_phonetic, *query_word_keys]:
        if len(candidate) < 4:
            continue
        best_match: Optional[tuple] = None
        for local_phrase, fr_word in inverted_dict.items():
            local_key = _phonetic_key(local_phrase)
            if len(local_key) < 4:
                continue
            score = SequenceMatcher(None, candidate, local_key).ratio()
            if best_match is None or score > best_match[0]:
                best_match = (score, fr_word)
        if best_match and best_match[0] >= 0.86:
            _add_match(best_match[1])

    # 3. Correspondance syllabique : retrouve des correspondances même quand
    #    le "mot" transcrit par le STT ne matche jamais tel quel. Seuil de
    #    longueur minimal indispensable : une syllabe de 2-3 lettres apparaît
    #    comme sous-chaîne dans une grande partie du dictionnaire et noierait
    #    le signal utile sous des centaines de faux positifs.
    for fragment in _syllable_fragments(query):
        if len(fragment) < 4:
            continue
        fragment_key = _phonetic_key(fragment)
        best_fragment_match: Optional[tuple] = None
        for local_phrase, fr_word in inverted_dict.items():
            local_key = _phonetic_key(local_phrase)
            if len(local_key) < 4:
                continue
            score = SequenceMatcher(None, fragment_key, local_key).ratio()
            if best_fragment_match is None or score > best_fragment_match[0]:
                best_fragment_match = (score, fr_word)
        if best_fragment_match and best_fragment_match[0] >= 0.88:
            _add_match(best_fragment_match[1])

    # 4. Reconstruction intelligente : TOUJOURS passer par Gemini (jamais une
    #    simple substitution mot-à-mot), en combinant les correspondances
    #    locales détectées avec un extrait plus large du dictionnaire local.
    lang_name = LANG_NAMES[source_lang]
    dict_context_str = ""
    if matched_french_words:
        dict_context_str = (
            "\nMots-clés locaux détectés (mot entier + syllabes) et leurs correspondances "
            "en français :\n"
            + ", ".join(matched_french_words[:20])
            + "\n"
        )

    broader_context = _build_dict_context(query, source_lang, max_entries=15)
    broader_context_str = ""
    if broader_context:
        broader_context_str = (
            "\nAutre extrait du dictionnaire local pouvant être pertinent (français -> "
            f"{lang_name}) :\n"
            + json.dumps(broader_context, ensure_ascii=False)
            + "\n"
        )

    # Suggestions phonétiques, même imparfaites, pour permettre au linguiste IA
    # de reconstruire « mam kooda » lorsque le STT français écrit « mon cours ».
    candidate_scores: List[tuple] = []
    for local_phrase, fr_word in inverted_dict.items():
        local_key = _phonetic_key(local_phrase)
        if len(local_key) < 3:
            continue
        score = max(
            [SequenceMatcher(None, candidate, local_key).ratio()
             for candidate in [query_phonetic, *query_word_keys] if candidate]
            or [0.0]
        )
        if score >= 0.42:
            candidate_scores.append((score, local_phrase, fr_word))
    candidate_scores.sort(key=lambda item: item[0], reverse=True)
    phonetic_candidates = [
        {"local": local, "francais": french, "score": round(score, 2)}
        for score, local, french in candidate_scores[:12]
    ]
    candidates_str = (
        "\nCandidats du dictionnaire ressemblant aux sons entendus :\n"
        + json.dumps(phonetic_candidates, ensure_ascii=False)
        + "\n"
        if phonetic_candidates else ""
    )

    prompt = (
        f"Tu es un traducteur et linguiste expert de la langue {lang_name} (Burkina Faso) vers le Français.\n"
        f"L'utilisateur rural a posé une question en {lang_name}, transcrite PHONÉTIQUEMENT par un moteur "
        f"de reconnaissance vocale francophone (donc potentiellement déformée ou mal découpée en mots).\n"
        f"Reconstitue le sens réel de la question à partir des sons/syllabes et des correspondances "
        f"locales ci-dessous (ne traduis jamais mot-à-mot une transcription déformée), puis traduis-la "
        f"en Français correct et fluide pour qu'elle puisse servir à interroger un système RAG sur les "
        f"maladies agricoles/d'élevage/urgences.\n\n"
        f"Requête utilisateur (transcription phonétique {lang_name}) : \"{query}\"\n"
        f"{dict_context_str}{broader_context_str}{candidates_str}\n"
        + (
            "Retourne uniquement ce JSON : "
            '{"reconstructed_local":"phrase locale reconstruite avec syllabes séparées par des espaces ou tirets",'
            '"french_query":"sens complet en français","confidence":0.0}. '
            "La transcription locale reconstruite ne doit jamais être la transcription française brute."
            if return_details else
            "Traduis directement en Français (ne donne que la phrase traduite, rien d'autre, pas d'explication)."
        )
    )

    # Respecter AI_PROVIDER (Groq en production), puis utiliser Gemini comme repli.
    translated_text = _call_openai_plain_text(prompt)
    if not translated_text and gemini_api_key:
        try:
            import urllib.request
            url = (
                "https://generativelanguage.googleapis.com/v1beta/models/"
                f"gemini-2.5-flash:generateContent?key={gemini_api_key}"
            )
            payload = {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.1},
            }
            req = urllib.request.Request(
                url, data=json.dumps(payload).encode("utf-8"),
                headers={"Content-Type": "application/json"}, method="POST",
            )
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                translated_text = data["candidates"][0]["content"]["parts"][0]["text"].strip()
        except Exception as e:
            print(f"[TRANSLATOR] Échec traduction requête Gemini: {e}")

    if translated_text:
        if return_details:
            try:
                cleaned = re.sub(r"```(?:json)?\s*|```", "", translated_text).strip()
                start, end = cleaned.find("{"), cleaned.rfind("}")
                parsed = json.loads(cleaned[start:end + 1])
                french_query = str(parsed.get("french_query") or "").strip()
                reconstructed = str(parsed.get("reconstructed_local") or "").strip()
                if french_query:
                    return {
                        "original_transcript": query,
                        "reconstructed_local": reconstructed or query,
                        "french_query": french_query,
                        "confidence": float(parsed.get("confidence") or 0.0),
                    }
            except Exception as exc:
                print(f"[TRANSLATOR] JSON reconstruction vocale invalide: {exc}")
        translated_text = re.sub(r'^["\']|["\']$', '', translated_text).strip()
        if translated_text and not return_details:
            return translated_text

    # Fallback local
    if matched_french_words:
        fallback_french = " ".join(matched_french_words)
        if return_details:
            return {
                "original_transcript": query,
                "reconstructed_local": query,
                "french_query": fallback_french,
                "confidence": 0.35,
            }
        return fallback_french
        
    if return_details:
        return {
            "original_transcript": query,
            "reconstructed_local": query,
            "french_query": query,
            "confidence": 0.0,
        }
    return query

